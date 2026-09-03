"""Raportul «facturi transmise in e-Factura» — trei seturi, web + Excel + PDF.

RO: datele vin DOAR din pachetul Oracle `EFA_REPORT` (sql/04_efa_report.sql):
header (filtrul si totalurile lui), master (o linie per e-factura), detail
(marfurile, legate de master prin EFA_ID). Aici doar le citim prin
functiile pipelined si le impachetam: JSON pentru pagina, .xlsx (openpyxl,
trei foi) si .pdf (reportlab; font DejaVu / Arial Unicode pentru
diacritice si chirilice, altfel Helvetica).
EN: reads the EFA_REPORT package's three sets; renders JSON, XLSX, PDF.
"""
from __future__ import annotations

import io
import os
from datetime import date, datetime
from typing import Any, Dict, List, Optional

FONT_CANDIDATES = (
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/Library/Fonts/Arial Unicode.ttf",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
)
STATUSES = ("", "SENT", "ACCEPTED", "SIGNED", "ERROR", "REJECTED", "NEW")

MASTER_COLS = [("nrmanual", "Nr. doc"), ("doc_date", "Data doc"),
               ("client_name", "Client"), ("client_idno", "IDNO"),
               ("status", "Statut"), ("sfs_seria", "Seria SFS"),
               ("sfs_number", "Nr. SFS"), ("sent_at", "Trimisa la"),
               ("total", "Total, lei"), ("rows_cnt", "Pozitii"),
               ("qty_sum", "Cantitate"), ("request_id", "RequestId"),
               ("err_msg", "Eroare")]
DETAIL_COLS = [("nrmanual", "Nr. doc"), ("row_no", "#"), ("code", "Cod"),
               ("name", "Denumirea"), ("um", "UM"), ("qty", "Cant."),
               ("price", "Pret"), ("suma", "Suma")]
HEADER_LABELS = [("filter_from", "Perioada de la"), ("filter_to", "pina la"),
                 ("filter_status", "Statut"), ("filter_client", "Client (cod)"),
                 ("generated_at", "Generat la"), ("endpoint", "Adresa SFS"),
                 ("docs_cnt", "E-facturi"), ("sent_cnt", "Trimise"),
                 ("accepted_cnt", "Acceptate"), ("error_cnt", "Cu eroare"),
                 ("total_sum", "Total, lei")]


def _norm_date(v: Optional[str]) -> Optional[str]:
    """RO: '2026-09-03' sau '03.09.2026' -> 'YYYY-MM-DD'; gol -> None."""
    v = (v or "").strip()
    if not v:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError("data invalida: %s" % v)


def parse_filters(args: Dict[str, Any]) -> Dict[str, Any]:
    today = date.today()
    f_from = _norm_date(args.get("from")) or today.replace(day=1).strftime("%Y-%m-%d")
    f_to = _norm_date(args.get("to")) or today.strftime("%Y-%m-%d")
    status = (args.get("status") or "").strip().upper()
    if status not in STATUSES:
        raise ValueError("statut necunoscut: %s" % status)
    client = (args.get("client") or "").strip()
    return {"from": f_from, "to": f_to, "status": status or None,
            "client": int(client) if client.isdigit() else None}


def fetch(filters: Dict[str, Any]) -> Dict[str, Any]:
    """RO: cele trei seturi din EFA_REPORT (functiile pipelined)."""
    from models.biro26_db import Biro26DB
    from models.biro26_oracle_store import _rows
    db = Biro26DB()
    binds = {"f": filters["from"], "t": filters["to"],
             "s": filters["status"], "c": filters["client"]}
    args = "TO_DATE(:f,'YYYY-MM-DD'), TO_DATE(:t,'YYYY-MM-DD'), :s, :c"
    fmt = "TO_CHAR(%s,'DD.MM.YYYY') %s"
    sets = {}
    for name, cols in (
        ("header", "FILTER_STATUS, FILTER_CLIENT, ENDPOINT, DOCS_CNT, SENT_CNT, "
                   "ACCEPTED_CNT, ERROR_CNT, TOTAL_SUM, "
                   + fmt % ("FILTER_FROM", "FILTER_FROM") + ", "
                   + fmt % ("FILTER_TO", "FILTER_TO") + ", "
                   "TO_CHAR(GENERATED_AT,'DD.MM.YYYY HH24:MI') GENERATED_AT"),
        ("master", "EFA_ID, DOC_COD, NRMANUAL, CLIENT_COD, CLIENT_NAME, CLIENT_IDNO, "
                   "STATUS, SFS_SERIA, SFS_NUMBER, REQUEST_ID, ERR_MSG, TOTAL, "
                   "ROWS_CNT, QTY_SUM, " + fmt % ("DOC_DATE", "DOC_DATE") + ", "
                   "TO_CHAR(SENT_AT,'DD.MM.YYYY HH24:MI') SENT_AT"),
        ("detail", "EFA_ID, DOC_COD, ROW_NO, GOODS_COD, CODE, NAME, UM, QTY, PRICE, SUMA"),
    ):
        r = db.execute_query(
            f"SELECT {cols} FROM TABLE(EFA_REPORT.{name}({args}))", binds)
        if not r.get("success"):
            raise RuntimeError("EFA_REPORT.%s: %s" % (name, r.get("message") or r.get("error")))
        sets[name] = _rows(r)
    # RO: legatura detail -> master (nrmanual pe fiecare pozitie, pentru export)
    # RO: documente fara numar manual (nici in ERP, nici in EFA_DOC) -> «cod N»
    for m in sets["master"]:
        if not m.get("nrmanual"):
            m["nrmanual"] = "cod %s" % m.get("doc_cod")
    nr = {m["efa_id"]: m.get("nrmanual") for m in sets["master"]}
    for d in sets["detail"]:
        d["nrmanual"] = nr.get(d["efa_id"])
    return {"success": True, "filters": filters,
            "header": sets["header"][0] if sets["header"] else {},
            "master": sets["master"], "detail": sets["detail"]}


# ── Excel ───────────────────────────────────────────────────────────────
def to_xlsx(data: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Header"
    ws.append(["Raport: facturi transmise in e-Factura"])
    ws["A1"].font = Font(bold=True, size=13)
    h = data.get("header") or {}
    for key, label in HEADER_LABELS:
        ws.append([label, h.get(key)])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50
    for title, cols, rows in (("Master", MASTER_COLS, data.get("master") or []),
                              ("Detail", DETAIL_COLS, data.get("detail") or [])):
        w = wb.create_sheet(title)
        w.append(["EFA_ID", "DOC_COD"] + [lbl for _, lbl in cols])
        for c in w[1]:
            c.font = Font(bold=True)
        for r in rows:
            w.append([r.get("efa_id"), r.get("doc_cod")] + [r.get(k) for k, _ in cols])
        w.freeze_panes = "A2"
        for i, (k, _) in enumerate(cols, start=3):
            w.column_dimensions[w.cell(row=1, column=i).column_letter].width = (
                40 if k in ("client_name", "name", "err_msg") else 14)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ─────────────────────────────────────────────────────────────────
def _font() -> str:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    for path in FONT_CANDIDATES:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("EfaSans", path))
                return "EfaSans"
            except Exception:                                # noqa: BLE001
                continue
    return "Helvetica"


def _s(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return ("%.3f" % v).rstrip("0").rstrip(".") if v != int(v) else str(int(v))
    return str(v)


def to_pdf(data: Dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)
    font = _font()
    st = ParagraphStyle("b", fontName=font, fontSize=9, leading=11)
    st_h = ParagraphStyle("h", fontName=font, fontSize=13, leading=16, spaceAfter=4)
    st_s = ParagraphStyle("s", fontName=font, fontSize=8, leading=10)
    grid = TableStyle([("FONTNAME", (0, 0), (-1, -1), font),
                       ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                       ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                       ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eef7")),
                       ("VALIGN", (0, 0), (-1, -1), "TOP")])
    h = data.get("header") or {}
    story = [Paragraph("Facturi transmise in e-Factura (SFS)", st_h)]
    story.append(Paragraph(" · ".join("%s: %s" % (lbl, _s(h.get(k)))
                                      for k, lbl in HEADER_LABELS if h.get(k) not in (None, "")), st_s))
    story.append(Spacer(1, 4 * mm))
    m_cols = [c for c in MASTER_COLS if c[0] not in ("request_id", "err_msg", "qty_sum")]
    rows = [[lbl for _, lbl in m_cols]]
    for m in data.get("master") or []:
        rows.append([Paragraph(_s(m.get(k)), st_s) if k == "client_name" else _s(m.get(k))
                     for k, _ in m_cols])
    t = Table(rows, repeatRows=1, colWidths=[22 * mm, 20 * mm, 70 * mm, 26 * mm, 20 * mm,
                                             18 * mm, 24 * mm, 28 * mm, 22 * mm, 16 * mm])
    t.setStyle(grid)
    story += [Paragraph("Master — e-facturi: %d" % len(data.get("master") or []), st), t,
              Spacer(1, 5 * mm)]
    by_doc: Dict[Any, List[Dict[str, Any]]] = {}
    for d in data.get("detail") or []:
        by_doc.setdefault(d.get("efa_id"), []).append(d)
    d_cols = [c for c in DETAIL_COLS if c[0] != "nrmanual"]
    for m in data.get("master") or []:
        lines = by_doc.get(m.get("efa_id")) or []
        if not lines:
            continue
        story.append(Paragraph("Detail — %s din %s, %s (%s pozitii, total %s lei)" % (
            _s(m.get("nrmanual")), _s(m.get("doc_date")), _s(m.get("client_name")),
            len(lines), _s(m.get("total"))), st))
        rows = [[lbl for _, lbl in d_cols]]
        for d in lines:
            rows.append([Paragraph(_s(d.get(k)), st_s) if k in ("name", "code")
                         else _s(d.get(k)) for k, _ in d_cols])
        t = Table(rows, repeatRows=1, colWidths=[10 * mm, 28 * mm, 130 * mm, 16 * mm,
                                                 22 * mm, 24 * mm, 26 * mm])
        t.setStyle(grid)
        story += [t, Spacer(1, 3 * mm)]
    buf = io.BytesIO()
    SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm,
                      topMargin=10 * mm, bottomMargin=10 * mm,
                      title="e-Factura — facturi transmise").build(story)
    return buf.getvalue()


def file_name(filters: Dict[str, Any], ext: str) -> str:
    return "efactura_%s_%s.%s" % (filters["from"], filters["to"], ext)

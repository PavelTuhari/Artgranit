#!/usr/bin/env python3
"""Autopark — выгрузка пакета отчётности (ТЗ §14) в XLSX и PDF.

    venv/bin/python modules/autopark/scripts/autopark_reports.py \
        --report all --date-from 2026-07-01 --date-to 2026-07-31 \
        --xlsx --pdf --out docs/Autopark/examples

XLSX — openpyxl: лист на отчёт, шапка (название + период), жирные
заголовки, автоширина колонок, числовой формат `# ##0.00`, строка
итогов, замороженная шапка.

PDF — без внешних сервисов: отчёт собирается в самодостаточный HTML
(встроенные стили, кириллица, деловая палитра #132038/#1d4ed8) и
конвертируется установленным LibreOffice:

    soffice --headless --convert-to pdf --outdir <dir> <file>.html

Путь к soffice ищется в PATH, затем в /Applications/LibreOffice.app.
reportlab как запасной путь НЕ используется: LibreOffice на этой машине
установлен и проверен вживую (см. docs/Autopark/REPORTS.md), а второй
рендерер означал бы два разных внешних вида одного отчёта.
"""
from __future__ import annotations

import argparse
import html
import os
import shutil
import subprocess
import sys
from datetime import date, datetime

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.abspath(__file__)))))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from modules.autopark.reports import REPORTS, Report  # noqa: E402

NUMBER_FORMAT = "# ##0.00"
SOFFICE_CANDIDATES = (
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
)


# ── XLSX ────────────────────────────────────────────────────────────────

def write_xlsx(reports: "list[Report]", path: str) -> str:
    """Один файл, лист на отчёт (контракт reports.py -> openpyxl)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="132038")
    title_font = Font(bold=True, size=13, color="132038")
    totals_font = Font(bold=True)

    for report in reports:
        # Имя листа Excel: максимум 31 символ, без спецсимволов.
        sheet_name = report["title"][:31].replace("/", "-")
        ws = wb.create_sheet(sheet_name)
        ncols = len(report["columns"])

        ws.cell(row=1, column=1, value=report["title"]).font = title_font
        ws.cell(row=2, column=1, value=f"Период: {report['period']}")
        for col, name in enumerate(report["columns"], start=1):
            c = ws.cell(row=4, column=col, value=name)
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(wrap_text=True, vertical="center")

        row_idx = 5
        for row in report["rows"]:
            for col, value in enumerate(row, start=1):
                c = ws.cell(row=row_idx, column=col, value=value)
                if isinstance(value, float):
                    c.number_format = NUMBER_FORMAT
            row_idx += 1
        if report.get("totals"):
            for col, value in enumerate(report["totals"], start=1):
                c = ws.cell(row=row_idx, column=col, value=value)
                c.font = totals_font
                if isinstance(value, float):
                    c.number_format = NUMBER_FORMAT
            row_idx += 1
        for note in report.get("notes") or []:
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="* " + note)

        # Автоширина: по самой длинной строке колонки (в пределах разумного).
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            width = len(str(report["columns"][col - 1]))
            for row in report["rows"] + ([report["totals"]]
                                         if report.get("totals") else []):
                value = row[col - 1]
                text = (f"{value:,.2f}" if isinstance(value, float)
                        else str(value if value is not None else ""))
                width = max(width, len(text))
            ws.column_dimensions[letter].width = min(width + 3, 45)
        ws.freeze_panes = "A5"

    wb.save(path)
    return path


# ── PDF (HTML -> LibreOffice) ───────────────────────────────────────────

def find_soffice() -> "str | None":
    found = shutil.which("soffice")
    if found:
        return found
    for candidate in SOFFICE_CANDIDATES:
        if os.path.exists(candidate):
            return candidate
    return None


def _fmt_cell(value) -> str:
    if value is None:
        return "—"
    if isinstance(value, float):
        return f"{value:,.2f}".replace(",", " ")
    return html.escape(str(value))


def report_html(report: Report) -> str:
    """Самодостаточная HTML-страница отчёта (кириллица, деловая палитра)."""
    # Legacy-атрибуты (bgcolor/border/cellpadding) вместе с CSS — не
    # архаика ради архаики: HTML-импорт LibreOffice Writer игнорирует
    # часть CSS-свойств ячеек (фоны, границы), а атрибуты честно
    # переносит в PDF (проверено глазами на station-отчёте за июль).
    head_cells = "".join(
        f"<th bgcolor=\"#132038\">{html.escape(c)}</th>"
        for c in report["columns"])

    def _tr(row, shade: str = "", bold: bool = False) -> str:
        cells = []
        for v in row:
            align = " align=\"right\"" if isinstance(v, (int, float)) else ""
            text = _fmt_cell(v)
            if bold:
                text = f"<b>{text}</b>"
            cells.append(f"<td{align}{shade}>{text}</td>")
        return "<tr>" + "".join(cells) + "</tr>"

    body = []
    for i, row in enumerate(report["rows"]):
        body.append(_tr(row, " bgcolor=\"#f2f5fa\"" if i % 2 else ""))
    if report.get("totals"):
        body.append(_tr(report["totals"], " bgcolor=\"#dbe4f5\"", bold=True))
    notes = "".join(f"<p class='note'>* {html.escape(n)}</p>"
                    for n in report.get("notes") or [])
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    return f"""<!DOCTYPE html>
<html lang="ru"><head><meta charset="utf-8">
<title>{html.escape(report['title'])}</title>
<style>
  @page {{ size: A4 landscape; margin: 1.2cm; }}
  body {{ font-family: 'Helvetica Neue', Arial, sans-serif; color: #132038;
         font-size: 9.5pt; margin: 0; }}
  h1 {{ font-size: 15pt; margin: 0 0 2pt 0; color: #132038; }}
  .meta {{ color: #55617a; margin: 0 0 10pt 0; font-size: 9pt; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th {{ background: #132038; color: #ffffff; padding: 5pt 6pt;
       text-align: left; font-size: 8.5pt; border: 0.5pt solid #132038; }}
  td {{ padding: 4pt 6pt; border: 0.5pt solid #c9d1e0; }}
  td.num {{ text-align: right; white-space: nowrap; }}
  tr:nth-child(even) td {{ background: #f2f5fa; }}
  tr.totals td {{ font-weight: bold; background: #dbe4f5;
                 border-top: 1.2pt solid #1d4ed8; }}
  .note {{ color: #55617a; font-size: 8pt; margin: 6pt 0 0 0; }}
  .footer {{ color: #8a93a8; font-size: 7.5pt; margin-top: 10pt; }}
</style></head><body>
<h1>{html.escape(report['title'])}</h1>
<p class="meta">Период: {html.escape(report['period'])} · Autopark (Bemol)</p>
<table border="1" cellspacing="0" cellpadding="4" width="100%"
       bordercolor="#c9d1e0">
<thead><tr>{head_cells}</tr></thead>
<tbody>{''.join(body)}</tbody></table>
{notes}
<p class="footer">Сформировано {generated} · modules/autopark · Oracle FLT_*</p>
</body></html>
"""


def write_pdf(report: Report, path: str) -> str:
    """HTML -> PDF через headless LibreOffice (см. модульный docstring)."""
    soffice = find_soffice()
    if not soffice:
        raise RuntimeError(
            "LibreOffice (soffice) не найден ни в PATH, ни в /Applications "
            "— PDF собрать нечем. Установите LibreOffice или снимите --pdf.")
    out_dir = os.path.dirname(os.path.abspath(path)) or "."
    base = os.path.splitext(os.path.basename(path))[0]
    html_path = os.path.join(out_dir, base + ".html")
    with open(html_path, "w", encoding="utf-8") as fh:
        fh.write(report_html(report))
    subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", out_dir,
         html_path],
        check=True, capture_output=True, timeout=180)
    produced = os.path.join(out_dir, base + ".pdf")
    if not os.path.exists(produced):
        raise RuntimeError(f"LibreOffice не создал {produced}")
    os.remove(html_path)  # промежуточный HTML — не артефакт
    if produced != os.path.abspath(path):
        os.replace(produced, path)
    return path


# ── CLI ─────────────────────────────────────────────────────────────────

def _parse_date(raw: str) -> date:
    return datetime.strptime(raw, "%Y-%m-%d").date()


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Выгрузка отчётности Autopark (ТЗ §14) в XLSX/PDF")
    parser.add_argument("--report", default="all",
                        choices=["all"] + sorted(REPORTS))
    parser.add_argument("--date-from", required=True, type=_parse_date)
    parser.add_argument("--date-to", required=True, type=_parse_date)
    parser.add_argument("--xlsx", action="store_true")
    parser.add_argument("--pdf", action="store_true")
    parser.add_argument("--out", default=".")
    args = parser.parse_args(argv)

    if not args.xlsx and not args.pdf:
        parser.error("укажите хотя бы один формат: --xlsx и/или --pdf")
    os.makedirs(args.out, exist_ok=True)

    names = sorted(REPORTS) if args.report == "all" else [args.report]
    suffix = f"{args.date_from.isoformat()}_{args.date_to.isoformat()}"
    for name in names:
        report = REPORTS[name](args.date_from, args.date_to)
        base = os.path.join(args.out, f"autopark_{name}_{suffix}")
        if args.xlsx:
            write_xlsx([report], base + ".xlsx")
            print(f"XLSX: {base}.xlsx ({len(report['rows'])} строк)")
        if args.pdf:
            write_pdf(report, base + ".pdf")
            print(f"PDF:  {base}.pdf")
    return 0


if __name__ == "__main__":
    sys.exit(main())

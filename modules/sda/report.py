"""SDA — рапорты модуля: PDF через боковой сервис и Excel в процессе.

Модуль ничего не оставляет в общем коде (docs/CORE_MODULES.md): шаблоны
Handlebars лежат здесь же, в `modules/sda/reports/`, и уходят в jsReport
ВНУТРИ тела запроса вместе с данными. Общий каталог `reports/templates/`
принадлежит боковому Node-сервису и модулю не принадлежит — класть туда
свои шаблоны значило бы нарушить изоляцию.

Из общей инфраструктуры берём ровно две вещи: `Config.JSREPORT_URL`
(адрес сервиса рендера) и — косвенно, через `SDAStore` — `models.database`.
Своего SQL здесь нет: данные читаются только методами хранилища модуля.

Наружу отдаётся контракт портала: {"success": bool, ...}. Недоступный
сервис рендера — это `{"success": False, "message": ...}`, а не исключение
и не битый файл: клиент и регулятор получают либо настоящий документ,
либо внятный отказ.
"""
from __future__ import annotations

import io
import os
from datetime import date
from typing import Any, Dict, List, Optional

import requests

from config import Config
from modules.sda.store import SDAStore

_TPL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")

# Три вида рапорта — и ровно три шаблона. Неизвестный вид отклоняется до
# любого обращения к базе и к сервису рендера.
REPORT_KINDS = {
    "conformitate": "sda_conformitate.hbs",
    "registru": "sda_registru.hbs",
    "dosar": "sda_dosar.hbs",
}

REPORT_TITLES = {
    "conformitate": "Harta de conformitate",
    "registru": "Registrul ambalajelor SD",
    "dosar": "Dosarul de înregistrare (pct. 78)",
}

# Порядок и подписи режимов. Группировка по режиму — не украшение:
# документ читает регулятор, и три режима означают три разных обязанности.
REGIM_ORDER = ["A_PUNCT_PROPRIU", "B_EXCEPTIE_APL", "C_HORECA"]
REGIM_LABEL = {
    "A_PUNCT_PROPRIU": "A — punct de preluare propriu obligatoriu",
    "B_EXCEPTIE_APL": "B — excepție (sub pragul de suprafață)",
    "C_HORECA": "C — HoReCa: predare directă către Administrator",
    "": "FĂRĂ REGIM — de completat înainte de depunerea dosarului",
}

TIP_LABEL = {
    "MAGAZIN": "Magazin",
    "TARABA": "Tarabă",
    "CHIOSC": "Chioșc",
    "BENZINARIE": "Benzinărie",
    "ALIMENTATIE_PUBLICA": "Alimentație publică",
}

MATERIAL_LABEL = {"PLASTIC": "Plastic", "STICLA": "Sticlă", "METAL": "Metal"}

XLSX_MIME = ("application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.sheet")


def _fail(message: str) -> Dict[str, Any]:
    return {"success": False, "data": None, "message": message}


def _read(fname: str) -> str:
    with open(os.path.join(_TPL_DIR, fname), encoding="utf-8") as fh:
        return fh.read()


def _da_nu(value: Any) -> str:
    return "Da" if str(value or "N").upper() == "D" else "Nu"


def _txt(value: Any) -> str:
    return "" if value in (None, "") else str(value)


def _num(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _partic_id(params: Dict[str, Any]) -> Optional[int]:
    raw = (params or {}).get("partic_id")
    if raw in (None, ""):
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        raise ValueError("Parametrul partic_id trebuie să fie un număr întreg")


class SDAReport:
    """Данные и два рендерера для трёх рапортов модуля."""

    # ── сбор данных ─────────────────────────────────────────────────

    @staticmethod
    def conformitate_data(params: Dict[str, Any] = None) -> Dict[str, Any]:
        """Карта соответствия: все точки, сгруппированные по режиму.

        Точки без режима идут ПЕРВОЙ группой и попадают в шапку отдельным
        числом: это остаток работы перед подачей досье, и прятать его в
        конце таблицы означало бы прятать сам смысл документа.
        """
        try:
            pid = _partic_id(params or {})
        except ValueError as exc:
            return _fail(str(exc))
        listed = SDAStore.list_units(pid)
        if not listed.get("success"):
            return listed
        units = listed.get("data") or []

        def _row(u: Dict[str, Any]) -> Dict[str, Any]:
            supr = _num(u.get("suprafata_mp"))
            return {
                "cod_erp": _txt(u.get("cod_erp")),
                "denumire": _txt(u.get("denumire")),
                "adresa": _txt(u.get("adresa")),
                "localitate": _txt(u.get("localitate")),
                "suprafata_mp": supr,
                "suprafata_text": ("—" if supr is None else f"{supr:g}"),
                "tip_amplasament": TIP_LABEL.get(
                    (u.get("tip_amplasament") or "").upper(),
                    _txt(u.get("tip_amplasament"))),
                "regim": _txt(u.get("regim")),
                "regim_label": REGIM_LABEL.get(_txt(u.get("regim")),
                                               _txt(u.get("regim"))),
                "regim_motiv": _txt(u.get("regim_motiv")),
            }

        rows = [_row(u) for u in units]
        # Сортировка внутри группы — по названию: список сверяют глазами
        # с кассовой системой, где точки тоже идут по названию.
        groups = []
        for regim in [""] + REGIM_ORDER:
            in_group = sorted((r for r in rows if r["regim"] == regim),
                              key=lambda r: r["denumire"])
            if in_group:
                groups.append({"regim": regim,
                               "label": REGIM_LABEL.get(regim, regim),
                               "fara_regim": regim == "",
                               "n": len(in_group),
                               "units": in_group})

        fara_regim = sum(1 for r in rows if not r["regim"])
        fara_suprafata = sum(1 for r in rows if r["suprafata_mp"] is None)
        return {"success": True, "message": "", "data": {
            "titlu": REPORT_TITLES["conformitate"],
            "data_raport": date.today().strftime("%d.%m.%Y"),
            "partic_id": pid,
            "total": len(rows),
            "fara_regim": fara_regim,
            "fara_suprafata": fara_suprafata,
            "sumar": [{"regim": r, "label": REGIM_LABEL[r],
                       "n": sum(1 for x in rows if x["regim"] == r)}
                      for r in REGIM_ORDER],
            "grupuri": groups,
            "rows": [r for g in groups for r in g["units"]],
        }}

    @staticmethod
    def registru_data(params: Dict[str, Any] = None) -> Dict[str, Any]:
        """Реестр упаковки со сводкой по материалам."""
        search = (params or {}).get("q") or None
        listed = SDAStore.list_packs(search)
        if not listed.get("success"):
            return listed
        packs = listed.get("data") or []

        rows = []
        for p in packs:
            volum = _num(p.get("volum_l"))
            greutate = _num(p.get("greutate_g"))
            material = (p.get("material") or "").upper()
            rows.append({
                "ean": _txt(p.get("ean")),
                "denumire": _txt(p.get("denumire")),
                "producator": _txt(p.get("producator")),
                "material": material,
                "material_label": MATERIAL_LABEL.get(material, material),
                "culoare": _txt(p.get("culoare")),
                "bariera_o2": _da_nu(p.get("bariera_o2")),
                "reutilizabil": _da_nu(p.get("reutilizabil")),
                "volum_l": volum,
                "volum_text": ("—" if volum is None else f"{volum:g}"),
                "greutate_g": greutate,
                "greutate_text": ("—" if greutate is None else f"{greutate:g}"),
                "cat_admin": _txt(p.get("cat_admin")),
                "cat_gest": _txt(p.get("cat_gest")),
            })
        rows.sort(key=lambda r: r["denumire"])

        materiale = []
        for material in ("PLASTIC", "STICLA", "METAL"):
            n = sum(1 for r in rows if r["material"] == material)
            if n:
                materiale.append({"material": material,
                                  "label": MATERIAL_LABEL[material], "n": n})
        altele = sum(1 for r in rows
                     if r["material"] not in ("PLASTIC", "STICLA", "METAL"))
        if altele:
            materiale.append({"material": "ALTE", "label": "Alte materiale",
                              "n": altele})

        return {"success": True, "message": "", "data": {
            "titlu": REPORT_TITLES["registru"],
            "data_raport": date.today().strftime("%d.%m.%Y"),
            "total": len(rows),
            "materiale": materiale,
            "rows": rows,
        }}

    @staticmethod
    def dosar_data(params: Dict[str, Any] = None) -> Dict[str, Any]:
        """Досье регистрации (пункт 78) — восемь блоков в порядке регламента.

        Вердикт `poate_fi_depus` выносится в шапку документа: отрицательный
        вердикт, спрятанный в подвале, читатель увидит после того, как
        отправит досье.
        """
        try:
            pid = _partic_id(params or {})
        except ValueError as exc:
            return _fail(str(exc))
        if not pid:
            return _fail("Parametrul partic_id este obligatoriu")
        res = SDAStore.registration_dossier(pid)
        if not res.get("success"):
            return res
        d = res.get("data") or {}

        unitati = []
        for u in (d.get("unitati") or []):
            supr = _num(u.get("suprafata_mp"))
            unitati.append({
                "denumire": _txt(u.get("denumire")),
                "adresa": _txt(u.get("adresa")),
                "suprafata_text": ("—" if supr is None else f"{supr:g}"),
                "suprafata_mp": supr,
                "tip_amplasament": TIP_LABEL.get(
                    (u.get("tip_amplasament") or "").upper(),
                    _txt(u.get("tip_amplasament"))),
                "regim": _txt(u.get("regim")),
                "regim_label": REGIM_LABEL.get(_txt(u.get("regim")),
                                               _txt(u.get("regim"))),
            })
        puncte = [{"adresa": _txt(p.get("adresa")),
                   "orar": _txt(p.get("orar")),
                   "tip": _txt(p.get("tip"))}
                  for p in (d.get("punct_preluare") or [])]
        exceptii = [{"denumire": _txt(x.get("denumire")),
                     "adresa": _txt(x.get("adresa"))}
                    for x in (d.get("exceptii") or [])]

        incomplet = int(d.get("incomplet") or 0)
        poate = bool(d.get("poate_fi_depus"))
        return {"success": True, "message": "", "data": {
            "titlu": REPORT_TITLES["dosar"],
            "data_raport": date.today().strftime("%d.%m.%Y"),
            "partic_id": pid,
            "identificare": d.get("identificare") or {},
            "contact": d.get("contact") or {},
            "unitati": unitati,
            "punct_preluare": puncte,
            "modalitate_preluare": d.get("modalitate_preluare") or [],
            "modalitate_text": (", ".join(d.get("modalitate_preluare") or [])
                                or "nedeclarată"),
            "vandut_an_anterior": d.get("vandut_an_anterior"),
            "estimare_an_curent": d.get("estimare_an_curent"),
            "exceptii": exceptii,
            "incomplet": incomplet,
            "poate_fi_depus": poate,
            "verdict": ("Dosarul poate fi depus la Administrator."
                        if poate else
                        "ATENȚIE: dosarul NU poate fi depus — există date "
                        "incomplete (vezi mai jos)."),
        }}

    _DATA = {
        "conformitate": "conformitate_data",
        "registru": "registru_data",
        "dosar": "dosar_data",
    }

    @staticmethod
    def data_for(kind: str, params: Dict[str, Any] = None) -> Dict[str, Any]:
        if kind not in REPORT_KINDS:
            return _fail(f"Raport necunoscut: {kind}")
        return getattr(SDAReport, SDAReport._DATA[kind])(params or {})

    # ── PDF ─────────────────────────────────────────────────────────

    @staticmethod
    def render_pdf(kind: str, params: Dict[str, Any] = None) -> Dict[str, Any]:
        """Шаблон и данные уходят ВНУТРИ POST-запроса — сервису рендера не
        нужно ничего знать про модуль и хранить у себя."""
        if kind not in REPORT_KINDS:
            return _fail(f"Raport necunoscut: {kind}")
        prepared = SDAReport.data_for(kind, params)
        if not prepared.get("success"):
            return prepared
        try:
            template = _read(REPORT_KINDS[kind])
            helpers = _read("helpers.js")
        except OSError as exc:
            return _fail(f"Șablonul raportului nu poate fi citit: {exc}")
        try:
            resp = requests.post(
                Config.JSREPORT_URL.rstrip("/") + "/api/report",
                json={
                    "template": {
                        "content": template,
                        "engine": "handlebars",
                        "recipe": "chrome-pdf",
                        "helpers": helpers,
                        "chrome": {"format": "A4",
                                   "marginTop": "10mm", "marginBottom": "12mm",
                                   "marginLeft": "8mm", "marginRight": "8mm"},
                    },
                    "data": prepared["data"],
                },
                timeout=90)
        except requests.ConnectionError:
            return _fail("Serviciul de rapoarte nu este disponibil "
                         "(jsReport nu rulează)")
        except Exception as exc:  # timeout и всё прочее — тоже отказ, не 500
            return _fail(f"Eroare la generarea PDF: {exc}")
        if getattr(resp, "status_code", None) != 200:
            return _fail(f"Serviciul de rapoarte a răspuns HTTP "
                         f"{getattr(resp, 'status_code', '?')}")
        return {"success": True, "message": "", "pdf": resp.content}

    # ── Excel ───────────────────────────────────────────────────────

    XLSX_HEADERS = {
        "conformitate": ["Cod ERP", "Denumire", "Adresă", "Localitate",
                         "Suprafață, m²", "Tip amplasament", "Regim",
                         "Motivul regimului"],
        "registru": ["EAN", "Denumire", "Producător", "Material", "Culoare",
                     "Barieră O₂", "Reutilizabil", "Volum, l", "Greutate, g",
                     "Categoria tarifară (a–g)", "Categoria de gestiune (a–e)"],
        "dosar": ["Bloc (pct. 78)", "Element", "Valoare"],
    }

    @staticmethod
    def render_xlsx(kind: str, params: Dict[str, Any] = None) -> Dict[str, Any]:
        if kind not in REPORT_KINDS:
            return _fail(f"Raport necunoscut: {kind}")
        prepared = SDAReport.data_for(kind, params)
        if not prepared.get("success"):
            return prepared
        try:
            builder = {"conformitate": SDAReport._xlsx_conformitate,
                       "registru": SDAReport._xlsx_registru,
                       "dosar": SDAReport._xlsx_dosar}[kind]
            return {"success": True, "message": "",
                    "xlsx": builder(prepared["data"])}
        except Exception as exc:
            return _fail(f"Eroare la generarea Excel: {exc}")

    # -- вспомогательное для openpyxl --

    @staticmethod
    def _new_sheet(title: str, headers: List[str], widths: List[int]):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"
        return wb, ws

    @staticmethod
    def _bytes(wb) -> bytes:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _xlsx_conformitate(data: Dict[str, Any]) -> bytes:
        from openpyxl.styles import Font

        wb, ws = SDAReport._new_sheet(
            "Conformitate", SDAReport.XLSX_HEADERS["conformitate"],
            [14, 34, 38, 20, 14, 20, 20, 46])
        # Шапка со сводкой живёт на отдельном листе: первая строка первого
        # листа обязана быть строкой заголовков таблицы, иначе фильтры и
        # выгрузка в другие системы ломаются.
        r = 2
        for row in data["rows"]:
            ws.cell(row=r, column=1, value=row["cod_erp"])
            ws.cell(row=r, column=2, value=row["denumire"])
            ws.cell(row=r, column=3, value=row["adresa"])
            ws.cell(row=r, column=4, value=row["localitate"])
            ws.cell(row=r, column=5, value=row["suprafata_mp"])
            ws.cell(row=r, column=6, value=row["tip_amplasament"])
            ws.cell(row=r, column=7, value=row["regim"] or "FĂRĂ REGIM")
            ws.cell(row=r, column=8, value=row["regim_motiv"])
            if not row["regim"]:
                for c in range(1, 9):
                    ws.cell(row=r, column=c).font = Font(bold=True, color="B00020")
            r += 1

        s = wb.create_sheet("Sumar")
        s.column_dimensions["A"].width = 54
        s.column_dimensions["B"].width = 12
        s["A1"] = "Harta de conformitate — sumar"
        s["A1"].font = Font(bold=True, size=13)
        s["A2"] = f"Data raportului: {data['data_raport']}"
        rr = 4
        for item in data["sumar"]:
            s.cell(row=rr, column=1, value=item["label"])
            s.cell(row=rr, column=2, value=item["n"])
            rr += 1
        s.cell(row=rr, column=1, value="Unități fără regim atribuit").font = \
            Font(bold=True, color="B00020")
        s.cell(row=rr, column=2, value=data["fara_regim"]).font = \
            Font(bold=True, color="B00020")
        rr += 1
        s.cell(row=rr, column=1, value="Unități fără suprafață declarată").font = \
            Font(bold=True, color="B00020")
        s.cell(row=rr, column=2, value=data["fara_suprafata"]).font = \
            Font(bold=True, color="B00020")
        rr += 1
        s.cell(row=rr, column=1, value="Total unități").font = Font(bold=True)
        s.cell(row=rr, column=2, value=data["total"]).font = Font(bold=True)
        return SDAReport._bytes(wb)

    @staticmethod
    def _xlsx_registru(data: Dict[str, Any]) -> bytes:
        from openpyxl.styles import Font

        wb, ws = SDAReport._new_sheet(
            "Registru", SDAReport.XLSX_HEADERS["registru"],
            [18, 34, 24, 12, 14, 12, 13, 11, 13, 22, 24])
        r = 2
        for row in data["rows"]:
            for col, value in enumerate([
                    row["ean"], row["denumire"], row["producator"],
                    row["material_label"], row["culoare"], row["bariera_o2"],
                    row["reutilizabil"], row["volum_l"], row["greutate_g"],
                    row["cat_admin"], row["cat_gest"]], 1):
                ws.cell(row=r, column=col, value=value)
            r += 1

        s = wb.create_sheet("Sumar")
        s.column_dimensions["A"].width = 30
        s["A1"] = "Registrul ambalajelor SD — sumar pe material"
        s["A1"].font = Font(bold=True, size=13)
        s["A2"] = f"Data raportului: {data['data_raport']}"
        rr = 4
        for item in data["materiale"]:
            s.cell(row=rr, column=1, value=item["label"])
            s.cell(row=rr, column=2, value=item["n"])
            rr += 1
        s.cell(row=rr, column=1, value="Total ambalaje").font = Font(bold=True)
        s.cell(row=rr, column=2, value=data["total"]).font = Font(bold=True)
        return SDAReport._bytes(wb)

    @staticmethod
    def _xlsx_dosar(data: Dict[str, Any]) -> bytes:
        from openpyxl.styles import Font

        wb, ws = SDAReport._new_sheet(
            "Dosar pct. 78", SDAReport.XLSX_HEADERS["dosar"], [34, 34, 60])
        rows: List[List[Any]] = []
        # Вердикт — первой строкой данных, до всех восьми блоков.
        rows.append(["Verdict", "Poate fi depus",
                     "DA" if data["poate_fi_depus"] else "NU"])
        rows.append(["Verdict", "Unități incomplete", data["incomplet"]])
        ident = data["identificare"]
        rows.append(["1. Identificare", "IDNO", _txt(ident.get("idno"))])
        rows.append(["1. Identificare", "Denumire", _txt(ident.get("denumire"))])
        contact = data["contact"]
        rows.append(["2. Contact", "Nume", _txt(contact.get("nume"))])
        rows.append(["2. Contact", "Telefon", _txt(contact.get("telefon"))])
        rows.append(["2. Contact", "E-mail", _txt(contact.get("email"))])
        for u in data["unitati"]:
            rows.append(["3. Unități", u["denumire"],
                         f"{u['adresa']} · {u['suprafata_text']} m² · "
                         f"{u['tip_amplasament']} · {u['regim'] or 'FĂRĂ REGIM'}"])
        for p in data["punct_preluare"]:
            rows.append(["4. Puncte de preluare", p["adresa"],
                         f"{p['tip']} · {p['orar']}"])
        rows.append(["5. Modalitate de preluare", "Metode",
                     data["modalitate_text"]])
        rows.append(["6. Vândut anul anterior", "Unități",
                     data["vandut_an_anterior"]])
        rows.append(["7. Estimare anul curent", "Unități",
                     data["estimare_an_curent"]])
        for x in data["exceptii"]:
            rows.append(["8. Excepții", x["denumire"], x["adresa"]])
        if not data["exceptii"]:
            rows.append(["8. Excepții", "—", "Nicio unitate în regim de excepție"])

        for i, values in enumerate(rows, 2):
            for col, value in enumerate(values, 1):
                ws.cell(row=i, column=col, value=value)
            if values[0] == "Verdict" and not data["poate_fi_depus"]:
                for col in range(1, 4):
                    ws.cell(row=i, column=col).font = Font(bold=True,
                                                           color="B00020")
        return SDAReport._bytes(wb)

    # ── имя файла ───────────────────────────────────────────────────

    @staticmethod
    def filename(kind: str, ext: str) -> str:
        return f"sda-{kind}-{date.today().isoformat()}.{ext}"

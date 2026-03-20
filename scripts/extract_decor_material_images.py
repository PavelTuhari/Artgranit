from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = Path("/Users/pt/Projects.AI/decor/docs/veranda")
OUT_DIR = ROOT / "static" / "decor" / "materials"
OUT_JSON = ROOT / "data" / "decor_material_images.json"

CODE_RE = re.compile(r"^[A-Z0-9][A-Z0-9.\\-]{2,}$")


def slug(s: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return s or "sheet"


def image_ext_from_bytes(data: bytes, fmt: Optional[str]) -> str:
    f = (fmt or "").lower().strip(".")
    if f in {"png", "jpg", "jpeg", "gif", "bmp", "webp"}:
        return "jpg" if f == "jpeg" else f
    if data.startswith(b"\x89PNG"):
        return "png"
    if data.startswith(b"\xff\xd8"):
        return "jpg"
    if data[:6] in {b"GIF87a", b"GIF89a"}:
        return "gif"
    return "bin"


def row_values(ws, row_idx: int, max_col: int = 12) -> List[Any]:
    return [ws.cell(row=row_idx, column=c).value for c in range(1, min(ws.max_column, max_col) + 1)]


def candidate_codes(values: List[Any]) -> List[str]:
    out: List[str] = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        # Skip formulas and generic labels
        if s.startswith("="):
            continue
        if not CODE_RE.match(s):
            continue
        if any(ch.isalpha() for ch in s) or "-" in s or "." in s:
            out.append(s)
    return out


def detect_anchor_row_col(img: Any) -> tuple[Optional[int], Optional[int]]:
    anc = getattr(img, "anchor", None)
    if hasattr(anc, "_from"):
        return anc._from.row + 1, anc._from.col + 1
    if hasattr(anc, "from_"):
        return anc.from_.row + 1, anc.from_.col + 1
    return None, None


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)

    by_code: Dict[str, Dict[str, Any]] = {}
    images: List[Dict[str, Any]] = []
    idx = 1

    for xlsx in sorted(SOURCE_DIR.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue
        wb = load_workbook(xlsx)
        for ws in wb.worksheets:
            ws_images = getattr(ws, "_images", []) or []
            if not ws_images:
                continue
            for img in ws_images:
                try:
                    data = img._data()
                except Exception:
                    continue
                row, col = detect_anchor_row_col(img)
                ext = image_ext_from_bytes(data, getattr(img, "format", None))
                filename = f"{idx:03d}_{slug(xlsx.stem)}__{slug(ws.title)}_r{row or 0}_c{col or 0}.{ext}"
                out_path = OUT_DIR / filename
                out_path.write_bytes(data)
                rel_url = f"/static/decor/materials/{filename}"
                idx += 1

                row_hits: List[Dict[str, Any]] = []
                matched_codes: List[str] = []
                if row:
                    for rr in range(max(1, row - 2), min(ws.max_row, row + 2) + 1):
                        vals = row_values(ws, rr)
                        codes = candidate_codes(vals)
                        row_hits.append({"row": rr, "values": vals, "codes": codes})
                        if rr in {row - 1, row, row + 1}:
                            matched_codes.extend(codes)

                # Deduplicate preserve order
                matched_codes = list(dict.fromkeys(matched_codes))

                meta = {
                    "file": xlsx.name,
                    "sheet": ws.title,
                    "anchor_row": row,
                    "anchor_col": col,
                    "url": rel_url,
                    "width": getattr(img, "width", None),
                    "height": getattr(img, "height", None),
                    "matched_codes": matched_codes,
                    "row_context": row_hits,
                }
                images.append(meta)
                for code in matched_codes:
                    by_code.setdefault(code, meta)

    payload = {"by_code": by_code, "images": images}
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"saved {OUT_JSON} ({len(images)} images, {len(by_code)} codes)")


if __name__ == "__main__":
    main()

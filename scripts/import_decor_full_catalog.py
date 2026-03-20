#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from decor_local_store import DecorLocalStore


SRC_DIR = Path("/Users/pt/Projects.AI/decor/docs/veranda")
VALID_UNITS = {"pc", "kg", "m", "m2", "set"}


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _f(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        if isinstance(v, str):
            x = v.strip().replace(",", ".")
            if not x:
                return None
            return float(x)
        return float(v)
    except Exception:
        return None


def normalize_unit(v: str) -> str:
    x = _s(v).upper()
    u = {
        "PC": "pc",
        "PCS": "pc",
        "AD": "pc",
        "KG": "kg",
        "MTR": "m",
        "M": "m",
        "M2": "m2",
        "SET": "set",
    }.get(x, _s(v).lower() or "pc")
    return u if u in VALID_UNITS else "pc"


def infer_category(code: str, name: str, source_sheet: str) -> str:
    c = (code or "").upper()
    n = (name or "").upper()
    sh = (source_sheet or "").upper()

    if c in {"GLASS-TEKCAM", "GLASS-ISICAM", "TEKCAM", "ISICAM"}:
        return "glass"
    if c.startswith("CAM.") or c.startswith("EM052") or c.startswith("EM053"):
        return "glass"
    if c == "INSTALL":
        return "service"
    if c.startswith("M") or c.startswith("O."):
        return "profile"
    if c in {"EM06100001", "BIO-0018"} or ("OLUK" in n and ("SU" in n or "DRAIN" in n)):
        return "drainage"
    if c.startswith("VRD-0004") or c.startswith("VRD-0006") or c.startswith("VRD-0007") or c.startswith("EM062"):
        return "consumable"
    if (
        "LED" in n
        or "DIMMER" in n
        or "ADAPT" in n
        or "KONTROL" in n
        or "RTS" in n
        or "TELIS" in n
        or "GIRO" in n
        or c.startswith(("PRG-0172", "PRG-0174", "PRG-0177", "PRG-0226", "PRG-0236", "PRG-0276", "PRG-2143", "PRG-2225", "A510073", "A530", "EM081", "EM082", "EM031", "EM032", "9027615"))
    ):
        return "option_led"
    if "PROFILE" in sh:
        return "profile"
    return "accessory"


def _valid_code(code: str) -> bool:
    c = _s(code)
    if not c:
        return False
    if c in {"NO", "HUUN CODE", "DESCRIPTION", "KOD", "MALZEME KODU", "MALZEME İSİM", "MALZEME ISIM", "#N/A"}:
        return False
    if c.startswith("#"):
        return False
    if c.replace(".", "", 1).isdigit() and len(c) < 6:
        return False
    if not (re.search(r"\d", c) or "-" in c or "." in c):
        return False
    return True


def _insert_or_merge(items: Dict[str, Dict[str, Any]], row: Dict[str, Any]) -> None:
    code = row["code"]
    if code not in items:
        items[code] = row
        return
    cur = items[code]
    # Prefer rows with better unit/name and non-zero price; preserve first source otherwise.
    if (not cur.get("name")) and row.get("name"):
        cur["name"] = row["name"]
    if (cur.get("unit") in {"", "pc"}) and row.get("unit") in VALID_UNITS and row.get("unit") != "pc":
        cur["unit"] = row["unit"]
    if (not cur.get("unit_price")) and row.get("unit_price"):
        cur["unit_price"] = row["unit_price"]
    if (not cur.get("source_sheet")) and row.get("source_sheet"):
        cur["source_sheet"] = row["source_sheet"]
    if (not cur.get("source_file")) and row.get("source_file"):
        cur["source_file"] = row["source_file"]


def collect_catalog_rows() -> Dict[str, Dict[str, Any]]:
    items: Dict[str, Dict[str, Any]] = {}

    huun = SRC_DIR / "HUUN- NEW VERANDA- Material Base Price List -2024.xlsx"
    yeni = next(SRC_DIR.glob("YENI*HESAPLAMA.xlsx"))

    wb_h = load_workbook(huun, data_only=True)
    # HUUN profile
    ws = wb_h["PROFILE (NEW SYSTEM)"]
    for r in ws.iter_rows(values_only=True):
        code = _s(r[1] if len(r) > 1 else None)
        name = _s(r[3] if len(r) > 3 else None)
        if not (_valid_code(code) and name):
            continue
        unit = "m"
        # HUUN sheet usually lacks direct computed prices in data_only; keep if available
        unit_price = _f(r[6] if len(r) > 6 else None)
        _insert_or_merge(items, {
            "code": code,
            "name": name,
            "category": "profile",
            "unit": unit,
            "unit_price": round(unit_price, 2) if unit_price and unit_price > 0 else 0.0,
            "currency": "USD",
            "source_file": huun.name,
            "source_sheet": ws.title,
            "active": "Y",
        })

    # HUUN accessory
    ws = wb_h["ACCESSORY (2)"]
    for r in ws.iter_rows(values_only=True):
        code = _s(r[1] if len(r) > 1 else None)
        name = _s(r[2] if len(r) > 2 else None)
        if not (_valid_code(code) and name):
            continue
        unit = normalize_unit(_s(r[5] if len(r) > 5 else None))
        unit_price = _f(r[6] if len(r) > 6 else None)
        _insert_or_merge(items, {
            "code": code,
            "name": name,
            "category": infer_category(code, name, ws.title),
            "unit": unit,
            "unit_price": round(unit_price, 2) if unit_price and unit_price > 0 else 0.0,
            "currency": "USD",
            "source_file": huun.name,
            "source_sheet": ws.title,
            "active": "Y",
        })

    wb_y = load_workbook(yeni, data_only=True)

    # YENI production profile
    ws = wb_y["ÜRETİM-PROFİL"]
    for r in ws.iter_rows(max_col=4, values_only=True):
        code = _s(r[1] if len(r) > 1 else None)
        name = _s(r[2] if len(r) > 2 else None)
        if not (_valid_code(code) and name):
            continue
        _insert_or_merge(items, {
            "code": code,
            "name": name,
            "category": infer_category(code, name, ws.title),
            "unit": "m",
            "unit_price": 0.0,
            "currency": "USD",
            "source_file": yeni.name,
            "source_sheet": ws.title,
            "active": "Y",
        })

    # YENI full material list
    ws = wb_y["VERANDA MALZEME LİSTESİ"]
    for r in ws.iter_rows(max_col=4, values_only=True):
        pairs = [
            (_s(r[0] if len(r) > 0 else None), _s(r[2] if len(r) > 2 else None), _s(r[3] if len(r) > 3 else None)),
            (_s(r[1] if len(r) > 1 else None), _s(r[2] if len(r) > 2 else None), _s(r[3] if len(r) > 3 else None)),
        ]
        for code, name, unit_raw in pairs:
            if not (_valid_code(code) and name):
                continue
            _insert_or_merge(items, {
                "code": code,
                "name": name,
                "category": infer_category(code, name, ws.title),
                "unit": normalize_unit(unit_raw),
                "unit_price": 0.0,
                "currency": "USD",
                "source_file": yeni.name,
                "source_sheet": ws.title,
                "active": "Y",
            })

    # YENI accessory assembly sheets
    for sheet_name in ["AKSESUAR SAYFASI", "BOYALI AKSESUAR"]:
        ws = wb_y[sheet_name]
        for r in ws.iter_rows(max_col=12, values_only=True):
            if sheet_name == "AKSESUAR SAYFASI":
                pairs = [
                    (_s(r[0] if len(r) > 0 else None), _s(r[8] if len(r) > 8 else None), ""),
                    (_s(r[3] if len(r) > 3 else None), _s(r[8] if len(r) > 8 else None), ""),
                ]
            else:
                pairs = [
                    (_s(r[1] if len(r) > 1 else None), _s(r[7] if len(r) > 7 else None), "AD"),
                    (_s(r[2] if len(r) > 2 else None), _s(r[7] if len(r) > 7 else None), "AD"),
                ]
            for code, name, unit_raw in pairs:
                if not (_valid_code(code) and name):
                    continue
                _insert_or_merge(items, {
                    "code": code,
                    "name": name,
                    "category": infer_category(code, name, ws.title),
                    "unit": normalize_unit(unit_raw),
                    "unit_price": 0.0,
                    "currency": "USD",
                    "source_file": yeni.name,
                    "source_sheet": ws.title,
                    "active": "Y",
                })

    # YENI reference sheet (glass systems, controls aliases)
    ws = wb_y["Veri Sayfası-2"]
    for r in ws.iter_rows(max_col=3, values_only=True):
        code = _s(r[0] if len(r) > 0 else None)
        name = _s(r[1] if len(r) > 1 else None)
        if _valid_code(code) and name:
            _insert_or_merge(items, {
                "code": code,
                "name": name,
                "category": infer_category(code, name, ws.title),
                "unit": "pc",
                "unit_price": 0.0,
                "currency": "USD",
                "source_file": yeni.name,
                "source_sheet": ws.title,
                "active": "Y",
            })

    return items


def main() -> int:
    catalog = collect_catalog_rows()

    state = DecorLocalStore._load()
    existing_rows = list(state.get("materials", []))
    existing_by_code = {str(m.get("code") or "").strip().upper(): m for m in existing_rows}

    imported_codes = set(catalog.keys())
    merged_rows: List[Dict[str, Any]] = []

    max_id = 0
    for code in sorted(imported_codes):
        base = catalog[code]
        ex = existing_by_code.get(code)
        row: Dict[str, Any] = dict(base)
        if ex:
            # Preserve IDs and user-maintained pricing/notes when present.
            row["id"] = ex.get("id")
            row["created_at"] = ex.get("created_at") or DecorLocalStore._now_iso()
            row["updated_at"] = DecorLocalStore._now_iso()
            if float(ex.get("unit_price") or 0) > 0 and float(row.get("unit_price") or 0) <= 0:
                row["unit_price"] = float(ex.get("unit_price") or 0)
            if ex.get("currency"):
                row["currency"] = ex.get("currency")
            if ex.get("notes"):
                row["notes"] = ex.get("notes")
            # keep explicit deactivation if user set it
            if ex.get("active") in {"Y", "N"}:
                row["active"] = ex.get("active")
        else:
            row["id"] = None
            row["created_at"] = DecorLocalStore._now_iso()
            row["updated_at"] = DecorLocalStore._now_iso()

        row = DecorLocalStore._normalize_material_row(row)
        merged_rows.append(row)
        try:
            max_id = max(max_id, int(row.get("id") or 0))
        except Exception:
            pass

    # Keep custom/non-import materials (e.g. synthetic GLASS-*/LED-*/INSTALL) and any manually added records.
    for ex in existing_rows:
        code = str(ex.get("code") or "").strip().upper()
        if not code or code in imported_codes:
            continue
        row = DecorLocalStore._normalize_material_row(dict(ex))
        merged_rows.append(row)
        max_id = max(max_id, int(row.get("id") or 0))

    # Assign IDs for newly imported rows lacking IDs
    next_id = max(max_id + 1, int(state.get("next_material_id") or 1000))
    for row in merged_rows:
        if not row.get("id"):
            row["id"] = next_id
            next_id += 1

    merged_rows.sort(key=lambda m: (str(m.get("category") or ""), str(m.get("code") or "")))
    state["materials"] = merged_rows
    state["next_material_id"] = next_id
    DecorLocalStore._save(state)

    with_images = sum(1 for m in merged_rows if m.get("image_url"))
    with_ro = sum(1 for m in merged_rows if m.get("name_ro"))
    print(f"Imported materials into DECOR store: total={len(merged_rows)} imported_excel={len(imported_codes)} images={with_images} name_ro={with_ro}")
    print("Sample:")
    for code in ["M6521", "M6283", "EM07100091", "EM06200040", "PRG-0172", "CAM.967", "LED-WHITE", "GLASS-ISICAM", "INSTALL"]:
        m = next((x for x in merged_rows if str(x.get('code') or '').upper() == code), None)
        if m:
            print(f" - {m['code']}: cat={m.get('category')} unit={m.get('unit')} image={'Y' if m.get('image_url') else 'N'} ro={m.get('name_ro')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

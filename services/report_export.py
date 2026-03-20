"""
Экспорт отчётов в CSV, Excel, PDF.
Общее ядро для всех отчётов.
"""
from io import BytesIO, StringIO
from typing import List, Dict, Any, Optional
import csv


def export_csv(data: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> str:
    """Экспорт в CSV. columns — порядок колонок (если None — из первого ряда)."""
    if not data:
        return ""
    cols = columns or list(data[0].keys())
    out = StringIO()
    writer = csv.writer(out, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(cols)
    for row in data:
        writer.writerow([row.get(c, "") for c in cols])
    return out.getvalue()


def export_excel(data: List[Dict[str, Any]], report_name: str = "Report", columns: Optional[List[str]] = None) -> bytes:
    """Экспорт в Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl не установлен. Выполните: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = (report_name or "Sheet")[:31]
    cols = columns or (list(data[0].keys()) if data else [])
    for j, col in enumerate(cols, 1):
        ws.cell(row=1, column=j, value=col)
    for i, row in enumerate(data, 2):
        for j, col in enumerate(cols, 1):
            v = row.get(col, "")
            ws.cell(row=i, column=j, value=v)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _get_cyrillic_font():
    """Регистрирует шрифт с поддержкой кириллицы."""
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
        base = os.path.dirname(os.path.abspath(__file__))
        for p in [
            os.path.join(base, "..", "static", "fonts", "DejaVuSans.ttf"),
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/TTF/DejaVuSans.ttf",
        ]:
            if os.path.exists(p):
                pdfmetrics.registerFont(TTFont("DejaVu", p))
                return "DejaVu"
    except Exception:
        pass
    return "Helvetica"


def export_pdf(data: List[Dict[str, Any]], report_name: str = "Report", columns: Optional[List[str]] = None) -> bytes:
    """Экспорт в PDF с поддержкой кириллицы."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
    except ImportError:
        raise RuntimeError("reportlab не установен. Выполните: pip install reportlab")
    font_name = _get_cyrillic_font()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = font_name
    styles["Normal"].fontName = font_name
    elements = []
    elements.append(Paragraph(str(report_name or "Отчёт"), styles["Title"]))
    elements.append(Spacer(1, 8))
    cols = columns or (list(data[0].keys()) if data else [])
    if not data:
        elements.append(Paragraph("Нет данных", styles["Normal"]))
    else:
        table_data = [[str(c) for c in cols]]
        for row in data:
            table_data.append([str(row.get(c, ""))[:80] for c in cols])
        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066CC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fa")]),
        ]))
        elements.append(t)
    doc.build(elements)
    return buf.getvalue()

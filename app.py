#!/usr/bin/env python3
from __future__ import annotations
"""
Главное приложение Oracle SQL Developer - UNA.md/orasldev
MVC архитектура с WebSockets для реального времени
"""
from flask import Flask, Response, render_template, jsonify, request, session, redirect, url_for, g, send_from_directory, send_file
from flask_socketio import SocketIO, emit
from flask_babel import Babel, _, lazy_gettext as _l
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from config import Config
from controllers.auth_controller import AuthController
from controllers.dashboard_controller import DashboardController
from controllers.sql_controller import SQLController
from controllers.objects_controller import ObjectsController
from controllers.combo_scenario_controller import ComboScenarioController
from controllers.credit_controller import CreditController
from controllers.nufarul_controller import NufarulController
from controllers.documentation_controller import DocumentationController
from controllers.shell_controller import ShellController
from controllers.digi_marketing_controller import DigiMarketingController
from controllers.tbcontrol_controller import TBControlController
from controllers.peco_controller import PecoController
from controllers.planogram_controller import PlanogramController
from controllers.plg_mobile_controller import PlgMobileController
from controllers.plg_ai_controller import PlgAiController
from controllers.peco_supply_controller import PecoSupplyController
from models.peco_gps import PecoGps
from controllers.colass_controller import ColassController
import threading
import time
import os
import sys
from pathlib import Path
from decor_local_store import DecorLocalStore
from version_registry import VersionRegistry
from scripts.import_decor_order_xml_sample import import_xml_orders_from_dir

# Создание приложения Flask
app = Flask(__name__)
app.config.from_object(Config)
Config.init_app(app)

# RO: protectia sesiunii si a incarcarilor — cerinta auditului GDPR/securitate
#     (documentele personale ale clientilor: buletin fata/verso in
#     TMS_MUNC_ADDFILES). Cookie-ul merge doar pe HTTPS, nu e vizibil din
#     JavaScript si nu pleaca la cereri cross-site (protectie CSRF de baza);
#     corpul cererii e plafonat ca o incarcare uriasa sa nu epuizeze memoria.
# EN: session/upload hardening required by the GDPR & security audit —
#     HTTPS-only, HttpOnly, SameSite=Lax cookies and a hard body-size cap.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=(os.environ.get('ENVIRONMENT', '').upper() == 'REMOTE'),
    MAX_CONTENT_LENGTH=12 * 1024 * 1024,          # 12 MB (fisier max 8 MB)
)

# Инициализация Babel для интернационализации
babel = Babel()

def get_locale():
    """Определяет язык из сессии или заголовков браузера"""
    # Проверяем язык в сессии
    if 'language' in session:
        if session['language'] in Config.SUPPORTED_LANGUAGES:
            return session['language']
    
    # Проверяем язык из query параметра
    language = request.args.get('lang', None)
    if language and language in Config.SUPPORTED_LANGUAGES:
        session['language'] = language
        return language
    
    # Используем язык браузера
    return request.accept_languages.best_match(Config.SUPPORTED_LANGUAGES) or Config.BABEL_DEFAULT_LOCALE

# Инициализируем Babel с приложением
babel.init_app(app, locale_selector=get_locale)

# Rate limiting (только для /api/*; BIRO26 и auth — exempt)
# IMPORTANT: behind nginx/gunicorn all connections look like 127.0.0.1 if we use
# get_remote_address alone — shop, backoffice and sync share one bucket → mass 429.
# Prefer X-Real-IP (set by nginx from $remote_addr), then X-Forwarded-For.
def _client_ip_for_limiter():
    xri = (request.headers.get('X-Real-IP') or '').strip()
    if xri:
        return xri
    xff = (request.headers.get('X-Forwarded-For') or '').split(',')[0].strip()
    if xff:
        return xff
    return get_remote_address()


limiter = Limiter(
    key_func=_client_ip_for_limiter,
    app=app,
    default_limits=[Config.RATELIMIT_DEFAULT] if Config.RATELIMIT_ENABLED else [],
    storage_uri=Config.RATELIMIT_STORAGE_URI,
    default_limits_per_method=True,
)


@limiter.request_filter
def _skip_limit_for_non_api():
    """Skip rate limit for non-API, BIRO26, and authenticated sessions.

    RO: BIRO26 (shop + backoffice) face sute de GET pe sesiune (grid, price-history,
    sync catalog) — 200/oră dă 429 false-positive. Auth e pe rute.
    EN: BIRO26 (shop + backoffice) issues hundreds of GETs per session; a low
    shared limit caused false 429s. Route-level auth still applies.
    """
    path = request.path or ''
    if not path.startswith('/api/'):
        return True
    # BIRO26: high-volume legitimate UI + catalog sync; never rate-limit here.
    # EXCEPTION: the public, unauthenticated credit-flow endpoints under
    # /api/biro26/shop/credit/api/* carry their own explicit @limiter.limit(...)
    # decorators (anti-abuse for anonymous requests) — request_filter exempts
    # ALL limits (default_limits AND per-route decorators, see
    # flask_limiter._extension.Limiter.__check_all_limits_exempt), so those
    # paths must be carved out here or the decorators would be silently inert.
    if path.startswith('/api/biro26/shop/credit/api/'):
        return False
    if path.startswith('/api/biro26/'):
        return True
    try:
        if AuthController.is_authenticated():
            return True
    except Exception:
        pass
    return False


@app.errorhandler(429)
def ratelimit_handler(e):
    """Ответ при превышении лимита запросов"""
    resp = jsonify({
        "success": False,
        "error": "Rate limit exceeded",
        "message": "Слишком много запросов. Попробуйте позже.",
    })
    resp.status_code = 429
    resp.headers['Retry-After'] = '60'
    return resp


# Инициализация SocketIO для WebSockets
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Контекстный процессор для шаблонов - делает функцию _() доступной везде
@app.context_processor
def inject_gettext():
    return dict(_=_, get_locale=get_locale, languages=Config.LANGUAGES, supported_languages=Config.SUPPORTED_LANGUAGES)


@app.context_processor
def inject_app_version():
    """RO: versiunea din TMS_WEBAPPVERS — se afiseaza in subsolul site-ului.
    EN: the release version shown in the site footer."""
    from models.biro26_version import current
    try:
        return {"app_version": current()}
    except Exception:                                  # noqa: BLE001
        return {"app_version": ""}


def _version_widget_snippet() -> str:
    return """
<!-- UNA_VERSION_WIDGET -->
<script>
(function(){
  if (window.__unaVersionWidgetLoaded) return;
  window.__unaVersionWidgetLoaded = true;
  function esc(s){return String(s==null?'':s).replace(/[&<>\"']/g,function(m){return {'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',\"'\":'&#39;'}[m];});}
  function el(tag, attrs){var x=document.createElement(tag); if(attrs){for(var k in attrs){x.setAttribute(k, attrs[k]);}} return x;}
  function render(d){
    var root = document.getElementById('una-version-widget');
    if (!root) {
      root = el('div', {id:'una-version-widget'});
      document.body.appendChild(root);
    }
    var css = ''+
      '#una-version-widget{position:fixed;right:10px;bottom:10px;z-index:99999;max-width:340px;font:12px/1.35 Menlo,Consolas,monospace;color:#eaf2f7;background:rgba(10,18,24,.92);border:1px solid rgba(255,255,255,.16);border-radius:10px;box-shadow:0 8px 28px rgba(0,0,0,.35)}'+
      '#una-version-widget .h{padding:8px 10px;border-bottom:1px solid rgba(255,255,255,.12);display:flex;justify-content:space-between;gap:8px;cursor:pointer}'+
      '#una-version-widget .t{font-weight:700;color:#fff}'+
      '#una-version-widget .b{padding:8px 10px;display:grid;gap:6px}'+
      '#una-version-widget .r{display:grid;grid-template-columns:88px 1fr;gap:6px;align-items:start}'+
      '#una-version-widget .k{opacity:.75}'+
      '#una-version-widget .v{word-break:break-word}'+
      '#una-version-widget .pill{display:inline-block;padding:1px 6px;border-radius:999px;background:#173445;color:#bfe8ff;border:1px solid rgba(191,232,255,.18)}'+
      '#una-version-widget .muted{opacity:.7;font-size:11px}'+
      '#una-version-widget.collapsed .b{display:none}';
    if (!document.getElementById('una-version-widget-style')) {
      var st = el('style', {id:'una-version-widget-style'});
      st.textContent = css; document.head.appendChild(st);
    }
    var mod = d.module || {};
    var shell = d.shell || {};
    var app = d.app || {};
    root.innerHTML =
      '<div class=\"h\" title=\"Click to collapse / expand\">'+
      '<div class=\"t\">Version</div>'+
      '<div class=\"pill\">'+esc((mod.name||d.module_key||'module'))+' '+esc(mod.version||'')+'</div>'+
      '</div>'+
      '<div class=\"b\">'+
        '<div class=\"r\"><div class=\"k\">Module</div><div class=\"v\">'+esc(mod.name||d.module_key||'')+' <strong>'+esc(mod.version||'')+'</strong><div class=\"muted\">'+esc(mod.updated_at||'')+'</div></div></div>'+
        '<div class=\"r\"><div class=\"k\">Shell</div><div class=\"v\">'+esc(shell.name||'Shell')+' <strong>'+esc(shell.version||'')+'</strong><div class=\"muted\">'+esc(shell.updated_at||'')+'</div></div></div>'+
        '<div class=\"r\"><div class=\"k\">App</div><div class=\"v\">'+esc(app.name||'App')+' <strong>'+esc(app.version||'')+'</strong><div class=\"muted\">'+esc(app.updated_at||'')+'</div></div></div>'+
        '<div class=\"r\"><div class=\"k\">Path</div><div class=\"v muted\">'+esc(d.path||location.pathname)+'</div></div>'+
      '</div>';
    var head = root.querySelector('.h');
    if (head) head.onclick = function(){ root.classList.toggle('collapsed'); };
  }
  fetch('/api/system/version-info?path='+encodeURIComponent(location.pathname), {credentials:'same-origin'})
    .then(function(r){ return r.json(); })
    .then(function(d){ if (d && d.success) render(d); })
    .catch(function(){});
})();
</script>
"""


@app.after_request
def _inject_version_widget(response):
    if not app.config.get("VERSION_WIDGET_ENABLED", False):
        return response
    try:
        ctype = (response.content_type or "").lower()
        if "text/html" not in ctype:
            return response
        if response.direct_passthrough:
            return response
        if response.status_code >= 400:
            return response
        data = response.get_data(as_text=True)
        if not data or "UNA_VERSION_WIDGET" in data:
            return response
        low = data.lower()
        idx = low.rfind("</body>")
        if idx == -1:
            return response
        patched = data[:idx] + _version_widget_snippet() + data[idx:]
        response.set_data(patched)
        response.headers["Content-Length"] = str(len(patched.encode("utf-8")))
    except Exception:
        return response
    return response


# ── atributia traficului social (fbclid/gclid/utm/referrer) ───────────────
# RO: nucleul site-ului prinde click-ID-urile TUTUROR retelelor pe paginile
#     vitrinei, tine prima/ultima atingere in cookie si scrie vizitele
#     atribuite + conversiile in baza WordPress (wp_op_social_*), unde le
#     analizeaza pluginul WP «OfficePlus Social Analytics».
#     Vezi models/biro26_social.py; fail-silent daca MySQL lipseste.
# EN: social attribution capture on storefront pages; data lands in the
#     WordPress MySQL schema, analysed by the WP admin plugin.
_SOCIAL_PREFIXES = ('/UNA.md/orasldev/biro26-site',
                    '/UNA.md/orasldev/biro26-1shop',
                    '/UNA.md/orasldev/biro26-shop')

@app.before_request
def _biro26_social_capture():
    try:
        if request.method != 'GET':
            return
        if not request.path.startswith(_SOCIAL_PREFIXES):
            return
        from models.biro26_social import Biro26Social
        g._op_social = Biro26Social.on_request(request)
    except Exception:                                        # noqa: BLE001
        pass

@app.after_request
def _biro26_social_cookies(response):
    try:
        res = getattr(g, '_op_social', None)
        if res and res.get('set_cookies'):
            from models.biro26_social import COOKIE_DAYS
            for name, val in res['set_cookies'].items():
                response.set_cookie(
                    name, val, max_age=COOKIE_DAYS * 86400,
                    samesite='Lax', secure=request.is_secure,
                    httponly=True, path='/')
    except Exception:                                        # noqa: BLE001
        pass
    return response


def _login_redirect():
    """Редирект на логин с сохранением текущего URL в next= для возврата после входа."""
    next_path = (request.full_path or request.path or "").strip()
    if next_path and next_path != "login" and not next_path.startswith("/login"):
        return redirect(url_for("login", next=next_path))
    return redirect(url_for("login"))


def _is_safe_redirect_url(url):
    """Проверка, что URL для редиректа внутренний (нет открытых редиректов)."""
    if not url or not isinstance(url, str):
        return False
    url = url.strip()
    if not url.startswith("/") or "//" in url:
        return False
    return True

# Хранилище активных подписок на метрики
active_subscriptions = {}


@app.route('/')
def index():
    """Главная страница - редирект на login или SQL Developer"""
    if AuthController.is_authenticated():
        return redirect(url_for('sqldeveloper'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Страница входа"""
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        
        if AuthController.login(username, password):
            AuthController.set_authenticated(True)
            return jsonify({"success": True, "redirect": url_for('sqldeveloper')})
        else:
            return jsonify({"success": False, "error": _("Invalid credentials")}), 401
    
    # GET запрос - показываем форму с предзаполненными данными (next — куда вернуться после входа)
    next_url = request.args.get('next', '')
    # RO: instalare dedicata (officeplus.md): login EXCLUSIV OfficePlus —
    #     fara selectorul de proiecte, intrare direct in backoffice.
    # EN: dedicated install: OfficePlus-only login — no project selector.
    exclusive = (Config.LOGIN_EXCLUSIVE == 'biro26')
    if exclusive and not next_url:
        next_url = '/UNA.md/orasldev/biro26-backoffice'
    return render_template('login.html',
                         default_username=Config.DEFAULT_USERNAME,
                         default_password=Config.DEFAULT_PASSWORD,
                         next_url=next_url,
                         login_exclusive=exclusive,
                         exclusive_app_name=Config.BIRO26_APP_NAME)


@app.route('/logout')
def logout():
    """Выход из системы"""
    AuthController.logout()
    return redirect(url_for('login'))


@app.route('/UNA.md/orasldev')
@app.route('/UNA.md/orasldev/')
def sqldeveloper():
    """Oracle SQL Developer интерфейс"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('sqldeveloper_mdi.html', username=AuthController.get_current_user())


@app.route('/UNA.md/orasldev/dashboard')
@app.route('/UNA.md/orasldev/dashboard/<dashboard_id>')
def dashboard(dashboard_id=None):
    """Dashboard с метриками БД"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    
    # Проверяем query параметр ?p=XX
    query_param = request.args.get('p', None)
    if query_param:
        dashboard_id = query_param
    
    # Определяем режим: fullscreen если указан dashboard_id
    is_fullscreen = dashboard_id is not None
    
    return render_template('dashboard_mdi.html', dashboard_id=dashboard_id, is_fullscreen=is_fullscreen)


@app.route('/UNA.md/orasldev/credit-admin')
def credit_admin():
    """Админ-панель настройки кредитных предложений (embed в дашборде 04)"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('credit_admin.html')


@app.route('/UNA.md/orasldev/credit-portfolio-bomba')
def credit_portfolio_bomba():
    """Кредитный портфель Бомба: пивот категории/товары × банки/программы + настройки."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('credit_portfolio_bomba.html')


@app.route('/UNA.md/orasldev/credit-operator')
def credit_operator():
    """Интерфейс оператора оформления кредитов (embed в дашборде 05)"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('credit_operator.html')


@app.route('/UNA.md/orasldev/nufarul-admin')
def nufarul_admin():
    """Админка Nufarul: услуги, заказы, отчёты"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('nufarul_admin.html')


@app.route('/UNA.md/orasldev/nufarul-operator')
def nufarul_operator():
    """Интерфейс оператора приёма заказов в зале (Nufarul)"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('nufarul_operator.html')


@app.route('/UNA.md/orasldev/nufarul-oper-ts')
def nufarul_oper_ts():
    """Touchscreen kiosk interface for Nufarul intake/issue"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('nufarul_oper_ts.html')


@app.route('/UNA.md/orasldev/nufarul-customer-screen')
def nufarul_customer_screen():
    """Customer-facing second display for Nufarul TS kiosk"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('nufarul_customer_screen.html')


@app.route('/UNA.md/orasldev/decor-admin')
def decor_admin():
    """DECOR: админка материалов, коэффициентов и заказов стеклянных крыш/веранд."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('decor_admin.html')


@app.route('/UNA.md/orasldev/decor')
def decor_main():
    """DECOR: основной entry-point модуля (ведёт сразу в операторский интерфейс)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return redirect(url_for('decor_operator'))


@app.route('/UNA.md/orasldev/decor-operator')
def decor_operator():
    """DECOR: оператор приёма заказов в зале (расчёт + заявка/смета)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('decor_operator.html')


@app.route('/UNA.md/orasldev/decor-operator/document/<int:order_id>')
def decor_operator_document(order_id):
    """DECOR: печатная форма коммерческого предложения / сметы по заказу."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    result = DecorLocalStore.get_order_by_id(order_id)
    if not result.get("success") or not result.get("data"):
        return f"<h1>Заказ не найден</h1><p>Order ID: {order_id}</p><p><a href='/UNA.md/orasldev/decor-operator'>← Оператор DECOR</a></p>", 404
    order = result["data"]
    q = order.get("quote") or {}
    metrics = q.get("metrics") or {}
    summary = q.get("summary") or {}
    created_at = str(order.get("created_at") or "")[:19]
    auto_print = str(request.args.get("print") or "").strip() in {"1", "true", "yes", "y"}
    return render_template(
        "decor/document_offer_quote.html",
        order=order,
        quote=q,
        metrics=metrics,
        summary=summary,
        created_at=created_at,
        auto_print=auto_print,
    )


@app.route('/UNA.md/orasldev/credit-easycredit')
def credit_easycredit():
    """Оформление кредита по EasyCredit API (sandbox). Preapproved → Request → Status."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('credit_easycredit.html')


@app.route('/UNA.md/orasldev/credit-iute')
def credit_iute():
    """Оформление кредита по Iute API. Check Auth → Create Order → Check Status."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('credit_iute.html')


# ──────────────────── Colass: page routes ────────────────────
@app.route('/UNA.md/orasldev/colass-catalog')
def colass_catalog():
    """Colass: каталог работ F5 и ресурсов F3."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('colass_catalog.html')


@app.route('/UNA.md/orasldev/colass-estimator')
def colass_estimator():
    """Colass: сметчик с добавлением работ из каталога."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('colass_estimator.html')


@app.route('/UNA.md/orasldev/colass-crm')
def colass_crm():
    """Colass CRM: лиды, воронка, активности."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('colass_crm.html')


@app.route('/UNA.md/orasldev/colass-contracts')
def colass_contracts():
    """Colass: реестр договоров."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('colass_contracts.html')


# ──────────────────── Colass: catalog API ────────────────────
@app.route('/api/colass/catalog/tree')
def api_colass_catalog_tree():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    search = request.args.get('search')
    return jsonify(ColassController.get_catalog_tree(search=search))


@app.route('/api/colass/catalog/works')
def api_colass_catalog_works():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    search = request.args.get('search')
    limit = request.args.get('limit', 500, type=int)
    return jsonify(ColassController.get_work_catalog(search=search, limit=limit))


@app.route('/api/colass/catalog/resources')
def api_colass_catalog_resources():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    type_code = request.args.get('type')
    limit = request.args.get('limit', 500, type=int)
    return jsonify(ColassController.get_resources(type_code=type_code, limit=limit))


@app.route('/api/colass/catalog/work/<int:work_id>/resources')
def api_colass_work_resources(work_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.get_work_resources(work_id))


# ──────────────────── Colass: projects & estimates API ────────────────────
@app.route('/api/colass/projects')
def api_colass_projects():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.get_projects())


@app.route('/api/colass/projects/<int:project_id>/estimates')
def api_colass_project_estimates(project_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.get_project_estimates(project_id))


@app.route('/api/colass/estimates/<int:estimate_id>/sections')
def api_colass_estimate_sections(estimate_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.get_estimate_sections(estimate_id))


@app.route('/api/colass/estimates/<int:estimate_id>/items', methods=['GET', 'POST'])
def api_colass_estimate_items(estimate_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    if request.method == 'POST':
        return jsonify(ColassController.add_estimate_item(estimate_id, request.get_json(force=True)))
    section_id = request.args.get('section_id', type=int)
    return jsonify(ColassController.get_estimate_items(estimate_id, section_id=section_id))


@app.route('/api/colass/estimates/<int:estimate_id>/add-work', methods=['POST'])
def api_colass_estimate_add_work(estimate_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True)
    return jsonify(ColassController.add_work_to_estimate(
        estimate_id=estimate_id,
        work_id=int(payload.get('work_id', 0)),
        section_id=payload.get('section_id'),
        multiplier=float(payload.get('multiplier', 1.0)),
    ))


@app.route('/api/colass/estimates/<int:estimate_id>/summary')
def api_colass_estimate_summary(estimate_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.get_estimate_summary(estimate_id))


@app.route('/api/colass/estimate-items/<int:item_id>', methods=['PUT', 'DELETE'])
def api_colass_estimate_item(item_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    if request.method == 'DELETE':
        return jsonify(ColassController.delete_estimate_item(item_id))
    return jsonify(ColassController.update_estimate_item(item_id, request.get_json(force=True)))


@app.route('/api/colass/ai-parse', methods=['POST'])
def api_colass_ai_parse():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True)
    text = payload.get('text', '')
    return jsonify(ColassController.ai_parse_estimate_oracle(text))


# ──────────────────── Colass: CRM API ────────────────────
@app.route('/api/colass/crm/stages')
def api_colass_crm_stages():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.get_crm_stages())


@app.route('/api/colass/crm/sources')
def api_colass_crm_sources():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.get_crm_sources())


@app.route('/api/colass/crm/leads', methods=['GET', 'POST'])
def api_colass_crm_leads():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    if request.method == 'POST':
        return jsonify(ColassController.create_crm_lead(request.get_json(force=True)))
    search = request.args.get('search')
    stage_code = request.args.get('stage_code')
    limit = request.args.get('limit', 300, type=int)
    return jsonify(ColassController.get_crm_leads(search=search, stage_code=stage_code, limit=limit))


@app.route('/api/colass/crm/leads/<int:lead_id>', methods=['PUT'])
def api_colass_crm_lead_update(lead_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.update_crm_lead(lead_id, request.get_json(force=True)))


@app.route('/api/colass/crm/leads/<int:lead_id>/activities', methods=['GET', 'POST'])
def api_colass_crm_activities(lead_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    if request.method == 'POST':
        return jsonify(ColassController.add_crm_activity(lead_id, request.get_json(force=True)))
    limit = request.args.get('limit', 100, type=int)
    return jsonify(ColassController.get_crm_activities(lead_id, limit=limit))


@app.route('/api/colass/crm/leads/<int:lead_id>/register-contract', methods=['POST'])
def api_colass_crm_register_contract(lead_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.register_crm_contract(lead_id, request.get_json(force=True)))


@app.route('/api/colass/crm/import-email', methods=['POST'])
def api_colass_crm_import_email():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.import_crm_leads_from_email())


# ──────────────────── Colass: contracts API ────────────────────
@app.route('/api/colass/contracts', methods=['GET', 'POST'])
def api_colass_contracts():
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    if request.method == 'POST':
        return jsonify(ColassController.create_contract(request.get_json(force=True)))
    search = request.args.get('search')
    limit = request.args.get('limit', 300, type=int)
    return jsonify(ColassController.get_contracts(search=search, limit=limit))


@app.route('/api/colass/contracts/<int:contract_id>', methods=['GET', 'PUT', 'DELETE'])
def api_colass_contract(contract_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    if request.method == 'DELETE':
        return jsonify(ColassController.delete_contract(contract_id))
    if request.method == 'PUT':
        return jsonify(ColassController.update_contract(contract_id, request.get_json(force=True)))
    return jsonify(ColassController.get_contract_detail(contract_id))


@app.route('/api/colass/contracts/<int:contract_id>/contacts', methods=['POST'])
def api_colass_contract_contacts(contract_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True)
    return jsonify(ColassController.add_contract_contact(
        contract_id, payload.get('kind', ''), payload.get('value', ''), payload.get('is_primary', 'N')
    ))


@app.route('/api/colass/contracts/<int:contract_id>/contacts/<kind>/<int:row_id>', methods=['DELETE'])
def api_colass_contract_contact_delete(contract_id, kind, row_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.delete_contract_contact(contract_id, kind, row_id))


@app.route('/api/colass/contracts/<int:contract_id>/attachments', methods=['GET', 'POST'])
def api_colass_contract_attachments(contract_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    if request.method == 'POST':
        f = request.files.get('file')
        if not f:
            return jsonify({"error": "no file"}), 400
        return jsonify(ColassController.add_contract_attachment(
            contract_id,
            type_code=request.form.get('type_code', 'OTHER'),
            file_name=f.filename,
            mime_type=f.mimetype or 'application/octet-stream',
            file_bytes=f.read(),
        ))
    return jsonify(ColassController.get_contract_attachments(contract_id))


@app.route('/api/colass/contracts/<int:contract_id>/attachments/<int:att_id>', methods=['DELETE'])
def api_colass_contract_attachment_delete(contract_id, att_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.delete_contract_attachment(contract_id, att_id))


@app.route('/api/colass/contracts/attachments/<int:att_id>/download')
def api_colass_contract_attachment_download(att_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    result = ColassController.get_contract_attachment_blob(att_id)
    if not result.get('success'):
        return jsonify(result), 404
    data = result['data']
    return Response(data['blob'], mimetype=data.get('mime_type', 'application/octet-stream'),
                    headers={'Content-Disposition': f'attachment; filename="{data.get("file_name", "file")}"'})


@app.route('/api/colass/contracts/<int:contract_id>/export-word')
def api_colass_contract_export_word(contract_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    result = ColassController.export_contract_to_word(contract_id)
    if not result.get('success'):
        return jsonify(result), 500
    return Response(result['data']['blob'], mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                    headers={'Content-Disposition': f'attachment; filename="{result["data"].get("filename", "contract.docx")}"'})


@app.route('/api/colass/contracts/<int:contract_id>/approval', methods=['GET'])
def api_colass_contract_approval(contract_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(ColassController.get_contract_approval(contract_id))


@app.route('/api/colass/contracts/<int:contract_id>/approval/start', methods=['POST'])
def api_colass_contract_approval_start(contract_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True)
    return jsonify(ColassController.start_contract_approval(contract_id, payload.get('approvers', payload.get('steps', [])), started_by=payload.get('started_by', 'ui'), comment_text=payload.get('comment')))


@app.route('/api/colass/contracts/approval/steps/<int:step_id>/decision', methods=['POST'])
def api_colass_contract_approval_decision(step_id):
    if not AuthController.is_authenticated():
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True)
    return jsonify(ColassController.decide_contract_approval_step(
        step_id, payload.get('decision', ''), decided_by=payload.get('decided_by', 'ui'), comment_text=payload.get('comment')
    ))


@app.route('/UNA.md/orasldev/docs')
@app.route('/UNA.md/orasldev/docs/')
def docs_index():
    """Главная страница документации"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('docs_index.html')


def _docs_slugify(value, separator='-'):
    """
    Якорь заголовка в стиле GitHub, с сохранением кириллицы.

    Стандартный slugify из markdown прогоняет текст через ASCII и от русского
    заголовка не оставляет ничего — все внутренние ссылки документа ведут
    в пустоту. Здесь режем только пунктуацию.
    """
    import re as _re
    import unicodedata as _ud
    text = _ud.normalize('NFKC', str(value)).strip().lower()
    text = _re.sub(r'[^\w\s-]', '', text, flags=_re.UNICODE)
    return _re.sub(r'[\s_]+', separator, text).strip(separator)


def _docs_md_to_html(markdown_content):
    """Конвертация Markdown в HTML для docs viewer."""
    try:
        import markdown
        from markdown.extensions import codehilite, fenced_code, tables
        # toc даёт заголовкам якоря по тексту: без него внутренние ссылки
        # вида [Раздел](#раздел) в документе никуда не ведут — просмотрщик
        # нумерует заголовки как h0, h1, … и о слагах не знает.
        md = markdown.Markdown(extensions=['codehilite', 'fenced_code', 'tables',
                                           'nl2br', 'toc'],
                               extension_configs={'toc': {'slugify': _docs_slugify}})
        return md.convert(markdown_content)
    except ImportError:
        import re
        html = markdown_content
        html = re.sub(r'```(\w+)?\n(.*?)```', r'<pre><code>\2</code></pre>', html, flags=re.DOTALL)
        html = re.sub(r'^# (.+)$', r'<h1>\1</h1>', html, flags=re.MULTILINE)
        html = re.sub(r'^## (.+)$', r'<h2>\1</h2>', html, flags=re.MULTILINE)
        html = re.sub(r'^### (.+)$', r'<h3>\1</h3>', html, flags=re.MULTILINE)
        html = re.sub(r'^#### (.+)$', r'<h4>\1</h4>', html, flags=re.MULTILINE)
        html = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', html)
        html = re.sub(r'\*(.+?)\*', r'<em>\1</em>', html)
        html = re.sub(r'(?<!`)(?<!<code>)`([^`\n]+)`(?!`)(?!</code>)', r'<code>\1</code>', html)
        html = re.sub(r'\[([^\]]+)\]\(([^\)]+)\)', r'<a href="\2">\1</a>', html)
        lines = html.split('\n')
        result, in_list = [], False
        for line in lines:
            m = re.match(r'^[-*]\s+(.+)', line)
            if m:
                if not in_list:
                    result.append('<ul>')
                    in_list = True
                result.append(f'<li>{m.group(1)}</li>')
            else:
                if in_list:
                    result.append('</ul>')
                    in_list = False
                if line.strip() and not line.strip().startswith('<'):
                    result.append(f'<p>{line.strip()}</p>')
                elif line.strip():
                    result.append(line)
        if in_list:
            result.append('</ul>')
        return '\n'.join(result)


def _render_doc_page(md_path, title):
    """Загружает .md, конвертирует в HTML, рендерит docs_viewer."""
    from pathlib import Path
    path = Path(md_path)
    if not path.exists():
        return None, ("<h1>Документация не найдена</h1>", 404)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            html_content = _docs_md_to_html(f.read())
        return render_template('docs_viewer.html', content=html_content, title=title), None
    except Exception as e:
        return None, (f"<h1>Ошибка</h1><p>{e}</p>", 500)


@app.route('/UNA.md/orasldev/docs/dashboard/<dashboard_id>')
def docs_dashboard(dashboard_id):
    """Документация по конкретному дашборду"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from pathlib import Path
    docs_path = Path(__file__).parent / "docs" / "dashboards" / f"dashboard_{dashboard_id}.md"
    if not docs_path.exists():
        return f"<h1>Документация не найдена</h1><p>Документация для дашборда {dashboard_id} не существует.</p>", 404
    resp, err = _render_doc_page(docs_path, f"Документация: Dashboard {dashboard_id}")
    if err:
        return err
    return resp


@app.route('/UNA.md/orasldev/docs/configuration')
def docs_configuration():
    """Страница «Конфигурация»"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from pathlib import Path
    p = Path(__file__).parent / "docs" / "CONFIGURATION.md"
    resp, err = _render_doc_page(p, "Конфигурация")
    if err:
        return err
    return resp


@app.route('/UNA.md/orasldev/docs/deployment')
def docs_deployment():
    """Страница «Развертывание»"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from pathlib import Path
    p = Path(__file__).parent / "docs" / "DEPLOYMENT.md"
    resp, err = _render_doc_page(p, "Развертывание")
    if err:
        return err
    return resp


@app.route('/UNA.md/orasldev/docs/widgets')
def docs_widgets():
    """Страница «Разработка виджетов»"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from pathlib import Path
    p = Path(__file__).parent / "docs" / "WIDGET_DEVELOPMENT.md"
    resp, err = _render_doc_page(p, "Разработка виджетов")
    if err:
        return err
    return resp


@app.route('/UNA.md/orasldev/docs/api')
def docs_api():
    """Страница «API документация»"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from pathlib import Path
    p = Path(__file__).parent / "docs" / "API.md"
    resp, err = _render_doc_page(p, "API документация")
    if err:
        return err
    return resp


@app.route('/UNA.md/orasldev/docs/easycredit')
def docs_easycredit():
    """Документация интеграции EasyCredit"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from pathlib import Path
    p = Path(__file__).parent / "docs" / "CREDITE" / "project_easycredit.html"
    if not p.exists():
        return "<h1>Документация не найдена</h1>", 404
    return p.read_text(encoding='utf-8')


@app.route('/UNA.md/orasldev/docs/iute')
def docs_iute():
    """Документация интеграции Iute"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from pathlib import Path
    p = Path(__file__).parent / "docs" / "CREDITE" / "project_iute.html"
    if not p.exists():
        return "<h1>Документация Iute не найдена</h1><p><a href='/UNA.md/orasldev/docs'>Назад</a></p>", 404
    return p.read_text(encoding='utf-8')


@app.route('/UNA.md/orasldev/docs/cred-reports')
def docs_cred_reports():
    """Документация настраиваемых отчётов"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from pathlib import Path
    p = Path(__file__).parent / "docs" / "CRED_REPORTS.md"
    resp, err = _render_doc_page(p, "Настраиваемые отчёты")
    if err:
        return err
    return resp


@app.route('/UNA.md/orasldev/docs/project-documentation')
def docs_project_documentation():
    """Полная документация проекта в HTML для передачи в Claude Code"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    p = Path(__file__).resolve().parent / "docs" / "PROJECT_DOCUMENTATION.html"
    if not p.is_file():
        return "<h1>Документация не найдена</h1><p>Файл docs/PROJECT_DOCUMENTATION.html отсутствует.</p>", 404
    return Response(p.read_text(encoding="utf-8"), mimetype="text/html; charset=utf-8")


@app.route('/UNA.md/orasldev/docs/developer-guide')
def docs_developer_guide():
    """Developer Guide v2.0 — архитектура, модули, Oracle ADB, deploy, howto"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    p = Path(__file__).resolve().parent / "docs" / "DEVELOPER_GUIDE.html"
    if not p.is_file():
        return "<h1>Developer Guide не найден</h1><p>Файл docs/DEVELOPER_GUIDE.html отсутствует.</p><p><a href='/UNA.md/orasldev/docs'>Назад</a></p>", 404
    return Response(p.read_text(encoding="utf-8"), mimetype="text/html; charset=utf-8")


@app.route('/UNA.md/orasldev/docs/sql')
def docs_sql():
    """Документация DDL скриптов"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from pathlib import Path
    p = Path(__file__).parent / "sql" / "README.md"
    if not p.exists():
        return "<h1>Документация SQL не найдена</h1><p><a href='/UNA.md/orasldev/docs'>Назад</a></p>", 404
    resp, err = _render_doc_page(p, "DDL скрипты")
    if err:
        return err
    return resp


# База для HTML-документов Nufarul (абсолютный путь)
_DOCS_NUFARUL_DIR = Path(__file__).resolve().parent / "docs" / "Nufarul"


@app.route('/UNA.md/orasldev/docs/nufarul')
@app.route('/UNA.md/orasldev/docs/nufarul/')
def docs_nufarul_index():
    """ТЗ Nufarul — список материалов (индекс)"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    p = _DOCS_NUFARUL_DIR / "index.html"
    if not p.is_file():
        return "<h1>Не найдено</h1><p><a href='/UNA.md/orasldev/docs'>Назад</a></p>", 404
    html = p.read_text(encoding='utf-8')
    return Response(html, mimetype='text/html; charset=utf-8')


# Разрешить скачивание/просмотр исходных файлов из docs/Nufarul (.doc, .xlsx, .pdf, .rtf, изображения)
_DOCS_NUFARUL_ALLOWED_EXT = {'.doc', '.docx', '.xlsx', '.xls', '.pdf', '.rtf', '.jpeg', '.jpg', '.png', '.gif'}


@app.route('/UNA.md/orasldev/docs/nufarul/file/<path:encoded_name>')
def docs_nufarul_download(encoded_name):
    """Скачать/открыть исходный файл из docs/Nufarul (имя в URL-кодировке)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from urllib.parse import unquote
    try:
        decoded = unquote(encoded_name, errors='strict')
    except Exception:
        return "<h1>Неверное имя файла</h1>", 400
    safe = Path(decoded).name
    if not safe or '..' in decoded or '/' in decoded or '\\' in decoded:
        return "<h1>Неверный путь</h1>", 400
    if Path(safe).suffix.lower() not in _DOCS_NUFARUL_ALLOWED_EXT:
        return "<h1>Тип файла не разрешён</h1>", 400
    p = (_DOCS_NUFARUL_DIR / safe).resolve()
    try:
        p.relative_to(_DOCS_NUFARUL_DIR.resolve())
    except ValueError:
        return "<h1>Неверный путь</h1>", 400
    if not p.is_file():
        return "<h1>Файл не найден</h1><p><a href='/UNA.md/orasldev/docs/nufarul/'>Назад</a></p>", 404
    from flask import send_file
    return send_file(
        p, as_attachment=False, download_name=safe,
        mimetype=None
    )


@app.route('/UNA.md/orasldev/docs/nufarul/view-xlsx/<path:encoded_name>')
def docs_nufarul_view_xlsx(encoded_name):
    """Просмотр .xlsx из docs/Nufarul в виде HTML-таблиц."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from urllib.parse import unquote
    try:
        decoded = unquote(encoded_name, errors='strict')
    except Exception:
        return "<h1>Неверное имя файла</h1>", 400
    safe = Path(decoded).name
    if not safe or Path(safe).suffix.lower() != '.xlsx':
        return "<h1>Только .xlsx</h1>", 400
    p = (_DOCS_NUFARUL_DIR / safe).resolve()
    try:
        p.relative_to(_DOCS_NUFARUL_DIR.resolve())
    except ValueError:
        return "<h1>Неверный путь</h1>", 400
    if not p.is_file():
        return "<h1>Файл не найден</h1><p><a href='/UNA.md/orasldev/docs/nufarul/'>Назад</a></p>", 404
    try:
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        return f"<h1>Ошибка чтения Excel</h1><p>{e}</p><p><a href='/UNA.md/orasldev/docs/nufarul/'>Назад</a></p>", 500
    from html import escape
    def _cell(v):
        if v is None:
            return ""
        return escape(str(v).strip())
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            parts.append(f"<h2>{escape(sheet_name)}</h2><p>Пустой лист</p>")
            continue
        parts.append(f"<h2>{escape(sheet_name)}</h2>")
        parts.append("<table border='1' cellpadding='6' cellspacing='0' style='border-collapse:collapse; width:100%;'>")
        for i, row in enumerate(rows):
            tag = "th" if i == 0 else "td"
            cells = "".join(f"<{tag}>{_cell(v)}</{tag}>" for v in (row or []))
            parts.append(f"<tr>{cells}</tr>")
        parts.append("</table>")
    wb.close()
    html = f"""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8"><title>{escape(safe)}</title>
<style>body{{font-family:Segoe UI,sans-serif;margin:24px;color:#134e4a;}} table{{margin-bottom:24px;}} th{{background:#0d9488;color:#fff;}}</style></head><body>
<h1>{escape(safe)}</h1>
{"".join(parts)}
<p><a href="/UNA.md/orasldev/docs/nufarul/">← К списку материалов</a></p>
</body></html>"""
    return Response(html, mimetype='text/html; charset=utf-8')


def _docs_nufarul_convert_doc_to_html(p: Path, safe: str) -> tuple[str | None, str | None]:
    """Конвертирует .doc/.docx в HTML. Возвращает (html_body, error_message)."""
    from html import escape
    suffix = p.suffix.lower()
    # 1) .docx через mammoth
    if suffix == '.docx':
        try:
            import mammoth
            with open(p, 'rb') as f:
                result = mammoth.convert_to_html(f)
            body = result.value or ''
            if not body.strip():
                body = '<p>Документ пуст или не удалось извлечь текст.</p>'
            return (f'<h1>{escape(safe)}</h1>{body}', None)
        except Exception as e:
            return (None, str(e))
    # 2) .doc — пробуем mammoth (иногда срабатывает), иначе LibreOffice
    if suffix == '.doc':
        try:
            import mammoth
            with open(p, 'rb') as f:
                result = mammoth.convert_to_html(f)
            body = (result.value or '').strip()
            if body:
                return (f'<h1>{escape(safe)}</h1>{body}', None)
        except Exception:
            pass
        # LibreOffice headless: soffice --headless --convert-to html --outdir <dir> <file>
        import subprocess
        import tempfile
        soffice = None
        for candidate in ('soffice', '/Applications/LibreOffice.app/Contents/MacOS/soffice', '/usr/bin/soffice'):
            try:
                r = subprocess.run([candidate, '--version'], capture_output=True, timeout=2)
                if r.returncode == 0:
                    soffice = candidate
                    break
            except (FileNotFoundError, subprocess.TimeoutExpired):
                continue
        if soffice:
            try:
                with tempfile.TemporaryDirectory() as tmp:
                    out_dir = Path(tmp)
                    r = subprocess.run(
                        [soffice, '--headless', '--convert-to', 'html', '--outdir', str(out_dir), str(p)],
                        capture_output=True, text=True, timeout=30, cwd=str(p.parent)
                    )
                    if r.returncode == 0:
                        # Ищем .html в tmp (имя без .doc + .html)
                        html_name = p.stem + '.html'
                        html_path = out_dir / html_name
                        if not html_path.exists():
                            for f in out_dir.glob('*.html'):
                                html_path = f
                                break
                        if html_path and html_path.is_file():
                            raw = html_path.read_text(encoding='utf-8', errors='replace')
                            # Внутреннее содержимое body (без тегов <body>...</body>)
                            if '<body' in raw and '</body>' in raw:
                                start = raw.find('<body')
                                start = raw.find('>', start) + 1
                                end = raw.find('</body>', start)
                                body = raw[start:end] if end > start else raw
                            else:
                                body = raw
                            return (f'<h1>{escape(safe)}</h1><div class="doc-body">{body}</div>', None)
            except Exception as e:
                return (None, str(e))
        return (None, 'Для просмотра .doc в браузере установите LibreOffice или скачайте файл и откройте в Word.')
    return (None, 'Поддерживаются только .doc и .docx')


@app.route('/UNA.md/orasldev/docs/nufarul/view-doc/<path:encoded_name>')
def docs_nufarul_view_doc(encoded_name):
    """Просмотр .doc/.docx из docs/Nufarul в виде HTML."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from urllib.parse import unquote, quote
    try:
        decoded = unquote(encoded_name, errors='strict')
    except Exception:
        return "<h1>Неверное имя файла</h1>", 400
    safe = Path(decoded).name
    if not safe or Path(safe).suffix.lower() not in ('.doc', '.docx'):
        return "<h1>Только .doc и .docx</h1>", 400
    p = (_DOCS_NUFARUL_DIR / safe).resolve()
    try:
        p.relative_to(_DOCS_NUFARUL_DIR.resolve())
    except ValueError:
        return "<h1>Неверный путь</h1>", 400
    if not p.is_file():
        return "<h1>Файл не найден</h1><p><a href='/UNA.md/orasldev/docs/nufarul/'>Назад</a></p>", 404
    from html import escape
    body_html, err = _docs_nufarul_convert_doc_to_html(p, safe)
    if err:
        back = '/UNA.md/orasldev/docs/nufarul/'
        file_url = f'/UNA.md/orasldev/docs/nufarul/file/{quote(safe, safe="")}'
        return Response(
            f"""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8"><title>{escape(safe)}</title>
<style>body{{font-family:Segoe UI,sans-serif;margin:24px;color:#134e4a;}} .err{{background:#fef2f2;padding:16px;border-radius:8px;}}</style></head><body>
<h1>{escape(safe)}</h1>
<p class="err">{escape(err)}</p>
<p><a href="{escape(file_url)}">Скачать исходный файл</a></p>
<p><a href="{back}">← К списку материалов</a></p>
</body></html>""",
            mimetype='text/html; charset=utf-8'
        )
    full_html = f"""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8"><title>{escape(safe)}</title>
<style>body{{font-family:Segoe UI,sans-serif;margin:24px;max-width:800px;color:#134e4a;}} .doc-body table{{border-collapse:collapse;}} .doc-body th,.doc-body td{{border:1px solid #99f6e4;padding:8px;}}</style></head><body>
{body_html}
<p><a href="/UNA.md/orasldev/docs/nufarul/">← К списку материалов</a></p>
</body></html>"""
    return Response(full_html, mimetype='text/html; charset=utf-8')


def _docs_nufarul_convert_pdf_to_html(p: Path, safe: str) -> tuple[str | None, str | None]:
    """Конвертирует PDF в HTML (текст/разметка через pdfminer.six). Возвращает (html_body, error_message)."""
    from html import escape
    try:
        from io import BytesIO
        from pdfminer.high_level import extract_text_to_fp
        from pdfminer.layout import LAParams
        out = BytesIO()
        with open(p, 'rb') as fin:
            extract_text_to_fp(fin, out, laparams=LAParams(), output_type='html', codec='utf-8')
        body = out.getvalue().decode('utf-8', errors='replace')
        if not (body and body.strip()):
            return (None, 'Не удалось извлечь текст из PDF.')
        # pdfminer HTML может содержать полную страницу или фрагмент — берём содержимое body
        if '<body' in body and '</body>' in body:
            start = body.find('<body')
            start = body.find('>', start) + 1
            end = body.find('</body>', start)
            body = body[start:end] if end > start else body
        return (f'<h1>{escape(safe)}</h1><div class="pdf-body">{body}</div>', None)
    except ImportError:
        return (None, 'Установите pdfminer.six: pip install pdfminer.six')
    except Exception as e:
        return (None, str(e))


@app.route('/UNA.md/orasldev/docs/nufarul/view-pdf/<path:encoded_name>')
def docs_nufarul_view_pdf(encoded_name):
    """Просмотр PDF из docs/Nufarul в виде HTML."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    from urllib.parse import unquote, quote
    try:
        decoded = unquote(encoded_name, errors='strict')
    except Exception:
        return "<h1>Неверное имя файла</h1>", 400
    safe = Path(decoded).name
    if not safe or Path(safe).suffix.lower() != '.pdf':
        return "<h1>Только .pdf</h1>", 400
    p = (_DOCS_NUFARUL_DIR / safe).resolve()
    try:
        p.relative_to(_DOCS_NUFARUL_DIR.resolve())
    except ValueError:
        return "<h1>Неверный путь</h1>", 400
    if not p.is_file():
        return "<h1>Файл не найден</h1><p><a href='/UNA.md/orasldev/docs/nufarul/'>Назад</a></p>", 404
    from html import escape
    body_html, err = _docs_nufarul_convert_pdf_to_html(p, safe)
    if err:
        back = '/UNA.md/orasldev/docs/nufarul/'
        file_url = f'/UNA.md/orasldev/docs/nufarul/file/{quote(safe, safe="")}'
        return Response(
            f"""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8"><title>{escape(safe)}</title>
<style>body{{font-family:Segoe UI,sans-serif;margin:24px;color:#134e4a;}} .err{{background:#fef2f2;padding:16px;border-radius:8px;}}</style></head><body>
<h1>{escape(safe)}</h1>
<p class="err">{escape(err)}</p>
<p><a href="{escape(file_url)}">Скачать исходный PDF</a></p>
<p><a href="{back}">← К списку материалов</a></p>
</body></html>""",
            mimetype='text/html; charset=utf-8'
        )
    full_html = f"""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8"><title>{escape(safe)}</title>
<style>body{{font-family:Segoe UI,sans-serif;margin:24px;max-width:900px;color:#134e4a;}} .pdf-body table{{border-collapse:collapse;}} .pdf-body th,.pdf-body td{{border:1px solid #99f6e4;padding:6px;}}</style></head><body>
{body_html}
<p><a href="/UNA.md/orasldev/docs/nufarul/">← К списку материалов</a></p>
</body></html>"""
    return Response(full_html, mimetype='text/html; charset=utf-8')


_DOCS_NUFARUL_JPG_DIR = _DOCS_NUFARUL_DIR / "docs_jpg"


@app.route('/UNA.md/orasldev/docs/nufarul/docs_jpg')
@app.route('/UNA.md/orasldev/docs/nufarul/docs_jpg/')
def docs_nufarul_docs_jpg_index():
    """Уточнение постановки задач — галерея JPG (docs_jpg)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    p = _DOCS_NUFARUL_JPG_DIR / "index.html"
    if not p.is_file():
        return "<h1>Не найдено</h1><p><a href='/UNA.md/orasldev/docs/nufarul/'>Назад</a></p>", 404
    return Response(p.read_text(encoding='utf-8'), mimetype='text/html; charset=utf-8')


@app.route('/UNA.md/orasldev/docs/nufarul/docs_jpg/<path:subpath>')
def docs_nufarul_docs_jpg_file(subpath):
    """Файлы из docs/Nufarul/docs_jpg (JPG, HTML и т.д.)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    safe = Path(subpath).name
    if not safe or ".." in subpath or subpath != subpath.strip():
        return "<h1>Неверный путь</h1>", 400
    p = (_DOCS_NUFARUL_JPG_DIR / subpath).resolve()
    try:
        p.relative_to(_DOCS_NUFARUL_JPG_DIR.resolve())
    except ValueError:
        return "<h1>Неверный путь</h1>", 400
    if not p.is_file():
        return "<h1>Не найдено</h1><p><a href='/UNA.md/orasldev/docs/nufarul/docs_jpg/'>Назад</a></p>", 404
    suffix = p.suffix.lower()
    if suffix == '.html':
        return Response(p.read_text(encoding='utf-8'), mimetype='text/html; charset=utf-8')
    if suffix in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
        from flask import send_file
        return send_file(p, mimetype=None, download_name=p.name)
    return "<h1>Тип файла не разрешён</h1>", 400


@app.route('/UNA.md/orasldev/docs/nufarul/<path:filename>')
def docs_nufarul_file(filename):
    """ТЗ Nufarul — HTML файлы (TZ.html, caiet_de_sacrini_2.html и т.д.)"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    safe = Path(filename).name
    if not safe or safe != filename.strip():
        return "<h1>Неверный путь</h1>", 400
    if Path(safe).suffix.lower() != '.html':
        return "<h1>Неверный тип файла</h1>", 400
    p = (_DOCS_NUFARUL_DIR / safe).resolve()
    try:
        p.relative_to(_DOCS_NUFARUL_DIR.resolve())
    except ValueError:
        return "<h1>Неверный путь</h1>", 400
    if not p.is_file():
        return "<h1>Не найдено</h1><p><a href='/UNA.md/orasldev/docs/nufarul/'>Назад</a></p>", 404
    html = p.read_text(encoding='utf-8')
    return Response(html, mimetype='text/html; charset=utf-8')


# База для HTML-документов AGRO (абсолютный путь)
_DOCS_AGRO_DIR = Path(__file__).resolve().parent / "docs" / "AGRO"


@app.route('/UNA.md/orasldev/docs/agro')
@app.route('/UNA.md/orasldev/docs/agro/')
def docs_agro_index():
    """Documentație AGRO — index."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    p = _DOCS_AGRO_DIR / "index.html"
    if not p.is_file():
        return "<h1>Nu s-a găsit</h1>", 404
    return Response(p.read_text(encoding='utf-8'), mimetype='text/html; charset=utf-8')


@app.route('/UNA.md/orasldev/docs/agro/<path:filename>')
def docs_agro_file(filename):
    """Documentație AGRO — fișiere HTML din docs/AGRO/ (inclusiv subdirectoare)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    if ".." in filename:
        return "<h1>Cale invalidă</h1>", 400
    p = (_DOCS_AGRO_DIR / filename).resolve()
    try:
        p.relative_to(_DOCS_AGRO_DIR.resolve())
    except ValueError:
        return "<h1>Cale invalidă</h1>", 400
    if not p.is_file():
        return "<h1>Nu s-a găsit</h1><p><a href='/UNA.md/orasldev/docs/agro/'>Înapoi</a></p>", 404
    suffix = p.suffix.lower()
    if suffix == '.html':
        return Response(p.read_text(encoding='utf-8'), mimetype='text/html; charset=utf-8')
    if suffix == '.xlsx':
        return _docs_agro_view_xlsx(p)
    return "<h1>Tip de fișier neacceptat</h1><p><a href='/UNA.md/orasldev/docs/agro/'>Înapoi</a></p>", 400


def _docs_agro_view_xlsx(p: Path):
    """Render .xlsx ca tabel HTML (similar cu Nufarul xlsx viewer)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(p, data_only=True)
        sheets_html = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            tbl = f'<h3 style="margin-top:24px;color:#0d9488;">{ws.title}</h3>'
            tbl += '<table style="width:100%;border-collapse:collapse;margin:8px 0 20px;">'
            for i, row in enumerate(rows):
                tag = 'th' if i == 0 else 'td'
                style_hdr = 'background:#0f172a;color:#fff;' if i == 0 else ''
                cells = ''.join(
                    f'<{tag} style="border:1px solid #d1d5db;padding:8px 10px;{style_hdr}">{v if v is not None else ""}</{tag}>'
                    for v in row
                )
                tbl += f'<tr>{cells}</tr>'
            tbl += '</table>'
            sheets_html.append(tbl)
        body = ''.join(sheets_html)
    except Exception as e:
        body = f'<p style="color:red;">Eroare la citirea fișierului: {e}</p>'
    html = f'''<!DOCTYPE html>
<html lang="ro"><head><meta charset="utf-8">
<title>{p.name}</title>
<style>body{{font-family:Arial,sans-serif;margin:32px;background:#f8fafc;color:#111827;line-height:1.45}}
.container{{max-width:1400px;margin:0 auto;background:#fff;padding:28px 36px;border:1px solid #e5e7eb;box-shadow:0 8px 24px rgba(0,0,0,.05)}}
h2{{color:#0f172a;font-size:22px;margin-bottom:6px}}
a.back{{display:inline-block;margin-bottom:16px;color:#0d9488;text-decoration:none;font-weight:600}}
a.back:hover{{text-decoration:underline}}</style></head>
<body><div class="container">
<a class="back" href="/UNA.md/orasldev/docs/agro/">&larr; Înapoi la documentație AGRO</a>
<h2>{p.name}</h2>
{body}
</div></body></html>'''
    return Response(html, mimetype='text/html; charset=utf-8')


# База для HTML-документов DECOR (абсолютный путь)
_DOCS_DECOR_DIR = Path(__file__).resolve().parent / "docs" / "DECOR"


@app.route('/UNA.md/orasldev/docs/decor')
@app.route('/UNA.md/orasldev/docs/decor/')
def docs_decor_index():
    """ТЗ DECOR — список материалов (индекс)"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    p = _DOCS_DECOR_DIR / "index.html"
    if not p.is_file():
        return "<h1>Не найдено</h1><p><a href='/UNA.md/orasldev/docs'>Назад</a></p>", 404
    return Response(p.read_text(encoding='utf-8'), mimetype='text/html; charset=utf-8')


@app.route('/UNA.md/orasldev/docs/decor/<path:filename>')
def docs_decor_file(filename):
    """ТЗ DECOR — файлы документации (HTML + вложенные ресурсы .fld/*)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    rel = Path(filename.strip())
    if not str(rel) or str(rel).startswith("/"):
        return "<h1>Неверный путь</h1>", 400
    p = (_DOCS_DECOR_DIR / rel).resolve()
    try:
        p.relative_to(_DOCS_DECOR_DIR.resolve())
    except ValueError:
        return "<h1>Неверный путь</h1>", 400
    if not p.is_file():
        return "<h1>Не найдено</h1><p><a href='/UNA.md/orasldev/docs/decor/'>Назад</a></p>", 404
    ext = p.suffix.lower()
    if ext in {".html", ".htm"}:
        return Response(p.read_text(encoding='utf-8'), mimetype='text/html; charset=utf-8')
    # Отдаём вложенные картинки/ресурсы конвертированных документов (.fld/*).
    return send_from_directory(str(_DOCS_DECOR_DIR), str(rel))


# ========== Shell: основное приложение без дочерних проектов, список проектов из таблицы ==========
SHELL_PREFIX = '/una.md/shell'
# Каталог статики приложения «Агенты» (Nuxt generate → копировать в static/agents/)
SHELL_AGENTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'agents')


@app.route(SHELL_PREFIX)
@app.route(SHELL_PREFIX + '/')
def shell_home():
    """Shell — главная: редирект на список проектов"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return redirect(url_for('shell_projects'))


@app.route(SHELL_PREFIX + '/projects')
def shell_projects():
    """Страница со списком ссылок на проекты: /una.md/shell/projects"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    links = ShellController.project_links(base_url=request.host_url.rstrip('/'))
    return render_template('shell_projects.html', project_links=links)


@app.route(SHELL_PREFIX + '/agents')
def shell_agents():
    """Приложение «Агенты» (Nuxt SPA) — редирект на слэш для корректных путей"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return redirect(SHELL_PREFIX + '/agents/')


@app.route(SHELL_PREFIX + '/agents/')
def shell_agents_index():
    """Приложение «Агенты» — index.html (статическая сборка в static/agents/)"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    if not os.path.isdir(SHELL_AGENTS_DIR):
        return "<h1>Агенты</h1><p>Сборка не найдена. Выполните в каталоге AI_v7: <code>npm run generate</code>, затем скопируйте <code>.output/public/</code> в <code>Artgranit/static/agents/</code>.</p>", 404
    return send_from_directory(SHELL_AGENTS_DIR, 'index.html')


@app.route(SHELL_PREFIX + '/agents/<path:path>')
def shell_agents_static(path):
    """Приложение «Агенты» — статика и SPA fallback (client-side routing)"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    if not os.path.isdir(SHELL_AGENTS_DIR):
        return "<h1>Агенты</h1><p>Сборка не найдена.</p>", 404
    file_path = os.path.join(SHELL_AGENTS_DIR, path)
    if os.path.isfile(file_path):
        return send_from_directory(SHELL_AGENTS_DIR, path)
    return send_from_directory(SHELL_AGENTS_DIR, 'index.html')


@app.route(SHELL_PREFIX + '/<project_slug>')
@app.route(SHELL_PREFIX + '/<project_slug>/')
def shell_project_home(project_slug):
    """Проект по slug: редирект на дашборд проекта"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    project = ShellController.get_project_by_slug(project_slug)
    if not project:
        return f"<h1>Проект не найден</h1><p>Slug: {project_slug}</p><a href='{url_for('shell_projects')}'>← Список проектов</a>", 404
    return redirect(url_for('shell_project_dashboard', project_slug=project_slug))


@app.route(SHELL_PREFIX + '/<project_slug>/dashboard')
@app.route(SHELL_PREFIX + '/<project_slug>/dashboard/')
@app.route(SHELL_PREFIX + '/<project_slug>/dashboard/<dashboard_id>')
def shell_project_dashboard(project_slug, dashboard_id=None):
    """Дашборд проекта: свой набор дашбордов под префиксом /una.md/shell/<slug>"""
    if not AuthController.is_authenticated():
        return _login_redirect()
    project = ShellController.get_project_by_slug(project_slug)
    if not project:
        return f"<h1>Проект не найден</h1><p>Slug: {project_slug}</p><a href='{url_for('shell_projects')}'>← Список проектов</a>", 404
    query_param = request.args.get('p')
    if query_param:
        dashboard_id = query_param
    is_fullscreen = dashboard_id is not None
    return render_template(
        'shell_dashboard_mdi.html',
        project_slug=project_slug,
        project_name=project.get('name', project_slug),
        dashboard_id=dashboard_id,
        is_fullscreen=is_fullscreen,
    )


@app.route('/api/shell/projects', methods=['GET'])
def api_shell_projects():
    """API: список проектов shell (для страницы /una.md/shell/projects)"""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    projects = ShellController.get_all_projects()
    return jsonify({"success": True, "projects": projects, "count": len(projects)})


@app.route('/api/shell/projects/<project_slug>', methods=['GET'])
def api_shell_project(project_slug):
    """API: один проект по slug"""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    project = ShellController.get_project_by_slug(project_slug)
    if not project:
        return jsonify({"success": False, "error": "Project not found"}), 404
    return jsonify({"success": True, "project": project})


@app.route('/api/test_connection')
def test_connection():
    """Тест подключения к БД (для test.html)"""
    try:
        from models.database import DatabaseModel
        import datetime
        start_time = time.time()
        
        with DatabaseModel() as db:
            with db.connection.cursor() as cursor:
                cursor.execute("SELECT 1 FROM DUAL")
                result = cursor.fetchone()
                
        duration = time.time() - start_time
        return jsonify({
            "success": True, 
            "message": f"Connected successfully! Result: {result[0]}",
            "duration": f"{duration:.3f}s",
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
    except Exception as e:
        return jsonify({
            "success": False, 
            "error": str(e)
        })


@app.route('/test.html')
def test_html():
    """Тестовая HTML страница"""
    return render_template('test.html')


# API Routes
@app.route('/api/login', methods=['POST'])
def api_login():
    """API endpoint для входа. Поддерживает next= для редиректа после входа."""
    data = request.get_json() or {}
    username = data.get('username', '')
    password = data.get('password', '')
    next_path = data.get('next', '').strip()
    
    if AuthController.login(username, password):
        AuthController.set_authenticated(True)
        # RO: pe instalarile dedicate destinatia implicita e backoffice-ul
        # EN: dedicated installs land in the backoffice by default
        redirect_url = ('/UNA.md/orasldev/biro26-backoffice'
                        if Config.LOGIN_EXCLUSIVE == 'biro26'
                        else url_for('sqldeveloper'))
        if next_path and _is_safe_redirect_url(next_path):
            redirect_url = next_path
        return jsonify({"success": True, "redirect": redirect_url})
    else:
        return jsonify({"success": False, "error": "Неверные учетные данные"}), 401


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """API endpoint для выхода"""
    AuthController.logout()
    return jsonify({"success": True})


@app.route('/api/status', methods=['GET'])
def api_status():
    """API endpoint для проверки статуса сервера"""
    return jsonify({
        "status": "running",
        "authenticated": AuthController.is_authenticated(),
        "username": AuthController.get_current_user() if AuthController.is_authenticated() else None,
        "timestamp": __import__('datetime').datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })


@app.route('/api/execute-sql', methods=['POST'])
def api_execute_sql():
    """API endpoint для выполнения SQL запросов"""
    try:
        if not AuthController.is_authenticated():
            return jsonify({"success": False, "error": "Authentication required"}), 401
        
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Invalid JSON in request body"}), 400
        
        sql_query = data.get('sql', '').strip()
        if not sql_query:
            return jsonify({"success": False, "error": "SQL query is empty"}), 400
        
        result = SQLController.execute(sql_query)
        # Убеждаемся, что результат всегда валидный JSON
        if not isinstance(result, dict):
            result = {"success": False, "error": "Invalid response format"}
        return jsonify(result)
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "traceback": error_trace
        }), 500


@app.route('/api/dashboard/metrics', methods=['GET'])
def api_dashboard_metrics():
    """API endpoint для получения всех метрик БД"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    result = DashboardController.get_all_metrics()
    return jsonify(result)


@app.route('/api/dashboard/metric/<metric_name>', methods=['GET'])
def api_dashboard_metric(metric_name):
    """API endpoint для получения конкретной метрики"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    result = DashboardController.get_metric(metric_name)
    return jsonify(result)


@app.route('/api/dashboard/list', methods=['GET'])
def api_dashboard_list():
    """API endpoint для получения списка доступных dashboard'ов. Для shell: ?project_slug=<slug>."""
    try:
        if not AuthController.is_authenticated():
            return jsonify({"success": False, "error": "Authentication required"}), 401
        
        project_slug = request.args.get('project_slug') or None
        result = DashboardController.get_dashboards_list(project_slug=project_slug)
        if not isinstance(result, dict):
            result = {"success": False, "error": "Invalid response format", "dashboards": [], "count": 0}
        return jsonify(result)
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "dashboards": [],
            "count": 0,
            "traceback": error_trace
        }), 500


@app.route('/api/dashboard/config/<dashboard_id>', methods=['GET'])
def api_dashboard_config(dashboard_id):
    """API endpoint для получения конфигурации dashboard'а по ID"""
    try:
        if not AuthController.is_authenticated():
            return jsonify({"success": False, "error": "Authentication required"}), 401
        
        result = DashboardController.get_dashboard_config(dashboard_id)
        if not isinstance(result, dict):
            result = {"success": False, "error": "Invalid response format"}
        return jsonify(result)
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "traceback": error_trace
        }), 500


@app.route('/api/dashboard/widget/custom-sql', methods=['POST'])
def api_dashboard_widget_custom_sql():
    """API endpoint для выполнения SQL запроса из custom_sql виджета"""
    try:
        if not AuthController.is_authenticated():
            return jsonify({"success": False, "error": "Authentication required"}), 401
        
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Invalid JSON in request body"}), 400
        
        database_type = data.get('database_type', 'oracle')
        sql_query = data.get('sql_query', '').strip()
        connection_params = data.get('connection_params', {})
        
        if not sql_query:
            return jsonify({"success": False, "error": "SQL query is empty"}), 400
        
        result = DashboardController.execute_custom_sql(database_type, sql_query, connection_params)
        if not isinstance(result, dict):
            result = {"success": False, "error": "Invalid response format"}
        return jsonify(result)
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "traceback": error_trace
        }), 500


@app.route('/api/objects/schemas', methods=['GET'])
def api_objects_schemas():
    """API endpoint для получения списка схем"""
    try:
        if not AuthController.is_authenticated():
            return jsonify({"success": False, "error": "Authentication required"}), 401
        
        result = ObjectsController.get_schemas()
        # Убеждаемся, что результат всегда валидный JSON
        if not isinstance(result, dict):
            result = {"success": False, "error": "Invalid response format", "schemas": [], "count": 0}
        return jsonify(result)
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "schemas": [],
            "count": 0,
            "traceback": error_trace
        }), 500


@app.route('/api/objects/tables', methods=['GET'])
def api_objects_tables():
    """API endpoint для получения списка таблиц"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_tables(schema)
    return jsonify(result)


@app.route('/api/objects/views', methods=['GET'])
def api_objects_views():
    """API endpoint для получения списка представлений"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_views(schema)
    return jsonify(result)


@app.route('/api/objects/procedures', methods=['GET'])
def api_objects_procedures():
    """API endpoint для получения списка процедур"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_procedures(schema)
    return jsonify(result)


@app.route('/api/objects/functions', methods=['GET'])
def api_objects_functions():
    """API endpoint для получения списка функций"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_functions(schema)
    return jsonify(result)


@app.route('/api/objects/packages', methods=['GET'])
def api_objects_packages():
    """API endpoint для получения списка пакетов"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_packages(schema)
    return jsonify(result)


@app.route('/api/objects/sequences', methods=['GET'])
def api_objects_sequences():
    """API endpoint для получения списка последовательностей"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_sequences(schema)
    return jsonify(result)


@app.route('/api/objects/synonyms', methods=['GET'])
def api_objects_synonyms():
    """API endpoint для получения списка синонимов"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_synonyms(schema)
    return jsonify(result)


@app.route('/api/objects/indexes', methods=['GET'])
def api_objects_indexes():
    """API endpoint для получения списка индексов"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_indexes(schema)
    return jsonify(result)


@app.route('/api/objects/triggers', methods=['GET'])
def api_objects_triggers():
    """API endpoint для получения списка триггеров"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_triggers(schema)
    return jsonify(result)


@app.route('/api/objects/types', methods=['GET'])
def api_objects_types():
    """API endpoint для получения списка типов"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_types(schema)
    return jsonify(result)


@app.route('/api/objects/materialized_views', methods=['GET'])
def api_objects_materialized_views():
    """API endpoint для получения списка материализованных представлений"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    schema = request.args.get('schema', None)
    result = ObjectsController.get_materialized_views(schema)
    return jsonify(result)


# ========== DIGI SM (Scale Management) Routes ==========

@app.route('/UNA.md/orasldev/digi-sm')
@app.route('/UNA.md/orasldev/digi-sm/')
def digi_sm():
    """Модуль управления весами DIGI SM"""
    if not AuthController.is_authenticated():
        return redirect(url_for('login'))
    return render_template('digi_marketing.html')

@app.route('/UNA.md/orasldev/digi-marketing')
@app.route('/UNA.md/orasldev/digi-marketing/')
def digi_marketing_redirect():
    return redirect('/UNA.md/orasldev/digi-sm')


# --- Dashboard & Stats ---
@app.route('/api/digi/stats', methods=['GET'])
def api_digi_stats():
    result = DigiMarketingController.get_dashboard_stats()
    return jsonify(result)


@app.route('/api/digi/events', methods=['GET'])
def api_digi_events():
    limit = request.args.get('limit', 100, type=int)
    entity_type = request.args.get('entity_type', None)
    result = DigiMarketingController.get_event_log(limit, entity_type)
    return jsonify(result)


@app.route('/api/digi/init-demo', methods=['POST'])
def api_digi_init_demo():
    result = DigiMarketingController.init_demo_data()
    return jsonify(result)


# --- Stores ---
@app.route('/api/digi/stores', methods=['GET'])
def api_digi_stores():
    result = DigiMarketingController.get_stores()
    return jsonify(result)


@app.route('/api/digi/stores/<store_id>', methods=['GET'])
def api_digi_store(store_id):
    result = DigiMarketingController.get_store(store_id)
    return jsonify(result)


@app.route('/api/digi/stores', methods=['POST'])
def api_digi_create_store():
    data = request.get_json()
    result = DigiMarketingController.create_store(data)
    return jsonify(result)


@app.route('/api/digi/stores/<store_id>', methods=['PUT'])
def api_digi_update_store(store_id):
    data = request.get_json()
    result = DigiMarketingController.update_store(store_id, data)
    return jsonify(result)


@app.route('/api/digi/stores/<store_id>', methods=['DELETE'])
def api_digi_delete_store(store_id):
    result = DigiMarketingController.delete_store(store_id)
    return jsonify(result)


# --- Departments ---
@app.route('/api/digi/departments', methods=['GET'])
def api_digi_departments():
    store_id = request.args.get('store_id', None)
    result = DigiMarketingController.get_departments(store_id)
    return jsonify(result)


@app.route('/api/digi/departments', methods=['POST'])
def api_digi_create_department():
    data = request.get_json()
    result = DigiMarketingController.create_department(data)
    return jsonify(result)


@app.route('/api/digi/departments/<dept_id>', methods=['DELETE'])
def api_digi_delete_department(dept_id):
    result = DigiMarketingController.delete_department(dept_id)
    return jsonify(result)


# --- Devices ---
@app.route('/api/digi/devices', methods=['GET'])
def api_digi_devices():
    store_id = request.args.get('store_id', None)
    department_id = request.args.get('department_id', None)
    result = DigiMarketingController.get_devices(store_id, department_id)
    return jsonify(result)


@app.route('/api/digi/devices/<device_id>', methods=['GET'])
def api_digi_device(device_id):
    result = DigiMarketingController.get_device(device_id)
    return jsonify(result)


@app.route('/api/digi/devices', methods=['POST'])
def api_digi_register_device():
    data = request.get_json()
    result = DigiMarketingController.register_device(data)
    return jsonify(result)


@app.route('/api/digi/devices/<device_id>', methods=['PUT'])
def api_digi_update_device(device_id):
    data = request.get_json()
    result = DigiMarketingController.update_device(device_id, data)
    return jsonify(result)


@app.route('/api/digi/devices/<device_id>', methods=['DELETE'])
def api_digi_delete_device(device_id):
    result = DigiMarketingController.delete_device(device_id)
    return jsonify(result)


# --- Media ---
@app.route('/api/digi/media', methods=['GET'])
def api_digi_media():
    media_type = request.args.get('type', None)
    resolution = request.args.get('resolution', None)
    result = DigiMarketingController.get_media_list(media_type, resolution)
    return jsonify(result)


@app.route('/api/digi/media/<media_id>', methods=['GET'])
def api_digi_media_item(media_id):
    result = DigiMarketingController.get_media(media_id)
    return jsonify(result)


@app.route('/api/digi/media', methods=['POST'])
def api_digi_upload_media():
    data = request.get_json()
    result = DigiMarketingController.upload_media(data)
    return jsonify(result)


@app.route('/api/digi/media/<media_id>', methods=['PUT'])
def api_digi_update_media(media_id):
    data = request.get_json()
    result = DigiMarketingController.update_media(media_id, data)
    return jsonify(result)


@app.route('/api/digi/media/<media_id>', methods=['DELETE'])
def api_digi_delete_media(media_id):
    result = DigiMarketingController.delete_media(media_id)
    return jsonify(result)


# --- Playlists ---
@app.route('/api/digi/playlists', methods=['GET'])
def api_digi_playlists():
    result = DigiMarketingController.get_playlists()
    return jsonify(result)


@app.route('/api/digi/playlists/<playlist_id>', methods=['GET'])
def api_digi_playlist(playlist_id):
    result = DigiMarketingController.get_playlist(playlist_id)
    return jsonify(result)


@app.route('/api/digi/playlists', methods=['POST'])
def api_digi_create_playlist():
    data = request.get_json()
    result = DigiMarketingController.create_playlist(data)
    return jsonify(result)


@app.route('/api/digi/playlists/<playlist_id>', methods=['PUT'])
def api_digi_update_playlist(playlist_id):
    data = request.get_json()
    result = DigiMarketingController.update_playlist(playlist_id, data)
    return jsonify(result)


@app.route('/api/digi/playlists/<playlist_id>', methods=['DELETE'])
def api_digi_delete_playlist(playlist_id):
    result = DigiMarketingController.delete_playlist(playlist_id)
    return jsonify(result)


# --- Campaigns ---
@app.route('/api/digi/campaigns', methods=['GET'])
def api_digi_campaigns():
    status = request.args.get('status', None)
    result = DigiMarketingController.get_campaigns(status)
    return jsonify(result)


@app.route('/api/digi/campaigns/<campaign_id>', methods=['GET'])
def api_digi_campaign(campaign_id):
    result = DigiMarketingController.get_campaign(campaign_id)
    return jsonify(result)


@app.route('/api/digi/campaigns', methods=['POST'])
def api_digi_create_campaign():
    data = request.get_json()
    result = DigiMarketingController.create_campaign(data)
    return jsonify(result)


@app.route('/api/digi/campaigns/<campaign_id>', methods=['PUT'])
def api_digi_update_campaign(campaign_id):
    data = request.get_json()
    result = DigiMarketingController.update_campaign(campaign_id, data)
    return jsonify(result)


@app.route('/api/digi/campaigns/<campaign_id>/publish', methods=['POST'])
def api_digi_publish_campaign(campaign_id):
    result = DigiMarketingController.publish_campaign(campaign_id)
    return jsonify(result)


@app.route('/api/digi/campaigns/<campaign_id>/pause', methods=['POST'])
def api_digi_pause_campaign(campaign_id):
    result = DigiMarketingController.pause_campaign(campaign_id)
    return jsonify(result)


@app.route('/api/digi/campaigns/<campaign_id>/stop', methods=['POST'])
def api_digi_stop_campaign(campaign_id):
    result = DigiMarketingController.stop_campaign(campaign_id)
    return jsonify(result)


@app.route('/api/digi/campaigns/<campaign_id>', methods=['DELETE'])
def api_digi_delete_campaign(campaign_id):
    result = DigiMarketingController.delete_campaign(campaign_id)
    return jsonify(result)


@app.route('/api/digi/campaigns/<campaign_id>/retry', methods=['POST'])
def api_digi_retry_campaign(campaign_id):
    device_id = request.args.get('device_id', None)
    result = DigiMarketingController.retry_delivery(campaign_id, device_id)
    return jsonify(result)


# --- Sync & Reports ---
@app.route('/api/digi/sync-log', methods=['GET'])
def api_digi_sync_log():
    campaign_id = request.args.get('campaign_id', None)
    device_id = request.args.get('device_id', None)
    limit = request.args.get('limit', 50, type=int)
    result = DigiMarketingController.get_sync_log(campaign_id, device_id, limit)
    return jsonify(result)


@app.route('/api/digi/delivery-report', methods=['GET'])
def api_digi_delivery_report():
    campaign_id = request.args.get('campaign_id', None)
    result = DigiMarketingController.get_delivery_report(campaign_id)
    return jsonify(result)


# --- Reference data ---
@app.route('/api/digi/ref/department-types', methods=['GET'])
def api_digi_department_types():
    return jsonify(DigiMarketingController.get_department_types())


@app.route('/api/digi/ref/device-types', methods=['GET'])
def api_digi_device_types():
    return jsonify(DigiMarketingController.get_device_types())


@app.route('/api/digi/ref/resolutions', methods=['GET'])
def api_digi_resolutions():
    return jsonify(DigiMarketingController.get_resolutions())


@app.route('/api/digi/ref/roles', methods=['GET'])
def api_digi_roles():
    return jsonify(DigiMarketingController.get_roles())


# ========== Хэш-инвайты: автологин по ссылке ?h=<hash> (INV_LINKS) ==========

@app.before_request
def _invite_autologin():
    """Ссылка вида /UNA.md/orasldev/tbcontrol?h=43hhjghj34g5jh345hj:
    хэш ищется в INV_LINKS, при успехе сессия логинится кредами инвайта."""
    try:
        if request.method != 'GET':
            return None
        inv_hash = request.args.get('h')
        if not inv_hash or AuthController.is_authenticated():
            return None
        if not request.path.startswith('/UNA.md/'):
            return None
        cred = TBControlController.resolve_invite(inv_hash)
        if cred and AuthController.login(cred['login'], cred['password']):
            AuthController.set_authenticated(True)
            session['username'] = cred['login']
            session['invite_login'] = True
            # Хэш убираем из адресной строки повторным заходом без ?h=
            clean = request.args.to_dict()
            clean.pop('h', None)
            from urllib.parse import urlencode
            qs = ('?' + urlencode(clean)) if clean else ''
            return redirect(request.path + qs)
    except Exception:
        pass
    return None


@app.route('/api/tbc/invites', methods=['GET'])
def api_tbc_invites():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(TBControlController.get_invites())


@app.route('/api/tbc/invites', methods=['POST'])
def api_tbc_create_invite():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(TBControlController.create_invite(request.get_json() or {}))


@app.route('/api/tbc/invites/<invite_id>', methods=['PUT'])
def api_tbc_update_invite(invite_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(TBControlController.update_invite(invite_id, request.get_json() or {}))


@app.route('/api/tbc/invites/<invite_id>', methods=['DELETE'])
def api_tbc_delete_invite(invite_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(TBControlController.delete_invite(invite_id))


# ========== TBControl (Front Office / POS / SCO / Android Operations) Routes ==========

@app.route('/UNA.md/orasldev/tbcontrol')
@app.route('/UNA.md/orasldev/tbcontrol/')
def tbcontrol():
    """Модуль TBControl — платформа контроля софта и оборудования магазинов"""
    if not AuthController.is_authenticated():
        return redirect(url_for('login'))
    return render_template('tbcontrol.html')


@app.route('/UNA.md/orasldev/tbcontrol/presentation')
def tbcontrol_presentation():
    """HTML-презентация TBControl с живыми ссылками на панели системы"""
    if not AuthController.is_authenticated():
        return redirect(url_for('login', next=request.path))
    from flask import send_file
    return send_file(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                  'docs', 'TBControl', 'presentation.html'))


_TBC_DOCS = {
    'INDEX': 'TBControl — полная документация',
    'TECHNICAL-OPS': 'TBControl — ТЗ (Technical Ops)',
    'TBCONTROL_MODULE': 'TBControl — справочник реализации',
    'SCENARIOS': 'TBControl — сценарии полезности',
    'PRESENTATION_GOOGLE_LM': 'TBControl — материал для NotebookLM',
}


@app.route('/UNA.md/orasldev/tbcontrol/docs')
@app.route('/UNA.md/orasldev/tbcontrol/docs/<name>')
def tbcontrol_docs(name='INDEX'):
    """Документация TBControl (MD → HTML); относительные ссылки между
    документами остаются рабочими."""
    if not AuthController.is_authenticated():
        return redirect(url_for('login', next=request.path))
    if name in ('presentation.html', 'presentation'):
        return redirect('/UNA.md/orasldev/tbcontrol/presentation')
    key = name[:-3] if name.endswith('.md') else name
    if key not in _TBC_DOCS:
        return "<h1>Документ не найден</h1>", 404
    md_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           'docs', 'TBControl', key + '.md')
    page, err = _render_doc_page(md_path, _TBC_DOCS[key])
    return page if page else err


# --- Dashboard & refs ---
@app.route('/api/tbc/stats', methods=['GET'])
def api_tbc_stats():
    return jsonify(TBControlController.get_dashboard_stats())


@app.route('/api/tbc/store-health', methods=['GET'])
def api_tbc_store_health():
    return jsonify(TBControlController.get_store_health())


@app.route('/api/tbc/refs', methods=['GET'])
def api_tbc_refs():
    return jsonify(TBControlController.get_refs())


@app.route('/api/tbc/init-demo', methods=['POST'])
def api_tbc_init_demo():
    return jsonify(TBControlController.init_demo_data())


@app.route('/api/tbc/audit', methods=['GET'])
def api_tbc_audit():
    limit = request.args.get('limit', 100, type=int)
    entity_type = request.args.get('entity_type', None)
    return jsonify(TBControlController.get_audit_log(limit, entity_type))


# --- Stores ---
@app.route('/api/tbc/stores', methods=['GET'])
def api_tbc_stores():
    return jsonify(TBControlController.get_stores())


@app.route('/api/tbc/stores', methods=['POST'])
def api_tbc_create_store():
    return jsonify(TBControlController.create_store(request.get_json() or {}))


@app.route('/api/tbc/stores/<store_id>', methods=['PUT'])
def api_tbc_update_store(store_id):
    return jsonify(TBControlController.update_store(store_id, request.get_json() or {}))


@app.route('/api/tbc/stores/<store_id>', methods=['DELETE'])
def api_tbc_delete_store(store_id):
    return jsonify(TBControlController.delete_store(store_id))


# --- Devices ---
@app.route('/api/tbc/devices', methods=['GET'])
def api_tbc_devices():
    return jsonify(TBControlController.get_devices(
        request.args.get('store_id'), request.args.get('device_type'), request.args.get('status')))


@app.route('/api/tbc/devices/<device_id>', methods=['GET'])
def api_tbc_device(device_id):
    return jsonify(TBControlController.get_device(device_id))


@app.route('/api/tbc/devices', methods=['POST'])
def api_tbc_register_device():
    return jsonify(TBControlController.register_device(request.get_json() or {}))


@app.route('/api/tbc/devices/<device_id>', methods=['PUT'])
def api_tbc_update_device(device_id):
    return jsonify(TBControlController.update_device(device_id, request.get_json() or {}))


@app.route('/api/tbc/devices/<device_id>', methods=['DELETE'])
def api_tbc_delete_device(device_id):
    return jsonify(TBControlController.delete_device(device_id))


@app.route('/api/tbc/devices/<device_id>/diagnostics', methods=['POST'])
def api_tbc_device_diagnostics(device_id):
    return jsonify(TBControlController.run_diagnostics(device_id))


# --- Agent Heartbeat (Zabbix Agent 2 / Android Monitoring Agent) ---
@app.route('/api/tbc/agent/heartbeat', methods=['POST'])
def api_tbc_agent_heartbeat():
    return jsonify(TBControlController.agent_heartbeat(request.get_json() or {}))


# --- Applications & versions ---
@app.route('/api/tbc/applications', methods=['GET'])
def api_tbc_applications():
    return jsonify(TBControlController.get_applications())


@app.route('/api/tbc/applications', methods=['POST'])
def api_tbc_create_application():
    return jsonify(TBControlController.create_application(request.get_json() or {}))


@app.route('/api/tbc/applications/<app_id>', methods=['PUT'])
def api_tbc_update_application(app_id):
    return jsonify(TBControlController.update_application(app_id, request.get_json() or {}))


@app.route('/api/tbc/applications/<app_id>', methods=['DELETE'])
def api_tbc_delete_application(app_id):
    return jsonify(TBControlController.delete_application(app_id))


@app.route('/api/tbc/versions', methods=['GET'])
def api_tbc_versions():
    return jsonify(TBControlController.get_versions(
        request.args.get('app_id'), request.args.get('status')))


# --- Events ---
@app.route('/api/tbc/events', methods=['GET'])
def api_tbc_events():
    return jsonify(TBControlController.get_events(
        request.args.get('status'), request.args.get('severity'),
        request.args.get('store_id'), request.args.get('limit', 200, type=int)))


@app.route('/api/tbc/events', methods=['POST'])
def api_tbc_create_event():
    return jsonify(TBControlController.create_event(request.get_json() or {}))


@app.route('/api/tbc/events/<event_id>/ack', methods=['POST'])
def api_tbc_ack_event(event_id):
    return jsonify(TBControlController.set_event_status(event_id, 'ack'))


@app.route('/api/tbc/events/<event_id>/resolve', methods=['POST'])
def api_tbc_resolve_event(event_id):
    return jsonify(TBControlController.set_event_status(event_id, 'resolved'))


@app.route('/api/tbc/events/<event_id>/incident', methods=['POST'])
def api_tbc_event_to_incident(event_id):
    return jsonify(TBControlController.create_incident_from_event(event_id, request.get_json() or {}))


# --- Incidents ---
@app.route('/api/tbc/incidents', methods=['GET'])
def api_tbc_incidents():
    return jsonify(TBControlController.get_incidents(
        request.args.get('status'), request.args.get('severity'),
        request.args.get('limit', 200, type=int)))


@app.route('/api/tbc/incidents/<incident_id>', methods=['PUT'])
def api_tbc_update_incident(incident_id):
    return jsonify(TBControlController.update_incident(incident_id, request.get_json() or {}))


# --- Changes / deployment ---
@app.route('/api/tbc/changes', methods=['GET'])
def api_tbc_changes():
    return jsonify(TBControlController.get_changes(request.args.get('status')))


@app.route('/api/tbc/changes/<change_id>', methods=['GET'])
def api_tbc_change(change_id):
    return jsonify(TBControlController.get_change(change_id))


@app.route('/api/tbc/changes', methods=['POST'])
def api_tbc_create_change():
    return jsonify(TBControlController.create_change(request.get_json() or {}))


@app.route('/api/tbc/changes/<change_id>/deploy', methods=['POST'])
def api_tbc_deploy_change(change_id):
    return jsonify(TBControlController.deploy_change(change_id))


@app.route('/api/tbc/changes/<change_id>/rollback', methods=['POST'])
def api_tbc_rollback_change(change_id):
    return jsonify(TBControlController.rollback_change(change_id))


# --- Monitoring Center (раздел 72 ТЗ) ---
@app.route('/api/tbc/monitor/overview', methods=['GET'])
def api_tbc_monitor_overview():
    return jsonify(TBControlController.monitor_overview(
        request.args.get('store_id'), request.args.get('device_type')))


@app.route('/api/tbc/monitor/series/<device_id>', methods=['GET'])
def api_tbc_monitor_series(device_id):
    return jsonify(TBControlController.monitor_series(
        device_id, request.args.get('scope', 'hw'),
        request.args.get('from'), request.args.get('to'),
        request.args.get('bucket', 'hour')))


# --- Processing Center (раздел 73 ТЗ) ---
@app.route('/api/tbc/proc/stats', methods=['GET'])
def api_tbc_proc_stats():
    return jsonify(TBControlController.get_proc_stats())


@app.route('/api/tbc/nodes', methods=['GET'])
def api_tbc_nodes():
    return jsonify(TBControlController.get_nodes(request.args.get('node_type')))


@app.route('/api/tbc/nodes', methods=['POST'])
def api_tbc_create_node():
    return jsonify(TBControlController.create_node(request.get_json() or {}))


@app.route('/api/tbc/nodes/heartbeat', methods=['POST'])
def api_tbc_node_heartbeat():
    return jsonify(TBControlController.node_heartbeat(request.get_json() or {}))


@app.route('/api/tbc/flows', methods=['GET'])
def api_tbc_flows():
    return jsonify(TBControlController.get_flows(
        request.args.get('status'), request.args.get('store_id')))


@app.route('/api/tbc/flows/<flow_id>/log', methods=['GET'])
def api_tbc_flow_log(flow_id):
    return jsonify(TBControlController.get_flow_log(flow_id, request.args.get('limit', 50, type=int)))


@app.route('/api/tbc/flows/<flow_id>/report', methods=['POST'])
def api_tbc_flow_report(flow_id):
    return jsonify(TBControlController.flow_report(flow_id, request.get_json() or {}))


@app.route('/api/tbc/flows/<flow_id>/retry', methods=['POST'])
def api_tbc_flow_retry(flow_id):
    return jsonify(TBControlController.retry_flow(flow_id))


# --- Действия персонала / тикеты / климат / отчёт ---
@app.route('/api/tbc/actions', methods=['GET'])
def api_tbc_actions():
    return jsonify(TBControlController.get_actions(
        request.args.get('store_id'), request.args.get('unjustified'),
        request.args.get('limit', 200, type=int)))


@app.route('/api/tbc/actions', methods=['POST'])
def api_tbc_create_action():
    return jsonify(TBControlController.create_action(request.get_json() or {}))


@app.route('/api/tbc/tickets', methods=['GET'])
def api_tbc_tickets():
    return jsonify(TBControlController.get_tickets(
        request.args.get('target'), request.args.get('status'),
        request.args.get('limit', 200, type=int)))


@app.route('/api/tbc/tickets', methods=['POST'])
def api_tbc_create_ticket():
    return jsonify(TBControlController.create_ticket(request.get_json() or {}))


@app.route('/api/tbc/tickets/<ticket_id>', methods=['PUT'])
def api_tbc_update_ticket(ticket_id):
    return jsonify(TBControlController.update_ticket(ticket_id, request.get_json() or {}))


@app.route('/api/tbc/env/report', methods=['POST'])
def api_tbc_env_report():
    return jsonify(TBControlController.env_report(request.get_json() or {}))


@app.route('/api/tbc/settings', methods=['GET'])
def api_tbc_settings():
    return jsonify(TBControlController.get_settings())


@app.route('/api/tbc/settings', methods=['PUT'])
def api_tbc_save_settings():
    return jsonify(TBControlController.save_settings(request.get_json() or {}))


# --- Эмулятор сценариев / Zabbix-коннектор (tbc_emulator.py) ---
@app.route('/api/tbc/emulator/status', methods=['GET'])
def api_tbc_emulator_status():
    from tbc_emulator import RUNTIME
    return jsonify({"success": True, "data": RUNTIME.status()})


@app.route('/api/tbc/emulator/start', methods=['POST'])
def api_tbc_emulator_start():
    from tbc_emulator import RUNTIME
    data = request.get_json() or {}
    mode = data.get('mode', 'emulator')
    interval = int(data.get('interval') or TBControlController.get_setting_raw('emulator_interval') or 60)
    zbx_url = data.get('zabbix_url') or ''
    zbx_token = data.get('zabbix_token') or ''
    zbx_user = data.get('zabbix_user') or ''
    zbx_password = data.get('zabbix_password') or ''
    # Сохраняем конфигурацию (маскированные секреты не перезапишутся)
    TBControlController.save_settings({'emulator_interval': interval,
                                       'zabbix_url': zbx_url or None,
                                       'zabbix_token': zbx_token or None,
                                       'zabbix_user': zbx_user or None,
                                       'zabbix_password': zbx_password or None})
    if mode == 'zabbix':
        zbx_url = (zbx_url or TBControlController.get_setting_raw('zabbix_url') or '').strip()
        if zbx_token.endswith('***') or not zbx_token:
            zbx_token = TBControlController.get_setting_raw('zabbix_token') or ''
        if zbx_password.endswith('***') or not zbx_password:
            zbx_password = TBControlController.get_setting_raw('zabbix_password') or ''
        zbx_user = zbx_user or TBControlController.get_setting_raw('zabbix_user') or ''
        if not zbx_url or not (zbx_token or (zbx_user and zbx_password)):
            return jsonify({"success": False,
                            "error": "Укажите Zabbix URL и API token (5.4+) либо логин/пароль (3.x/4.x)"})
    result = RUNTIME.start(
        mode=mode,
        base_url=f'http://127.0.0.1:{Config.SERVER_PORT}',
        username=Config.DEFAULT_USERNAME, password=Config.DEFAULT_PASSWORD,
        interval=max(15, interval), zabbix_url=zbx_url, zabbix_token=zbx_token,
        zabbix_user=zbx_user, zabbix_password=zbx_password)
    if result.get('success'):
        TBControlController._add_audit('start', 'emulator', None, f'Запущен режим {mode}, интервал {interval}с')
    return jsonify(result)


@app.route('/api/tbc/emulator/stop', methods=['POST'])
def api_tbc_emulator_stop():
    from tbc_emulator import RUNTIME
    result = RUNTIME.stop()
    TBControlController._add_audit('stop', 'emulator', None, 'Эмулятор/коннектор остановлен')
    return jsonify(result)


@app.route('/api/tbc/env/series', methods=['GET'])
def api_tbc_env_series():
    return jsonify(TBControlController.env_series(
        request.args.get('store_id'), request.args.get('node_id'),
        request.args.get('metric'), request.args.get('hours', 48, type=int)))


@app.route('/api/tbc/report/ops', methods=['GET'])
def api_tbc_report_ops():
    return jsonify(TBControlController.report_ops(request.args.get('days', 14, type=int)))


# --- AI Diagnostic Dossiers (раздел 74 ТЗ) ---
@app.route('/api/tbc/ai/dossiers', methods=['GET'])
def api_tbc_dossiers():
    return jsonify(TBControlController.get_dossiers(request.args.get('limit', 100, type=int)))


@app.route('/api/tbc/ai/dossiers/generate', methods=['POST'])
def api_tbc_generate_dossier():
    data = request.get_json() or {}
    return jsonify(TBControlController.generate_dossier(
        data.get('source_type', 'event'), data.get('ref_id', 0)))


@app.route('/api/tbc/ai/dossiers/<dossier_id>', methods=['PUT'])
def api_tbc_update_dossier(dossier_id):
    return jsonify(TBControlController.update_dossier(dossier_id, request.get_json() or {}))


@app.route('/api/tbc/ai/dossier/<code>.md', methods=['GET'])
def api_tbc_dossier_md(code):
    """Выдача MD-досье внешнему AI-провайдеру по секретному токену
    (или авторизованному пользователю UI)."""
    result = TBControlController.get_dossier_md(
        code, request.args.get('token'),
        authenticated=AuthController.is_authenticated())
    if not result.get('success'):
        return jsonify(result), result.get('status', 400)
    from flask import Response
    return Response(result['md'], mimetype='text/markdown; charset=utf-8')


# --- SLA ---
@app.route('/api/tbc/sla', methods=['GET'])
def api_tbc_sla():
    return jsonify(TBControlController.get_sla())


@app.route('/api/tbc/sla/<sla_id>', methods=['PUT'])
def api_tbc_update_sla(sla_id):
    return jsonify(TBControlController.update_sla(sla_id, request.get_json() or {}))


# ========== Планограммы (Planograms) Routes ==========

def _plg_lang():
    """Язык интерфейса модуля планограмм: ?lang=ru|ro|en."""
    return PlanogramController.lang(request.args.get('lang'))


@app.route('/UNA.md/orasldev/planograms')
@app.route('/UNA.md/orasldev/planograms/')
def planograms():
    """Модуль «Планограммы» — выкладка товара, зоны зала, проходимость.

    Без входа модуль открывается в демо-режиме: все разделы видны, но любые
    изменения запрещены (см. _plg_block_anonymous_writes). Это нужно, чтобы
    живые ссылки из презентации и документации вели в работающую систему,
    а не на форму входа. Данные модуля — синтетический тестовый набор.
    """
    return render_template('planograms.html',
                           plg_demo=not AuthController.is_authenticated())


@app.before_request
def _plg_block_anonymous_writes():
    """Демо-режим модуля планограмм — только чтение.

    Страница модуля открыта анонимно, поэтому запись должна быть закрыта на
    сервере, а не только скрытием кнопок в интерфейсе: иначе демо-набор
    испортит любой, кто отправит POST руками.
    """
    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return None
    if not request.path.startswith('/api/plg/'):
        return None
    if request.path.startswith('/api/plg/mobile/'):
        # У мобильного контура своя авторизация — токен устройства.
        # Проверяет её сам маршрут, здесь пропускаем.
        return None
    if AuthController.is_authenticated():
        return None
    return jsonify({
        'success': False,
        'demo': True,
        'error': 'Демо-режим: изменение данных доступно после входа в систему',
    }), 403


# --- Публичная документация модуля ---
#
# Реестр документов. Флаг public=True открывает документ без входа в систему.
# Руководство пользователя и презентация показывают только работу с интерфейсом
# и потому открыты. Техническое описание и инструкция по установке содержат
# пути на сервере, размещение wallet, имя systemd-юнита и перечень ключей
# окружения — их анонимный доступ был бы разведданными для атакующего,
# поэтому они закрыты входом. Чтобы открыть их тоже, достаточно поставить
# 'public': True в соответствующей строке.
PLG_DOCS = [
    {'slug': 'user-guide', 'file': 'USER_GUIDE.md', 'public': True,
     'icon': '📘', 'cls': 'g', 'title': 'Руководство пользователя',
     'audience': 'для сотрудников сети',
     'descr': 'Как работать в модуле: карта зала, планограммы и согласование, '
              'прогноз заказов, диаграмма Ганта, поставщики, конкуренты, рынки. '
              'Пять типовых сценариев и пояснения, как читать цвета и метрики.'},
    {'slug': 'module', 'file': 'PLANOGRAMS_MODULE.md', 'public': False,
     'icon': '⚙', 'cls': '', 'title': 'Техническое описание',
     'audience': 'для разработчиков',
     'descr': 'Модель данных, представления, API, алгоритмы генерации и прогноза, '
              'инженерные решения и найденные при тестировании дефекты.'},
    {'slug': 'install', 'file': 'INSTALL.md', 'public': False,
     'icon': '⬇', 'cls': 'c', 'title': 'Установка и развёртывание',
     'audience': 'для администраторов',
     'descr': 'Все скрипты установки, порядок SQL-файлов, флаги деплоя, '
              'конфигурация окружения и обязательные проверки после релиза.'},
    {'slug': 'auto-order', 'file': 'AUTO_ORDER_GUIDE.md', 'public': True,
     'icon': '∿', 'cls': 'v', 'title': 'Методичка по автозаказу',
     'audience': 'для категорийного менеджера и закупщика',
     'descr': 'Все алгоритмы автозаказа со скриншотами и ссылками в работающую '
              'систему, фреш через РЦ и прямой поставкой, голосовой дозаказ, '
              'дорожная карта ИИ, векторные возможности Oracle 26ai '
              'и сравнение методологических подходов.'},
    {'slug': 'presentation-plan', 'file': 'PRESENTATION_PLAN.md', 'public': True,
     'icon': '◐', 'cls': 'w', 'title': 'План презентации',
     'audience': 'для докладчика',
     'descr': 'Сценарий показа на 25 минут: структура слайдов с таймингом, '
              'что демонстрировать вживую, ожидаемые вопросы и план пилота.'},
]

PLG_DOCS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'docs', 'Planograms')


def _plg_doc_by_slug(slug):
    return next((d for d in PLG_DOCS if d['slug'] == slug), None)


@app.route('/UNA.md/orasldev/planograms/docs')
@app.route('/UNA.md/orasldev/planograms/docs/')
def planograms_docs_index():
    """Хаб документации модуля — открыт без входа."""
    return render_template('planograms_docs.html', docs=PLG_DOCS, doc=None,
                           page_title='Планограммы — документация модуля')


@app.route('/UNA.md/orasldev/planograms/docs/<slug>')
def planograms_doc(slug):
    """Отдельный документ модуля, отрендеренный из Markdown."""
    doc = _plg_doc_by_slug(slug)
    if not doc:
        return render_template('planograms_docs.html', docs=PLG_DOCS, doc=None,
                               page_title='Документ не найден'), 404
    if not doc['public'] and not AuthController.is_authenticated():
        return redirect(url_for('login'))

    path = os.path.join(PLG_DOCS_DIR, doc['file'])
    if not os.path.isfile(path):
        return render_template('planograms_docs.html', docs=PLG_DOCS, doc=None,
                               page_title='Документ не найден'), 404
    with open(path, 'r', encoding='utf-8') as f:
        source = f.read()

    # Ссылки между файлами переписываем на маршруты приложения,
    # иначе из браузера они ведут в никуда.
    for other in PLG_DOCS:
        source = source.replace(f"]({other['file']})",
                                f"](/UNA.md/orasldev/planograms/docs/{other['slug']})")
    source = source.replace('](presentation.html)',
                            '](/UNA.md/orasldev/planograms/presentation)')

    return render_template('planograms_docs.html', docs=PLG_DOCS, doc=doc,
                           content=_docs_md_to_html(source),
                           page_title=f"{doc['title']} — Планограммы")


@app.route('/UNA.md/orasldev/planograms/presentation')
def planograms_presentation():
    """Презентация модуля — самостоятельная HTML-страница, открыта без входа."""
    path = os.path.join(PLG_DOCS_DIR, 'presentation.html')
    if not os.path.isfile(path):
        return '<h1>Презентация не найдена</h1>', 404
    from flask import Response
    with open(path, 'r', encoding='utf-8') as f:
        return Response(f.read(), mimetype='text/html; charset=utf-8')


# --- Язык, словарь, справочники ---
@app.route('/api/plg/langs', methods=['GET'])
def api_plg_langs():
    return jsonify(PlanogramController.get_langs(_plg_lang()))


@app.route('/api/plg/i18n', methods=['GET'])
def api_plg_i18n():
    return jsonify(PlanogramController.get_i18n(_plg_lang()))


@app.route('/api/plg/refs', methods=['GET'])
def api_plg_refs():
    return jsonify(PlanogramController.get_refs(_plg_lang()))


@app.route('/api/plg/stores', methods=['GET'])
def api_plg_stores():
    return jsonify(PlanogramController.get_stores(
        _plg_lang(), request.args.get('dataset_id', type=int)))


# --- Дашборд и карта ---
@app.route('/api/plg/dashboard', methods=['GET'])
def api_plg_dashboard():
    return jsonify(PlanogramController.get_dashboard(
        request.args.get('store_id', type=int), _plg_lang(),
        request.args.get('days', 14, type=int)))


@app.route('/api/plg/map', methods=['GET'])
def api_plg_map():
    return jsonify(PlanogramController.get_store_map(
        request.args.get('store_id', type=int), _plg_lang()))


@app.route('/api/plg/analytics', methods=['GET'])
def api_plg_analytics():
    return jsonify(PlanogramController.get_analytics(
        request.args.get('store_id', type=int),
        request.args.get('days', 14, type=int), _plg_lang()))


# --- Зоны ---
@app.route('/api/plg/zones', methods=['GET'])
def api_plg_zones():
    return jsonify(PlanogramController.get_zones(
        request.args.get('store_id', type=int), _plg_lang()))


@app.route('/api/plg/zones', methods=['POST'])
def api_plg_create_zone():
    return jsonify(PlanogramController.save_zone(request.get_json() or {}))


@app.route('/api/plg/zones/<int:zone_id>', methods=['PUT'])
def api_plg_update_zone(zone_id):
    return jsonify(PlanogramController.save_zone(request.get_json() or {}, zone_id))


@app.route('/api/plg/zones/<int:zone_id>', methods=['DELETE'])
def api_plg_delete_zone(zone_id):
    return jsonify(PlanogramController.delete_zone(zone_id))


# --- Оборудование ---
@app.route('/api/plg/fixtures', methods=['GET'])
def api_plg_fixtures():
    return jsonify(PlanogramController.get_fixtures(
        request.args.get('store_id', type=int),
        request.args.get('zone_id', type=int), _plg_lang()))


@app.route('/api/plg/fixtures', methods=['POST'])
def api_plg_create_fixture():
    return jsonify(PlanogramController.save_fixture(request.get_json() or {}))


@app.route('/api/plg/fixtures/<int:fixture_id>', methods=['PUT'])
def api_plg_update_fixture(fixture_id):
    return jsonify(PlanogramController.save_fixture(request.get_json() or {}, fixture_id))


@app.route('/api/plg/fixtures/<int:fixture_id>', methods=['DELETE'])
def api_plg_delete_fixture(fixture_id):
    return jsonify(PlanogramController.delete_fixture(fixture_id))


# --- Товары ---
@app.route('/api/plg/products', methods=['GET'])
def api_plg_products():
    return jsonify(PlanogramController.get_products(
        request.args.get('category_id', type=int),
        request.args.get('q'), _plg_lang()))


@app.route('/api/plg/products', methods=['POST'])
def api_plg_create_product():
    return jsonify(PlanogramController.save_product(request.get_json() or {}))


@app.route('/api/plg/products/<int:product_id>', methods=['PUT'])
def api_plg_update_product(product_id):
    return jsonify(PlanogramController.save_product(request.get_json() or {}, product_id))


@app.route('/api/plg/products/<int:product_id>', methods=['DELETE'])
def api_plg_delete_product(product_id):
    return jsonify(PlanogramController.delete_product(product_id))


# --- Планограммы ---
@app.route('/api/plg/planograms', methods=['GET'])
def api_plg_planograms():
    return jsonify(PlanogramController.get_planograms(
        request.args.get('store_id', type=int), request.args.get('status'),
        request.args.get('zone_id', type=int), _plg_lang()))


@app.route('/api/plg/planograms/<int:planogram_id>', methods=['GET'])
def api_plg_planogram(planogram_id):
    return jsonify(PlanogramController.get_planogram(planogram_id, _plg_lang()))


@app.route('/api/plg/planograms', methods=['POST'])
def api_plg_create_planogram():
    return jsonify(PlanogramController.save_planogram(request.get_json() or {}))


@app.route('/api/plg/planograms/<int:planogram_id>', methods=['PUT'])
def api_plg_update_planogram(planogram_id):
    return jsonify(PlanogramController.save_planogram(request.get_json() or {}, planogram_id))


@app.route('/api/plg/planograms/<int:planogram_id>', methods=['DELETE'])
def api_plg_delete_planogram(planogram_id):
    return jsonify(PlanogramController.delete_planogram(planogram_id))


@app.route('/api/plg/planograms/<int:planogram_id>/status', methods=['POST'])
def api_plg_planogram_status(planogram_id):
    data = request.get_json() or {}
    return jsonify(PlanogramController.set_planogram_status(planogram_id, data.get('status', '')))


# --- Позиции планограммы ---
@app.route('/api/plg/planograms/<int:planogram_id>/items', methods=['POST'])
def api_plg_create_item(planogram_id):
    return jsonify(PlanogramController.save_planogram_item(planogram_id, request.get_json() or {}))


@app.route('/api/plg/planograms/<int:planogram_id>/items/<int:item_id>', methods=['PUT'])
def api_plg_update_item(planogram_id, item_id):
    return jsonify(PlanogramController.save_planogram_item(
        planogram_id, request.get_json() or {}, item_id))


@app.route('/api/plg/items/<int:item_id>', methods=['DELETE'])
def api_plg_delete_item(item_id):
    return jsonify(PlanogramController.delete_planogram_item(item_id))


# --- История изменений ---
@app.route('/api/plg/history', methods=['GET'])
def api_plg_history():
    return jsonify(PlanogramController.get_history(
        request.args.get('store_id', type=int),
        request.args.get('planogram_id', type=int),
        request.args.get('limit', 200, type=int), _plg_lang()))


# --- Акции ---
@app.route('/api/plg/promos', methods=['GET'])
def api_plg_promos():
    return jsonify(PlanogramController.get_promos(
        request.args.get('store_id', type=int),
        request.args.get('active') == '1', _plg_lang()))


@app.route('/api/plg/promos', methods=['POST'])
def api_plg_create_promo():
    return jsonify(PlanogramController.save_promo(request.get_json() or {}))


@app.route('/api/plg/promos/<int:promo_id>', methods=['PUT'])
def api_plg_update_promo(promo_id):
    return jsonify(PlanogramController.save_promo(request.get_json() or {}, promo_id))


@app.route('/api/plg/promos/<int:promo_id>', methods=['DELETE'])
def api_plg_delete_promo(promo_id):
    return jsonify(PlanogramController.delete_promo(promo_id))


# --- Задачи ---
@app.route('/api/plg/tasks', methods=['GET'])
def api_plg_tasks():
    return jsonify(PlanogramController.get_tasks(
        request.args.get('store_id', type=int), request.args.get('status'), _plg_lang()))


@app.route('/api/plg/tasks', methods=['POST'])
def api_plg_create_task():
    return jsonify(PlanogramController.save_task(request.get_json() or {}))


@app.route('/api/plg/tasks/<int:task_id>', methods=['PUT'])
def api_plg_update_task(task_id):
    return jsonify(PlanogramController.save_task(request.get_json() or {}, task_id))


@app.route('/api/plg/tasks/<int:task_id>', methods=['DELETE'])
def api_plg_delete_task(task_id):
    return jsonify(PlanogramController.delete_task(task_id))


# --- Документы ---
@app.route('/api/plg/documents', methods=['GET'])
def api_plg_documents():
    return jsonify(PlanogramController.get_documents(
        request.args.get('store_id', type=int),
        request.args.get('planogram_id', type=int), _plg_lang()))


@app.route('/api/plg/documents', methods=['POST'])
def api_plg_create_document():
    return jsonify(PlanogramController.save_document(request.get_json() or {}))


@app.route('/api/plg/documents/<int:document_id>', methods=['PUT'])
def api_plg_update_document(document_id):
    return jsonify(PlanogramController.save_document(request.get_json() or {}, document_id))


@app.route('/api/plg/documents/<int:document_id>', methods=['DELETE'])
def api_plg_delete_document(document_id):
    return jsonify(PlanogramController.delete_document(document_id))


# --- Уведомления ---
@app.route('/api/plg/notifications', methods=['GET'])
def api_plg_notifications():
    return jsonify(PlanogramController.get_notifications(
        request.args.get('store_id', type=int),
        request.args.get('unread') == '1', _plg_lang()))


@app.route('/api/plg/notifications/<int:notification_id>/read', methods=['POST'])
def api_plg_notification_read(notification_id):
    return jsonify(PlanogramController.mark_notification_read(notification_id))


@app.route('/api/plg/notifications/read-all', methods=['POST'])
def api_plg_notifications_read_all():
    return jsonify(PlanogramController.mark_notification_read(
        None, request.args.get('store_id', type=int)))


# --- Настройки и аудит ---
@app.route('/api/plg/settings', methods=['GET'])
def api_plg_settings():
    return jsonify(PlanogramController.get_settings(_plg_lang()))


@app.route('/api/plg/settings', methods=['POST'])
def api_plg_save_setting():
    data = request.get_json() or {}
    return jsonify(PlanogramController.save_setting(
        data.get('param_code', ''), data.get('param_value', '')))


@app.route('/api/plg/audit', methods=['GET'])
def api_plg_audit():
    return jsonify(PlanogramController.get_audit(request.args.get('limit', 200, type=int)))


# --- Тестовые наборы данных ---
@app.route('/api/plg/datasets', methods=['GET'])
def api_plg_datasets():
    return jsonify(PlanogramController.get_datasets(_plg_lang()))


@app.route('/api/plg/datasets', methods=['POST'])
def api_plg_create_dataset():
    return jsonify(PlanogramController.create_dataset(request.get_json() or {}))


@app.route('/api/plg/datasets/<int:dataset_id>', methods=['DELETE'])
def api_plg_delete_dataset(dataset_id):
    return jsonify(PlanogramController.delete_dataset(dataset_id))


# --- Генерация тестовых данных ---
@app.route('/api/plg/gen/algorithms', methods=['GET'])
def api_plg_gen_algorithms():
    return jsonify(PlanogramController.get_gen_algorithms(_plg_lang()))


@app.route('/api/plg/gen/start', methods=['POST'])
def api_plg_gen_start():
    return jsonify(PlanogramController.start_generation(request.get_json() or {}))


@app.route('/api/plg/gen/runs', methods=['GET'])
def api_plg_gen_runs():
    return jsonify(PlanogramController.get_gen_runs(
        request.args.get('dataset_id', type=int),
        request.args.get('limit', 50, type=int), _plg_lang()))


@app.route('/api/plg/gen/runs/<int:run_id>', methods=['GET'])
def api_plg_gen_run(run_id):
    return jsonify(PlanogramController.get_gen_run(run_id, _plg_lang()))


@app.route('/api/plg/gen/runs/<int:run_id>/cancel', methods=['POST'])
def api_plg_gen_cancel(run_id):
    return jsonify(PlanogramController.cancel_generation(run_id))


# --- Прогноз заказов: алгоритмы и модели ---
@app.route('/api/plg/forecast/algorithms', methods=['GET'])
def api_plg_fct_algorithms():
    return jsonify(PlanogramController.get_fct_algorithms(_plg_lang()))


@app.route('/api/plg/forecast/models', methods=['GET'])
def api_plg_fct_models():
    return jsonify(PlanogramController.get_fct_models(_plg_lang()))


@app.route('/api/plg/forecast/models', methods=['POST'])
def api_plg_create_fct_model():
    return jsonify(PlanogramController.save_fct_model(request.get_json() or {}))


@app.route('/api/plg/forecast/models/<int:model_id>', methods=['PUT'])
def api_plg_update_fct_model(model_id):
    return jsonify(PlanogramController.save_fct_model(request.get_json() or {}, model_id))


@app.route('/api/plg/forecast/models/<int:model_id>', methods=['DELETE'])
def api_plg_delete_fct_model(model_id):
    return jsonify(PlanogramController.delete_fct_model(model_id))


# --- Прогноз заказов: прогоны ---
@app.route('/api/plg/forecast/start', methods=['POST'])
def api_plg_fct_start():
    return jsonify(PlanogramController.start_forecast(request.get_json() or {}))


@app.route('/api/plg/forecast/runs', methods=['GET'])
def api_plg_fct_runs():
    return jsonify(PlanogramController.get_fct_runs(
        request.args.get('model_id', type=int),
        request.args.get('dataset_id', type=int),
        request.args.get('limit', 50, type=int), _plg_lang()))


@app.route('/api/plg/forecast/runs/<int:run_id>', methods=['GET'])
def api_plg_fct_run(run_id):
    return jsonify(PlanogramController.get_fct_run(run_id, _plg_lang()))


@app.route('/api/plg/forecast/runs/<int:run_id>/cancel', methods=['POST'])
def api_plg_fct_cancel(run_id):
    return jsonify(PlanogramController.cancel_forecast(run_id))


@app.route('/api/plg/forecast/runs/<int:run_id>/orders', methods=['GET'])
def api_plg_order_proposal(run_id):
    return jsonify(PlanogramController.get_order_proposal(
        run_id, request.args.get('store_id', type=int),
        request.args.get('limit', 200, type=int), _plg_lang()))


# --- Логистика: РЦ, транспорт, рейсы, Гант ---
@app.route('/api/plg/dc', methods=['GET'])
def api_plg_dc():
    return jsonify(PlanogramController.get_dc(
        request.args.get('dataset_id', type=int), _plg_lang()))


@app.route('/api/plg/dc', methods=['POST'])
def api_plg_create_dc():
    return jsonify(PlanogramController.save_dc(request.get_json() or {}))


@app.route('/api/plg/dc/<int:dc_id>', methods=['PUT'])
def api_plg_update_dc(dc_id):
    return jsonify(PlanogramController.save_dc(request.get_json() or {}, dc_id))


@app.route('/api/plg/dc/<int:dc_id>', methods=['DELETE'])
def api_plg_delete_dc(dc_id):
    return jsonify(PlanogramController.delete_dc(dc_id))


@app.route('/api/plg/vehicles', methods=['GET'])
def api_plg_vehicles():
    return jsonify(PlanogramController.get_vehicles(
        request.args.get('dataset_id', type=int), _plg_lang()))


@app.route('/api/plg/vehicles', methods=['POST'])
def api_plg_create_vehicle():
    return jsonify(PlanogramController.save_vehicle(request.get_json() or {}))


@app.route('/api/plg/vehicles/<int:vehicle_id>', methods=['PUT'])
def api_plg_update_vehicle(vehicle_id):
    return jsonify(PlanogramController.save_vehicle(request.get_json() or {}, vehicle_id))


@app.route('/api/plg/vehicles/<int:vehicle_id>', methods=['DELETE'])
def api_plg_delete_vehicle(vehicle_id):
    return jsonify(PlanogramController.delete_vehicle(vehicle_id))


@app.route('/api/plg/shipments', methods=['GET'])
def api_plg_shipments():
    return jsonify(PlanogramController.get_shipments(
        request.args.get('dataset_id', type=int), request.args.get('date_from'),
        request.args.get('days', 7, type=int), request.args.get('store_id', type=int),
        request.args.get('type'), request.args.get('limit', 1000, type=int), _plg_lang()))


@app.route('/api/plg/shipments', methods=['POST'])
def api_plg_create_shipment():
    return jsonify(PlanogramController.save_shipment(request.get_json() or {}))


@app.route('/api/plg/shipments/<int:shipment_id>', methods=['PUT'])
def api_plg_update_shipment(shipment_id):
    return jsonify(PlanogramController.save_shipment(request.get_json() or {}, shipment_id))


@app.route('/api/plg/shipments/<int:shipment_id>', methods=['DELETE'])
def api_plg_delete_shipment(shipment_id):
    return jsonify(PlanogramController.delete_shipment(shipment_id))


@app.route('/api/plg/gantt', methods=['GET'])
def api_plg_gantt():
    return jsonify(PlanogramController.get_gantt(
        request.args.get('dataset_id', type=int), request.args.get('date_from'),
        request.args.get('days', 3, type=int), request.args.get('group_by', 'vehicle'),
        _plg_lang()))


@app.route('/api/plg/logistics/stats', methods=['GET'])
def api_plg_logistics_stats():
    return jsonify(PlanogramController.get_logistics_stats(
        request.args.get('dataset_id', type=int),
        request.args.get('days', 7, type=int), _plg_lang()))


# --- Поставщики: карточка, контакты, контракты, граф связей ---
@app.route('/api/plg/suppliers', methods=['GET'])
def api_plg_suppliers():
    return jsonify(PlanogramController.get_suppliers(
        request.args.get('dataset_id', type=int), request.args.get('q'), _plg_lang()))


@app.route('/api/plg/suppliers/<int:supplier_id>', methods=['GET'])
def api_plg_supplier(supplier_id):
    return jsonify(PlanogramController.get_supplier(supplier_id, _plg_lang()))


@app.route('/api/plg/suppliers', methods=['POST'])
def api_plg_create_supplier():
    return jsonify(PlanogramController.save_supplier(request.get_json() or {}))


@app.route('/api/plg/suppliers/<int:supplier_id>', methods=['PUT'])
def api_plg_update_supplier(supplier_id):
    return jsonify(PlanogramController.save_supplier(request.get_json() or {}, supplier_id))


@app.route('/api/plg/suppliers/<int:supplier_id>', methods=['DELETE'])
def api_plg_delete_supplier(supplier_id):
    return jsonify(PlanogramController.delete_supplier(supplier_id))


@app.route('/api/plg/suppliers/<int:supplier_id>/contacts', methods=['POST'])
def api_plg_create_contact(supplier_id):
    return jsonify(PlanogramController.save_contact(supplier_id, request.get_json() or {}))


@app.route('/api/plg/suppliers/<int:supplier_id>/contacts/<int:contact_id>', methods=['PUT'])
def api_plg_update_contact(supplier_id, contact_id):
    return jsonify(PlanogramController.save_contact(supplier_id, request.get_json() or {}, contact_id))


@app.route('/api/plg/contacts/<int:contact_id>', methods=['DELETE'])
def api_plg_delete_contact(contact_id):
    return jsonify(PlanogramController.delete_contact(contact_id))


@app.route('/api/plg/contracts', methods=['GET'])
def api_plg_contracts():
    return jsonify(PlanogramController.get_contracts(
        request.args.get('dataset_id', type=int), request.args.get('supplier_id', type=int),
        request.args.get('expiring') == '1', _plg_lang()))


@app.route('/api/plg/suppliers/<int:supplier_id>/contracts', methods=['POST'])
def api_plg_create_contract(supplier_id):
    return jsonify(PlanogramController.save_contract(supplier_id, request.get_json() or {}))


@app.route('/api/plg/suppliers/<int:supplier_id>/contracts/<int:contract_id>', methods=['PUT'])
def api_plg_update_contract(supplier_id, contract_id):
    return jsonify(PlanogramController.save_contract(supplier_id, request.get_json() or {}, contract_id))


@app.route('/api/plg/contracts/<int:contract_id>', methods=['DELETE'])
def api_plg_delete_contract(contract_id):
    return jsonify(PlanogramController.delete_contract(contract_id))


@app.route('/api/plg/suppliers/graph', methods=['GET'])
def api_plg_supplier_graph():
    return jsonify(PlanogramController.get_supplier_graph(
        request.args.get('dataset_id', type=int),
        request.args.get('top', 18, type=int), _plg_lang()))


# --- Конкуренты ---
@app.route('/api/plg/competitors', methods=['GET'])
def api_plg_competitors():
    return jsonify(PlanogramController.get_competitors(
        request.args.get('dataset_id', type=int), _plg_lang()))


@app.route('/api/plg/competitors', methods=['POST'])
def api_plg_create_competitor():
    return jsonify(PlanogramController.save_competitor(request.get_json() or {}))


@app.route('/api/plg/competitors/<int:competitor_id>', methods=['PUT'])
def api_plg_update_competitor(competitor_id):
    return jsonify(PlanogramController.save_competitor(request.get_json() or {}, competitor_id))


@app.route('/api/plg/competitors/<int:competitor_id>', methods=['DELETE'])
def api_plg_delete_competitor(competitor_id):
    return jsonify(PlanogramController.delete_competitor(competitor_id))


@app.route('/api/plg/competitors/price-index', methods=['GET'])
def api_plg_price_index():
    return jsonify(PlanogramController.get_price_index(
        request.args.get('dataset_id', type=int), _plg_lang()))


@app.route('/api/plg/competitors/prices', methods=['GET'])
def api_plg_price_compare():
    return jsonify(PlanogramController.get_price_compare(
        request.args.get('dataset_id', type=int), request.args.get('competitor_id', type=int),
        request.args.get('category_id', type=int), request.args.get('position'),
        request.args.get('limit', 300, type=int), _plg_lang()))


@app.route('/api/plg/competitors/suppliers', methods=['GET'])
def api_plg_competitor_suppliers():
    return jsonify(PlanogramController.get_competitor_suppliers(
        request.args.get('dataset_id', type=int),
        request.args.get('competitor_id', type=int), _plg_lang()))


@app.route('/api/plg/competitors/prices/import', methods=['POST'])
def api_plg_import_prices():
    data = request.get_json() or {}
    return jsonify(PlanogramController.import_competitor_prices(
        data.get('csv', ''), data.get('dataset_id')))


# --- Рынки других стран ---
@app.route('/api/plg/markets', methods=['GET'])
def api_plg_markets():
    return jsonify(PlanogramController.get_markets(
        request.args.get('dataset_id', type=int), _plg_lang()))


@app.route('/api/plg/markets', methods=['POST'])
def api_plg_create_market():
    return jsonify(PlanogramController.save_market(request.get_json() or {}))


@app.route('/api/plg/markets/<int:market_id>', methods=['PUT'])
def api_plg_update_market(market_id):
    return jsonify(PlanogramController.save_market(request.get_json() or {}, market_id))


@app.route('/api/plg/markets/<int:market_id>', methods=['DELETE'])
def api_plg_delete_market(market_id):
    return jsonify(PlanogramController.delete_market(market_id))


@app.route('/api/plg/markets/chains', methods=['GET'])
def api_plg_market_chains():
    return jsonify(PlanogramController.get_market_chains(
        request.args.get('dataset_id', type=int),
        request.args.get('market_id', type=int), _plg_lang()))


@app.route('/api/plg/markets/<int:market_id>/chains', methods=['POST'])
def api_plg_create_chain(market_id):
    return jsonify(PlanogramController.save_market_chain(market_id, request.get_json() or {}))


@app.route('/api/plg/markets/<int:market_id>/chains/<int:chain_id>', methods=['PUT'])
def api_plg_update_chain(market_id, chain_id):
    return jsonify(PlanogramController.save_market_chain(market_id, request.get_json() or {}, chain_id))


@app.route('/api/plg/chains/<int:chain_id>', methods=['DELETE'])
def api_plg_delete_chain(chain_id):
    return jsonify(PlanogramController.delete_market_chain(chain_id))


@app.route('/api/plg/markets/benchmark', methods=['GET'])
def api_plg_market_benchmark():
    return jsonify(PlanogramController.get_market_benchmark(
        request.args.get('dataset_id', type=int), _plg_lang()))


# --- Бизнес-процессы модуля (схемы draw.io) ---
@app.route('/api/plg/processes', methods=['GET'])
def api_plg_processes():
    return jsonify(PlanogramController.get_processes(_plg_lang()))


@app.route('/api/plg/processes/<code>', methods=['GET'])
def api_plg_process(code):
    return jsonify(PlanogramController.get_process(code, _plg_lang()))


@app.route('/api/plg/processes', methods=['POST'])
def api_plg_create_process():
    return jsonify(PlanogramController.save_process(request.get_json() or {}))


@app.route('/api/plg/processes/<int:process_id>', methods=['PUT'])
def api_plg_update_process(process_id):
    return jsonify(PlanogramController.save_process(request.get_json() or {}, process_id))


@app.route('/api/plg/processes/<int:process_id>', methods=['DELETE'])
def api_plg_delete_process(process_id):
    return jsonify(PlanogramController.delete_process(process_id))


@app.route('/api/plg/processes/<code>/drawio', methods=['GET'])
def api_plg_process_drawio(code):
    """Выгрузка схемы файлом .drawio — открывается в diagrams.net как есть."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется вход"}), 401
    result = PlanogramController.get_process(code, 'ru')
    if not result.get('success'):
        return jsonify(result), 404
    xml = result['data'].get('diagram_xml') or ''
    return Response(xml, mimetype='application/xml; charset=utf-8',
                    headers={'Content-Disposition':
                             f'attachment; filename="{code}.drawio"'})


# ========== Мобильное приложение и голосовой заказ ==========
#
# Авторизация здесь — по токену устройства в заголовке X-PLG-Device-Token,
# а не по сессии браузера: телефон менеджера в зале не должен переспрашивать
# пароль у полки. Демо-блокировка записи (_plg_block_anonymous_writes) эти
# маршруты пропускает — доступ закрывает сам токен.

def _plg_device():
    """Устройство по токену запроса. None = доступа нет."""
    token = (request.headers.get('X-PLG-Device-Token')
             or request.headers.get('Authorization', '').replace('Bearer ', '').strip()
             or request.args.get('token'))
    return PlgMobileController.device_by_token(token)


def _plg_device_or_401():
    device = _plg_device()
    if not device:
        return None, (jsonify({'success': False,
                               'error': 'Устройство не авторизовано'}), 401)
    return device, None


@app.route('/api/plg/mobile/pair', methods=['POST'])
def api_plg_mobile_pair():
    """Обмен кода сопряжения на токен. Единственный маршрут без токена."""
    result = PlgMobileController.pair(request.get_json() or {})
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/mobile/session', methods=['GET'])
def api_plg_mobile_session():
    device, err = _plg_device_or_401()
    if err:
        return err
    return jsonify(PlgMobileController.session(device))


@app.route('/api/plg/mobile/catalog', methods=['GET'])
def api_plg_mobile_catalog():
    device, err = _plg_device_or_401()
    if err:
        return err
    return jsonify(PlgMobileController.catalog(
        int(device['store_id']),
        request.args.get('lang') or device.get('lang') or 'ru',
        request.args.get('q', ''), request.args.get('limit', 50, type=int)))


@app.route('/api/plg/mobile/voice', methods=['POST'])
def api_plg_mobile_voice():
    """Распознанная на устройстве фраза → разбор → черновик заказа."""
    device, err = _plg_device_or_401()
    if err:
        return err
    result = PlgMobileController.voice(device, request.get_json() or {})
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/mobile/orders', methods=['GET'])
def api_plg_mobile_orders():
    device, err = _plg_device_or_401()
    if err:
        return err
    return jsonify(PlgMobileController.list_orders(device, request.args.get('status')))


@app.route('/api/plg/mobile/orders', methods=['POST'])
def api_plg_mobile_create_order():
    device, err = _plg_device_or_401()
    if err:
        return err
    result = PlgMobileController.create_order(device, request.get_json() or {})
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/mobile/orders/<int:order_id>', methods=['GET'])
def api_plg_mobile_get_order(order_id):
    device, err = _plg_device_or_401()
    if err:
        return err
    result = PlgMobileController.get_order(device, order_id)
    return jsonify(result), (200 if result.get('success') else result.get('status', 404))


@app.route('/api/plg/mobile/orders/<int:order_id>/items/<int:item_id>', methods=['PUT'])
def api_plg_mobile_update_item(order_id, item_id):
    device, err = _plg_device_or_401()
    if err:
        return err
    result = PlgMobileController.update_item(device, order_id, item_id,
                                             request.get_json() or {})
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/mobile/orders/<int:order_id>/items/<int:item_id>', methods=['DELETE'])
def api_plg_mobile_delete_item(order_id, item_id):
    device, err = _plg_device_or_401()
    if err:
        return err
    return jsonify(PlgMobileController.remove_item(device, order_id, item_id))


@app.route('/api/plg/mobile/orders/<int:order_id>/submit', methods=['POST'])
def api_plg_mobile_submit(order_id):
    device, err = _plg_device_or_401()
    if err:
        return err
    result = PlgMobileController.submit_order(device, order_id)
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/mobile/orders/<int:order_id>/cancel', methods=['POST'])
def api_plg_mobile_cancel(order_id):
    device, err = _plg_device_or_401()
    if err:
        return err
    return jsonify(PlgMobileController.cancel_order(device, order_id))


# --- Бэк-офис: устройства, приёмка заказов из зала, словарь и журнал ---

@app.route('/api/plg/devices', methods=['GET'])
def api_plg_devices():
    return jsonify(PlgMobileController.list_devices(
        request.args.get('store_id', type=int), _plg_lang()))


@app.route('/api/plg/devices', methods=['POST'])
def api_plg_create_device():
    result = PlgMobileController.create_device(request.get_json() or {},
                                               session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else 400)


@app.route('/api/plg/devices/<int:device_id>/revoke', methods=['POST'])
def api_plg_revoke_device(device_id):
    return jsonify(PlgMobileController.revoke_device(device_id))


@app.route('/api/plg/floor-orders', methods=['GET'])
def api_plg_floor_orders():
    return jsonify(PlgMobileController.office_orders(
        request.args.get('store_id', type=int), request.args.get('status'), _plg_lang()))


@app.route('/api/plg/floor-orders/<int:order_id>/review', methods=['POST'])
def api_plg_review_floor_order(order_id):
    body = request.get_json() or {}
    result = PlgMobileController.review_order(order_id, body.get('decision'),
                                              body.get('note'), session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/voice/log', methods=['GET'])
def api_plg_voice_log():
    return jsonify(PlgMobileController.voice_log(
        request.args.get('store_id', type=int), _plg_lang(),
        request.args.get('limit', 200, type=int)))


@app.route('/api/plg/voice/synonyms', methods=['GET'])
def api_plg_voice_synonyms():
    return jsonify(PlgMobileController.synonyms(request.args.get('lang')))


@app.route('/api/plg/voice/synonyms', methods=['POST'])
def api_plg_add_voice_synonym():
    result = PlgMobileController.save_synonym(request.get_json() or {},
                                              session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/voice/synonyms/<int:syn_id>', methods=['DELETE'])
def api_plg_delete_voice_synonym(syn_id):
    return jsonify(PlgMobileController.delete_synonym(syn_id))


# --- Фреш: маршруты поставки и профили категорий ---

@app.route('/api/plg/fresh/routes', methods=['GET'])
def api_plg_fresh_routes():
    return jsonify(PlanogramController.get_fresh_routes(
        _plg_lang(), request.args.get('store_id', type=int)))


@app.route('/api/plg/fresh/routes/<int:route_id>', methods=['PUT'])
def api_plg_update_fresh_route(route_id):
    return jsonify(PlanogramController.save_fresh_route(route_id, request.get_json() or {}))


@app.route('/api/plg/fresh/profiles', methods=['GET'])
def api_plg_fresh_profiles():
    return jsonify(PlanogramController.get_fresh_profiles(_plg_lang()))


@app.route('/api/plg/fresh/profiles/<int:profile_id>', methods=['PUT'])
def api_plg_update_fresh_profile(profile_id):
    return jsonify(PlanogramController.save_fresh_profile(profile_id, request.get_json() or {}))


@app.route('/api/plg/fresh/order', methods=['GET'])
def api_plg_fresh_order():
    """Рекомендуемый заказ фреш по прогону: маршрут, покрытие, ожидаемое списание."""
    return jsonify(PlanogramController.get_fresh_order(
        _plg_lang(), request.args.get('run_id', type=int),
        request.args.get('store_id', type=int)))


# ========== ИИ-мониторинг продаж и витрина признаков ==========

@app.route('/api/plg/ai/monitor/start', methods=['POST'])
def api_plg_ai_start():
    return jsonify(PlgAiController.start_monitor(request.get_json() or {},
                                                 session.get('username', 'user')))


@app.route('/api/plg/ai/monitor/runs', methods=['GET'])
def api_plg_ai_runs():
    return jsonify(PlgAiController.monitor_runs(request.args.get('limit', 20, type=int)))


@app.route('/api/plg/ai/signals', methods=['GET'])
def api_plg_ai_signals():
    return jsonify(PlgAiController.signals(
        _plg_lang(), request.args.get('store_id', type=int),
        request.args.get('type'), request.args.get('run_id', type=int)))


@app.route('/api/plg/ai/signals/<int:signal_id>/ack', methods=['POST'])
def api_plg_ai_ack(signal_id):
    return jsonify(PlgAiController.ack_signal(signal_id, session.get('username', 'user')))


@app.route('/api/plg/ai/features', methods=['GET'])
def api_plg_ai_features():
    return jsonify(PlgAiController.features(
        _plg_lang(), request.args.get('store_id', type=int),
        request.args.get('run_id', type=int),
        request.args.get('limit', 200, type=int)))


@app.route('/api/plg/ai/similar', methods=['GET'])
def api_plg_ai_similar():
    """Похожие по поведению SKU — векторный поиск по HNSW-индексу 26ai."""
    result = PlgAiController.similar_skus(
        _plg_lang(), request.args.get('store_id', type=int),
        request.args.get('product_id', type=int),
        request.args.get('limit', 8, type=int))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/ai/features/export', methods=['GET'])
def api_plg_ai_export():
    """Выгрузка массива признаков для обучения моделей: CSV или JSON."""
    result = PlgAiController.export_features(
        request.args.get('fmt', 'csv'), request.args.get('run_id', type=int),
        request.args.get('store_id', type=int))
    if not result.get('success'):
        return jsonify(result), result.get('status', 400)
    return Response(result['content'], mimetype=result['mimetype'],
                    headers={'Content-Disposition':
                             f'attachment; filename="{result["filename"]}"'})


# ========== Автозаказ: корректировка на лету и пакет документов ==========

@app.route('/api/plg/orders/runs', methods=['GET'])
def api_plg_order_runs():
    return jsonify(PlgAiController.order_runs(_plg_lang()))


@app.route('/api/plg/orders/proposal', methods=['GET'])
def api_plg_adjusted_proposal():
    return jsonify(PlgAiController.order_proposal(
        _plg_lang(), request.args.get('run_id', type=int),
        request.args.get('store_id', type=int)))


@app.route('/api/plg/orders/adjust', methods=['POST'])
def api_plg_order_adjust():
    result = PlgAiController.adjust_order(request.get_json() or {},
                                          session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/orders/adjust/reset', methods=['POST'])
def api_plg_order_adjust_reset():
    return jsonify(PlgAiController.reset_adjustment(request.get_json() or {}))


@app.route('/UNA.md/orasldev/planograms/order-package')
def plg_order_package():
    """Печатный пакет документов заказа: спецификации по поставщикам."""
    lang = _plg_lang()
    run_id = request.args.get('run_id', type=int)
    store_id = request.args.get('store_id', type=int)
    result = PlgAiController.order_package(lang, run_id, store_id)
    if not result.get('success'):
        return jsonify(result), 400
    return render_template('plg_order_package.html', pkg=result, lang=lang,
                           generated_by=session.get('username', 'Гость'))


# ========== Заказы импорта ==========

@app.route('/api/plg/imports', methods=['GET'])
def api_plg_imports():
    return jsonify(PlgAiController.import_orders(_plg_lang(), request.args.get('status')))


@app.route('/api/plg/imports', methods=['POST'])
def api_plg_create_import():
    result = PlgAiController.create_import(request.get_json() or {},
                                           session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else 400)


@app.route('/api/plg/imports/<int:order_id>', methods=['GET'])
def api_plg_import_order(order_id):
    result = PlgAiController.import_order(_plg_lang(), order_id)
    return jsonify(result), (200 if result.get('success') else result.get('status', 404))


@app.route('/api/plg/imports/<int:order_id>/stage', methods=['POST'])
def api_plg_import_stage(order_id):
    result = PlgAiController.advance_stage(order_id, request.get_json() or {},
                                           session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/imports/<int:order_id>/docs/<int:doc_id>', methods=['PUT'])
def api_plg_import_doc(order_id, doc_id):
    result = PlgAiController.set_import_doc(order_id, doc_id, request.get_json() or {},
                                            session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/imports/delays', methods=['GET'])
def api_plg_import_delays():
    return jsonify(PlgAiController.import_delay_stats(_plg_lang()))


# ========== Автозаказ топлива: карта АЗС, нефтебаза, рейсы, GPS ==========
#
# Контур живёт в разделе #fuel модуля «Планограммы», а данными опирается
# на PECO: станции и резервуары уже описаны там, дублировать их в PLG_*
# запрещено правилами проекта.

@app.route('/api/plg/fuel/stations', methods=['GET'])
def api_plg_fuel_stations():
    return jsonify(PecoSupplyController.stations(_plg_lang()))


@app.route('/api/plg/fuel/stations/<int:station_id>/geo', methods=['PUT'])
def api_plg_fuel_station_geo(station_id):
    """Координаты станции: перетащили маркер либо нашли по адресу."""
    result = PecoSupplyController.save_geo(station_id, request.get_json() or {},
                                           session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/fuel/tanks', methods=['GET'])
def api_plg_fuel_tanks():
    return jsonify(PecoSupplyController.tanks(request.args.get('station_id', type=int)))


@app.route('/api/plg/fuel/params', methods=['GET'])
def api_plg_fuel_params():
    return jsonify(PecoSupplyController.params())


@app.route('/api/plg/fuel/autoorder', methods=['POST'])
def api_plg_fuel_autoorder():
    result = PecoSupplyController.run_autoorder(request.get_json() or {},
                                                session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/fuel/runs', methods=['GET'])
def api_plg_fuel_runs():
    return jsonify(PecoSupplyController.runs(request.args.get('limit', 15, type=int)))


@app.route('/api/plg/fuel/orders', methods=['GET'])
def api_plg_fuel_orders():
    return jsonify(PecoSupplyController.orders(
        _plg_lang(), request.args.get('status'), request.args.get('run_id', type=int)))


@app.route('/api/plg/fuel/orders/items/<int:item_id>', methods=['PUT'])
def api_plg_fuel_adjust_item(item_id):
    result = PecoSupplyController.adjust_item(item_id, request.get_json() or {},
                                              session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/fuel/orders/<int:order_id>/status', methods=['POST'])
def api_plg_fuel_order_status(order_id):
    body = request.get_json() or {}
    result = PecoSupplyController.set_order_status(order_id, body.get('status'),
                                                   session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/fuel/orders/approve-all', methods=['POST'])
def api_plg_fuel_approve_all():
    body = request.get_json() or {}
    return jsonify(PecoSupplyController.approve_all(body.get('run_id'),
                                                    session.get('username', 'user')))


@app.route('/api/plg/fuel/depot', methods=['GET'])
def api_plg_fuel_depot():
    return jsonify(PecoSupplyController.depot(
        _plg_lang(), request.args.get('import_lead_days', 9.0, type=float)))


@app.route('/api/plg/fuel/trips', methods=['GET'])
def api_plg_fuel_trips():
    return jsonify(PecoSupplyController.trips(_plg_lang(), request.args.get('status')))


@app.route('/api/plg/fuel/trips/plan', methods=['POST'])
def api_plg_fuel_plan_trips():
    result = PecoSupplyController.plan_trips(request.get_json() or {},
                                             session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/fuel/trips/<int:trip_id>/status', methods=['POST'])
def api_plg_fuel_trip_status(trip_id):
    body = request.get_json() or {}
    result = PecoSupplyController.set_trip_status(trip_id, body.get('status'),
                                                  session.get('username', 'user'))
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/api/plg/fuel/trips/<int:trip_id>/track', methods=['GET'])
def api_plg_fuel_track(trip_id):
    return jsonify(PecoSupplyController.track(trip_id))


@app.route('/api/plg/fuel/gps/events', methods=['GET'])
def api_plg_fuel_gps_events():
    return jsonify(PecoSupplyController.gps_events(
        _plg_lang(), request.args.get('status', 'new'),
        request.args.get('limit', 100, type=int)))


@app.route('/api/plg/fuel/gps/analyze', methods=['POST'])
def api_plg_fuel_gps_analyze():
    body = request.get_json() or {}
    return jsonify(PecoSupplyController.analyze(body.get('trip_id')))


@app.route('/api/peco/gps/ping', methods=['POST'])
def api_peco_gps_ping():
    """
    Приём телеметрии от внешнего GPS-провайдера (датчики на аутсорсе).

    Авторизация — токеном провайдера в заголовке X-PECO-GPS-Token,
    а не сессией: пинги шлёт сервер провайдера пачками.
    """
    token = (request.headers.get('X-PECO-GPS-Token')
             or request.headers.get('Authorization', '').replace('Bearer ', '').strip())
    provider = PecoGps.provider_by_token(token)
    if not provider:
        return jsonify({'success': False, 'error': 'Провайдер не авторизован'}), 401
    result = PecoGps.ingest(provider, request.get_json() or {})
    return jsonify(result), (200 if result.get('success') else result.get('status', 400))


@app.route('/UNA.md/orasldev/planograms/import-package/<int:order_id>')
def plg_import_package(order_id):
    """Печатный пакет импортного заказа: реквизиты, позиции с ТН ВЭД, чек-лист."""
    lang = _plg_lang()
    result = PlgAiController.import_order(lang, order_id)
    if not result.get('success'):
        return jsonify(result), 404
    return render_template('plg_import_package.html', o=result['data'], lang=lang,
                           generated_by=session.get('username', 'Гость'))


# WebSocket Events
@socketio.on('connect')
def handle_connect():
    """Обработка подключения WebSocket"""
    if not AuthController.is_authenticated():
        emit('error', {'message': 'Authentication required'})
        return False
    emit('connected', {'message': 'Connected to server'})


@socketio.on('disconnect')
def handle_disconnect():
    """Обработка отключения WebSocket"""
    # Удаляем все подписки пользователя
    sid = request.sid
    for metric in list(active_subscriptions.keys()):
        if sid in active_subscriptions[metric]:
            active_subscriptions[metric].remove(sid)
            if not active_subscriptions[metric]:
                del active_subscriptions[metric]


@socketio.on('subscribe_metric')
def handle_subscribe_metric(data):
    """Подписка на обновление конкретной метрики"""
    if not AuthController.is_authenticated():
        emit('error', {'message': 'Authentication required'})
        return
    
    metric_name = data.get('metric')
    if not metric_name:
        emit('error', {'message': 'Metric name required'})
        return
    
    sid = request.sid
    
    # Добавляем подписку
    if metric_name not in active_subscriptions:
        active_subscriptions[metric_name] = []
    
    if sid not in active_subscriptions[metric_name]:
        active_subscriptions[metric_name].append(sid)
    
    # Отправляем начальные данные
    result = DashboardController.get_metric(metric_name)
    emit('metric_update', result)


@socketio.on('unsubscribe_metric')
def handle_unsubscribe_metric(data):
    """Отписка от обновления метрики"""
    metric_name = data.get('metric')
    sid = request.sid
    
    if metric_name in active_subscriptions and sid in active_subscriptions[metric_name]:
        active_subscriptions[metric_name].remove(sid)
        if not active_subscriptions[metric_name]:
            del active_subscriptions[metric_name]


def background_metric_updater():
    """Фоновая задача для обновления метрик через WebSocket"""
    while True:
        try:
            time.sleep(Config.DASHBOARD_UPDATE_INTERVAL)
            
            # Обновляем каждую подписанную метрику
            for metric_name in list(active_subscriptions.keys()):
                if active_subscriptions[metric_name]:
                    try:
                        result = DashboardController.get_metric(metric_name)
                        # Отправляем обновление всем подписчикам
                        for sid in active_subscriptions[metric_name]:
                            socketio.emit('metric_update', result, room=sid)
                    except Exception as e:
                        print(f"Error updating metric {metric_name}: {e}")
        except Exception as e:
            print(f"Error in background updater: {e}")
            time.sleep(5)


@app.route('/api/ai-generate-table', methods=['POST'])
def api_ai_generate_table():
    """API endpoint для генерации SQL скрипта создания таблицы через ИИ"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    try:
        # Пытаемся использовать основной ai_helper, если не получается - используем серверный
        try:
            from ai_helper import generate_table_sql, is_ai_available
        except ImportError:
            # На сервере может не быть ai_helper, используем серверный вариант
            try:
                from ai_helper_server import generate_table_sql, is_ai_available
            except ImportError:
                return jsonify({
                    "success": False,
                    "error": "AI helper module not found"
                }), 500
        
        data = request.get_json()
        description = data.get('description', '').strip()
        use_ai = data.get('use_ai', True)
        
        if not description:
            return jsonify({
                "success": False,
                "error": "Описание таблицы не может быть пустым"
            }), 400
        
        # Генерируем SQL
        result = generate_table_sql(description, use_ai=use_ai and is_ai_available())
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/combo-scenario/execute', methods=['POST'])
def api_combo_scenario_execute():
    """API endpoint для выполнения комбинированного сценария AI -> SQL"""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Invalid JSON in request body"}), 400
        
        main_task = data.get('main_task', '').strip()
        iterations_count = data.get('iterations_count', 1)
        iterative_task = data.get('iterative_task', '').strip()
        
        if not main_task:
            return jsonify({"success": False, "error": "Главное задание не может быть пустым"}), 400
        
        if iterations_count < 1:
            return jsonify({"success": False, "error": "Количество итераций должно быть больше 0"}), 400
        
        # Функция для генерации SQL через AI
        def ai_generate_func(description: str):
            try:
                from ai_helper import generate_table_sql, is_ai_available
            except ImportError:
                from ai_helper_server import generate_table_sql, is_ai_available
            
            return generate_table_sql(description, use_ai=is_ai_available())
        
        # Выполняем сценарий
        result = ComboScenarioController.execute_scenario(
            main_task=main_task,
            iterations_count=iterations_count,
            iterative_task=iterative_task,
            ai_generate_func=ai_generate_func
        )
        
        return jsonify(result)
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "traceback": error_trace
        }), 500


@app.route('/api/combo-scenario/execute-iteration', methods=['POST'])
def api_combo_scenario_execute_iteration():
    """API endpoint для выполнения одной итерации комбинированного сценария"""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Invalid JSON in request body"}), 400
        
        main_task = data.get('main_task', '').strip()
        iterations_count = data.get('iterations_count', 1)
        iterative_task = data.get('iterative_task', '').strip()
        current_iteration = data.get('current_iteration', 1)
        previous_iterations = data.get('previous_iterations', [])
        
        if not main_task:
            return jsonify({"success": False, "error": "Главное задание не может быть пустым"}), 400
        
        # Функция для генерации SQL через AI
        def ai_generate_func(description: str):
            try:
                from ai_helper import generate_table_sql, is_ai_available
            except ImportError:
                from ai_helper_server import generate_table_sql, is_ai_available
            
            return generate_table_sql(description, use_ai=is_ai_available())
        
        # Выполняем одну итерацию
        result = ComboScenarioController.execute_single_iteration_step(
            main_task=main_task,
            iterations_count=iterations_count,
            iterative_task=iterative_task,
            current_iteration=current_iteration,
            previous_iterations=previous_iterations,
            ai_generate_func=ai_generate_func
        )
        
        return jsonify(result)
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "traceback": error_trace
        }), 500


@app.route('/api/dashboard/documentation/<dashboard_id>', methods=['GET'])
def api_dashboard_documentation(dashboard_id):
    """API endpoint для получения документации по дашборду"""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(DocumentationController.get_dashboard_documentation(dashboard_id))


@app.route('/api/dashboard/ddl/<dashboard_id>', methods=['GET'])
def api_dashboard_ddl(dashboard_id):
    """API endpoint для генерации DDL скрипта для дашборда"""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    result = DocumentationController.get_ddl_script(dashboard_id)
    return jsonify(result)


@app.route('/api/dashboard/dml/<dashboard_id>', methods=['GET'])
def api_dashboard_dml(dashboard_id):
    """API endpoint для генерации DML скрипта (демо-данные) для дашборда"""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    result = DocumentationController.get_dml_script(dashboard_id)
    return jsonify(result)


@app.route('/api/dashboard/documentation/list', methods=['GET'])
def api_dashboard_documentation_list():
    """API endpoint для получения списка всех дашбордов"""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(DocumentationController.get_all_dashboards_list())


@app.route('/api/credit-admin/programs', methods=['GET'])
def api_credit_admin_programs():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    bank = request.args.get('bank') or None
    term = request.args.get('term', type=int) or None
    active = request.args.get('active') or None
    result = CreditController.get_programs(bank=bank, term=term, active=active)
    return jsonify(result)


@app.route('/api/credit-admin/programs/<int:program_id>', methods=['GET'])
def api_credit_admin_program_get(program_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    try:
        result = CreditController.get_program_by_id(program_id)
        return jsonify(result)
    except Exception as e:
        import traceback
        print(f"Error in api_credit_admin_program_get: {e}")
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e), "data": None}), 500


@app.route('/api/credit-admin/programs', methods=['POST'])
def api_credit_admin_programs_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    return jsonify(CreditController.upsert_program(data))


@app.route('/api/credit-admin/programs/<int:program_id>', methods=['DELETE'])
def api_credit_admin_program_delete(program_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.delete_program(program_id))


@app.route('/api/credit-admin/banks', methods=['GET'])
def api_credit_admin_banks():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_banks())


@app.route('/api/credit-admin/categories', methods=['GET'])
def api_credit_admin_categories():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_categories())


@app.route('/api/credit-admin/brands', methods=['GET'])
def api_credit_admin_brands():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_brands())


@app.route('/api/credit-admin/matrix', methods=['GET'])
def api_credit_admin_matrix():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_matrix())


@app.route('/api/credit-admin/matrix', methods=['POST'])
def api_credit_admin_matrix_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    pid = data.get('program_id')
    cid = data.get('category_id')
    enabled = bool(data.get('enabled', True))
    if pid is None or cid is None:
        return jsonify({"success": False, "error": "program_id and category_id required"}), 400
    return jsonify(CreditController.set_matrix_row(int(pid), int(cid), enabled))


@app.route('/api/credit-admin/matrix/products', methods=['GET'])
def api_credit_admin_matrix_products():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    pid = request.args.get('program_id', type=int)
    cid = request.args.get('category_id', type=int)
    if pid is None or cid is None:
        return jsonify({"success": False, "error": "program_id and category_id required"}), 400
    search = request.args.get('search') or None
    limit = request.args.get('limit', default=500, type=int)
    return jsonify(CreditController.get_matrix_products(program_id=pid, category_id=cid, search=search, limit=limit))


@app.route('/api/credit-admin/matrix/products', methods=['POST'])
def api_credit_admin_matrix_products_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    pid = data.get('program_id')
    cid = data.get('category_id')
    product_ids = data.get('product_ids') or []
    if pid is None or cid is None:
        return jsonify({"success": False, "error": "program_id and category_id required"}), 400
    try:
        ids = [int(x) for x in product_ids if x is not None and str(x).strip() != '']
    except (TypeError, ValueError):
        ids = []
    return jsonify(CreditController.set_matrix_products(program_id=int(pid), category_id=int(cid), product_ids=ids))


@app.route('/api/credit-admin/pivot/meta', methods=['GET'])
def api_credit_admin_pivot_meta():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_pivot_meta())


@app.route('/api/credit-admin/pivot/matrix', methods=['GET'])
def api_credit_admin_pivot_matrix():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_pivot_matrix())


@app.route('/api/credit-admin/pivot/products', methods=['GET'])
def api_credit_admin_pivot_products():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    cids = request.args.get('category_ids')
    category_ids = [int(x) for x in cids.split(',')] if cids else None
    search = request.args.get('search') or None
    limit = request.args.get('limit', default=500, type=int)
    return jsonify(CreditController.get_pivot_products(category_ids=category_ids, search=search, limit=limit))


@app.route('/api/credit-admin/easycredit-settings', methods=['GET'])
def api_credit_admin_easycredit_settings():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    from config import Config
    env = Config.easycredit_env()
    base_url = Config.easycredit_base_url()
    user = Config.easycredit_api_user()
    pwd = Config.easycredit_api_password()
    return jsonify({
        "success": True,
        "data": {
            "env": env,
            "base_url": base_url,
            "api_user": user,
            "api_password_masked": "********" if pwd else "",
        },
    })


@app.route('/api/credit-admin/easycredit-settings', methods=['POST'])
def api_credit_admin_easycredit_settings_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    from config import save_easycredit_settings
    data = request.get_json() or {}
    env = (data.get("env") or "sandbox").strip().lower()
    base_url = (data.get("base_url") or "").strip()
    api_user = (data.get("api_user") or "").strip()
    api_password = (data.get("api_password") or "").strip()
    try:
        save_easycredit_settings(env=env, base_url=base_url, api_user=api_user, api_password=api_password)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": True})


@app.route('/api/credit-admin/iute-settings', methods=['GET'])
def api_credit_admin_iute_settings():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    from config import Config
    env = Config.iute_env()
    base_url = Config.iute_base_url()
    api_key = Config.iute_api_key()
    pos_identifier = Config.iute_pos_identifier()
    salesman_identifier = Config.iute_salesman_identifier()
    # Never return real secrets - only masked indicators
    return jsonify({
        "success": True,
        "data": {
            "env": env,
            "base_url": base_url,
            "api_key_masked": bool(api_key),
            "pos_identifier_masked": bool(pos_identifier),
            "salesman_identifier_masked": bool(salesman_identifier),
        },
    })


@app.route('/api/credit-admin/reports', methods=['GET'])
def api_credit_admin_reports():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_reports())


@app.route('/api/credit-admin/reports/<int:report_id>', methods=['GET'])
def api_credit_admin_report_get(report_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_report_by_id(report_id))


@app.route('/api/credit-admin/reports/<int:report_id>/params', methods=['GET'])
def api_credit_admin_report_params(report_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_report_params(report_id))


@app.route('/api/credit-admin/reports/<int:report_id>/template', methods=['PUT', 'POST'])
def api_credit_admin_report_template(report_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    return jsonify(CreditController.update_report_template(
        report_id,
        name=data.get("name"),
        description=data.get("description"),
        template_html=data.get("template_html"),
    ))


@app.route('/api/credit-admin/reports/<int:report_id>/execute', methods=['POST'])
def api_credit_admin_report_execute(report_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    params = data.get("params") or data
    return jsonify(CreditController.execute_report(report_id, params))


@app.route('/api/credit-admin/reports/<int:report_id>/export', methods=['POST'])
def api_credit_admin_report_export(report_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    params = data.get("params") or data
    fmt = (data.get("format") or "csv").lower()
    report_name = data.get("report_name") or "Report"
    result = CreditController.execute_report(report_id, params)
    if not result.get("success"):
        return jsonify(result), 400
    rows = result.get("data") or []
    cols = list(rows[0].keys()) if rows else []
    try:
        from services.report_export import export_csv, export_excel, export_pdf
        if fmt == "csv":
            content = export_csv(rows, cols)
            from flask import Response
            return Response(content, mimetype="text/csv; charset=utf-8",
                           headers={"Content-Disposition": f"attachment; filename={report_name}.csv"})
        if fmt == "excel" or fmt == "xlsx":
            content = export_excel(rows, report_name, cols)
            from flask import Response
            return Response(content, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": f"attachment; filename={report_name}.xlsx"})
        if fmt == "pdf":
            content = export_pdf(rows, report_name, cols)
            from flask import Response
            return Response(content, mimetype="application/pdf",
                           headers={"Content-Disposition": f"attachment; filename={report_name}.pdf"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": False, "error": f"Формат {fmt} не поддерживается"}), 400


@app.route('/api/credit-admin/reports/export-pdf', methods=['POST'])
def api_credit_admin_export_pdf_direct():
    """Экспорт в PDF по переданным данным (без перезапуска отчёта)."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    rows = data.get("data") or data.get("rows") or []
    cols = data.get("columns") or data.get("cols")
    report_name = data.get("report_name") or "Report"
    if not rows:
        return jsonify({"success": False, "error": "Нет данных"}), 400
    if not cols:
        cols = list(rows[0].keys())
    try:
        from services.report_export import export_pdf
        from flask import Response
        content = export_pdf(rows, report_name, cols)
        return Response(content, mimetype="application/pdf",
                       headers={"Content-Disposition": f"attachment; filename={report_name}.pdf"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/credit-admin/iute-settings', methods=['POST'])
def api_credit_admin_iute_settings_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    from config import save_iute_settings
    data = request.get_json() or {}
    env = (data.get("env") or "sandbox").strip().lower()
    base_url = (data.get("base_url") or "").strip()
    api_key = (data.get("api_key") or "").strip()
    pos_identifier = (data.get("pos_identifier") or "").strip()
    salesman_identifier = (data.get("salesman_identifier") or "").strip()
    try:
        save_iute_settings(env=env, base_url=base_url, api_key=api_key, pos_identifier=pos_identifier, salesman_identifier=salesman_identifier)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": True})


# ========== Credit Testing — единый API для всех провайдеров ==========

@app.route('/api/credit-testing/providers', methods=['GET'])
def api_credit_testing_providers():
    """Список всех зарегистрированных кредитных провайдеров."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    # Убедимся что провайдеры зарегистрированы
    import integrations  # noqa: F401  — авто-регистрация
    from controllers.credit_testing_controller import CreditTestingController
    return jsonify(CreditTestingController.get_providers())


@app.route('/api/credit-testing/provider/<provider_id>', methods=['GET'])
def api_credit_testing_provider_info(provider_id):
    """Информация о провайдере."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    import integrations  # noqa: F401
    from controllers.credit_testing_controller import CreditTestingController
    return jsonify(CreditTestingController.get_provider_info(provider_id))


@app.route('/api/credit-testing/search-client', methods=['GET'])
def api_credit_testing_search_client():
    """Поиск клиента через провайдер. ?provider=...&uin=...&phone=..."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    import integrations  # noqa: F401
    from controllers.credit_testing_controller import CreditTestingController
    pid = request.args.get('provider', '')
    kwargs = {k: v for k, v in request.args.items() if k != 'provider'}
    _credit_log(f"{pid}.search_client", f"Поиск клиента: {kwargs}", "INFO", kwargs)
    result = CreditTestingController.search_client(pid, **kwargs)
    _credit_log(f"{pid}.search_client", f"Результат: success={result.get('success')}", "INFO" if result.get("success") else "WARN")
    return jsonify(result)


@app.route('/api/credit-testing/preapproved', methods=['POST'])
def api_credit_testing_preapproved():
    """Preapproved через провайдер."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    import integrations  # noqa: F401
    from controllers.credit_testing_controller import CreditTestingController
    data = request.get_json() or {}
    pid = data.pop('provider', '')
    _credit_log(f"{pid}.preapproved", f"Запрос preapproved", "INFO", {k: v for k, v in data.items() if k != 'password'})
    result = CreditTestingController.preapproved(pid, **data)
    _credit_log(f"{pid}.preapproved", f"Результат: success={result.get('success')}", "INFO" if result.get("success") else "WARN")
    return jsonify(result)


@app.route('/api/credit-testing/submit', methods=['POST'])
def api_credit_testing_submit():
    """Отправка заявки через провайдер."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    import integrations  # noqa: F401
    from controllers.credit_testing_controller import CreditTestingController
    data = request.get_json() or {}
    pid = data.pop('provider', '')
    _credit_log(f"{pid}.submit", f"Отправка заявки", "INFO", {k: v for k, v in data.items() if k not in ('password',)})
    result = CreditTestingController.submit(pid, **data)
    level = "INFO" if result.get("success") else "ERROR"
    _credit_log(f"{pid}.submit", f"Результат: {result.get('data', {})}", level)
    return jsonify(result)


@app.route('/api/credit-testing/status', methods=['GET'])
def api_credit_testing_status():
    """Проверка статуса через провайдер. ?provider=...&urn=...&order_id=..."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    import integrations  # noqa: F401
    from controllers.credit_testing_controller import CreditTestingController
    pid = request.args.get('provider', '')
    kwargs = {k: v for k, v in request.args.items() if k != 'provider'}
    _credit_log(f"{pid}.status", f"Проверка статуса: {kwargs}", "INFO")
    result = CreditTestingController.check_status(pid, **kwargs)
    _credit_log(f"{pid}.status", f"Результат: {result.get('data', {})}", "INFO" if result.get("success") else "WARN")
    return jsonify(result)


@app.route('/api/credit-testing/check-auth', methods=['GET'])
def api_credit_testing_check_auth():
    """Проверка авторизации провайдера. ?provider=..."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    import integrations  # noqa: F401
    from controllers.credit_testing_controller import CreditTestingController
    pid = request.args.get('provider', '')
    _credit_log(f"{pid}.check_auth", "Проверка авторизации", "INFO")
    result = CreditTestingController.check_auth(pid)
    _credit_log(f"{pid}.check_auth", f"Результат: success={result.get('success')}", "INFO" if result.get("success") else "WARN")
    return jsonify(result)


@app.route('/api/credit-testing/create-order', methods=['POST'])
def api_credit_testing_create_order():
    """Создание заказа через провайдер."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    import integrations  # noqa: F401
    from controllers.credit_testing_controller import CreditTestingController
    data = request.get_json() or {}
    pid = data.pop('provider', '')
    _credit_log(f"{pid}.create_order", f"Создание заказа", "INFO", data)
    result = CreditTestingController.create_order(pid, **data)
    _credit_log(f"{pid}.create_order", f"Результат: {result.get('data', {})}", "INFO" if result.get("success") else "ERROR")
    return jsonify(result)


@app.route('/api/credit-testing/order-status', methods=['GET'])
def api_credit_testing_order_status():
    """Статус заказа через провайдер. ?provider=...&order_id=..."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    import integrations  # noqa: F401
    from controllers.credit_testing_controller import CreditTestingController
    pid = request.args.get('provider', '')
    kwargs = {k: v for k, v in request.args.items() if k != 'provider'}
    _credit_log(f"{pid}.order_status", f"Проверка заказа: {kwargs}", "INFO")
    result = CreditTestingController.order_status(pid, **kwargs)
    _credit_log(f"{pid}.order_status", f"Результат: {result.get('data', {})}", "INFO" if result.get("success") else "WARN")
    return jsonify(result)


@app.route('/api/credit-logs', methods=['GET'])
def api_credit_logs():
    """Лог EasyCredit и кредитных операций для виджета Output (как SQL Developer)."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    from services.credit_logger import get as log_get
    limit = request.args.get('limit', default=500, type=int)
    since = request.args.get('since') or None
    entries = log_get(limit=min(limit, 2000), since_ts=since)
    return jsonify({"success": True, "data": entries})


@app.route('/api/credit-logs/clear', methods=['POST'])
def api_credit_logs_clear():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    from services.credit_logger import clear as log_clear
    log_clear()
    return jsonify({"success": True})


@app.route('/api/credit-operator/products', methods=['GET'])
def api_credit_operator_products():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    search = request.args.get('search') or None
    barcode = request.args.get('barcode') or None
    limit = request.args.get('limit', default=10, type=int)
    result = CreditController.get_products(search=search, barcode=barcode, limit=limit)
    return jsonify(result)


@app.route('/api/credit-operator/products/<int:product_id>', methods=['GET'])
def api_credit_operator_product_get(product_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_product_by_id(product_id))


@app.route('/api/credit-operator/programs-for-product/<int:product_id>', methods=['GET'])
def api_credit_operator_programs_for_product(product_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(CreditController.get_programs_for_product(product_id))


@app.route('/api/credit-operator/application', methods=['POST'])
def api_credit_operator_application():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    product_id = data.get('product_id')
    program_id = data.get('program_id')
    fio = (data.get('client_fio') or data.get('fio') or "").strip()
    phone = (data.get('client_phone') or data.get('phone') or "").strip()
    idn = (data.get('client_idn') or data.get('idn') or "").strip()
    if not all([product_id, program_id, fio, phone]):
        return jsonify({"success": False, "error": "product_id, program_id, client_fio, client_phone required"}), 400
    _credit_log(
        "credit.operator.application",
        f"Заявка product_id={product_id} program_id={program_id} fio={fio!r}",
        "INFO",
        {"product_id": product_id, "program_id": program_id},
    )
    out = CreditController.create_application(int(product_id), int(program_id), fio, phone, idn)
    if out.get("success"):
        app_id = out.get("application_id") or out.get("id") or "—"
        _credit_log("credit.operator.application", f"OK application_id={app_id}", "INFO", out)
    else:
        _credit_log("credit.operator.application", f"Ошибка: {out.get('error', '')}", "ERROR", out)
    return jsonify(out)


@app.route('/api/credit-operator/recent-applications', methods=['GET'])
def api_credit_operator_recent():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    limit = request.args.get('limit', default=5, type=int)
    return jsonify(CreditController.get_recent_applications(limit=limit))


# ---------- Nufarul: админка ----------
@app.route('/api/nufarul-admin/services', methods=['GET'])
def api_nufarul_admin_services():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(NufarulController.get_services(active_only=False))


@app.route('/api/nufarul-admin/services', methods=['POST'])
def api_nufarul_admin_services_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    return jsonify(NufarulController.upsert_service(data))


@app.route('/api/nufarul-admin/services/<int:service_id>', methods=['GET'])
def api_nufarul_admin_service_get(service_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(NufarulController.get_service_by_id(service_id))


@app.route('/api/nufarul-admin/services/<int:service_id>', methods=['DELETE'])
def api_nufarul_admin_service_delete(service_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(NufarulController.delete_service(service_id))


@app.route('/api/nufarul-admin/statuses', methods=['GET'])
def api_nufarul_admin_statuses():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(NufarulController.get_statuses())


@app.route('/api/nufarul-admin/orders', methods=['GET'])
def api_nufarul_admin_orders():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    status_id = request.args.get('status_id', type=int) or None
    date_from = request.args.get('date_from') or None
    date_to = request.args.get('date_to') or None
    search = request.args.get('search') or None
    limit = request.args.get('limit', default=200, type=int)
    return jsonify(NufarulController.get_orders(status_id=status_id, date_from=date_from, date_to=date_to, search=search, limit=limit))


@app.route('/api/nufarul-admin/orders/<int:order_id>/status', methods=['PUT', 'POST'])
def api_nufarul_admin_order_status(order_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    status_id = data.get('status_id') or data.get('statusId')
    if status_id is None:
        return jsonify({"success": False, "error": "status_id required"}), 400
    return jsonify(NufarulController.update_order_status(order_id, int(status_id)))


@app.route('/api/nufarul-admin/report-by-day', methods=['GET'])
def api_nufarul_admin_report_by_day():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    date_from = request.args.get('date_from') or None
    date_to = request.args.get('date_to') or None
    return jsonify(NufarulController.report_orders_by_day(date_from=date_from, date_to=date_to))


@app.route('/api/nufarul-admin/system-settings', methods=['GET'])
def api_nufarul_admin_system_settings_get():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(NufarulController.get_system_settings())


@app.route('/api/nufarul-admin/system-settings/<key>', methods=['PUT'])
def api_nufarul_admin_system_settings_put(key):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    value = str(data.get('value', '')).strip()
    if not value:
        return jsonify({"success": False, "error": "value required"}), 400
    return jsonify(NufarulController.update_system_setting(key, value))


@app.route('/api/nufarul-ts/system-settings', methods=['GET'])
def api_nufarul_ts_system_settings():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(NufarulController.get_system_settings())


# ---------- Nufarul: оператор ----------
@app.route('/api/nufarul-operator/services', methods=['GET'])
def api_nufarul_operator_services():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(NufarulController.get_services(active_only=True))


@app.route('/api/nufarul-operator/order', methods=['POST'])
def api_nufarul_operator_order():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    client_name = (data.get('client_name') or "").strip()
    client_phone = (data.get('client_phone') or "").strip()
    items = data.get('items') or []
    notes = (data.get('notes') or "").strip() or None
    if not client_name or not client_phone:
        return jsonify({"success": False, "error": "client_name and client_phone required"}), 400
    if not items:
        return jsonify({"success": False, "error": "items required"}), 400
    return jsonify(NufarulController.create_order(client_name, client_phone, items, notes))


@app.route('/api/nufarul-operator/recent-orders', methods=['GET'])
def api_nufarul_operator_recent_orders():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    limit = request.args.get('limit', default=20, type=int)
    return jsonify(NufarulController.get_recent_orders(limit=limit))


@app.route('/api/nufarul-operator/order-by-barcode', methods=['GET'])
def api_nufarul_operator_order_by_barcode():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    barcode = request.args.get('barcode') or ""
    return jsonify(NufarulController.get_order_by_barcode(barcode))


# ---------- Nufarul: AI parse order (shared by operator + TS kiosk) ----------
@app.route('/api/nufarul-operator/ai-parse-order', methods=['POST'])
def api_nufarul_operator_ai_parse_order():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    text = (data.get('text') or '').strip()
    backend = (data.get('backend') or 'oracle').lower()
    try:
        threshold = int(data.get('threshold') or 40)
    except (ValueError, TypeError):
        threshold = 40
    if not text:
        return jsonify({"success": False, "error": "text required"}), 400
    if backend == 'oracle':
        matches = NufarulController.ai_parse_order_oracle(text, threshold=threshold)
        return jsonify({"success": True, "matches": matches, "backend": "oracle"})
    # local rapidfuzz fallback
    try:
        from nufarul_ai_parser import parse_order as _local_parse
        svc_result = NufarulController.get_services(active_only=True)
        services = svc_result.get('data') or []
        matches = _local_parse(text, services, threshold=threshold)
        return jsonify({"success": True, "matches": matches, "backend": "local"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ---------- Nufarul: touchscreen kiosk API ----------
@app.route('/api/nufarul-ts/group-params', methods=['GET'])
def api_nufarul_ts_group_params():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(NufarulController.get_group_params())


@app.route('/api/nufarul-ts/group-params/<group_key>', methods=['GET'])
def api_nufarul_ts_group_params_single(group_key):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(NufarulController.get_group_params(group_key=group_key))


@app.route('/api/nufarul-ts/order', methods=['POST'])
def api_nufarul_ts_order():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    client_name = (data.get('client_name') or 'Аноним').strip() or 'Аноним'
    client_phone = (data.get('client_phone') or '').strip()
    payment_method = (data.get('payment_method') or 'cash').strip().lower()
    if payment_method not in ('cash', 'card', 'mia', 'mpay'):
        payment_method = 'cash'
    items = data.get('items') or []
    notes = (data.get('notes') or '').strip() or None
    ready_date = (data.get('ready_date') or '').strip() or None
    if not items:
        return jsonify({"success": False, "error": "items required"}), 400
    if not isinstance(items, list) or not all(isinstance(i, dict) for i in items):
        return jsonify({"success": False, "error": "items must be a list of objects"}), 400
    if len(items) > 50:
        return jsonify({"success": False, "error": "Too many items (max 50)"}), 400
    for i in items:
        if not i.get("service_id") or int(i.get("service_id", 0)) <= 0:
            return jsonify({"success": False, "error": "Each item must have a valid service_id"}), 400
        if float(i.get("qty", 0)) <= 0:
            return jsonify({"success": False, "error": "qty must be > 0"}), 400
        if float(i.get("price", -1)) < 0:
            return jsonify({"success": False, "error": "price must be >= 0"}), 400
    return jsonify(NufarulController.create_order_with_params(client_name, client_phone, items, notes, payment_method, ready_date))


# ========== DECOR: админка + оператор (локальное JSON-хранилище, fallback без Oracle) ==========

@app.route('/api/decor-admin/materials', methods=['GET'])
def api_decor_admin_materials():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    active_only = (request.args.get('active_only') or '').upper() == 'Y'
    return jsonify(DecorLocalStore.get_materials(active_only=active_only))


@app.route('/api/decor-admin/materials', methods=['POST'])
def api_decor_admin_materials_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    resp = DecorLocalStore.upsert_material(data)
    return jsonify(resp), (200 if resp.get("success") else 400)


@app.route('/api/decor-admin/materials/<int:material_id>', methods=['DELETE'])
def api_decor_admin_material_delete(material_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    resp = DecorLocalStore.delete_material(material_id)
    return jsonify(resp), (200 if resp.get("success") else 404)


@app.route('/api/decor-admin/settings', methods=['GET'])
def api_decor_admin_settings():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(DecorLocalStore.get_settings())


@app.route('/api/decor-admin/settings', methods=['POST'])
def api_decor_admin_settings_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    resp = DecorLocalStore.update_settings(data)
    return jsonify(resp), (200 if resp.get("success") else 400)


@app.route('/api/decor-admin/statuses', methods=['GET'])
def api_decor_admin_statuses():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(DecorLocalStore.get_statuses())


@app.route('/api/decor-admin/orders', methods=['GET'])
def api_decor_admin_orders():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    status_id = request.args.get('status_id', type=int)
    date_from = request.args.get('date_from') or None
    date_to = request.args.get('date_to') or None
    search = request.args.get('search') or None
    limit = request.args.get('limit', default=200, type=int)
    return jsonify(DecorLocalStore.get_orders(status_id=status_id, date_from=date_from, date_to=date_to, search=search, limit=limit))


@app.route('/api/decor-admin/orders/<int:order_id>/status', methods=['POST', 'PUT'])
def api_decor_admin_order_status(order_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    status_id = int(data.get('status_id') or 0)
    resp = DecorLocalStore.update_order_status(order_id, status_id)
    return jsonify(resp), (200 if resp.get("success") else 400)


@app.route('/api/decor-admin/report-by-day', methods=['GET'])
def api_decor_admin_report_by_day():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    date_from = request.args.get('date_from') or None
    date_to = request.args.get('date_to') or None
    return jsonify(DecorLocalStore.report_by_day(date_from=date_from, date_to=date_to))


@app.route('/api/decor-admin/import-xml-orders', methods=['POST'])
def api_decor_admin_import_xml_orders():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    raw_dir = str(data.get("dir") or "").strip()
    base_dir = Path(__file__).resolve().parent
    xml_dir = (Path(raw_dir).expanduser() if raw_dir else (base_dir / "docs" / "DECOR"))
    if not xml_dir.is_absolute():
        xml_dir = (base_dir / xml_dir).resolve()
    else:
        xml_dir = xml_dir.resolve()
    result = import_xml_orders_from_dir(xml_dir=xml_dir, static_root=base_dir / "static")
    return jsonify(result), (200 if result.get("success") else 400)


@app.route('/api/decor-operator/catalog', methods=['GET'])
def api_decor_operator_catalog():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(DecorLocalStore.get_catalog(active_only=True))


@app.route('/api/decor-operator/ai-parse-items', methods=['POST'])
def api_decor_operator_ai_parse_items():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    text = str(data.get("text") or "").strip()
    backend = str(data.get("backend") or "local").strip().lower()
    try:
        threshold = int(data.get("threshold") or 40)
    except Exception:
        threshold = 40
    if not text:
        return jsonify({"success": False, "error": "text is required"}), 400
    return jsonify(DecorLocalStore.ai_parse_items(text=text, threshold=threshold, backend=backend))


@app.route('/api/decor-operator/calculate', methods=['POST'])
def api_decor_operator_calculate():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    resp = DecorLocalStore.calculate_quote(data)
    return jsonify(resp), (200 if resp.get("success") else 400)


@app.route('/api/decor-operator/order', methods=['POST'])
def api_decor_operator_order():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    resp = DecorLocalStore.create_order(data)
    return jsonify(resp), (200 if resp.get("success") else 400)


@app.route('/api/decor-operator/recent-orders', methods=['GET'])
def api_decor_operator_recent_orders():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    limit = request.args.get('limit', default=20, type=int)
    return jsonify(DecorLocalStore.get_recent_orders(limit=limit))


@app.route('/api/decor-operator/order-by-number', methods=['GET'])
def api_decor_operator_order_by_number():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    num = request.args.get('number') or request.args.get('barcode') or ""
    resp = DecorLocalStore.get_order_by_number(num)
    return jsonify(resp), (200 if resp.get("success") else 404)


# ── Sliding API ────────────────────────────────────────────

@app.route('/api/decor-admin/sliding-materials', methods=['GET'])
def api_decor_admin_sliding_materials():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(DecorLocalStore.get_sliding_materials())

@app.route('/api/decor-admin/sliding-materials/<int:material_id>', methods=['POST', 'PUT'])
def api_decor_admin_sliding_material_update(material_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    resp = DecorLocalStore.update_sliding_material(material_id, data)
    return jsonify(resp), (200 if resp.get("success") else 400)

@app.route('/api/decor-admin/sliding-settings', methods=['GET'])
def api_decor_admin_sliding_settings():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(DecorLocalStore.get_sliding_settings())

@app.route('/api/decor-admin/sliding-settings', methods=['POST'])
def api_decor_admin_sliding_settings_update():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    resp = DecorLocalStore.update_sliding_settings(data)
    return jsonify(resp), (200 if resp.get("success") else 400)

@app.route('/api/decor-admin/sliding-variants', methods=['GET'])
def api_decor_admin_sliding_variants():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    return jsonify(DecorLocalStore.get_sliding_variants())

@app.route('/api/decor-operator/calculate-sliding', methods=['POST'])
def api_decor_operator_calculate_sliding():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    resp = DecorLocalStore.calculate_sliding_quote(data)
    return jsonify(resp), (200 if resp.get("success") else 400)

@app.route('/api/decor-operator/sliding-cutting-list', methods=['POST'])
def api_decor_operator_sliding_cutting_list():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json(silent=True) or {}
    resp = DecorLocalStore.calculate_sliding_cutting_list(data)
    return jsonify(resp), (200 if resp.get("success") else 400)


@app.route('/UNA.md/orasldev/nufarul-operator/document/jurnal')
def nufarul_operator_document_jurnal():
    """Журнал регистраций заказов (Jurnal Registru): последние заказы таблицей."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    limit = min(int(request.args.get('limit', 50)), 200)
    result = NufarulController.get_recent_orders(limit=limit)
    if not result.get('success'):
        return f"<h1>Ошибка</h1><p>{result.get('error', '')}</p>", 500
    orders = result.get('data') or []
    entries = []
    for i, o in enumerate(orders, start=1):
        created_time = o.get('created_time')
        if not created_time and o.get('created_at'):
            ct = o['created_at']
            created_time = ct.strftime('%Y-%m-%d %H:%M') if hasattr(ct, 'strftime') else str(ct)[:16]
        entries.append({
            "row_num": i,
            "order_number": o.get('order_number') or '—',
            "client_name": o.get('client_name') or '—',
            "client_phone": o.get('client_phone') or '',
            "total_amount": o.get('total_amount') or 0,
            "created_time": created_time or '—',
        })
    from datetime import datetime
    period_label = f"Ultimele {len(entries)} comenzi / Последние {len(entries)} заказов (generat {datetime.now().strftime('%d.%m.%Y %H:%M')})"
    return render_template(
        "nufarul/document_jurnal_registru.html",
        entries=entries,
        period_label=period_label,
    )


@app.route('/UNA.md/orasldev/nufarul-operator/document/<int:order_id>')
def nufarul_operator_document(order_id):
    """Печать первичного документа по заказу: Bon de Comandă или Comandă (type=bon_comanda|comanda)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    doc_type = request.args.get('type', 'bon_comanda').strip().lower()
    if doc_type not in ('bon_comanda', 'comanda'):
        doc_type = 'bon_comanda'
    result = NufarulController.get_order_by_id(order_id)
    if not result.get('success') or not result.get('data'):
        return f"<h1>Заказ не найден</h1><p>Order ID: {order_id}</p><p><a href='/UNA.md/orasldev/nufarul-operator'>← Оператор</a></p>", 404
    order = result["data"]
    created_at = order.get('created_at')
    if hasattr(created_at, 'strftime'):
        created_at = created_at.strftime('%d.%m.%Y %H:%M')
    else:
        created_at = str(created_at)[:16] if created_at else '—'
    template_name = f"nufarul/document_{doc_type}.html"
    return render_template(
        template_name,
        order_number=order.get('order_number') or '—',
        barcode=order.get('barcode') or order.get('order_number'),
        client_name=order.get('client_name') or '—',
        client_phone=order.get('client_phone') or '',
        created_at=created_at,
        notes=order.get('notes') or '',
        items=order.get('items') or [],
        total_amount=order.get('total_amount') or 0,
    )


def _easycredit_mock_preapproved(amount: int, idn: str):
    return {
        "success": True,
        "data": {
            "preapproved": True,
            "max_amount": max(int(amount or 10000) * 2, 50000),
            "message": "Mock: предодобрение (тест).",
        },
        "fallback": True,
    }


def _easycredit_mock_submit(fio: str, phone: str):
    import uuid
    urn = f"EC-MOCK-{uuid.uuid4().hex[:12].upper()}"
    return {
        "success": True,
        "data": {"urn": urn, "message": "Mock: заявка (тест)."},
        "fallback": True,
    }


def _easycredit_mock_status(urn: str):
    return {
        "success": True,
        "data": {"urn": urn, "status": "Approved", "message": "Mock: статус (тест)."},
        "fallback": True,
    }


def _credit_log(source: str, message: str, level: str = "INFO", payload: dict = None):
    try:
        from services.credit_logger import append as _append
        _append(source, message, level=level, payload=payload or {})
    except Exception:
        pass


@app.route('/api/credit-easycredit/preapproved', methods=['POST'])
def api_credit_easycredit_preapproved():
    """Preapproved (EasyCredit): проверка предодобренной суммы. Реальный EC при user+pass."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    amount = int(data.get("amount") or 10000)
    idn = (data.get("idn") or "12345678901234").strip()
    idn_masked = ("***" + idn[-4:]) if len(idn) >= 4 else "***"
    _credit_log("easycredit.preapproved", f"Запрос Preapproved amount={amount} idn={idn_masked}", "INFO", {"amount": amount, "idn_masked": idn_masked})
    base_url = Config.easycredit_base_url()
    user = Config.easycredit_api_user()
    passwd = Config.easycredit_api_password()
    if user and passwd:
        try:
            from integrations.easycredit_client import preapproved as ec_preapproved
            verify_ssl = Config.easycredit_env() == "production"
            out = ec_preapproved(base_url, user, passwd, idn=idn, amount=amount, verify_ssl=verify_ssl)
            if out.get("success"):
                d = out["data"] or {}
                _credit_log("easycredit.preapproved", f"OK preapproved={d.get('preapproved')} max_amount={d.get('max_amount')} {d.get('message', '')}", "INFO", d)
                return jsonify({"success": True, "data": out["data"]})
            err = out.get("error") or out.get("data", {}).get("message") or "unknown"
            _credit_log("easycredit.preapproved", f"EC error, fallback to mock: {err}", "WARN", out)
            return jsonify(_easycredit_mock_preapproved(amount, idn))
        except Exception as e:
            _credit_log("easycredit.preapproved", f"Exception, fallback to mock: {e}", "ERROR", {"error": str(e)})
            return jsonify(_easycredit_mock_preapproved(amount, idn))
    _credit_log("easycredit.preapproved", "Нет user/pass, mock", "INFO", {})
    return jsonify(_easycredit_mock_preapproved(amount, idn))


@app.route('/api/credit-easycredit/submit', methods=['POST'])
def api_credit_easycredit_submit():
    """Request (EasyCredit): отправка заявки. Реальный EC при user+pass."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    product_id = data.get("product_id")
    program_id = data.get("program_id")
    try:
        pid = int(product_id) if product_id is not None else None
    except (TypeError, ValueError):
        pid = None
    try:
        prid = int(program_id) if program_id is not None else None
    except (TypeError, ValueError):
        prid = None
    amount = int(data.get("amount") or 10000)
    fio = (data.get("fio") or "Тест Тестович Тестов").strip()
    phone = (data.get("phone") or "+37369123456").strip()
    idn = (data.get("idn") or "12345678901234").strip()
    product_name = (data.get("product_name") or "").strip()
    program_name = (data.get("program_name") or "").strip()
    if not product_name and pid:
        pr = CreditController.get_product_by_id(pid)
        if pr.get("success") and pr.get("data"):
            product_name = (pr["data"].get("name") or "Тестовый товар").strip()
    if not product_name:
        product_name = "Тестовый товар"
    if not program_name and prid and pid:
        prog = CreditController.get_programs_for_product(pid)
        if prog.get("success") and prog.get("data"):
            for p in prog["data"]:
                if (p.get("program_id") or p.get("id")) == prid:
                    program_name = (p.get("program_name") or p.get("name") or "0-0-12").strip()
                    break
    if not program_name:
        program_name = "0-0-12"
    _credit_log("easycredit.submit", f"Запрос Submit amount={amount} fio={fio!r} product={product_name!r} program={program_name!r}", "INFO", {"amount": amount, "product_name": product_name, "program_name": program_name})
    base_url = Config.easycredit_base_url()
    user = Config.easycredit_api_user()
    passwd = Config.easycredit_api_password()
    if user and passwd:
        try:
            from integrations.easycredit_client import submit_request as ec_submit
            verify_ssl = Config.easycredit_env() == "production"
            out = ec_submit(
                base_url, user, passwd,
                amount=amount, fio=fio, phone=phone, idn=idn,
                product_name=product_name, program_name=program_name,
                verify_ssl=verify_ssl,
            )
            if out.get("success") and out.get("data", {}).get("urn"):
                urn = out["data"]["urn"]
                _credit_log("easycredit.submit", f"OK urn={urn}", "INFO", {"urn": urn})
                return jsonify({"success": True, "data": out["data"]})
            err = out.get("error") or (out.get("data") or {}).get("message") or "unknown"
            _credit_log("easycredit.submit", f"EC error, fallback to mock: {err}", "WARN", out)
            return jsonify(_easycredit_mock_submit(fio, phone))
        except Exception as e:
            _credit_log("easycredit.submit", f"Exception, fallback to mock: {e}", "ERROR", {"error": str(e)})
            return jsonify(_easycredit_mock_submit(fio, phone))
    _credit_log("easycredit.submit", "Нет user/pass, mock", "INFO", {})
    return jsonify(_easycredit_mock_submit(fio, phone))


@app.route('/api/credit-easycredit/status', methods=['GET'])
def api_credit_easycredit_status():
    """Status (EasyCredit): статус заявки по URN. Реальный EC при user+pass."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    urn = (request.args.get("urn") or "").strip()
    if not urn:
        return jsonify({"success": False, "error": "urn required"}), 400
    _credit_log("easycredit.status", f"Запрос Status urn={urn}", "INFO", {"urn": urn})
    base_url = Config.easycredit_base_url()
    user = Config.easycredit_api_user()
    passwd = Config.easycredit_api_password()
    if user and passwd:
        try:
            from integrations.easycredit_client import status as ec_status
            verify_ssl = Config.easycredit_env() == "production"
            out = ec_status(base_url, user, passwd, urn, verify_ssl=verify_ssl)
            if out.get("success"):
                d = out.get("data") or {}
                st = d.get("status") or ""
                _credit_log("easycredit.status", f"OK status={st}", "INFO", d)
                return jsonify({"success": True, "data": out["data"]})
            err = out.get("error") or (out.get("data") or {}).get("message") or "unknown"
            _credit_log("easycredit.status", f"EC error, fallback to mock: {err}", "WARN", out)
            return jsonify(_easycredit_mock_status(urn))
        except Exception as e:
            _credit_log("easycredit.status", f"Exception, fallback to mock: {e}", "ERROR", {"error": str(e)})
            return jsonify(_easycredit_mock_status(urn))
    _credit_log("easycredit.status", "Нет user/pass, mock", "INFO", {})
    return jsonify(_easycredit_mock_status(urn))


@app.route('/api/credit-easycredit/client-by-phone', methods=['GET'])
def api_credit_easycredit_client_by_phone():
    """Получить информацию о клиенте по телефону (EasyCredit ECM_GetClientInfoByPhone)."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    phone = (request.args.get("phone") or "").strip()
    if not phone:
        return jsonify({"success": False, "error": "phone required"}), 400
    _credit_log("easycredit.client-by-phone", f"Запрос по телефону {phone[:4]}***", "INFO", {"phone_prefix": phone[:4]})
    base_url = Config.easycredit_base_url()
    user = Config.easycredit_api_user()
    passwd = Config.easycredit_api_password()
    if user and passwd:
        try:
            from integrations.easycredit_client import get_client_info_by_phone
            verify_ssl = Config.easycredit_env() == "production"
            out = get_client_info_by_phone(base_url, user, passwd, phone, verify_ssl=verify_ssl)
            if out.get("success"):
                _credit_log("easycredit.client-by-phone", f"OK", "INFO", out.get("data", {}))
                return jsonify(out)
            err = out.get("error") or "unknown"
            _credit_log("easycredit.client-by-phone", f"EC error: {err}", "WARN", out)
            return jsonify(out)
        except Exception as e:
            _credit_log("easycredit.client-by-phone", f"Exception: {e}", "ERROR", {"error": str(e)})
            return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": False, "error": "Нет user/pass"}), 400


@app.route('/api/credit-easycredit/client-by-uin', methods=['GET'])
def api_credit_easycredit_client_by_uin():
    """Получить информацию о клиенте по UIN (IDNP) (EasyCredit eShopClientInfo_v3)."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    uin = (request.args.get("uin") or "").strip()
    if not uin:
        return jsonify({"success": False, "error": "uin required"}), 400
    uin_masked = ("***" + uin[-4:]) if len(uin) >= 4 else "***"
    _credit_log("easycredit.client-by-uin", f"Запрос по UIN {uin_masked}", "INFO", {"uin_masked": uin_masked})
    base_url = Config.easycredit_base_url()
    user = Config.easycredit_api_user()
    passwd = Config.easycredit_api_password()
    if user and passwd:
        try:
            from integrations.easycredit_client import get_client_info
            verify_ssl = Config.easycredit_env() == "production"
            out = get_client_info(base_url, user, passwd, uin, verify_ssl=verify_ssl)
            if out.get("success"):
                _credit_log("easycredit.client-by-uin", f"OK", "INFO", out.get("data", {}))
                return jsonify(out)
            err = out.get("error") or "unknown"
            _credit_log("easycredit.client-by-uin", f"EC error: {err}", "WARN", out)
            return jsonify(out)
        except Exception as e:
            _credit_log("easycredit.client-by-uin", f"Exception: {e}", "ERROR", {"error": str(e)})
            return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": False, "error": "Нет user/pass"}), 400


@app.route('/api/credit-easycredit/urns', methods=['GET'])
def api_credit_easycredit_urns():
    """Получить список заявок (URN) клиента по UIN (EasyCredit ECM_GetUrnPerUin_V2)."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    uin = (request.args.get("uin") or "").strip()
    if not uin:
        return jsonify({"success": False, "error": "uin required"}), 400
    uin_masked = ("***" + uin[-4:]) if len(uin) >= 4 else "***"
    group = request.args.get("group") or ""
    status_filter = request.args.get("status") or ""
    mode = request.args.get("mode") or ""
    _credit_log("easycredit.urns", f"Запрос URNs по UIN {uin_masked}", "INFO", {"uin_masked": uin_masked, "group": group, "status": status_filter, "mode": mode})
    base_url = Config.easycredit_base_url()
    user = Config.easycredit_api_user()
    passwd = Config.easycredit_api_password()
    if user and passwd:
        try:
            from integrations.easycredit_client import get_urns_per_uin
            verify_ssl = Config.easycredit_env() == "production"
            out = get_urns_per_uin(base_url, user, passwd, uin, group=group, status_filter=status_filter, mode=mode, verify_ssl=verify_ssl)
            if out.get("success"):
                _credit_log("easycredit.urns", f"OK", "INFO", out.get("data", {}))
                return jsonify(out)
            err = out.get("error") or "unknown"
            _credit_log("easycredit.urns", f"EC error: {err}", "WARN", out)
            return jsonify(out)
        except Exception as e:
            _credit_log("easycredit.urns", f"Exception: {e}", "ERROR", {"error": str(e)})
            return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": False, "error": "Нет user/pass"}), 400


@app.route('/api/credit-easycredit/test-submit', methods=['POST'])
def api_credit_easycredit_test_submit():
    """Тестовая заявка (EasyCredit): отправка заявки с тестовыми данными."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    # Тестовые данные клиента
    uin = data.get("uin") or "2000000000001"
    fio = data.get("fio") or "Тестов Тест Тестович"
    phone = data.get("phone") or "+37369000001"
    amount = int(data.get("amount") or 15000)
    goods_name = data.get("goods_name") or "Тестовый товар"
    goods_price = int(data.get("goods_price") or amount)
    
    uin_masked = ("***" + uin[-4:]) if len(uin) >= 4 else "***"
    _credit_log("easycredit.test-submit", f"Тестовая заявка UIN={uin_masked} FIO={fio} amount={amount}", "INFO", {
        "uin_masked": uin_masked, "fio": fio, "phone": phone[:6] + "***", "amount": amount, "goods_name": goods_name
    })
    
    base_url = Config.easycredit_base_url()
    user = Config.easycredit_api_user()
    passwd = Config.easycredit_api_password()
    if user and passwd:
        try:
            from integrations.easycredit_client import submit_request
            verify_ssl = Config.easycredit_env() == "production"
            out = submit_request(
                base_url, user, passwd,
                amount=amount, fio=fio, phone=phone, idn=uin,
                product_name=goods_name, program_name="Test", product_id=0, goods_price=goods_price,
                verify_ssl=verify_ssl
            )
            if out.get("success"):
                d = out.get("data") or {}
                urn = d.get("urn") or d.get("URN") or ""
                st = d.get("status") or ""
                _credit_log("easycredit.test-submit", f"OK urn={urn} status={st}", "INFO", d)
                return jsonify(out)
            err = out.get("error") or (out.get("data") or {}).get("message") or "unknown"
            _credit_log("easycredit.test-submit", f"EC error: {err}", "WARN", out)
            return jsonify(out)
        except Exception as e:
            _credit_log("easycredit.test-submit", f"Exception: {e}", "ERROR", {"error": str(e)})
            return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": False, "error": "Нет user/pass"}), 400


def _iute_mock_check_auth():
    return {
        "success": True,
        "data": {
            "partnerId": "mock-partner-id",
            "posId": "mock-pos-id",
            "products": [{"name": "Flexi", "description": "Mock product"}],
        },
        "fallback": True,
    }


def _iute_mock_create_order(order_id: str, phone: str):
    return {
        "success": True,
        "data": {
            "status": "PENDING",
            "message": "Mock: заказ создан (тест).",
            "myiuteCustomer": True,
        },
        "fallback": True,
    }


def _iute_mock_order_status(order_id: str):
    return {
        "success": True,
        "data": {
            "orderId": order_id,
            "status": "PENDING",
            "productName": None,
            "loanDuration": None,
        },
        "fallback": True,
    }


@app.route('/api/credit-iute/check-auth', methods=['GET'])
def api_credit_iute_check_auth():
    """Check Auth (Iute): проверка авторизации и получение информации о партнёре."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    _credit_log("iute.check-auth", "Запрос Check Auth", "INFO", {})
    base_url = Config.iute_base_url()
    api_key = Config.iute_api_key()
    if api_key:
        try:
            from integrations.iute_client import check_auth as iute_check_auth
            out = iute_check_auth(base_url, api_key)
            if out.get("success"):
                d = out["data"] or {}
                _credit_log("iute.check-auth", f"OK partnerId={d.get('partnerId')} posId={d.get('posId')}", "INFO", d)
                return jsonify({"success": True, "data": out["data"]})
            err = out.get("error") or "unknown"
            _credit_log("iute.check-auth", f"Iute error, fallback to mock: {err}", "WARN", out)
            return jsonify(_iute_mock_check_auth())
        except Exception as e:
            _credit_log("iute.check-auth", f"Exception, fallback to mock: {e}", "ERROR", {"error": str(e)})
            return jsonify(_iute_mock_check_auth())
    _credit_log("iute.check-auth", "Нет API key, mock", "INFO", {})
    return jsonify(_iute_mock_check_auth())


@app.route('/api/credit-iute/create-order', methods=['POST'])
def api_credit_iute_create_order():
    """Create Order (Iute): создание или обновление заказа."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    data = request.get_json() or {}
    order_id = (data.get("order_id") or f"order-{int(__import__('time').time() * 1000)}").strip()
    myiute_phone = (data.get("myiute_phone") or "+37369123456").strip()
    total_amount = int(data.get("total_amount") or 1000)
    currency = (data.get("currency") or "EUR").strip()
    pos_identifier = Config.iute_pos_identifier()
    salesman_identifier = Config.iute_salesman_identifier()
    user_pin = (data.get("user_pin") or "").strip() or None
    birthday = (data.get("birthday") or "").strip() or None
    gender = (data.get("gender") or "").strip() or None
    items = data.get("items") or []
    if not items and data.get("product_id"):
        # Попытка получить информацию о товаре
        try:
            pid = int(data.get("product_id"))
            pr = CreditController.get_product_by_id(pid)
            if pr.get("success") and pr.get("data"):
                product_name = pr["data"].get("name") or "Товар"
                product_price = pr["data"].get("price") or total_amount
                items = [{
                    "displayName": product_name,
                    "id": str(pid),
                    "sku": None,
                    "unitPrice": product_price,
                    "qty": 1,
                    "itemImageUrl": None,
                    "itemUrl": None,
                }]
        except Exception:
            pass
    _credit_log("iute.create-order", f"Запрос Create Order order_id={order_id} phone={myiute_phone} amount={total_amount} currency={currency}", "INFO", {"order_id": order_id, "myiute_phone": myiute_phone, "total_amount": total_amount, "currency": currency})
    base_url = Config.iute_base_url()
    api_key = Config.iute_api_key()
    if api_key and pos_identifier and salesman_identifier:
        try:
            from integrations.iute_client import create_order as iute_create_order
            out = iute_create_order(
                base_url, api_key,
                order_id=order_id,
                myiute_phone=myiute_phone,
                total_amount=total_amount,
                currency=currency,
                pos_identifier=pos_identifier,
                salesman_identifier=salesman_identifier,
                user_pin=user_pin,
                birthday=birthday,
                gender=gender,
                items=items,
            )
            if out.get("success"):
                d = out["data"] or {}
                status = d.get("status", "PENDING")
                _credit_log("iute.create-order", f"OK order_id={order_id} status={status} customer={d.get('myiuteCustomer')}", "INFO", d)
                return jsonify({"success": True, "data": out["data"]})
            err = out.get("error") or (out.get("data") or {}).get("message") or "unknown"
            _credit_log("iute.create-order", f"Iute error, fallback to mock: {err}", "WARN", out)
            return jsonify(_iute_mock_create_order(order_id, myiute_phone))
        except Exception as e:
            _credit_log("iute.create-order", f"Exception, fallback to mock: {e}", "ERROR", {"error": str(e)})
            return jsonify(_iute_mock_create_order(order_id, myiute_phone))
    _credit_log("iute.create-order", "Нет API key/POS/Salesman, mock", "INFO", {})
    return jsonify(_iute_mock_create_order(order_id, myiute_phone))


@app.route('/api/credit-iute/order-status', methods=['GET'])
def api_credit_iute_order_status():
    """Order Status (Iute): проверка статуса заказа."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Authentication required"}), 401
    order_id = (request.args.get("order_id") or "").strip()
    if not order_id:
        return jsonify({"success": False, "error": "order_id required"}), 400
    _credit_log("iute.order-status", f"Запрос Order Status order_id={order_id}", "INFO", {"order_id": order_id})
    base_url = Config.iute_base_url()
    api_key = Config.iute_api_key()
    if api_key:
        try:
            from integrations.iute_client import get_order_status as iute_get_status
            out = iute_get_status(base_url, api_key, order_id)
            if out.get("success"):
                d = out["data"] or {}
                status = d.get("status", "")
                _credit_log("iute.order-status", f"OK order_id={order_id} status={status}", "INFO", d)
                return jsonify({"success": True, "data": out["data"]})
            err = out.get("error") or "unknown"
            _credit_log("iute.order-status", f"Iute error, fallback to mock: {err}", "WARN", out)
            return jsonify(_iute_mock_order_status(order_id))
        except Exception as e:
            _credit_log("iute.order-status", f"Exception, fallback to mock: {e}", "ERROR", {"error": str(e)})
            return jsonify(_iute_mock_order_status(order_id))
    _credit_log("iute.order-status", "Нет API key, mock", "INFO", {})
    return jsonify(_iute_mock_order_status(order_id))


@app.route('/api/ai-status', methods=['GET'])
def api_ai_status():
    """API endpoint для проверки доступности ИИ"""
    if not AuthController.is_authenticated():
        return jsonify({"error": "Authentication required"}), 401
    
    try:
        # Пытаемся использовать основной ai_helper, если не получается - используем серверный
        is_server = False
        ai_available = False
        
        try:
            from ai_helper import is_ai_available, IS_SERVER
            is_server = IS_SERVER
            ai_available = is_ai_available()
        except ImportError:
            # На сервере может не быть ai_helper, используем серверный вариант
            try:
                from ai_helper_server import is_ai_available
                is_server = True
                ai_available = is_ai_available()  # Всегда False для сервера
            except ImportError:
                return jsonify({
                    "success": False,
                    "error": "AI helper module not found"
                }), 500
        
        return jsonify({
            "success": True,
            "ai_available": ai_available,
            "is_server": is_server
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/set-language', methods=['POST'])
def set_language():
    """Устанавливает язык интерфейса"""
    data = request.get_json()
    language = data.get('lang', 'ru')
    
    if language in Config.SUPPORTED_LANGUAGES:
        session['language'] = language
        return jsonify({"success": True, "language": language})
    else:
        return jsonify({"success": False, "error": "Unsupported language"}), 400


@app.route('/api/get-language', methods=['GET'])
def get_language():
    """Получает текущий язык интерфейса"""
    current_lang = session.get('language', Config.BABEL_DEFAULT_LOCALE)
    return jsonify({
        "success": True,
        "language": current_lang,
        "supported_languages": Config.SUPPORTED_LANGUAGES,
        "languages": Config.LANGUAGES
    })


@app.route('/api/system/version-info', methods=['GET'])
def api_system_version_info():
    path = request.args.get('path') or request.path or ''
    return jsonify(VersionRegistry.for_path(path))


@app.route('/api/restart', methods=['POST'])
def restart_server():
    """Перезапускает сервер"""
    def restart():
        time.sleep(1)
        print("Перезапуск сервера по запросу пользователя...")
        import sys
        import os
        os.execv(sys.executable, [sys.executable] + sys.argv)

    threading.Thread(target=restart).start()
    return jsonify({"success": True, "message": "Server is restarting..."})


# ---------------------------------------------------------------------------
# AGRO module — imports
# ---------------------------------------------------------------------------
from controllers.agro_admin_controller import AgroAdminController
from controllers.agro_field_controller import AgroFieldController
from controllers.agro_warehouse_controller import AgroWarehouseController
from controllers.agro_qa_controller import AgroQaController
from controllers.agro_sales_controller import AgroSalesController
from models.agro_oracle_store import AgroStore


# ---------------------------------------------------------------------------
# AGRO — UI routes
# ---------------------------------------------------------------------------

@app.route('/UNA.md/orasldev/agro')
def agro_mdi():
    """AGRO: MDI shell — all AGRO modules in tabbed interface."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('agro_mdi.html')


@app.route('/UNA.md/orasldev/agro-admin')
def agro_admin():
    """AGRO: admin — references, settings, reports."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    embed = request.args.get('embed') == '1'
    return render_template('agro_admin.html', embed=embed)


@app.route('/UNA.md/orasldev/agro-field')
def agro_field():
    """AGRO: field — purchases, barcodes, crates, offline sync."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    embed = request.args.get('embed') == '1'
    return render_template('agro_field.html', embed=embed)


@app.route('/UNA.md/orasldev/agro-warehouse')
def agro_warehouse():
    """AGRO: warehouse — stock, movements, readings, tasks."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    embed = request.args.get('embed') == '1'
    return render_template('agro_warehouse.html', embed=embed)


@app.route('/UNA.md/orasldev/agro-qa')
def agro_qa():
    """AGRO: QA / HACCP — checklists, checks, batch blocks."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    embed = request.args.get('embed') == '1'
    return render_template('agro_qa.html', embed=embed)


@app.route('/UNA.md/orasldev/agro-sales')
def agro_sales():
    """AGRO: sales — shipments, export, batch allocation."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    embed = request.args.get('embed') == '1'
    return render_template('agro_sales.html', embed=embed)


@app.route('/UNA.md/orasldev/agro-document/<int:doc_id>')
def agro_document_print(doc_id):
    """AGRO: Print document — renders A4-optimized document templates."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    doc_type = request.args.get('type', 'purchase_act')
    allowed_types = [
        'purchase_act', 'weight_ticket', 'shipping_note', 'invoice',
        'export_decl', 'qa_protocol', 'gmp_checklist', 'haccp_report',
        'mass_balance',
    ]
    if doc_type not in allowed_types:
        return jsonify({"success": False, "error": f"Unknown document type: {doc_type}"}), 400
    data = AgroStore.get_document_data(doc_type, doc_id) if hasattr(AgroStore, 'get_document_data') else {}
    if not isinstance(data, dict):
        data = {}
    template = f"agro/document_{doc_type}.html"
    return render_template(template, **data)


# ---------------------------------------------------------------------------
# AGRO — Admin API: reference tables CRUD
# ---------------------------------------------------------------------------

# --- Suppliers ---
@app.route('/api/agro-admin/suppliers', methods=['GET'])
def api_agro_admin_suppliers():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    active_only = request.args.get('active_only', '0') == '1'
    return jsonify(AgroAdminController.get_suppliers(active_only))

@app.route('/api/agro-admin/suppliers', methods=['POST'])
def api_agro_admin_suppliers_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_supplier(request.get_json() or {}))

@app.route('/api/agro-admin/suppliers/<int:record_id>', methods=['DELETE'])
def api_agro_admin_supplier_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_supplier(record_id))

# --- Customers ---
@app.route('/api/agro-admin/customers', methods=['GET'])
def api_agro_admin_customers():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    active_only = request.args.get('active_only', '0') == '1'
    return jsonify(AgroAdminController.get_customers(active_only))

@app.route('/api/agro-admin/customers', methods=['POST'])
def api_agro_admin_customers_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_customer(request.get_json() or {}))

@app.route('/api/agro-admin/customers/<int:record_id>', methods=['DELETE'])
def api_agro_admin_customer_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_customer(record_id))

# --- Warehouses ---
@app.route('/api/agro-admin/warehouses', methods=['GET'])
def api_agro_admin_warehouses():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    active_only = request.args.get('active_only', '0') == '1'
    return jsonify(AgroAdminController.get_warehouses(active_only))

@app.route('/api/agro-admin/warehouses', methods=['POST'])
def api_agro_admin_warehouses_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_warehouse(request.get_json() or {}))

@app.route('/api/agro-admin/warehouses/<int:record_id>', methods=['DELETE'])
def api_agro_admin_warehouse_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_warehouse(record_id))

# --- Storage Cells ---
@app.route('/api/agro-admin/storage-cells', methods=['GET'])
def api_agro_admin_storage_cells():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    wh_id = request.args.get('warehouse_id', type=int)
    return jsonify(AgroAdminController.get_storage_cells(wh_id))

@app.route('/api/agro-admin/storage-cells', methods=['POST'])
def api_agro_admin_storage_cells_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_storage_cell(request.get_json() or {}))

@app.route('/api/agro-admin/storage-cells/<int:record_id>', methods=['DELETE'])
def api_agro_admin_storage_cell_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_storage_cell(record_id))

# --- Items ---
@app.route('/api/agro-admin/items', methods=['GET'])
def api_agro_admin_items():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    active_only = request.args.get('active_only', '0') == '1'
    return jsonify(AgroAdminController.get_items(active_only))

@app.route('/api/agro-admin/items', methods=['POST'])
def api_agro_admin_items_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_item(request.get_json() or {}))

@app.route('/api/agro-admin/items/<int:record_id>', methods=['DELETE'])
def api_agro_admin_item_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_item(record_id))

# --- Packaging Types ---
@app.route('/api/agro-admin/packaging-types', methods=['GET'])
def api_agro_admin_packaging_types():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    active_only = request.args.get('active_only', '0') == '1'
    return jsonify(AgroAdminController.get_packaging_types(active_only))

@app.route('/api/agro-admin/packaging-types', methods=['POST'])
def api_agro_admin_packaging_types_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_packaging_type(request.get_json() or {}))

@app.route('/api/agro-admin/packaging-types/<int:record_id>', methods=['DELETE'])
def api_agro_admin_packaging_type_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_packaging_type(record_id))

# --- Vehicles ---
@app.route('/api/agro-admin/vehicles', methods=['GET'])
def api_agro_admin_vehicles():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    active_only = request.args.get('active_only', '0') == '1'
    return jsonify(AgroAdminController.get_vehicles(active_only))

@app.route('/api/agro-admin/vehicles', methods=['POST'])
def api_agro_admin_vehicles_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_vehicle(request.get_json() or {}))

@app.route('/api/agro-admin/vehicles/<int:record_id>', methods=['DELETE'])
def api_agro_admin_vehicle_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_vehicle(record_id))

# --- Currencies ---
@app.route('/api/agro-admin/currencies', methods=['GET'])
def api_agro_admin_currencies():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    active_only = request.args.get('active_only', '0') == '1'
    return jsonify(AgroAdminController.get_currencies(active_only))

@app.route('/api/agro-admin/currencies', methods=['POST'])
def api_agro_admin_currencies_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_currency(request.get_json() or {}))

@app.route('/api/agro-admin/currencies/<int:record_id>', methods=['DELETE'])
def api_agro_admin_currency_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_currency(record_id))

# --- Exchange Rates ---
@app.route('/api/agro-admin/exchange-rates', methods=['GET'])
def api_agro_admin_exchange_rates():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.get_exchange_rates())

@app.route('/api/agro-admin/exchange-rates', methods=['POST'])
def api_agro_admin_exchange_rates_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_exchange_rate(request.get_json() or {}))

@app.route('/api/agro-admin/exchange-rates/<int:record_id>', methods=['DELETE'])
def api_agro_admin_exchange_rate_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_exchange_rate(record_id))

# --- Formula Params ---
@app.route('/api/agro-admin/formula-params', methods=['GET'])
def api_agro_admin_formula_params():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.get_formula_params())

@app.route('/api/agro-admin/formula-params', methods=['POST'])
def api_agro_admin_formula_params_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_formula_param(request.get_json() or {}))

@app.route('/api/agro-admin/formula-params/<int:record_id>', methods=['DELETE'])
def api_agro_admin_formula_param_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_formula_param(record_id))

# --- Module Config ---
@app.route('/api/agro-admin/module-config', methods=['GET'])
def api_agro_admin_module_config():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.get_module_config())

@app.route('/api/agro-admin/module-config', methods=['POST'])
def api_agro_admin_module_config_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_module_config(request.get_json() or {}))

@app.route('/api/agro-admin/module-config/<int:record_id>', methods=['DELETE'])
def api_agro_admin_module_config_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_module_config(record_id))

# --- AGRO Admin: Item Varieties ---
@app.route('/api/agro-admin/item-varieties', methods=['GET'])
def api_agro_admin_item_varieties():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    item_id = request.args.get('item_id', type=int)
    active_only = request.args.get('active_only', '0') == '1'
    return jsonify(AgroAdminController.get_item_varieties(item_id, active_only))

@app.route('/api/agro-admin/item-varieties', methods=['POST'])
def api_agro_admin_item_varieties_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_item_variety(request.get_json() or {}))

@app.route('/api/agro-admin/item-varieties/<int:record_id>', methods=['DELETE'])
def api_agro_admin_item_variety_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_item_variety(record_id))

# --- AGRO Admin: Acceptance Profiles ---
@app.route('/api/agro-admin/acceptance-profiles', methods=['GET'])
def api_agro_admin_acceptance_profiles():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    active_only = request.args.get('active_only', '0') == '1'
    return jsonify(AgroAdminController.get_acceptance_profiles(active_only))

@app.route('/api/agro-admin/acceptance-profiles', methods=['POST'])
def api_agro_admin_acceptance_profiles_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.upsert_acceptance_profile(request.get_json() or {}))

@app.route('/api/agro-admin/acceptance-profiles/<int:record_id>', methods=['DELETE'])
def api_agro_admin_acceptance_profile_delete(record_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroAdminController.delete_acceptance_profile(record_id))

# ---------------------------------------------------------------------------
# AGRO — Field API: barcodes, crates, purchases, offline sync
# ---------------------------------------------------------------------------

# --- AGRO Field API ---
@app.route('/api/agro-field/barcodes/generate', methods=['GET'])
def api_agro_barcodes_gen():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    count = int(request.args.get('count', 10))
    bc_type = request.args.get('type', 'internal')
    return jsonify(AgroFieldController.generate_barcodes(count, bc_type))

@app.route('/api/agro-field/barcodes/print-batch', methods=['GET'])
def api_agro_barcodes_print():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    ids = request.args.getlist('ids')
    return jsonify(AgroFieldController.get_barcode_print_batch(ids))

@app.route('/api/agro-field/crates/scan', methods=['POST'])
def api_agro_crate_scan():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.get_json() or {}
    return jsonify(AgroFieldController.scan_crate(data.get('barcode', '')))

@app.route('/api/agro-field/crates/weigh', methods=['POST'])
def api_agro_crate_weigh():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.register_crate(request.get_json() or {}))

@app.route('/api/agro-field/purchases', methods=['GET'])
def api_agro_field_purchases():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.get_purchases(request.args.to_dict()))

@app.route('/api/agro-field/purchases', methods=['POST'])
def api_agro_field_purchase_create():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.create_purchase(request.get_json() or {}))

@app.route('/api/agro-field/purchases/<int:doc_id>', methods=['GET'])
def api_agro_field_purchase_get(doc_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.get_purchase_by_id(doc_id))

@app.route('/api/agro-field/purchases/<int:doc_id>', methods=['PUT'])
def api_agro_field_purchase_update(doc_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.get_json() or {}
    data['id'] = doc_id
    return jsonify(AgroFieldController.update_purchase(data))

@app.route('/api/agro-field/purchases/<int:doc_id>/confirm', methods=['PUT'])
def api_agro_field_purchase_confirm(doc_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.confirm_purchase(doc_id))

@app.route('/api/agro-field/purchases/<int:doc_id>/cancel', methods=['PUT'])
def api_agro_field_purchase_cancel(doc_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.cancel_purchase(doc_id))

@app.route('/api/agro-field/sync', methods=['POST'])
def api_agro_field_sync():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.sync_offline_queue(request.get_json() or {}))

@app.route('/api/agro-field/sync/references', methods=['GET'])
def api_agro_field_sync_refs():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.get_sync_references())

# --- AGRO Field: Field Requests ---
@app.route('/api/agro-field/requests', methods=['GET'])
def api_agro_field_requests():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.get_field_requests(request.args.to_dict()))

@app.route('/api/agro-field/requests', methods=['POST'])
def api_agro_field_request_create():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.create_field_request(request.get_json() or {}))

@app.route('/api/agro-field/requests/<int:request_id>', methods=['GET'])
def api_agro_field_request_get(request_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.get_field_request_by_id(request_id))

@app.route('/api/agro-field/requests/<int:request_id>', methods=['PUT'])
def api_agro_field_request_update(request_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.get_json() or {}
    data['id'] = request_id
    return jsonify(AgroFieldController.update_field_request(data))

@app.route('/api/agro-field/requests/<int:request_id>/approve', methods=['PUT'])
def api_agro_field_request_approve(request_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.get_json() or {}
    return jsonify(AgroFieldController.approve_field_request(request_id, data.get('approved_by')))

@app.route('/api/agro-field/requests/<int:request_id>/cancel', methods=['PUT'])
def api_agro_field_request_cancel(request_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.cancel_field_request(request_id))

# --- AGRO Field: Batch Inspections ---
@app.route('/api/agro-field/inspections', methods=['GET'])
def api_agro_field_inspections():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    batch_id = request.args.get('batch_id', type=int)
    return jsonify(AgroFieldController.get_batch_inspections(batch_id))

@app.route('/api/agro-field/inspections', methods=['POST'])
def api_agro_field_inspection_perform():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.perform_batch_inspection(request.get_json() or {}))

@app.route('/api/agro-field/inspections/<int:inspection_id>', methods=['GET'])
def api_agro_field_inspection_detail(inspection_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroFieldController.get_batch_inspection_detail(inspection_id))


# ============================================================
# AGRO Scale Emulator API
# ============================================================
from services.scale_emulator import get_scale, list_scales, create_scale


@app.route('/api/agro-scale/read', methods=['GET'])
@limiter.exempt
def api_agro_scale_read():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    scale_id = request.args.get('scale_id', 'default')
    scale = get_scale(scale_id)
    return jsonify({"success": True, "data": scale.read().to_dict()})


@app.route('/api/agro-scale/zero', methods=['POST'])
def api_agro_scale_zero():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    scale_id = (request.json or {}).get('scale_id', 'default')
    return jsonify(get_scale(scale_id).zero())


@app.route('/api/agro-scale/tare', methods=['POST'])
def api_agro_scale_tare():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    scale_id = (request.json or {}).get('scale_id', 'default')
    return jsonify(get_scale(scale_id).tare())


@app.route('/api/agro-scale/capture', methods=['POST'])
@limiter.exempt
def api_agro_scale_capture():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    scale_id = (request.json or {}).get('scale_id', 'default')
    return jsonify(get_scale(scale_id).capture())


@app.route('/api/agro-scale/simulate', methods=['POST'])
@limiter.exempt
def api_agro_scale_simulate():
    """Place a simulated weight on the scale (emulator only)."""
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.json or {}
    scale_id = data.get('scale_id', 'default')
    scale = get_scale(scale_id)
    weight = data.get('weight_kg')
    if weight is not None:
        return jsonify(scale.simulate_load(float(weight)))
    if data.get('random'):
        return jsonify(scale.simulate_random_load(
            min_kg=float(data.get('min_kg', 5)),
            max_kg=float(data.get('max_kg', 50)),
        ))
    if data.get('remove'):
        return jsonify(scale.simulate_remove())
    return jsonify({"success": False, "error": "Provide weight_kg, random, or remove"})


@app.route('/api/agro-scale/config', methods=['GET'])
def api_agro_scale_config():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    scale_id = request.args.get('scale_id', 'default')
    return jsonify({"success": True, "data": get_scale(scale_id).get_config()})


@app.route('/api/agro-scale/list', methods=['GET'])
def api_agro_scale_list():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify({"success": True, "data": list_scales()})


# ============================================================
# AGRO Warehouse API
# ============================================================

@app.route('/api/agro-warehouse/stock', methods=['GET'])
def api_agro_wh_stock():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    filters = {}
    if request.args.get('warehouse_id'):
        filters['warehouse_id'] = int(request.args['warehouse_id'])
    if request.args.get('item_id'):
        filters['item_id'] = int(request.args['item_id'])
    if request.args.get('status'):
        filters['status'] = request.args['status']
    return jsonify(AgroWarehouseController.get_stock_balance(filters or None))

@app.route('/api/agro-warehouse/batches/<int:batch_id>', methods=['GET'])
def api_agro_wh_batch(batch_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroWarehouseController.get_batch_by_id(batch_id))

@app.route('/api/agro-warehouse/batches/<int:batch_id>/history', methods=['GET'])
def api_agro_wh_batch_history(batch_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroWarehouseController.get_batch_history(batch_id))

@app.route('/api/agro-warehouse/movements', methods=['POST'])
def api_agro_wh_movement():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroWarehouseController.create_movement(request.json))

@app.route('/api/agro-warehouse/receive', methods=['POST'])
def api_agro_wh_receive():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroWarehouseController.receive_crates(request.json))

@app.route('/api/agro-warehouse/readings', methods=['GET'])
def api_agro_wh_readings_get():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    cell_id = request.args.get('cell_id', type=int)
    return jsonify(AgroWarehouseController.get_readings(
        cell_id, request.args.get('date_from'), request.args.get('date_to')
    ))

@app.route('/api/agro-warehouse/readings', methods=['POST'])
def api_agro_wh_readings_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroWarehouseController.add_reading(request.json))

@app.route('/api/agro-warehouse/alerts', methods=['GET'])
def api_agro_wh_alerts():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    ack = request.args.get('acknowledged')
    return jsonify(AgroWarehouseController.get_alerts(ack))

@app.route('/api/agro-warehouse/alerts/<int:alert_id>/ack', methods=['PUT'])
def api_agro_wh_alert_ack(alert_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.json or {}
    return jsonify(AgroWarehouseController.acknowledge_alert(alert_id, data.get('user')))

@app.route('/api/agro-warehouse/tasks', methods=['GET'])
def api_agro_wh_tasks_get():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    filters = {}
    if request.args.get('status'):
        filters['status'] = request.args['status']
    if request.args.get('batch_id'):
        filters['batch_id'] = int(request.args['batch_id'])
    if request.args.get('task_type'):
        filters['task_type'] = request.args['task_type']
    return jsonify(AgroWarehouseController.get_processing_tasks(filters or None))

@app.route('/api/agro-warehouse/tasks', methods=['POST'])
def api_agro_wh_tasks_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroWarehouseController.create_processing_task(request.json))

@app.route('/api/agro-warehouse/tasks/<int:task_id>/status', methods=['PUT'])
def api_agro_wh_task_status(task_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.json or {}
    return jsonify(AgroWarehouseController.update_task_status(
        task_id, data.get('status', ''),
        data.get('output_qty'), data.get('waste_qty')
    ))


# ============================================================
# AGRO Sales API
# ============================================================

@app.route('/api/agro-sales/documents', methods=['GET'])
def api_agro_sales_docs():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    filters = {}
    if request.args.get('status'):
        filters['status'] = request.args['status']
    if request.args.get('customer_id'):
        filters['customer_id'] = int(request.args['customer_id'])
    if request.args.get('date_from'):
        filters['date_from'] = request.args['date_from']
    if request.args.get('date_to'):
        filters['date_to'] = request.args['date_to']
    return jsonify(AgroSalesController.get_sales_docs(filters or None))

@app.route('/api/agro-sales/documents', methods=['POST'])
def api_agro_sales_doc_create():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.create_sales_doc(request.json))

@app.route('/api/agro-sales/documents/<int:doc_id>', methods=['GET'])
def api_agro_sales_doc_get(doc_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.get_sales_doc_by_id(doc_id))

@app.route('/api/agro-sales/documents/<int:doc_id>/confirm', methods=['PUT'])
def api_agro_sales_doc_confirm(doc_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.confirm_sales_doc(doc_id))

@app.route('/api/agro-sales/allocate', methods=['POST'])
def api_agro_sales_allocate():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.allocate_batches(request.json))

@app.route('/api/agro-sales/available-stock', methods=['GET'])
def api_agro_sales_avail_stock():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.get_available_stock(
        request.args.get('item_id', type=int),
        request.args.get('warehouse_id', type=int)
    ))

@app.route('/api/agro-sales/export-decl', methods=['POST'])
def api_agro_sales_export_create():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.create_export_decl(request.json))

@app.route('/api/agro-sales/export-decl/<int:decl_id>', methods=['GET'])
def api_agro_sales_export_get(decl_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.get_export_decl(decl_id))

@app.route('/api/agro-sales/export-decl/<int:decl_id>', methods=['PUT'])
def api_agro_sales_export_update(decl_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.json or {}
    data['id'] = decl_id
    return jsonify(AgroSalesController.update_export_decl(data))

# --- Weight Tickets ---

@app.route('/api/agro-sales/weight-tickets', methods=['GET'])
def api_agro_sales_wt_list():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    filters = {}
    for k in ('status', 'date_from', 'date_to'):
        if request.args.get(k):
            filters[k] = request.args[k]
    return jsonify(AgroSalesController.get_weight_tickets(filters or None))

@app.route('/api/agro-sales/weight-tickets', methods=['POST'])
def api_agro_sales_wt_create():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.create_weight_ticket(request.json))

@app.route('/api/agro-sales/weight-tickets/<int:tid>', methods=['GET'])
def api_agro_sales_wt_get(tid):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.get_weight_ticket_by_id(tid))

@app.route('/api/agro-sales/weight-tickets/<int:tid>', methods=['PUT'])
def api_agro_sales_wt_update(tid):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.update_weight_ticket(tid, request.json))

@app.route('/api/agro-sales/weight-tickets/<int:tid>/lines', methods=['POST'])
def api_agro_sales_wt_add_line(tid):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.add_weight_line(tid, request.json))

@app.route('/api/agro-sales/weight-tickets/<int:tid>/lines/<int:lid>', methods=['DELETE'])
def api_agro_sales_wt_del_line(tid, lid):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.remove_weight_line(tid, lid))

@app.route('/api/agro-sales/weight-tickets/<int:tid>/finalize', methods=['POST'])
def api_agro_sales_wt_finalize(tid):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.finalize_weight_ticket(tid))

@app.route('/api/agro-field/scoring-config', methods=['GET'])
def api_agro_field_scoring_config():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroSalesController.get_scoring_config())

# ============================================================
# AGRO QA / HACCP API
# ============================================================

@app.route('/api/agro-qa/checklists', methods=['GET'])
def api_agro_qa_checklists():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    ctype = request.args.get('type')
    return jsonify(AgroQaController.get_checklists(ctype))

@app.route('/api/agro-qa/checklists', methods=['POST'])
def api_agro_qa_checklists_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.upsert_checklist(request.json or {}))

@app.route('/api/agro-qa/checklists/<int:cl_id>', methods=['GET'])
def api_agro_qa_checklist_detail(cl_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.get_checklist_by_id(cl_id))

@app.route('/api/agro-qa/checklists/<int:cl_id>', methods=['DELETE'])
def api_agro_qa_checklist_delete(cl_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.delete_checklist(cl_id))

@app.route('/api/agro-qa/checks', methods=['GET'])
def api_agro_qa_checks():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    batch_id = request.args.get('batch_id', type=int)
    return jsonify(AgroQaController.get_checks(batch_id))

@app.route('/api/agro-qa/checks', methods=['POST'])
def api_agro_qa_checks_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.perform_check(request.json or {}))

@app.route('/api/agro-qa/checks/<int:check_id>', methods=['GET'])
def api_agro_qa_check_detail(check_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.get_check_detail(check_id))

@app.route('/api/agro-qa/batches/<int:batch_id>/block', methods=['POST'])
def api_agro_qa_block(batch_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.json or {}
    return jsonify(AgroQaController.block_batch(batch_id, data.get('reason', ''), data.get('blocked_by')))

@app.route('/api/agro-qa/batches/<int:batch_id>/unblock', methods=['POST'])
def api_agro_qa_unblock(batch_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    data = request.json or {}
    return jsonify(AgroQaController.unblock_batch(batch_id, data.get('unblocked_by'), data.get('resolution')))

@app.route('/api/agro-qa/blocks', methods=['GET'])
def api_agro_qa_blocks():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    active_only = request.args.get('active_only', '1') == '1'
    return jsonify(AgroQaController.get_batch_blocks(active_only))

@app.route('/api/agro-qa/haccp/plans', methods=['GET'])
def api_agro_qa_haccp_plans():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.get_haccp_plans())

@app.route('/api/agro-qa/haccp/plans', methods=['POST'])
def api_agro_qa_haccp_plans_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.upsert_haccp_plan(request.json or {}))

@app.route('/api/agro-qa/haccp/plans/<int:plan_id>/ccps', methods=['GET'])
def api_agro_qa_ccps(plan_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.get_ccps(plan_id))

@app.route('/api/agro-qa/haccp/ccps', methods=['POST'])
def api_agro_qa_ccps_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.upsert_ccp(request.json or {}))

@app.route('/api/agro-qa/haccp/records', methods=['POST'])
def api_agro_qa_haccp_record():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.record_haccp_measurement(request.json or {}))

@app.route('/api/agro-qa/haccp/deviations', methods=['GET'])
def api_agro_qa_deviations():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    return jsonify(AgroQaController.get_haccp_deviations(
        request.args.get('date_from'), request.args.get('date_to')))


# ============================================================
# AGRO Reports API
# ============================================================

@app.route('/api/agro-admin/reports/<report_type>', methods=['GET'])
def api_agro_report(report_type):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    allowed = ['purchases', 'sales', 'mass_balance', 'stock', 'expiry']
    if report_type not in allowed:
        return jsonify({"success": False, "error": "Unknown report type"}), 404
    filters = {k: v for k, v in request.args.items()}
    method = getattr(AgroStore, f'report_{report_type}', None)
    if not method:
        return jsonify({"success": False, "error": "Not implemented"}), 404
    return jsonify(method(**filters))

@app.route('/api/agro-admin/reports/export/<report_type>', methods=['GET'])
def api_agro_report_export(report_type):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Auth required"}), 401
    allowed = ['purchases', 'sales', 'mass_balance', 'stock', 'expiry']
    if report_type not in allowed:
        return jsonify({"success": False, "error": "Unknown report type"}), 404
    fmt = request.args.get('format', 'xlsx')
    filters = {k: v for k, v in request.args.items() if k != 'format'}
    data = AgroStore.export_report(report_type, fmt, filters)
    if data is None:
        return jsonify({"success": False, "error": "Export failed"}), 500
    import io as _io
    mime = {'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'csv': 'text/csv'}.get(fmt, 'application/octet-stream')
    return send_file(_io.BytesIO(data), mimetype=mime,
                     as_attachment=True,
                     download_name=f'agro_{report_type}.{fmt}')


# ============================================================
# AGRO Socket.io Events
# ============================================================

def agro_emit(event_type, data):
    """Broadcast AGRO event to all connected clients."""
    try:
        socketio.emit('agro_event', {
            'type': event_type,
            'data': data,
            'timestamp': __import__('datetime').datetime.now().isoformat()
        })
    except Exception:
        pass  # Don't fail operations if socket emit fails

# Monkey-patch store methods to emit events after successful operations
_orig_add_reading = AgroStore.add_reading.__func__ if hasattr(AgroStore.add_reading, '__func__') else AgroStore.add_reading

@staticmethod
def _patched_add_reading(data):
    result = _orig_add_reading(data)
    if result.get('success'):
        alerts = result.get('data', {}).get('alerts', [])
        if alerts:
            agro_emit('temp_alert', {
                'cell_id': data.get('cell_id'),
                'alerts_count': len(alerts),
                'message': f"Температурный алерт! {len(alerts)} нарушений"
            })
    return result
AgroStore.add_reading = _patched_add_reading

_orig_block_batch = AgroStore.block_batch.__func__ if hasattr(AgroStore.block_batch, '__func__') else AgroStore.block_batch

@staticmethod
def _patched_block_batch(batch_id, reason, blocked_by=None):
    result = _orig_block_batch(batch_id, reason, blocked_by)
    if result.get('success'):
        agro_emit('batch_blocked', {
            'batch_id': batch_id,
            'reason': reason,
            'blocked_by': blocked_by or 'operator',
            'message': f"Партия #{batch_id} заблокирована: {reason}"
        })
    return result
AgroStore.block_batch = _patched_block_batch

_orig_confirm_sales = AgroStore.confirm_sales_doc.__func__ if hasattr(AgroStore.confirm_sales_doc, '__func__') else AgroStore.confirm_sales_doc

@staticmethod
def _patched_confirm_sales(doc_id):
    result = _orig_confirm_sales(doc_id)
    if result.get('success'):
        agro_emit('sales_confirmed', {
            'doc_id': doc_id,
            'message': f"Документ отгрузки #{doc_id} подтверждён"
        })
    return result
AgroStore.confirm_sales_doc = _patched_confirm_sales


# ---------------------------------------------------------------------------
# AEI — Asociații de Economii și Împrumut
# ---------------------------------------------------------------------------
from controllers.aei_controller import AEIController


@app.route('/UNA.md/orasldev/aei')
@app.route('/UNA.md/orasldev/aei-admin')
def aei_admin():
    """AEÎ: main admin — deposits, loans, members, accounting."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('aei_admin.html')


@app.route('/UNA.md/orasldev/aei-operator')
def aei_operator():
    """AEÎ: Loan officer interface — issue credits, record payments."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('aei/operator.html')


@app.route('/UNA.md/orasldev/aei-backoffice')
def aei_backoffice():
    """AEÎ: Back-office interface — BI reports, settings, accounts, journal."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('aei/backoffice.html')


@app.route('/UNA.md/orasldev/aei-tz')
def aei_tz():
    """AEÎ: Technical Specification (TZ) document."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    import os
    tz_path = os.path.join(os.path.dirname(__file__), 'docs', 'AEI.md', 'TZ_AEI.html')
    with open(tz_path, 'r', encoding='utf-8') as f:
        return f.read()


@app.route('/UNA.md/orasldev/aei-simulator')
def aei_simulator():
    """AEÎ: Credit simulation with commission calculator."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    import os
    sim_path = os.path.join(os.path.dirname(__file__), 'docs', 'AEI.md', 'simulator_credit.html')
    with open(sim_path, 'r', encoding='utf-8') as f:
        return f.read()


@app.route('/UNA.md/orasldev/aei-docs')
def aei_docs():
    """AEÎ: Module documentation."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    import os
    doc_path = os.path.join(os.path.dirname(__file__), 'docs', 'AEI.md', 'README_AEI.html')
    with open(doc_path, 'r', encoding='utf-8') as f:
        return f.read()


# AEI — API routes

@app.route('/api/aei/dashboard', methods=['GET'])
def api_aei_dashboard():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_dashboard())


@app.route('/api/aei/members', methods=['GET'])
def api_aei_members_get():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_members())


@app.route('/api/aei/members', methods=['POST'])
def api_aei_members_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.upsert_member())


@app.route('/api/aei/members/<int:member_id>', methods=['GET'])
def api_aei_member_get(member_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_member(member_id))


@app.route('/api/aei/members/<int:member_id>', methods=['DELETE'])
def api_aei_member_delete(member_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.delete_member(member_id))


@app.route('/api/aei/deposits', methods=['GET'])
def api_aei_deposits_get():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_deposits())


@app.route('/api/aei/deposits', methods=['POST'])
def api_aei_deposits_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.upsert_deposit())


@app.route('/api/aei/deposits/<int:deposit_id>', methods=['GET'])
def api_aei_deposit_get(deposit_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_deposit(deposit_id))


@app.route('/api/aei/deposits/<int:deposit_id>/flows', methods=['GET'])
def api_aei_deposit_flows(deposit_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_deposit_flows(deposit_id))


@app.route('/api/aei/loans', methods=['GET'])
def api_aei_loans_get():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_loans())


@app.route('/api/aei/loans', methods=['POST'])
def api_aei_loans_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.upsert_loan())


@app.route('/api/aei/loans/<int:loan_id>', methods=['GET'])
def api_aei_loan_get(loan_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_loan(loan_id))


@app.route('/api/aei/loans/<int:loan_id>/flows', methods=['GET'])
def api_aei_loan_flows(loan_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_loan_flows(loan_id))


@app.route('/api/aei/loans/<int:loan_id>/generate-schedule', methods=['POST'])
def api_aei_loan_generate_schedule(loan_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.generate_loan_schedule(loan_id))


@app.route('/api/aei/loans/payment', methods=['POST'])
def api_aei_loan_payment():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.record_loan_payment())


@app.route('/api/aei/deposits/<int:deposit_id>/calculate-interest', methods=['GET'])
def api_aei_deposit_calc_interest(deposit_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.calculate_deposit_interest(deposit_id))


@app.route('/api/aei/journal', methods=['GET'])
def api_aei_journal_get():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_journal())


@app.route('/api/aei/journal', methods=['POST'])
def api_aei_journal_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.insert_journal_entry())


@app.route('/api/aei/accounts', methods=['GET'])
def api_aei_accounts():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_accounts())


@app.route('/api/aei/settings', methods=['GET'])
def api_aei_settings_get():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_settings())


@app.route('/api/aei/settings', methods=['POST'])
def api_aei_settings_post():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.update_setting())


@app.route('/api/aei/reports/trial-balance', methods=['GET'])
def api_aei_trial_balance():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_trial_balance())


@app.route('/api/aei/reports/olap-loans', methods=['GET'])
def api_aei_olap_loans():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(AEIController.get_olap_loans())


# ---------------------------------------------------------------------------
# ServOuts26 — CRM/SaaS servicii contabile (schema UNITEST, Oracle 11g)
# Same thick-mode subprocess worker pattern as Biro26; main app stays thin.
# ---------------------------------------------------------------------------
from controllers.servouts26_controller import ServOuts26Controller


@app.route('/UNA.md/orasldev/servouts26')
@app.route('/UNA.md/orasldev/servouts26-admin')
def servouts26_admin():
    """ServOuts26: trilingual (RU/RO/EN) admin — nomenclator, prices, import."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('servouts26_admin.html',
                           app_name=Config.SERVOUTS26_APP_NAME)


@app.route('/UNA.md/orasldev/servouts26-tz')
def servouts26_tz():
    """ServOuts26: Technical Specification (TZ) document."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    import os
    tz_path = os.path.join(os.path.dirname(__file__), 'docs', 'ServOuts26',
                           'TZ_Servouts26_App.md')
    with open(tz_path, 'r', encoding='utf-8') as f:
        body = f.read()
    from markupsafe import escape
    return ('<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8">'
            '<title>ServOuts26 — TZ</title></head>'
            '<body style="background:#1a1a2e;color:#e2e8f0;font-family:monospace">'
            f'<pre style="white-space:pre-wrap;max-width:980px;margin:24px auto">'
            f'{escape(body)}</pre></body></html>')


@app.route('/UNA.md/orasldev/servouts26-docs')
def servouts26_docs():
    """ServOuts26: module documentation."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    import os
    doc_path = os.path.join(os.path.dirname(__file__), 'docs', 'ServOuts26',
                            'README_ServOuts26.html')
    with open(doc_path, 'r', encoding='utf-8') as f:
        return f.read()


# ServOuts26 — API routes

@app.route('/api/servouts26/connection/test', methods=['GET'])
def api_srvo_conn_test():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.test_connection())


@app.route('/api/servouts26/dashboard', methods=['GET'])
def api_srvo_dashboard():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_dashboard())


@app.route('/api/servouts26/univers', methods=['GET'])
def api_srvo_univers():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_univers())


@app.route('/api/servouts26/univers/filters', methods=['GET'])
def api_srvo_univers_filters():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_univers_filters())


@app.route('/api/servouts26/univers/<int:cod>', methods=['GET'])
def api_srvo_univers_card(cod):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_univers_card(cod))


@app.route('/api/servouts26/univers/<int:cod>', methods=['POST'])
def api_srvo_univers_update(cod):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.update_univers(cod))


@app.route('/api/servouts26/univers/<int:cod>/archive', methods=['POST'])
def api_srvo_univers_archive(cod):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.archive_univers(cod))


@app.route('/api/servouts26/groups', methods=['GET'])
def api_srvo_groups():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_groups())


@app.route('/api/servouts26/groups/rename', methods=['POST'])
def api_srvo_groups_rename():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.rename_group())


@app.route('/api/servouts26/groups/merge', methods=['POST'])
def api_srvo_groups_merge():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.merge_groups())


@app.route('/api/servouts26/systree', methods=['GET'])
def api_srvo_systree():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_systree())


@app.route('/api/servouts26/orgs', methods=['GET'])
def api_srvo_orgs():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_orgs())


@app.route('/api/servouts26/orgs/<int:cod>', methods=['GET'])
def api_srvo_org_card(cod):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_org_card(cod))


@app.route('/api/servouts26/pricelists', methods=['GET'])
def api_srvo_pricelists():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_pricelists())


@app.route('/api/servouts26/prices', methods=['GET'])
def api_srvo_prices():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_prices())


@app.route('/api/servouts26/prices/update', methods=['POST'])
def api_srvo_prices_update():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.update_price())


@app.route('/api/servouts26/pricelists/rollback', methods=['POST'])
def api_srvo_pricelists_rollback():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.rollback_pricelist())


@app.route('/api/servouts26/staging', methods=['GET'])
def api_srvo_staging():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_staging())


@app.route('/api/servouts26/staging', methods=['POST'])
def api_srvo_staging_load():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.load_staging())


@app.route('/api/servouts26/staging/clear', methods=['POST'])
def api_srvo_staging_clear():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.clear_staging())


@app.route('/api/servouts26/config', methods=['GET'])
def api_srvo_config():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_config())


@app.route('/api/servouts26/import/run', methods=['POST'])
def api_srvo_import_run():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.run_step())


@app.route('/api/servouts26/profiles', methods=['GET'])
def api_srvo_profiles():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_profiles())


@app.route('/api/servouts26/profiles', methods=['POST'])
def api_srvo_profiles_save():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.save_profile())


@app.route('/api/servouts26/profiles/delete', methods=['POST'])
def api_srvo_profiles_delete():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.delete_profile())


@app.route('/api/servouts26/journal', methods=['GET'])
def api_srvo_journal():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.get_journal())


# ── ServOuts26 web-shop: PUBLIC front-office for accounting-outsourcing
#    services (client self-registration, catalog, orders, personal cabinet) ──

@app.route('/UNA.md/orasldev/servouts26-shop')
def servouts26_shop():
    """Public storefront: accounting outsourcing services + client cabinet."""
    return render_template('servouts26/shop.html',
                           app_name=Config.SERVOUTS26_APP_NAME)


@app.route('/api/servouts26/shop/catalog', methods=['GET'])
def api_srvo_shop_catalog():
    # public read-only services catalog (module pricelist, today's prices)
    return jsonify(ServOuts26Controller.shop_catalog())


@app.route('/api/servouts26/shop/register', methods=['POST'])
def api_srvo_shop_register():
    return jsonify(ServOuts26Controller.shop_register())


@app.route('/api/servouts26/shop/login', methods=['POST'])
def api_srvo_shop_login():
    return jsonify(ServOuts26Controller.shop_login())


@app.route('/api/servouts26/shop/logout', methods=['POST'])
def api_srvo_shop_logout():
    return jsonify(ServOuts26Controller.shop_logout())


@app.route('/api/servouts26/shop/me', methods=['GET'])
def api_srvo_shop_me():
    return jsonify(ServOuts26Controller.shop_me())


@app.route('/api/servouts26/shop/order', methods=['POST'])
def api_srvo_shop_order():
    # auth enforced inside (shop client session)
    return jsonify(ServOuts26Controller.shop_order())


@app.route('/api/servouts26/shop/my-orders', methods=['GET'])
def api_srvo_shop_my_orders():
    return jsonify(ServOuts26Controller.shop_my_orders())


@app.route('/api/servouts26/shop/order/<int:order_id>', methods=['GET'])
def api_srvo_shop_order_detail(order_id):
    # shop client sees only own orders; backoffice session sees any
    return jsonify(ServOuts26Controller.shop_order_detail(order_id))


# ── ServOuts26 orders (back-office) ──

@app.route('/api/servouts26/orders', methods=['GET'])
def api_srvo_orders_admin():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.orders_admin())


@app.route('/api/servouts26/orders/status', methods=['POST'])
def api_srvo_orders_status():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return jsonify(ServOuts26Controller.order_set_status())


# ---------------------------------------------------------------------------
# Biro26 — Nomenclator / Listă de prețuri / Import (OfficePlus ERP, Oracle 11g)
# Reaches officeplus via an isolated thick-mode subprocess worker; main app stays thin.
# ---------------------------------------------------------------------------
from controllers.biro26_controller import Biro26Controller


def _biro26_api_guard():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "auth"}), 401
    return None


@app.route('/UNA.md/orasldev/biro26')
@app.route('/UNA.md/orasldev/biro26-admin')
def biro26_admin():
    """Biro26: launcher / landing."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26_admin.html', app_name=Config.BIRO26_APP_NAME)


@app.route('/UNA.md/orasldev/biro26-backoffice')
def biro26_backoffice():
    """Biro26: main trilingual back-office (source, dictionary, groups, prices, mapping)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/backoffice.html', app_name=Config.BIRO26_APP_NAME)


@app.route('/UNA.md/orasldev/biro26-tz')
def biro26_tz():
    """Biro26: Technical Specification (TZ) — rendered from Markdown."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    import os
    tz_path = os.path.join(os.path.dirname(__file__), 'docs', 'Biro26', 'TZ_BIRO26_App.md')
    with open(tz_path, 'r', encoding='utf-8') as f:
        text = f.read()
    try:
        import markdown as _md
        body = _md.markdown(text, extensions=['tables', 'fenced_code'])
    except Exception:
        from markupsafe import escape
        body = '<pre>' + str(escape(text)) + '</pre>'
    return ('<!doctype html><meta charset="utf-8"><title>TZ Biro26</title>'
            '<body style="max-width:900px;margin:2rem auto;font-family:system-ui;'
            'line-height:1.5;padding:0 1rem">' + body + '</body>')


@app.route('/UNA.md/orasldev/biro26-docs')
def biro26_docs():
    """Biro26: module documentation."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    import os
    doc_path = os.path.join(os.path.dirname(__file__), 'docs', 'Biro26', 'README_BIRO26.html')
    with open(doc_path, 'r', encoding='utf-8') as f:
        return f.read()


# Biro26 — API routes
def _b26(fn):
    """Run a Biro26Controller call behind the auth guard, return jsonify."""
    g = _biro26_api_guard()
    if g is not None:
        return g
    return jsonify(fn())


@app.route('/api/biro26/connection/test', methods=['GET'])
def api_biro26_conn_test():
    return _b26(Biro26Controller.connection_test)

@app.route('/api/biro26/mapping/g-params', methods=['GET'])
def api_biro26_gparams():
    return _b26(Biro26Controller.list_g_params)

@app.route('/api/biro26/mapping/profiles', methods=['GET'])
def api_biro26_profiles_get():
    return _b26(Biro26Controller.get_profiles)

@app.route('/api/biro26/mapping/profiles', methods=['POST'])
def api_biro26_profiles_post():
    return _b26(Biro26Controller.create_profile)

@app.route('/api/biro26/mapping/profiles/<int:pid>', methods=['GET'])
def api_biro26_profile_get(pid):
    return _b26(lambda: Biro26Controller.get_profile(pid))

@app.route('/api/biro26/mapping/profiles/<int:pid>', methods=['PUT'])
def api_biro26_profile_put(pid):
    return _b26(lambda: Biro26Controller.update_profile(pid))

@app.route('/api/biro26/mapping/profiles/<int:pid>/activate', methods=['POST'])
def api_biro26_profile_activate(pid):
    return _b26(lambda: Biro26Controller.activate_profile(pid))

@app.route('/api/biro26/goods', methods=['GET'])
def api_biro26_goods():
    return _b26(Biro26Controller.get_goods)

@app.route('/api/biro26/goods/brands', methods=['GET'])
def api_biro26_goods_brands():
    return _b26(Biro26Controller.goods_brands)

@app.route('/api/biro26/goods/count', methods=['GET'])
def api_biro26_goods_count():
    return _b26(Biro26Controller.goods_count)

@app.route('/api/biro26/goods/validate', methods=['POST'])
def api_biro26_goods_validate():
    return _b26(Biro26Controller.validate_input)

@app.route('/api/biro26/goods/prepare', methods=['POST'])
def api_biro26_goods_prepare():
    return _b26(Biro26Controller.prepare_input)

@app.route('/api/biro26/goods/assign-keys', methods=['POST'])
def api_biro26_goods_assign():
    return _b26(Biro26Controller.assign_keys)

@app.route('/api/biro26/source/columns', methods=['GET'])
def api_biro26_source_columns():
    return _b26(Biro26Controller.source_columns)

@app.route('/api/biro26/source/sample', methods=['GET'])
def api_biro26_source_sample():
    return _b26(Biro26Controller.source_sample)

@app.route('/api/biro26/univers', methods=['GET'])
def api_biro26_univers():
    return _b26(Biro26Controller.get_univers)

@app.route('/api/biro26/univers/<int:cod>', methods=['GET'])
def api_biro26_univers_card(cod):
    return _b26(lambda: Biro26Controller.get_univers_card(cod))

@app.route('/api/biro26/univers/import', methods=['POST'])
def api_biro26_univers_import():
    return _b26(Biro26Controller.import_univers)

@app.route('/api/biro26/images/import', methods=['POST'])
def api_biro26_images_import():
    return _b26(Biro26Controller.import_images)

@app.route('/api/biro26/univers/archive', methods=['POST'])
def api_biro26_univers_archive():
    return _b26(Biro26Controller.archive_univers)

@app.route('/api/biro26/univers/fix-confusables', methods=['POST'])
def api_biro26_univers_fix():
    return _b26(Biro26Controller.fix_confusables)

@app.route('/api/biro26/groups', methods=['GET'])
def api_biro26_groups_get():
    return _b26(Biro26Controller.get_groups)

@app.route('/api/biro26/groups', methods=['PUT'])
def api_biro26_groups_put():
    return _b26(Biro26Controller.update_group)

@app.route('/api/biro26/groups/import', methods=['POST'])
def api_biro26_groups_import():
    return _b26(Biro26Controller.import_groups)

@app.route('/api/biro26/groups/merge', methods=['POST'])
def api_biro26_groups_merge():
    return _b26(Biro26Controller.merge_groups)

@app.route('/api/biro26/categories', methods=['GET'])
def api_biro26_categories():
    return _b26(Biro26Controller.get_categories)

@app.route('/api/biro26/suppliers', methods=['GET'])
def api_biro26_suppliers():
    return _b26(Biro26Controller.get_suppliers)

@app.route('/api/biro26/suppliers/furnizori', methods=['GET'])
def api_biro26_furnizori():
    return _b26(Biro26Controller.get_furnizori)

@app.route('/api/biro26/prices', methods=['GET'])
def api_biro26_prices_get():
    return _b26(Biro26Controller.get_prices)

@app.route('/api/biro26/prices', methods=['PUT'])
def api_biro26_prices_put():
    return _b26(Biro26Controller.update_price)

@app.route('/api/biro26/prices/lists', methods=['GET'])
def api_biro26_prices_lists():
    return _b26(Biro26Controller.get_pricelists)

@app.route('/api/biro26/prices/dates', methods=['GET'])
def api_biro26_prices_dates_get():
    return _b26(Biro26Controller.get_dates)

@app.route('/api/biro26/prices/import-dates', methods=['POST'])
def api_biro26_prices_dates():
    return _b26(Biro26Controller.import_dates)

@app.route('/api/biro26/prices/import', methods=['POST'])
def api_biro26_prices_import():
    return _b26(Biro26Controller.import_prices)

@app.route('/api/biro26/prices/by-article', methods=['GET'])
def api_biro26_price_by_article_get():
    return _b26(Biro26Controller.price_by_article_get)

@app.route('/api/biro26/prices/by-article', methods=['PUT'])
def api_biro26_price_by_article_set():
    return _b26(Biro26Controller.price_by_article_set)

@app.route('/api/biro26/prices/rollback', methods=['POST'])
def api_biro26_prices_rollback():
    return _b26(Biro26Controller.rollback_pricelist)

@app.route('/api/biro26/sources', methods=['GET'])
def api_biro26_sources_list():
    return _b26(Biro26Controller.list_sources)

@app.route('/api/biro26/sources', methods=['POST'])
def api_biro26_sources_create():
    return _b26(Biro26Controller.create_source)

@app.route('/api/biro26/sources/sample', methods=['POST'])
def api_biro26_sources_sample():
    return _b26(Biro26Controller.sample_select)

@app.route('/api/biro26/sources/ai-draft-md', methods=['POST'])
def api_biro26_sources_ai_md():
    return _b26(Biro26Controller.ai_draft_md)

@app.route('/api/biro26/sources/ai-suggest-mapping', methods=['POST'])
def api_biro26_sources_ai_map():
    return _b26(Biro26Controller.ai_suggest_mapping)

@app.route('/api/biro26/stock/calculate', methods=['POST'])
def api_biro26_stock_calculate():
    return _b26(Biro26Controller.calc_stock)

@app.route('/api/biro26/stock/latest', methods=['GET'])
def api_biro26_stock_latest():
    return _b26(Biro26Controller.get_latest_stock_calc)

@app.route('/api/biro26/stock/items', methods=['GET'])
def api_biro26_stock_items():
    return _b26(Biro26Controller.get_stock_items)

@app.route('/api/biro26/products', methods=['GET'])
def api_biro26_products():
    return _b26(Biro26Controller.get_products_stock)

@app.route('/api/biro26/products/brands', methods=['GET'])
def api_biro26_products_brands():
    return _b26(Biro26Controller.get_product_brands)

@app.route('/api/biro26/products/categories', methods=['GET'])
def api_biro26_products_categories():
    return _b26(Biro26Controller.get_product_categories)

@app.route('/api/biro26/products/tree', methods=['GET'])
def api_biro26_products_tree():
    return _b26(Biro26Controller.get_product_tree)

@app.route('/api/biro26/products/<int:cod>', methods=['PUT'])
def api_biro26_product_update(cod):
    return _b26(lambda: Biro26Controller.update_product(cod))

@app.route('/api/biro26/products/tree/rename', methods=['POST'])
def api_biro26_tree_rename():
    return _b26(Biro26Controller.tree_rename)

@app.route('/api/biro26/products/tree/move', methods=['POST'])
def api_biro26_tree_move():
    return _b26(Biro26Controller.tree_move)

# ── price periods on Marfă/Stoc (split on change, merge on delete) ──
@app.route('/api/biro26/products/price-history', methods=['GET'])
def api_biro26_price_history():
    return _b26(Biro26Controller.product_price_history)

@app.route('/api/biro26/products/price', methods=['POST'])
def api_biro26_price_set():
    return _b26(Biro26Controller.product_price_set)

@app.route('/api/biro26/products/price/delete', methods=['POST'])
def api_biro26_price_delete():
    return _b26(Biro26Controller.product_price_delete)

# ── BIRO26PT: universal file/zip import (2-phase: dry-run -> commit) ──
@app.route('/UNA.md/orasldev/biro26-import-pt')
def biro26_import_pt():
    """Web UI over the BIRO26PT_importData package (upload -> analyze -> commit)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/import_pt.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/pt/uploads', methods=['POST'])
def api_biro26_pt_uploads():
    return _b26(Biro26Controller.pt_upload)

@app.route('/api/biro26/pt/analyze', methods=['POST'])
def api_biro26_pt_analyze():
    return _b26(Biro26Controller.pt_analyze)

@app.route('/api/biro26/pt/preview/<int:load_id>', methods=['GET'])
def api_biro26_pt_preview(load_id):
    return _b26(lambda: Biro26Controller.pt_preview(load_id))

@app.route('/api/biro26/pt/commit', methods=['POST'])
def api_biro26_pt_commit():
    return _b26(Biro26Controller.pt_commit)

@app.route('/api/biro26/pt/remap', methods=['POST'])
def api_biro26_pt_remap():
    return _b26(Biro26Controller.pt_remap)

# ── health: ce versiune ruleaza REAL pe server ────────────────────────────
# RO: raspunsul la intrebarea "ce commit e desfasurat?" dintr-un singur curl.
#     Commit-ul vine din fisierul DEPLOY_COMMIT scris de deploy (serverul nu
#     e un checkout git); local se citeste din git. route_check listeaza
#     referintele Biro26Controller.<metoda> din app.py care NU exista in
#     controller — exact rasincronul care a dat 500 in loc de 401 la
#     /api/biro26/pt/sources (fisiere din commit-uri diferite pe server).
# EN: deployment health — running commit + controller-reference smoke test;
#     a file/commit mismatch shows up here instead of as a prod 500.
import datetime as _dt
_HEALTH_STARTED = _dt.datetime.now().strftime('%Y-%m-%dT%H:%M:%S')

def _biro26_route_check():
    import re as _re
    try:
        src = open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                'app.py'), encoding='utf-8').read()
        missing = sorted({m for m in _re.findall(
            r'Biro26Controller\.(\w+)', src)
            if not hasattr(Biro26Controller, m)})
        return missing
    except Exception:                                        # noqa: BLE001
        return []

@app.route('/api/biro26/health', methods=['GET'])
def api_biro26_health():
    base = os.path.dirname(os.path.abspath(__file__))
    commit = None
    try:
        with open(os.path.join(base, 'DEPLOY_COMMIT'), encoding='utf-8') as f:
            commit = f.read().strip()[:40] or None
    except OSError:
        try:
            import subprocess as _sp
            commit = _sp.run(['git', 'rev-parse', '--short', 'HEAD'],
                             cwd=base, capture_output=True, text=True,
                             timeout=3).stdout.strip() or None
        except Exception:                                    # noqa: BLE001
            commit = None
    missing = _biro26_route_check()
    return jsonify({'commit': commit, 'started_at': _HEALTH_STARTED,
                    'routes': len(list(app.url_map.iter_rules())),
                    'missing_controller_refs': missing,
                    'ok': not missing})

# RO: smoke-test la pornire — un rasincron de fisiere se vede in journalctl
#     imediat, nu ca 500 in productie cind cineva atinge endpoint-ul.
# EN: startup smoke test — file desync shows in the log at boot, not as a 500.
_missing_refs = _biro26_route_check()
if _missing_refs:
    print('[biro26-health] AVERTISMENT: app.py foloseste metode inexistente '
          'in Biro26Controller (fisiere din commit-uri diferite?): '
          + ', '.join(_missing_refs))

# RO: incalzirea cache-ului /api/biro26/site/config in fundal la pornire —
#     altfel PRIMUL vizitator dupa restart astepta ~15s (3-4 interogari prin
#     subprocesul Oracle thick). Dupa incalzire raspunsul e din cache (<0.3s),
#     iar expirarile se reimprospateaza tot in fundal (stale-while-revalidate).
# EN: warm the site-config cache at boot so the first visitor never pays.
def _biro26_warm_site_config():
    try:
        from models.biro26_site import Biro26Site
        Biro26Site.config()
    except Exception:                                        # noqa: BLE001
        pass

threading.Thread(target=_biro26_warm_site_config, daemon=True).start()

@app.route('/api/biro26/img', methods=['GET'])
def api_biro26_img():
    """RO: serveste pe HTTPS o imagine gazduita doar pe HTTP (impreso.md).
    Fara asta, browserul o blocheaza ca "mixed content" si produsul apare
    fara poza, desi URL-ul din baza e corect.
    EN: re-serve an HTTP-only image over HTTPS; without this the browser blocks
    it as mixed content."""
    from models.biro26_imgproxy import fetch
    u = request.args.get('u', '')
    try:
        data, ctype = fetch(u)
    except Exception as e:
        return (str(e), 400)
    # RO: imaginile de produs nu se schimba des / EN: product images rarely change
    return Response(data, mimetype=ctype,
                    headers={'Cache-Control': 'public, max-age=86400'})

@app.route('/api/biro26/pt/algorithms', methods=['GET'])
def api_biro26_pt_algorithms():
    return _b26(Biro26Controller.pt_algorithms)

@app.route('/api/biro26/pt/sources', methods=['GET'])
def api_biro26_pt_sources():
    return _b26(Biro26Controller.pt_sources)

@app.route('/api/biro26/pt/sources/<src_code>/files', methods=['GET'])
def api_biro26_pt_source_files(src_code):
    return _b26(lambda: Biro26Controller.pt_source_files(src_code))

@app.route('/api/biro26/pt/help', methods=['GET'])
def api_biro26_pt_help():
    return _b26(Biro26Controller.pt_help)

# ── notification settings: email / Telegram / WhatsApp on new invoices ──
@app.route('/UNA.md/orasldev/biro26-notify-settings')
def biro26_notify_settings():
    """Admin: notification channels for new orders/invoices."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/notify_settings.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/notify-settings', methods=['GET'])
def api_biro26_notify_get():
    return _b26(Biro26Controller.notify_settings_get)

@app.route('/api/biro26/notify-settings', methods=['PUT'])
def api_biro26_notify_save():
    return _b26(Biro26Controller.notify_settings_save)

@app.route('/api/biro26/notify-test', methods=['POST'])
def api_biro26_notify_test():
    return _b26(Biro26Controller.notify_test)

# ── report template admin (simple editor for reports/templates/*) ──
@app.route('/UNA.md/orasldev/biro26-report-templates')
def biro26_report_templates():
    """Simple admin: edit the jsReport templates (invoice/order/helpers)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/report_templates.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/report-templates', methods=['GET'])
def api_biro26_rtpl_list():
    return _b26(Biro26Controller.report_templates_list)

@app.route('/api/biro26/report-templates/<name>', methods=['GET'])
def api_biro26_rtpl_get(name):
    return _b26(lambda: Biro26Controller.report_template_get(name))

@app.route('/api/biro26/report-templates/<name>', methods=['PUT'])
def api_biro26_rtpl_save(name):
    return _b26(lambda: Biro26Controller.report_template_save(name))

@app.route('/api/biro26/report-engines', methods=['GET'])
def api_biro26_rtpl_engines_get():
    return _b26(Biro26Controller.report_engines_get)

@app.route('/api/biro26/report-engines', methods=['PUT'])
def api_biro26_rtpl_engines_set():
    return _b26(Biro26Controller.report_engines_set)

@app.route('/UNA.md/orasldev/biro26-pdfme-designer')
def biro26_pdfme_designer():
    """Visual pdfme Designer for the pdfme_*.json templates."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/pdfme_designer.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/report-templates/preview', methods=['POST'])
def api_biro26_rtpl_preview():
    g = _biro26_api_guard()
    if g is not None:
        return g
    r = Biro26Controller.report_template_preview()
    if not r.get('success'):
        return jsonify(r), 400
    resp = app.response_class(r['pdf'], mimetype='application/pdf')
    resp.headers['Content-Disposition'] = 'inline; filename="preview.pdf"'
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp

# ── product variants (BIRO26_VARIANTS): family detail + editing ──
@app.route('/api/biro26/univers/<int:cod>/variants', methods=['GET'])
def api_biro26_variants_get(cod):
    return _b26(lambda: Biro26Controller.get_variants(cod))

@app.route('/api/biro26/variants/<int:cod>', methods=['PUT'])
def api_biro26_variants_put(cod):
    return _b26(lambda: Biro26Controller.update_variant(cod))


# ── Biro26 web-shop: PUBLIC page + API (client self-registration,
#    catalog browsing, invoice creation via package y_ai_BIRO26) ──

# RO: cache-ul paginilor informative WP (slug -> (expira, titlu, html));
#     redate in interiorul magazinului ca meniul sa ramana vizibil pe tot site-ul.
# EN: WP info-page cache (slug -> (expiry, title, html)); pages render inside
#     the shop so the top menu stays visible across the whole site.
_BIRO26_WP_CACHE = {}
_BIRO26_WP_TTL = 300  # seconds

def _biro26_wp_page(slug):
    """Fetch a WP page (title, rendered HTML) via REST; '' base = feature off."""
    import re as _re, time as _time, json as _json
    import urllib.request as _rq
    if not Config.BIRO26_SHOP_WP_API or not _re.match(r'^[a-z0-9-]{1,80}$', slug):
        return None, None
    hit = _BIRO26_WP_CACHE.get(slug)
    if hit and hit[0] > _time.time():
        return hit[1], hit[2]
    try:
        url = (Config.BIRO26_SHOP_WP_API.rstrip('/')
               + '/wp/v2/pages?slug=' + slug + '&_fields=title,content')
        with _rq.urlopen(url, timeout=6) as resp:
            pages = _json.loads(resp.read().decode('utf-8'))
        if not pages:
            return None, None
        title = pages[0].get('title', {}).get('rendered') or slug
        html = pages[0].get('content', {}).get('rendered') or ''
        _BIRO26_WP_CACHE[slug] = (_time.time() + _BIRO26_WP_TTL, title, html)
        return title, html
    except Exception:
        # RO: WP indisponibil -> cade inapoi pe catalog / EN: WP down -> catalog
        return None, None

def _biro26_rate_plans():
    """RO: pachetele de rate FARA dobinda, pentru eticheta «Preț în rate».

    Un singur numar poate rezuma corect DOAR un pachet fara dobinda: acolo
    pretul finantat este tot ce plateste clientul. Pachetele cu dobinda
    (Microinvest Standard — 39% anual, MAIB Credit de consum — 10,5%) au un
    pret finantat MAI MIC, dar in total costa mai mult; daca intra in minimul
    afisat, clientul vede un numar pe care nu-l plateste nimeni. Limitele de
    suma se verifica pe suma FINANTATA, exact ca in crTiles() si in
    models/biro26_credit.py calc().

    Intoarce (liber_pct, liber_min, rate_plans) — procentul si minimul raman
    ca rezerva pentru cazul in care ofertele nu se pot citi.
    EN: only 0%-interest plans can be summarised by a single "rate price".
    """
    from models.biro26_oracle_store import Biro26Store
    try:
        liber_pct = float(Biro26Store.get_setting('RATE_LIBER_PCT', '5'))
        liber_min = float(Biro26Store.get_setting('RATE_LIBER_MIN', '100'))
    except Exception:                                        # noqa: BLE001
        liber_pct, liber_min = 5.0, 100.0
    rate_plans = []
    try:
        from models.biro26_credit import Biro26Credit
        best = None
        for o in (Biro26Credit.public_offers().get("data") or []):
            tm = float(o.get('transport_markup_pct') or 0)
            for pl in (o.get('plans') or []):
                if (float(pl.get('annual_pct') or 0) != 0
                        or float(pl.get('monthly_fee_pct') or 0) != 0):
                    continue
                eff = float(pl.get('markup_pct') or 0) + tm
                rate_plans.append({
                    'p': eff,
                    'mn': float(pl.get('amount_min') or 0),
                    'mx': float(pl.get('amount_max') or 0) or 1e12})
                if best is None or eff < best:
                    best = eff
        if best is not None:
            liber_pct = best
    except Exception:                                        # noqa: BLE001
        rate_plans = []
    return liber_pct, liber_min, rate_plans


def _biro26_site_ctx():
    """RO: contextul comun al paginilor noului site Figma.
    EN: shared context for the new-site pages."""
    from models.biro26_oracle_store import Biro26Store
    liber_pct, liber_min, rate_plans = _biro26_rate_plans()
    # RO: suma MINIMA a comenzii de la care se poate achita in rate/credit.
    #     Conditia se scrie cu ROSU pe card, pe pagina produsului si in cos
    #     (cerinta proprietar 18.08.2026). Se schimba din YBIRO_SETTINGS,
    #     fara atingerea codului.
    # EN: minimum order total that unlocks instalments; shown in red.
    try:
        credit_min_order = float(Biro26Store.get_setting('CREDIT_MIN_ORDER', '1500'))
    except Exception:                                        # noqa: BLE001
        credit_min_order = 1500.0
    try:
        brand_filter = Biro26Store.get_setting('SHOP_BRAND_FILTER', '0')
    except Exception:
        brand_filter = '0'
    try:
        fmt_html = Biro26Store.get_setting('SHOP_FMT_HTML', '1')
        fmt_xlsx = Biro26Store.get_setting('SHOP_FMT_XLSX', '1')
    except Exception:
        fmt_html, fmt_xlsx = '1', '1'
    # RO: coloana de pret dupa TIPUL clientului logat (fizica/juridica);
    #     vizitatorii vad preturile pentru persoane fizice
    try:
        from flask import session as _s
        _cl = _s.get('biro26_client')
        price_field = (Biro26Store.client_price_field(_cl['univers_cod'])
                       if _cl else Biro26Store.get_setting('SHOP_PRICE_FIZ',
                                                           'retail1'))
    except Exception:
        price_field = 'retail1'
    # RO: siglele de plata DISPONIBILE pe disc — subsolul cere <img> doar
    #     pentru ele, restul raman badge text. Altfel browserul incerca sa
    #     incarce fisiere inexistente si consola se umplea de 404
    #     (siglele oficiale se adauga in /static/biro26/pay/ sau din WP).
    # EN: which payment logos actually exist, so the footer never requests a
    #     missing file (404 noise); the rest fall back to a text badge.
    try:
        _paydir = os.path.join(app.static_folder, 'biro26', 'pay')
        pay_logos = sorted(f.rsplit('.', 1)[0].lower()
                           for f in os.listdir(_paydir)
                           if f.lower().endswith(('.svg', '.png'))
                           and not f.startswith(('.', '_')))
    except Exception:
        pay_logos = []
    return {'app_name': Config.BIRO26_APP_NAME,
            'liber_pct': liber_pct, 'liber_min': liber_min,
            'credit_min_order': credit_min_order,
            'rate_plans': rate_plans,
            'brand_filter': brand_filter,
            'fmt_html': fmt_html, 'fmt_xlsx': fmt_xlsx,
            'price_field': price_field,
            'pay_logos': pay_logos}

@app.route('/UNA.md/orasldev/biro26-site')
# RO: alias '1shop' — acelasi site nou si pe instantele FARA nginx pretty-URLs
#     (ex. nufarul); navigarea e tradusa client-side de siteURL() din site.js.
@app.route('/UNA.md/orasldev/biro26-1shop')
def biro26_site():
    """RO: pagina principala LIVE dupa Figma (landingfigma1) — vitrina noului
    site (TZ OFFICEPLUS_AI_SITE_PROJECT.md); pe shop1 nginx o serveste la '/'.
    EN: live Figma homepage for the new site; shop1 nginx maps it to '/'."""
    return render_template('biro26/site_home.html', **_biro26_site_ctx())

@app.route('/UNA.md/orasldev/biro26-site/catalog')
@app.route('/UNA.md/orasldev/biro26-1shop/catalog')
def biro26_site_catalog():
    # RO: catalog (PLP) in stilul Figma; filtrele vin din URL (deep-link)
    return render_template('biro26/site_catalog.html', **_biro26_site_ctx())

@app.route('/UNA.md/orasldev/biro26-site/product/<int:cod>')
@app.route('/UNA.md/orasldev/biro26-1shop/product/<int:cod>')
def biro26_site_product(cod):
    # RO: fisa produsului (PDP) — datele se incarca client-side dupa COD
    return render_template('biro26/site_product.html', cod=cod,
                           **_biro26_site_ctx())

@app.route('/UNA.md/orasldev/biro26-site/cart')
@app.route('/UNA.md/orasldev/biro26-1shop/cart')
def biro26_site_cart():
    # RO: cos + checkout pe API-urile existente /api/biro26/shop/*
    return render_template('biro26/site_cart.html', **_biro26_site_ctx())

@app.route('/UNA.md/orasldev/biro26-site/payment-result')
@app.route('/UNA.md/orasldev/biro26-1shop/payment-result')
def biro26_site_payment_result():
    """RO: pagina de retur dupa plata, cu detaliile comenzii — cerinta maib
    (docs.maibmerchants.md/main/ro/integration/requirements).
    EN: post-payment return page with the order details (maib requirement)."""
    return render_template('biro26/site_payment_result.html', **_biro26_site_ctx())

@app.route('/UNA.md/orasldev/biro26-site/account')
@app.route('/UNA.md/orasldev/biro26-1shop/account')
def biro26_site_account():
    return render_template('biro26/site_account.html', **_biro26_site_ctx())

@app.route('/UNA.md/orasldev/biro26-site/page/<slug>')
@app.route('/UNA.md/orasldev/biro26-1shop/page/<slug>')
def biro26_site_page(slug):
    """RO: pagina informativa din WordPress REST, redata in chrome-ul noului
    site (WP ramine DOAR CMS de continut — TZ §7). ?lang=ru|en -> slug-ru."""
    lang = request.args.get('lang')
    real = slug + ('-' + lang if lang in ('ru', 'en') else '')
    title, html = _biro26_wp_page(real)
    if html is None and lang:                       # RO: fallback pe RO
        title, html = _biro26_wp_page(slug)
    if html is None:
        return redirect('/UNA.md/orasldev/biro26-site')
    return render_template('biro26/site_page.html', slug=slug,
                           page_title=title, page_html=html,
                           **_biro26_site_ctx())

@app.route('/UNA.md/orasldev/biro26-site-admin')
def biro26_site_admin():
    """RO: LIMITED ADMIN al vitrinei (TZ §6): hero, produsul zilei, sectiuni.
    Marfa/preturile ramin in ERP; textele informative ramin in WordPress."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/site_admin.html',
                           app_name=Config.BIRO26_APP_NAME)

# ── API-ul vitrinei noului site (config public + admin CRUD) ──────────
@app.route('/api/biro26/site/config', methods=['GET'])
def api_biro26_site_config():
    from models.biro26_site import Biro26Site
    return jsonify(Biro26Site.config())

@app.route('/api/biro26/site/hero', methods=['GET', 'POST'])
def api_biro26_site_hero():
    g = _biro26_api_guard()
    if g is not None:
        return g
    from models.biro26_site import Biro26Site
    if request.method == 'GET':
        return jsonify(Biro26Site.hero_list())
    return jsonify(Biro26Site.hero_save(request.get_json(silent=True) or {}))

@app.route('/api/biro26/site/hero/<int:hid>', methods=['DELETE'])
def api_biro26_site_hero_del(hid):
    g = _biro26_api_guard()
    if g is not None:
        return g
    from models.biro26_site import Biro26Site
    return jsonify(Biro26Site.hero_delete(hid))

@app.route('/api/biro26/site/deal', methods=['GET', 'POST'])
def api_biro26_site_deal():
    g = _biro26_api_guard()
    if g is not None:
        return g
    from models.biro26_site import Biro26Site
    if request.method == 'GET':
        return jsonify(Biro26Site.deal_get())
    return jsonify(Biro26Site.deal_save(request.get_json(silent=True) or {}))

@app.route('/api/biro26/site/subscribe', methods=['POST'])
def api_biro26_site_subscribe():
    # RO: abonare newsletter (public) — stocata in YBIRO_SITE_SUBSCRIBER
    from models.biro26_site import Biro26Site
    return jsonify(Biro26Site.subscribe(request.get_json(silent=True) or {}))

@app.route('/api/biro26/site/subscribers', methods=['GET'])
def api_biro26_site_subscribers():
    g = _biro26_api_guard()
    if g is not None:
        return g
    from models.biro26_site import Biro26Site
    return jsonify(Biro26Site.subscribers_list())

@app.route('/UNA.md/orasldev/biro26-site/favorites')
@app.route('/UNA.md/orasldev/biro26-1shop/favorites')
def biro26_site_favorites():
    # RO: lista de produse favorite (inimioarele de pe carduri)
    return render_template('biro26/site_favorites.html', **_biro26_site_ctx())

@app.route('/UNA.md/orasldev/biro26-site/compare')
@app.route('/UNA.md/orasldev/biro26-1shop/compare')
def biro26_site_compare():
    # RO: compararea produselor (max 4, alese de pe fisele PDP)
    return render_template('biro26/site_compare.html', **_biro26_site_ctx())

@app.route('/api/biro26/site/info/<slug>', methods=['GET'])
def api_biro26_site_info(slug):
    """RO: continutul paginilor WP "site-*" pentru blocurile vitrinei
    (contact, despre, tipuri de plata) — sursa de adevar editabila SIMPLU
    de non-admin in WP Admin (TZ §7). EN: WP-sourced storefront blocks."""
    if not slug.startswith('site-'):
        return jsonify({'success': False, 'error': 'slug invalid'}), 400
    title, html = _biro26_wp_page(slug)
    if html is None:
        return jsonify({'success': False, 'error': 'pagina lipseste'}), 404
    return jsonify({'success': True, 'data': {'title': title, 'html': html}})

# RO: lista de branduri (nume | logo) vine din pagina WP "site-branduri";
#     cache 10 min (pagina are ~1800 rinduri)
_BIRO26_BRANDS_CACHE = {'exp': 0, 'data': []}

@app.route('/api/biro26/site/brands', methods=['GET'])
def api_biro26_site_brands():
    import time as _t, re as _re, html as _h
    if _BIRO26_BRANDS_CACHE['exp'] > _t.time():
        return jsonify({'success': True, 'data': _BIRO26_BRANDS_CACHE['data']})
    title, html = _biro26_wp_page('site-branduri')
    rows = []
    if html:
        text = _re.sub(r'<br\s*/?>', '\n', html)
        text = _re.sub(r'<[^>]+>', '\n', text)
        for line in _h.unescape(text).splitlines():
            if '|' in line:
                name, url = line.split('|', 1)
                if name.strip() and url.strip().startswith('http'):
                    rows.append({'brand': name.strip(), 'img': url.strip()})
    _BIRO26_BRANDS_CACHE.update(exp=_t.time() + 600, data=rows)
    return jsonify({'success': True, 'data': rows})

@app.route('/UNA.md/orasldev/biro26-site/brands')
@app.route('/UNA.md/orasldev/biro26-1shop/brands')
def biro26_site_brands():
    # RO: pagina "Branduri" din meniul magazinului (lista + logo din WP)
    return render_template('biro26/site_brands.html', **_biro26_site_ctx())

@app.route('/api/biro26/site/featured', methods=['GET', 'POST'])
def api_biro26_site_featured():
    g = _biro26_api_guard()
    if g is not None:
        return g
    from models.biro26_site import Biro26Site
    if request.method == 'GET':
        return jsonify(Biro26Site.featured_list())
    return jsonify(Biro26Site.featured_save(request.get_json(silent=True) or {}))

@app.route('/api/biro26/site/section', methods=['POST'])
def api_biro26_site_section():
    g = _biro26_api_guard()
    if g is not None:
        return g
    from models.biro26_site import Biro26Site
    return jsonify(Biro26Site.section_save(request.get_json(silent=True) or {}))

@app.route('/UNA.md/orasldev/biro26-shop')
def biro26_shop():
    """Public self-service Marfă/Stoc page for individual clients."""
    # topbar theming + nav links are per-deployment (.env); a light
    # background flips the button styling inside the template
    bg = Config.BIRO26_SHOP_TOPBAR_BG
    try:
        h = bg.lstrip('#')
        lum = (0.299 * int(h[0:2], 16) + 0.587 * int(h[2:4], 16)
               + 0.114 * int(h[4:6], 16))
    except Exception:
        lum = 0
    nav = [tuple(p.split('|', 1)) for p in Config.BIRO26_SHOP_NAV.split(';')
           if '|' in p]
    # RO: pagina informativa in interiorul magazinului (nav "info:<slug>")
    # EN: info page inside the shop (nav "info:<slug>")
    info_slug = (request.args.get('info') or '').strip()
    # RO: limba paginilor informative: paginile traduse traiesc in WP cu
    #     sufix de slug (<slug>-ru / <slug>-en, editabile in admin WP);
    #     fara traducere -> fallback la varianta RO.
    # EN: info-page language: translations live in WP as suffixed slugs
    #     (<slug>-ru / <slug>-en); missing translation -> RO fallback.
    lang = (request.args.get('lang') or '').strip().lower()
    if lang not in ('ru', 'en'):
        lang = ''
    info_title = info_html = None
    if info_slug:
        if lang:
            info_title, info_html = _biro26_wp_page(f"{info_slug}-{lang}")
        if not info_html:
            info_title, info_html = _biro26_wp_page(info_slug)
    # RO: produse pe pagina — setabil in admin (YBIRO_SETTINGS SHOP_PAGE_SIZE)
    # EN: products per page — admin-configurable (SHOP_PAGE_SIZE setting)
    try:
        from models.biro26_oracle_store import Biro26Store
        page_size = int(Biro26Store.get_setting('SHOP_PAGE_SIZE', '24'))
        page_size = max(1, min(page_size, 200))
    except Exception:
        page_size = 24
    # RO: "Pret oferta in rate" — acelasi calcul ca pe noul site: pachetele
    #     de rate FARA dobinda, cu limitele lor de suma (vezi
    #     _biro26_rate_plans). EN: same rate-price source as the new site.
    liber_pct, liber_min, rate_plans = _biro26_rate_plans()
    try:
        fmt_html = Biro26Store.get_setting('SHOP_FMT_HTML', '1')
        fmt_xlsx = Biro26Store.get_setting('SHOP_FMT_XLSX', '1')
    except Exception:
        fmt_html, fmt_xlsx = '1', '1'
    return render_template('biro26/shop.html', app_name=Config.BIRO26_APP_NAME,
                           topbar_bg=bg, topbar_fg=Config.BIRO26_SHOP_TOPBAR_FG,
                           topbar_light=(lum > 140), shop_nav=nav,
                           info_slug=info_slug, info_title=info_title,
                           info_html=info_html, page_size=page_size,
                           liber_pct=liber_pct, liber_min=liber_min,
                           rate_plans=rate_plans,
                           fmt_html=fmt_html, fmt_xlsx=fmt_xlsx,
                           price_field=(Biro26Store.client_price_field(
                               session['biro26_client']['univers_cod'])
                               if session.get('biro26_client')
                               else Biro26Store.get_setting('SHOP_PRICE_FIZ',
                                                            'retail1')),
                           cur_lang=(lang or 'ro'))

# ── credit payment: admin page + orgs/plans API + public offers/calc ──
@app.route('/UNA.md/orasldev/biro26-credit-admin')
def biro26_credit_admin():
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/credit_admin.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/shop/credit/offers', methods=['GET'])
def api_biro26_credit_offers():
    # public: enabled credit organizations + plans (shop cart / product page)
    return jsonify(Biro26Controller.credit_offers())

@app.route('/api/biro26/shop/credit/calc', methods=['POST'])
def api_biro26_credit_calc():
    # public: estimative simulation {amount, plan_id, months?, avans?}
    return jsonify(Biro26Controller.credit_calc())

@app.route('/api/biro26/shop/credit/request', methods=['POST'])
def api_biro26_credit_request():
    # public: the product-page loan request form (name+phone required)
    r = Biro26Controller.credit_request()
    _biro26_social_conv('credit_req', r)
    return jsonify(r)

@app.route('/api/biro26/credit/requests', methods=['GET'])
def api_biro26_credit_requests():
    return _b26(Biro26Controller.credit_requests_list)

@app.route('/api/biro26/credit/requests/<int:req_id>', methods=['PUT'])
def api_biro26_credit_request_status(req_id):
    return _b26(lambda: Biro26Controller.credit_request_status(req_id))

@app.route('/api/biro26/credit/orgs', methods=['GET'])
def api_biro26_credit_orgs():
    return _b26(Biro26Controller.credit_orgs)

@app.route('/api/biro26/credit/orgs', methods=['PUT'])
def api_biro26_credit_org_save():
    return _b26(Biro26Controller.credit_org_save)

@app.route('/api/biro26/credit/plans', methods=['GET'])
def api_biro26_credit_plans():
    return _b26(Biro26Controller.credit_plans)

@app.route('/api/biro26/credit/plans', methods=['PUT'])
def api_biro26_credit_plan_save():
    return _b26(Biro26Controller.credit_plan_save)

@app.route('/api/biro26/credit/plans/<int:plan_id>', methods=['DELETE'])
def api_biro26_credit_plan_delete(plan_id):
    return _b26(lambda: Biro26Controller.credit_plan_delete(plan_id))

# ── credit: provideri API (admin, auth) ──
@app.route('/UNA.md/orasldev/biro26-clients')
def biro26_clients_page():
    """RO: clientii magazinului + marcajul lor (admin/test/trusted)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/clients.html', app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/shop-clients', methods=['GET'])
def api_biro26_shop_clients():
    return _b26(Biro26Controller.shop_clients)

@app.route('/UNA.md/orasldev/biro26-contragenti.zip')
def biro26_contragenti_download():
    """RO: utilitarul LOCAL «Contragenti» (preluarea datelor din date.gov.md),
    impachetat la cerere din tools/contragenti — operatorul il descarca direct
    din back-office cind nu ruleaza pe calculatorul lui."""
    import io
    import zipfile
    if not AuthController.is_authenticated():
        return _login_redirect()
    src = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       'tools', 'contragenti')
    if not os.path.isdir(src):
        return jsonify({'success': False, 'error': 'utilitarul nu este în proiect'}), 404
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        for name in sorted(os.listdir(src)):
            f = os.path.join(src, name)
            if os.path.isfile(f) and not name.startswith('.'):
                z.write(f, f'contragenti/{name}')
    buf.seek(0)
    resp = app.response_class(buf.read(), mimetype='application/zip')
    resp.headers['Content-Disposition'] = 'attachment; filename="contragenti.zip"'
    return resp

@app.route('/UNA.md/orasldev/biro26-gov-return')
def biro26_gov_return():
    """RO: pagina de INTOARCERE pentru utilitarul Contragenti: primeste datele
    prin query (302 din utilitar), le trimite ferestrei-parinte si se inchide.
    Asa fluxul nu se mai opreste pe pagina utilitarului."""
    return render_template('biro26/gov_return.html')

@app.route('/api/biro26/shop-clients', methods=['POST'])
def api_biro26_shop_client_add():
    """RO: client NOU inregistrat de operator (minim: denumire + tip)."""
    r = Biro26Controller.client_quick_add()
    return jsonify(r), (200 if r.get('success')
                        else 401 if r.get('error') == 'auth required' else 400)

@app.route('/api/biro26/shop-clients/mark', methods=['PUT'])
def api_biro26_shop_client_mark():
    return _b26(Biro26Controller.shop_client_mark_set)

@app.route('/UNA.md/orasldev/biro26-integration-log')
def biro26_integration_log():
    """RO: jurnal TEHNIC al integrarilor (creditare + plati) — diagnostic.
    EN: technical integration log (credit providers + payments)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/integration_log.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/integration-log', methods=['GET'])
def api_biro26_integration_log():
    from models.biro26_credit import Biro26Credit
    try:
        n = int(request.args.get('limit') or 200)
    except (TypeError, ValueError):
        n = 200
    return _b26(lambda: Biro26Credit.integration_log(n))

@app.route('/UNA.md/orasldev/biro26-credite-docs')
def biro26_credite_docs_page():
    """RO: documentele de credit (TMDB_CREDITE_M/D) — master + detail,
    cite un tab per organizatie de creditare (EasyCredit / Liber Card)."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/credite_docs.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/credite-docs', methods=['GET'])
def api_biro26_credite_docs():
    from models.biro26_credit import Biro26Credit
    try:
        org = int(request.args.get('org_id') or 0) or None
    except (TypeError, ValueError):
        org = None
    return _b26(lambda: Biro26Credit.documents(org))

@app.route('/api/biro26/credite-docs/anketa/<int:req_id>', methods=['GET'])
def api_biro26_credite_anketa(req_id):
    """RO: ancheta completa a cererii (pentru copiere in cererea la banca)."""
    from models.biro26_credit import Biro26Credit
    return _b26(lambda: Biro26Credit.request_anketa(req_id))

@app.route('/api/biro26/credite-docs/gsheets', methods=['GET'])
def api_biro26_credite_gsheets_status():
    from models.biro26_gsheets import Biro26GSheets
    return _b26(Biro26GSheets.status)

@app.route('/api/biro26/credite-docs/gsheets', methods=['POST'])
def api_biro26_credite_gsheets_sync():
    from models.biro26_gsheets import Biro26GSheets
    return _b26(Biro26GSheets.sync)

@app.route('/api/biro26/credite-docs/<int:cod>/lines', methods=['GET'])
def api_biro26_credite_doc_lines(cod):
    from models.biro26_credit import Biro26Credit
    return _b26(lambda: Biro26Credit.document_lines(cod))

@app.route('/api/biro26/credit/providers', methods=['GET'])
def api_biro26_credit_providers():
    return _b26(Biro26Controller.credit_providers)

@app.route('/api/biro26/credit/providers', methods=['PUT'])
def api_biro26_credit_provider_save():
    return _b26(Biro26Controller.credit_provider_save)

@app.route('/api/biro26/credit/providers/<code>/test', methods=['POST'])
def api_biro26_credit_provider_test(code):
    return _b26(lambda: Biro26Controller.credit_provider_test(code))

@app.route('/api/biro26/credit/requests/<int:req_id>/events', methods=['GET'])
def api_biro26_credit_request_events(req_id):
    return _b26(lambda: Biro26Controller.credit_request_events(req_id))

@app.route('/api/biro26/credit/requests/<int:req_id>/refresh', methods=['POST'])
def api_biro26_credit_request_refresh(req_id):
    return _b26(lambda: Biro26Controller.credit_request_refresh(req_id))

# ── credit: fluxul API al clientului (public, rate-limited) ──
@app.route('/api/biro26/shop/credit/api/preapproved', methods=['POST'])
@limiter.limit("10 per minute")
def api_biro26_shop_credit_preapproved():
    return jsonify(Biro26Controller.credit_api_preapproved())

@app.route('/api/biro26/shop/credit/api/submit', methods=['POST'])
@limiter.limit("5 per minute")
def api_biro26_shop_credit_submit():
    return jsonify(Biro26Controller.credit_api_submit())

@app.route('/api/biro26/shop/credit/api/status', methods=['GET'])
@limiter.limit("60 per minute")
def api_biro26_shop_credit_api_status():
    return jsonify(Biro26Controller.credit_api_status())

# ── translations management page + API (grouping RU/EN dictionary) ──
@app.route('/UNA.md/orasldev/biro26-translations')
def biro26_translations():
    if not AuthController.is_authenticated():
        return _login_redirect()
    from models.biro26_i18n import Biro26I18n
    return render_template('biro26/translations.html',
                           app_name=Config.BIRO26_APP_NAME,
                           last_job=Biro26I18n.last_job())

@app.route('/api/biro26/i18n/groups', methods=['GET'])
def api_biro26_i18n_groups():
    return _b26(Biro26Controller.i18n_groups)

@app.route('/api/biro26/i18n/groups', methods=['PUT'])
def api_biro26_i18n_save():
    return _b26(Biro26Controller.i18n_save)

@app.route('/api/biro26/i18n/groups.csv', methods=['GET'])
def api_biro26_i18n_export():
    g = _biro26_api_guard()
    if g is not None:
        return g
    from models.biro26_i18n import Biro26I18n
    only_missing = request.args.get('only_missing') == '1'
    resp = app.response_class(Biro26I18n.export_csv(only_missing),
                              mimetype='text/csv')
    resp.headers['Content-Disposition'] = \
        'attachment; filename="grupare_traduceri.csv"'
    return resp

@app.route('/api/biro26/i18n/import', methods=['POST'])
def api_biro26_i18n_import():
    return _b26(Biro26Controller.i18n_import)

@app.route('/api/biro26/i18n/auto', methods=['POST'])
def api_biro26_i18n_auto():
    return _b26(Biro26Controller.i18n_auto_start)

@app.route('/api/biro26/i18n/auto/<job_id>', methods=['GET'])
def api_biro26_i18n_auto_status(job_id):
    return _b26(lambda: Biro26Controller.i18n_auto_status(job_id))

# ── soft-delete (native ISARHIV): deactivate/reactivate a product card ──
@app.route('/api/biro26/products/<int:cod>/archive', methods=['PUT'])
def api_biro26_product_archive(cod):
    return _b26(lambda: Biro26Controller.product_archive(cod))

# ── online payments (MAIB card / MIA QR): shop + callbacks + admin ──
@app.route('/api/biro26/shop/pay/methods', methods=['GET'])
def api_biro26_pay_methods():
    return jsonify(Biro26Controller.pay_methods())

@app.route('/api/biro26/shop/pay/<method>', methods=['POST'])
def api_biro26_pay_create(method):
    # auth is enforced inside (shop client session or backoffice session)
    return jsonify(Biro26Controller.pay_create(method))

@app.route('/api/biro26/shop/pay/mia-status', methods=['GET'])
def api_biro26_pay_mia_status():
    return jsonify(Biro26Controller.pay_mia_check())

@app.route('/api/biro26/shop/order/<int:cod>', methods=['GET'])
def api_biro26_shop_order_view(cod):
    # RO/EN: detaliile comenzii pentru pagina de retur dupa plata (maib)
    return jsonify(Biro26Controller.shop_order_view(cod))

@app.route('/api/biro26/pay/maib-callback', methods=['GET', 'POST'])
def api_biro26_pay_maib_callback():
    """RO: okUrl/failUrl/callbackUrl de la MAIB — statusul se verifica
    server-side prin pay-info; browserul e intors in magazin.
    EN: MAIB return/callback — verified server-side via pay-info."""
    from models.biro26_pay import Biro26Pay
    body = request.get_json(silent=True) or {}
    order_key = request.args.get('orderKey') or body.get('orderId') or ''
    pay_id = (request.args.get('payId') or request.form.get('payId')
              or (body.get('result') or {}).get('payId')
              or body.get('payId') or '')
    typeurl = (request.args.get('typeurl') or 'callbackurl').lower()
    r = Biro26Pay.maib_callback(order_key, pay_id, typeurl)
    if request.method == 'POST' or typeurl == 'callbackurl':
        return jsonify(r)
    # RO: retur in browser (okUrl/failUrl) -> pagina cu DETALIILE comenzii,
    #     nu doar un flag; e cerinta maib pentru e-commerce.
    # EN: browser return -> order-details page, not just a flag (maib rule).
    # RO: caile scurte (/cos, /catalog) le mapeaza nginx-ul de pe officeplus.md;
    #     folosim calea completa, care e deja proxata pe ambele contururi —
    #     un alias /plata-rezultat se poate adauga ulterior in nginx.
    # EN: short paths are nginx-mapped on officeplus.md; use the full path,
    #     already proxied on both contours.
    return redirect('/UNA.md/orasldev/biro26-site/payment-result?pay='
                    + ('ok' if r.get('paid') else 'fail')
                    + '&cod=' + str(r.get('doc_cod') or 0))

@app.route('/api/biro26/pay/mia-callback', methods=['GET', 'POST'])
def api_biro26_pay_mia_callback():
    """RO: callbackEchoUrl de la MIA — statusul se reverifica prin API.
    EN: MIA echo callback — re-verified through the status API."""
    from models.biro26_pay import Biro26Pay
    order = request.args.get('orderKey') or ''
    return jsonify(Biro26Pay.mia_check(order) if order else {"success": True})

@app.route('/api/biro26/pay/refund', methods=['POST'])
def api_biro26_pay_refund():
    return _b26(Biro26Controller.pay_refund)

# ── admin test page: create ad-hoc MAIB test links, verify, refund ──
@app.route('/UNA.md/orasldev/biro26-pay-test')
def biro26_pay_test():
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/pay_test.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/pay/test-checkout', methods=['POST'])
def api_biro26_pay_test_checkout():
    return _b26(Biro26Controller.pay_test_create)

@app.route('/api/biro26/pay/list', methods=['GET'])
def api_biro26_pay_list():
    return _b26(Biro26Controller.pay_list)

@app.route('/api/biro26/pay/verify', methods=['POST'])
def api_biro26_pay_verify():
    return _b26(Biro26Controller.pay_verify)

@app.route('/api/biro26/pay-settings', methods=['GET'])
def api_biro26_pay_settings_get():
    return _b26(Biro26Controller.pay_settings_get)

@app.route('/api/biro26/pay-settings', methods=['PUT'])
def api_biro26_pay_settings_put():
    return _b26(Biro26Controller.pay_settings_put)

# ── shop display settings (admin): products per page ──
@app.route('/api/biro26/shop-settings', methods=['GET'])
def api_biro26_shop_settings_get():
    return _b26(Biro26Controller.shop_settings_get)

@app.route('/api/biro26/shop-settings', methods=['PUT'])
def api_biro26_shop_settings_put():
    return _b26(Biro26Controller.shop_settings_put)

# ── product window: description (public read) + client comments ──
@app.route('/api/biro26/shop/product/<int:cod>', methods=['GET'])
def api_biro26_shop_product(cod):
    return jsonify(Biro26Controller.shop_product_info(cod))

@app.route('/api/biro26/shop/product/<int:cod>/comment', methods=['POST'])
def api_biro26_shop_product_comment(cod):
    # auth is enforced inside (shop client session or backoffice session)
    return jsonify(Biro26Controller.shop_product_comment(cod))

@app.route('/api/biro26/product-desc/<int:cod>', methods=['PUT'])
def api_biro26_product_desc(cod):
    return _b26(lambda: Biro26Controller.set_product_desc(cod))

# ── Arhivele SITE-ului (saptaminal, surse+metadate, fara marfa ERP) ──
@app.route('/UNA.md/orasldev/biro26-backups')
def biro26_backups():
    """RO: pagina «Arhive site» — descarcare arhive + setari FTP/SFTP."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/backup_admin.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/backup/archives', methods=['GET'])
def api_biro26_backup_archives():
    from models.biro26_backup import Biro26Backup
    return _b26(lambda: Biro26Backup.archives())

@app.route('/api/biro26/backup/download/<name>', methods=['GET'])
def api_biro26_backup_download(name):
    if not AuthController.is_authenticated():
        return _login_redirect()
    from models.biro26_backup import Biro26Backup
    p = Biro26Backup.archive_path(name)
    if not p:
        return jsonify({'success': False, 'error': 'arhiva inexistenta'}), 404
    from flask import send_file
    return send_file(p, as_attachment=True, download_name=name)

@app.route('/api/biro26/backup/run', methods=['POST'])
def api_biro26_backup_run():
    from models.biro26_backup import Biro26Backup
    return _b26(lambda: Biro26Backup.run_now())

@app.route('/api/biro26/backup/dest', methods=['GET', 'POST'])
def api_biro26_backup_dest():
    from models.biro26_backup import Biro26Backup
    if request.method == 'GET':
        return _b26(lambda: Biro26Backup.dest_list())
    return _b26(lambda: Biro26Backup.dest_save(
        request.get_json(silent=True) or {}))

@app.route('/api/biro26/backup/dest/<int:did>', methods=['DELETE'])
def api_biro26_backup_dest_del(did):
    from models.biro26_backup import Biro26Backup
    return _b26(lambda: Biro26Backup.dest_delete(did))

@app.route('/api/biro26/backup/log', methods=['GET'])
def api_biro26_backup_log():
    from models.biro26_backup import Biro26Backup
    return _b26(lambda: Biro26Backup.log_list())

# ── TMS_MPT_WEBATTR: atribute web multilingve (BLOB = original cu
#    diacritice; copiile de cautare le intretine triggerul) ──
@app.route('/api/biro26/webattr/<int:cod>', methods=['GET'])
def api_biro26_webattr_get(cod):
    return _b26(lambda: Biro26Controller.webattr_get(cod))

@app.route('/api/biro26/webattr/<int:cod>', methods=['PUT'])
def api_biro26_webattr_put(cod):
    return _b26(lambda: Biro26Controller.webattr_save(cod))

@app.route('/api/biro26/product-comment/<int:cid>', methods=['DELETE'])
def api_biro26_product_comment_del(cid):
    return _b26(lambda: Biro26Controller.delete_product_comment(cid))

@app.route('/api/biro26/shop/register', methods=['POST'])
def api_biro26_shop_register():
    return jsonify(Biro26Controller.shop_register())

@app.route('/api/biro26/shop/login', methods=['POST'])
def api_biro26_shop_login():
    return jsonify(Biro26Controller.shop_login())

@app.route('/api/biro26/shop/logout', methods=['POST'])
def api_biro26_shop_logout():
    return jsonify(Biro26Controller.shop_logout())

@app.route('/api/biro26/shop/me/type', methods=['PUT'])
def api_biro26_shop_me_type():
    # RO: tip client (fizica/juridica) din cabinet -> schimba preturile
    return jsonify(Biro26Controller.shop_set_client_type())

@app.route('/api/biro26/shop/me/credit-save', methods=['PUT'])
def api_biro26_shop_me_credit_save():
    # RO: memorarea datelor formularului de credit — pornita/oprita din cabinet
    return jsonify(Biro26Controller.shop_credit_profile_set())

@app.route('/api/biro26/shop/me/fmt', methods=['PUT'])
def api_biro26_shop_me_fmt():
    # RO: constanta personala — formatele contului (pdf/html/xlsx)
    return jsonify(Biro26Controller.shop_set_invoice_fmt())

@app.route('/api/biro26/shop/me', methods=['GET'])
def api_biro26_shop_me():
    return jsonify(Biro26Controller.shop_me())

@app.route('/api/biro26/shop/products', methods=['GET'])
def api_biro26_shop_products():
    # public read-only catalog (same grid data as Marfă/Stoc)
    return jsonify(Biro26Controller.get_products_stock())

@app.route('/api/biro26/shop/tree', methods=['GET'])
def api_biro26_shop_tree():
    # public read-only grupa->categorie facet tree (Amazon-style sidebar)
    return jsonify(Biro26Controller.get_product_tree())

@app.route('/api/biro26/shop/brands', methods=['GET'])
def api_biro26_shop_brands():
    # public read-only brand facet with counts
    return jsonify(Biro26Controller.get_product_brands())

@app.route('/api/biro26/shop/transport', methods=['GET'])
def api_biro26_shop_transport():
    # public read-only: round-trip transport tariff grid (TMS_MPT_DISTANTE)
    return jsonify(Biro26Controller.shop_transport())

@app.route('/api/biro26/shop/logistics', methods=['GET'])
def api_biro26_shop_logistics():
    # public read-only: ACTIVE logistics centers (TMS_MPT_CENTRE_LOG) —
    # the transport distance is measured from the chosen center
    return jsonify(Biro26Controller.shop_logistics())

@app.route('/api/biro26/shop/services', methods=['GET'])
def api_biro26_shop_services():
    # public read-only: optional services for the cart (group from
    # YBIRO_SETTINGS.SHOP_SERVICES_GRUPA)
    return jsonify(Biro26Controller.shop_services())

@app.route('/api/biro26/shop/variants', methods=['GET'])
def api_biro26_shop_variants():
    # public read-only variant family (choose a characteristic in the shop)
    return jsonify(Biro26Controller.shop_variants())

# ── external-app API: customer document list + PDFs by NUMBER (#NRMANUAL) ──
@app.route('/api/biro26/docs', methods=['GET'])
def api_biro26_docs_list():
    # ?client=<nume|cod|#nr>&limit= — X-API-Key token or backoffice session
    r = Biro26Controller.docs_list()
    return jsonify(r), (200 if r.get('success')
                        else 401 if r.get('error') == 'login required' else 400)

@app.route('/api/biro26/report-by-nr/<kind>/<path:nr>', methods=['GET'])
def api_biro26_report_by_nr(kind, nr):
    """RO: PDF dupa numarul documentului: /report-by-nr/invoice/%23338
    (sau 338). Auth: X-API-Key, ?sig= sau sesiune backoffice.
    EN: PDF by the document number (hashtag form supported)."""
    r = Biro26Controller.report_by_nr(kind, nr)
    if not r.get('success'):
        return jsonify(r), (401 if r.get('error') == 'login required'
                            else 404 if 'not found' in str(r.get('error'))
                            else 400)
    names = {'invoice': 'Cont_de_plata', 'order': 'Comanda'}
    resp = app.response_class(r['pdf'], mimetype='application/pdf')
    resp.headers['Content-Disposition'] = (
        f'inline; filename="{names.get(kind, kind)}_'
        f'{str(nr).lstrip("#")}.pdf"')
    # RO/EN: avoid CDN/browser serving a stale PDF after NRMANUAL fix
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp

@app.route('/api/biro26/doc/<int:cod>', methods=['GET'])
def api_biro26_doc_json(cod):
    # document data as JSON for desktop/integration layers
    # (X-API-Key token or backoffice session)
    r = Biro26Controller.doc_json(cod)
    return jsonify(r), (200 if r.get('success')
                        else 401 if r.get('error') == 'login required' else 400)

@app.route('/api/biro26/shop/report/<kind>/<int:cod>', methods=['GET'])
def api_biro26_shop_report(kind, cod):
    # PDF via the jsReport sidecar; shop clients only for their own docs,
    # backoffice sessions for any (see Biro26Controller.shop_report)
    r = Biro26Controller.shop_report(kind, cod)
    if not r.get('success'):
        return jsonify(r), (401 if r.get('error') == 'login required' else 400)
    names = {'invoice': 'Cont_de_plata', 'order': 'Comanda'}
    resp = app.response_class(r['pdf'], mimetype='application/pdf')
    resp.headers['Content-Disposition'] = \
        f'inline; filename="{names.get(kind, kind)}_{cod}.pdf"'
    # RO/EN: avoid CDN/browser serving a stale PDF after NRMANUAL fix
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp

@app.route('/api/biro26/gen-docs-by-nr/<path:nr>', methods=['GET', 'POST'])
def api_biro26_gen_docs_by_nr(nr):
    """RO: genereaza + ataseaza contul de plata si comanda la un document
    EXISTENT (dupa NRMANUAL) — apelabil din Oracle (y_ai_BIRO26.gen_conturi,
    UTL_HTTP pe http://officeplus.md) sau desktop (?api_key=)."""
    r = Biro26Controller.gen_docs_by_nr(nr)
    return jsonify(r), (200 if r.get('success')
                        else 401 if r.get('error') == 'login required' else 400)

@app.route('/api/biro26/shop/report-html/<kind>/<int:cod>', methods=['GET'])
def api_biro26_shop_report_html(kind, cod):
    """RO: formularul in varianta HTML (stil site/una.md, modelul aprobat) —
    aceeasi paza ca la PDF; se deschide direct in browser."""
    r = Biro26Controller.shop_report_html(kind, cod)
    if not r.get('success'):
        return jsonify(r), (401 if r.get('error') == 'login required' else 400)
    resp = app.response_class(r['html'], mimetype='text/html')
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return resp

@app.route('/api/biro26/shop/report-xlsx/invoice/<int:cod>', methods=['GET'])
def api_biro26_shop_report_xlsx(cod):
    """RO: contul de plata in EXCEL (bifa «si Excel» din cos): tabel Excel
    adevarat, Suma=formula cant*pret, TOTAL=formula SUM, logo inclus."""
    r = Biro26Controller.shop_report_xlsx(cod)
    if not r.get('success'):
        return jsonify(r), (401 if r.get('error') == 'login required' else 400)
    resp = app.response_class(
        r['xlsx'],
        mimetype='application/vnd.openxmlformats-officedocument.'
                 'spreadsheetml.sheet')
    resp.headers['Content-Disposition'] = \
        f'attachment; filename="Cont_de_plata_{cod}.xlsx"'
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return resp

def _biro26_social_conv(kind, r, amount=None):
    """RO: inregistreaza conversia in atributia sociala (fail-silent);
    suma se calculeaza best-effort din payload cind nu e data explicit."""
    try:
        if not (r or {}).get('success'):
            return
        if amount is None:
            items = (request.get_json(silent=True) or {}).get('items') or []
            amount = sum(float(i.get('qty') or 0) * float(i.get('price') or 0)
                         for i in items if isinstance(i, dict)) or None
        doc = ((r.get('data') or {}).get('cod')
               if isinstance(r.get('data'), dict) else None)
        from models.biro26_social import Biro26Social
        Biro26Social.conversion(request, kind, doc=doc, amount=amount)
    except Exception:                                        # noqa: BLE001
        pass

@app.route('/api/biro26/shop/invoice', methods=['POST'])
def api_biro26_shop_invoice():
    r = Biro26Controller.shop_invoice()
    _biro26_social_conv('invoice', r)
    return jsonify(r)

@app.route('/api/biro26/b2b/order', methods=['POST'])
def api_biro26_b2b_order():
    """RO: comanda B2B — clienti autentificati SAU integrari/angajati cu
    cheie de incredere (X-API-Key + client_cod). Raspunsul include mostra
    contului: linkuri semnate PDF/HTML + JSON-ul documentului.
    EN: B2B order placement; the reply carries the invoice sample links."""
    r = Biro26Controller.b2b_order()
    _biro26_social_conv('b2b', r)
    return jsonify(r), (200 if r.get('success')
                        else 401 if r.get('error') == 'login required' else 400)

# ── Biro26: JURNAL UNIVERSAL (back-office) ─────────────────────────────────
# RO: o singura forma pentru documente — filtre, grila master si patru file
#     (contari / marfuri / fisiere / LOG). Doar tabele Oracle EXISTENTE.

@app.route('/UNA.md/orasldev/biro26-journal')
def biro26_journal_page():
    """RO: jurnalul universal de documente + casa de operator."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    return render_template('biro26/journal.html',
                           app_name=Config.BIRO26_APP_NAME)

@app.route('/api/biro26/journal/docs', methods=['GET'])
def api_biro26_journal_docs():
    if not AuthController.is_authenticated():
        return jsonify({'success': False, 'error': 'auth required'}), 401
    from models.biro26_journal import Biro26Journal
    return jsonify(Biro26Journal.docs(
        request.args.get('from', ''), request.args.get('to', ''),
        request.args.get('q', ''), request.args.get('limit', 200, type=int)))

@app.route('/api/biro26/journal/doc/<int:cod>', methods=['GET'])
def api_biro26_journal_detail(cod):
    if not AuthController.is_authenticated():
        return jsonify({'success': False, 'error': 'auth required'}), 401
    from models.biro26_journal import Biro26Journal
    return jsonify(Biro26Journal.detail(cod))

@app.route('/api/biro26/journal/clients', methods=['GET'])
def api_biro26_journal_clients():
    if not AuthController.is_authenticated():
        return jsonify({'success': False, 'error': 'auth required'}), 401
    from models.biro26_journal import Biro26Journal
    return jsonify(Biro26Journal.clients(request.args.get('q', ''),
                                         request.args.get('limit', 30, type=int)))

@app.route('/api/biro26/journal/client', methods=['POST'])
def api_biro26_journal_client_add():
    """RO: inregistrare RAPIDA a clientului de catre operator (minim:
    denumirea + tipul fizica/juridica). Ajunge in aceleasi tabele ca
    inregistrarea din cabinetul clientului."""
    if not AuthController.is_authenticated():
        return jsonify({'success': False, 'error': 'auth required'}), 401
    from models.biro26_journal import Biro26Journal
    d = request.get_json(silent=True) or {}
    r = Biro26Journal.client_quick_add(
        d.get('name', ''), bool(d.get('is_company')), d.get('idno', ''),
        d.get('phone', ''), d.get('email', ''), d.get('address', ''))
    return jsonify(r), (200 if r.get('success') else 400)

@app.route('/UNA.md/orasldev/biro26-site/credit-form')
@app.route('/UNA.md/orasldev/biro26-1shop/credit-form')
def biro26_site_credit_form():
    """RO: cererea de credit dupa macheta owner-ului (2 pasi + acte)."""
    return render_template('biro26/site_credit_form.html', **_biro26_site_ctx())

@app.route('/api/biro26/shop/credit/apply', methods=['POST'])
def api_biro26_credit_apply():
    r = Biro26Controller.credit_apply()
    _biro26_social_conv('credit', r)
    return jsonify(r), (200 if r.get('success') else 400)

@app.route('/api/biro26/shop/my-files', methods=['GET'])
def api_biro26_client_files_list():
    """RO: actele personale ale clientului (cabinet) / dosarul unui client (operator)."""
    r = Biro26Controller.client_files_list()
    return jsonify(r), (200 if r.get('success')
                        else 401 if r.get('error') == 'login required' else 400)

@app.route('/api/biro26/shop/my-files', methods=['POST'])
def api_biro26_client_files_upload():
    r = Biro26Controller.client_files_upload()
    return jsonify(r), (200 if r.get('success')
                        else 401 if r.get('error') == 'login required' else 400)

@app.route('/api/biro26/shop/my-files/<int:file_id>', methods=['GET'])
def api_biro26_client_file_get(file_id):
    r = Biro26Controller.client_files_get(file_id)
    if not r.get('success'):
        return jsonify(r), (401 if r.get('error') == 'login required' else 404)
    d = r['data']
    resp = app.response_class(d['content'], mimetype=d['mime'])
    # RO/EN: date personale — fara cache in browser/CDN
    resp.headers['Content-Disposition'] = f'inline; filename="{d["file_name"]}"'
    resp.headers['Cache-Control'] = 'no-store, private, max-age=0'
    return resp

@app.route('/api/biro26/shop/my-files/<int:file_id>', methods=['DELETE'])
def api_biro26_client_file_delete(file_id):
    r = Biro26Controller.client_files_delete(file_id)
    return jsonify(r), (200 if r.get('success')
                        else 401 if r.get('error') == 'login required' else 400)

@app.route('/api/biro26/shop/my-invoices', methods=['GET'])
def api_biro26_shop_my_invoices():
    """RO: cabinet client — lista propriilor conturi («Comenzile mele»)."""
    r = Biro26Controller.shop_my_invoices()
    return jsonify(r), (200 if r.get('success')
                        else 401 if r.get('error') == 'login required' else 400)


# ── Biro26: service (maintenance) functions — dynamic registry ─────────────
# RO: lista vine din YBIRO_SERVICE_FUNCTIONS; o functie noua = un simplu INSERT.
# EN: the list comes from YBIRO_SERVICE_FUNCTIONS; a new function = one INSERT.

@app.route('/api/biro26/services', methods=['GET'])
def api_biro26_services():
    if not AuthController.is_authenticated():
        return jsonify({'success': False, 'error': 'auth required'}), 401
    from models.biro26_services import Biro26Services
    return jsonify(Biro26Services.list_functions(request.args.get('lang', 'ro')))


@app.route('/api/biro26/services/<code>/count', methods=['GET'])
def api_biro26_services_count(code):
    if not AuthController.is_authenticated():
        return jsonify({'success': False, 'error': 'auth required'}), 401
    from models.biro26_services import Biro26Services
    return jsonify(Biro26Services.count(code))


@app.route('/api/biro26/services/<code>/csv', methods=['GET'])
def api_biro26_services_csv(code):
    if not AuthController.is_authenticated():
        return jsonify({'success': False, 'error': 'auth required'}), 401
    from models.biro26_services import Biro26Services
    res = Biro26Services.to_csv(code)
    if not res.get('success'):
        return jsonify(res), 400
    # RO: BOM => Excel (RO/RU) recunoaste UTF-8 / EN: BOM so Excel detects UTF-8
    return Response(
        '﻿' + res['csv'],
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition':
                 f'attachment; filename="{res["file_name"]}"'})


# ========== PECO (розничная продажа топлива в сети АЗС) Routes ==========

def _peco_station_id():
    """Станция из query-параметра; по умолчанию первая активная."""
    raw = request.args.get('station_id') or request.form.get('station_id')
    if raw:
        try:
            return int(raw)
        except (TypeError, ValueError):
            pass
    res = PecoController.default_station_id()
    return res.get('station_id') if res.get('success') else None


@app.route('/UNA.md/orasldev/peco-pump')
@app.route('/UNA.md/orasldev/peco-pump/')
def peco_pump():
    """Фронт-офис колонки: самообслуживание и отпуск сотрудником."""
    if not AuthController.is_authenticated():
        return redirect(url_for('login', next=request.path))
    return render_template('peco_pump.html')


@app.route('/UNA.md/orasldev/peco-shift')
@app.route('/UNA.md/orasldev/peco-shift/')
def peco_shift_console():
    """Консоль оператора АЗС: смена, счётчики, приём цистерн, касса."""
    if not AuthController.is_authenticated():
        return redirect(url_for('login', next=request.path))
    return render_template('peco_shift.html')


@app.route('/UNA.md/orasldev/peco-admin')
@app.route('/UNA.md/orasldev/peco-admin/')
def peco_admin():
    """Бэк-офис: сеть АЗС, цены, остатки, расхождения."""
    if not AuthController.is_authenticated():
        return redirect(url_for('login', next=request.path))
    return render_template('peco_admin.html')


@app.route('/UNA.md/orasldev/peco')
@app.route('/UNA.md/orasldev/peco/')
def peco_index():
    """Главная страница PECO — единая точка доступа со ссылками на все функции."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    p = Path(os.path.dirname(os.path.abspath(__file__))) / 'docs' / 'PECO' / 'index.html'
    if not p.is_file():
        return "<h1>Не найдено</h1><p><a href='/UNA.md/orasldev/docs'>Назад</a></p>", 404
    from flask import send_file
    return send_file(str(p))


@app.route('/UNA.md/orasldev/docs/peco/TZ.html')
@app.route('/UNA.md/orasldev/docs/peco/')
def peco_tz():
    """Страница технического задания PECO с кнопками входа в интерфейсы."""
    if not AuthController.is_authenticated():
        return _login_redirect()
    p = Path(os.path.dirname(os.path.abspath(__file__))) / 'docs' / 'PECO' / 'TZ.html'
    if not p.is_file():
        return "<h1>Не найдено</h1><p><a href='/UNA.md/orasldev/docs'>Назад</a></p>", 404
    from flask import send_file
    return send_file(str(p))


# ---------- PECO API ----------

@app.route('/api/peco/pump/state')
def api_peco_pump_state():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    station_id = _peco_station_id()
    if station_id is None:
        return jsonify({"success": False, "error": "Нет активных станций"})
    return jsonify(PecoController.pump_state(station_id))


@app.route('/api/peco/txn/authorize', methods=['POST'])
def api_peco_txn_authorize():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.authorize(request.get_json(silent=True) or {}))


@app.route('/api/peco/txn/start', methods=['POST'])
def api_peco_txn_start():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.start(request.get_json(silent=True) or {}))


@app.route('/api/peco/txn/finish', methods=['POST'])
def api_peco_txn_finish():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.finish(request.get_json(silent=True) or {}))


@app.route('/api/peco/txn/pay', methods=['POST'])
def api_peco_txn_pay():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.pay(request.get_json(silent=True) or {}))


@app.route('/api/peco/txn/void', methods=['POST'])
def api_peco_txn_void():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.void(request.get_json(silent=True) or {}))


@app.route('/api/peco/shift/open', methods=['POST'])
def api_peco_shift_open():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.shift_open(request.get_json(silent=True) or {}))


@app.route('/api/peco/shift/<int:shift_id>/meters')
def api_peco_shift_meters(shift_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.shift_meters(shift_id))


@app.route('/api/peco/shift/meter', methods=['POST'])
def api_peco_shift_meter():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.shift_save_meter(request.get_json(silent=True) or {}))


@app.route('/api/peco/shift/close', methods=['POST'])
def api_peco_shift_close():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.shift_close(request.get_json(silent=True) or {}))


@app.route('/api/peco/shift/approve', methods=['POST'])
def api_peco_shift_approve():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.shift_approve(request.get_json(silent=True) or {}))


@app.route('/api/peco/delivery', methods=['POST'])
def api_peco_delivery():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.delivery_receive(request.get_json(silent=True) or {}))


@app.route('/api/peco/tanks')
def api_peco_tanks():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    station_id = _peco_station_id()
    if station_id is None:
        return jsonify({"success": False, "error": "Нет активных станций"})
    return jsonify(PecoController.tank_levels(station_id))


@app.route('/api/peco/prices')
def api_peco_prices():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    station_id = _peco_station_id()
    if station_id is None:
        return jsonify({"success": False, "error": "Нет активных станций"})
    return jsonify(PecoController.prices(station_id))


@app.route('/api/peco/admin/overview')
def api_peco_admin_overview():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.admin_overview())


@app.route('/api/peco/admin/price', methods=['POST'])
def api_peco_admin_price():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.set_price(request.get_json(silent=True) or {}))


@app.route('/api/peco/employees')
def api_peco_employees():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    station_id = _peco_station_id() if request.args.get('station_id') else None
    return jsonify(PecoController.employees(station_id))


@app.route('/api/peco/employee/pin', methods=['POST'])
def api_peco_employee_pin():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.set_pin(request.get_json(silent=True) or {}))


@app.route('/api/peco/shifts/disputed')
def api_peco_shifts_disputed():
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    station_id = _peco_station_id() if request.args.get('station_id') else None
    return jsonify(PecoController.disputed_shifts(station_id))


@app.route('/api/peco/shift/<int:shift_id>/summary')
def api_peco_shift_summary(shift_id):
    if not AuthController.is_authenticated():
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401
    return jsonify(PecoController.shift_summary(shift_id))


if __name__ == '__main__':
    # Запускаем фоновый поток для обновления метрик
    updater_thread = threading.Thread(target=background_metric_updater, daemon=True)
    updater_thread.start()
    
    # Запускаем приложение с SocketIO
    # Используем параметры из конфигурации (поддерживает локальный и удаленный режимы)
    print(f"🚀 Запуск сервера в режиме: {Config.ENVIRONMENT}")
    
    # Получаем локальный IP адрес для отображения
    local_ip = None
    if Config.IS_LOCAL:
        try:
            import socket
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(('8.8.8.8', 80))
            local_ip = s.getsockname()[0]
            s.close()
        except Exception:
            pass
    
    print(f"📍 Адрес:")
    print(f"   • Localhost: http://localhost:{Config.SERVER_PORT}")
    if local_ip:
        print(f"   • Локальная сеть: http://{local_ip}:{Config.SERVER_PORT}")
    print(f"")
    print(f"🌐 Dashboard:")
    print(f"   • http://localhost:{Config.SERVER_PORT}/UNA.md/orasldev/dashboard")
    if local_ip:
        print(f"   • http://{local_ip}:{Config.SERVER_PORT}/UNA.md/orasldev/dashboard")
        print(f"   • Fullscreen: http://{local_ip}:{Config.SERVER_PORT}/UNA.md/orasldev/dashboard/01")
    print(f"")
    print(f"📂 Shell (список проектов из UNA_SHELL_PROJECTS):")
    print(f"   • http://localhost:{Config.SERVER_PORT}/una.md/shell/projects")
    if local_ip:
        print(f"   • http://{local_ip}:{Config.SERVER_PORT}/una.md/shell/projects")
    print(f"")
    print(f"⚖️ DIGI SM (управление весами):")
    print(f"   • http://localhost:{Config.SERVER_PORT}/UNA.md/orasldev/digi-sm")
    if local_ip:
        print(f"   • http://{local_ip}:{Config.SERVER_PORT}/UNA.md/orasldev/digi-sm")
    
    use_reloader = Config.ENVIRONMENT != "REMOTE"
    is_debug = Config.ENVIRONMENT != "REMOTE"
    socketio.run(app, host=Config.SERVER_HOST, port=Config.SERVER_PORT, debug=is_debug, use_reloader=use_reloader, allow_unsafe_werkzeug=True)

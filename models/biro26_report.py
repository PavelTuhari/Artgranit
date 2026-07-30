"""Biro26 printable reports via the jsReport sidecar (reports/, Node.js).

RO: Randare PDF pentru "cont de plata" si "comanda cumparatorului" din
    documentele create de y_ai_BIRO26 (vizibile in VMDB_DOCS_WORK /
    VMDB_ST201M / VMDB_ST201D). Sablonul Handlebars + datele se trimit
    inline la POST {JSREPORT_URL}/api/report (recipe chrome-pdf), deci nu
    depindem de store-ul jsReport.
EN: PDF rendering for the invoice ("cont de plata") and customer order
    forms of documents created by y_ai_BIRO26. The Handlebars template +
    data go inline to POST {JSREPORT_URL}/api/report (chrome-pdf recipe),
    so no jsReport store configuration is required.
"""
from __future__ import annotations

import os
from typing import Any, Dict, Optional

import requests

from config import Config
from models.biro26_db import Biro26DB
from models.biro26_oracle_store import _rows

_TPL_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "reports", "templates")

REPORT_KINDS = {"invoice": "biro26_invoice.hbs", "order": "biro26_order.hbs"}
# RO: al 3-lea sablon pdfme (curat) — camp invoice_nr pentru numarul contului
# EN: 3rd pdfme template (clean) — invoice_nr field for the invoice number
PDFME_KINDS = {
    "invoice": "pdfme_cont_plata.json",
    "order": "pdfme_order.json",
}
ENGINES_FILE = "engines.json"          # {"invoice": "jsreport"|"pdfme", ...}

_RO_MONTHS = ["", "ianuarie", "februarie", "martie", "aprilie", "mai", "iunie",
              "iulie", "august", "septembrie", "octombrie", "noiembrie", "decembrie"]


def _read(fname: str) -> str:
    with open(os.path.join(_TPL_DIR, fname), encoding="utf-8") as f:
        return f.read()


def _fmt(v) -> str:
    """1234.5 -> '1 234,50' (same as the JS helper)."""
    s = f"{float(v or 0):,.2f}"
    return s.replace(",", " ").replace(".", ",")


# RO: suma in litere in romana — portul Python al helpers.js (pdfme nu are
# helper-e, deci textul vine gata calculat in inputs).
# EN: Romanian amount-in-words — Python port of helpers.js (pdfme has no
# helpers, the text arrives precomputed in the inputs).
def _ro_words(n_raw) -> str:
    n = int(abs(float(n_raw or 0)))
    if n == 0:
        return "zero"
    uni = ["", "unu", "doi", "trei", "patru", "cinci", "șase", "șapte", "opt", "nouă"]
    spr = ["zece", "unsprezece", "doisprezece", "treisprezece", "paisprezece",
           "cincisprezece", "șaisprezece", "șaptesprezece", "optsprezece", "nouăsprezece"]

    def sub1000(x):
        parts = []
        h, r = divmod(x, 100)
        if h == 1:
            parts.append("o sută")
        elif h == 2:
            parts.append("două sute")
        elif h > 2:
            parts.append(uni[h] + " sute")
        if 10 <= r <= 19:
            parts.append(spr[r - 10])
        else:
            t, u = divmod(r, 10)
            if t == 2:
                parts.append("douăzeci și " + uni[u] if u else "douăzeci")
            elif t > 2:
                parts.append(uni[t] + "zeci și " + uni[u] if u else uni[t] + "zeci")
            elif u:
                parts.append(uni[u])
        return " ".join(parts)

    def scale(x, one, few, many):
        if x == 1:
            return one
        if x == 2:
            return "două " + few
        if x < 20:
            return sub1000(x) + " " + few
        return sub1000(x) + " de " + many

    out = []
    mil, n = divmod(n, 1000000)
    mii, n = divmod(n, 1000)
    if mil:
        out.append(scale(mil, "un milion", "milioane", "milioane"))
    if mii:
        out.append(scale(mii, "o mie", "mii", "mii"))
    if n:
        out.append(sub1000(n))
    s = " ".join(out)
    return s[:1].upper() + s[1:]


_LOGO_FILE = "logo.jpg"     # RO: logo-ul clientului / EN: the customer logo


def _logo_data_uri() -> Optional[str]:
    """Customer logo as a data URI (both engines embed it inline);
    replace reports/templates/logo.jpg to change the logo everywhere."""
    import base64
    p = os.path.join(_TPL_DIR, _LOGO_FILE)
    if not os.path.exists(p):
        return None
    with open(p, "rb") as fh:
        return "data:image/jpeg;base64," + base64.b64encode(fh.read()).decode()


def _ro_amount(total) -> str:
    v = float(total or 0)
    bani = round((v - int(v)) * 100)
    return f"{_ro_words(v)}, {bani:02d} ({_fmt(v)}) lei"


class Biro26Report:

    @staticmethod
    def docs_list(client: str = "", limit: int = 50) -> Dict[str, Any]:
        """RO: lista documentelor (conturi de plata web, SYSFID=12280) pentru
        aplicatiile EXTERNE (API): nr. documentului (#NRMANUAL — numarul
        vizibil), data, clientul, totalul si COD-ul intern.
        EN: document list for EXTERNAL apps: #NRMANUAL (visible number),
        date, client, total and the internal COD."""
        try:
            sql = ("SELECT * FROM ("
                   "SELECT d.COD, d.NRMANUAL, d.NRSET, "
                   "TO_CHAR(d.DATAMANUAL,'DD.MM.YYYY') DDATE, "
                   "m.DTDEP CLIENT_COD, u.DENUMIREA CLIENT_NAME, "
                   "(SELECT ROUND(SUM(l.SUMA),2) FROM VMDB_ST201D l "
                   " WHERE l.NRDOC = d.COD) TOTAL "
                   "FROM TMDB_DOCS d "
                   "JOIN VMDB_ST201M m ON m.NRDOC = d.COD "
                   "LEFT JOIN TMS_UNIVERS u ON u.COD = m.DTDEP "
                   "WHERE d.SYSFID = 12280")
            params: Dict[str, Any] = {}
            if client:
                sql += (" AND (UPPER(u.DENUMIREA) LIKE UPPER(:cl) "
                        "OR TO_CHAR(m.DTDEP) = :cl2 "
                        "OR TRIM(d.NRMANUAL) = :cl3 "
                        "OR TO_CHAR(d.NRSET) = :cl3)")
                params.update({"cl": f"%{client}%",
                               "cl2": client.lstrip('#'),
                               "cl3": client.lstrip('#')})
            sql += " ORDER BY d.COD DESC) WHERE ROWNUM <= :n"
            params["n"] = max(1, min(int(limit), 500))
            rows = _rows(Biro26DB().execute_query(sql, params))
            for r in rows:
                # RO/EN: visible invoice number is NRMANUAL (not NRSET subset)
                vis = r.get("nrmanual") or r.get("nrset")
                r["nr"] = f"#{vis}" if vis is not None and vis != "" else ""
            return {"success": True, "data": rows}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @staticmethod
    def resolve_nr(nr) -> Optional[int]:
        """RO: '#338' / '338' / 'A-23' -> COD-ul intern (cel mai recent
        document cu acest NRMANUAL). Suporta numerele cu SERIE (A-1..Z-999).
        EN: hashtag/number/series-number -> latest internal COD."""
        import re
        s = str(nr or "").strip().lstrip("#").strip()
        m = re.fullmatch(r"([A-Za-z])\s*-\s*(\d+)", s)
        if m:
            # RO: numar cu serie — potrivire exacta 'A-23'
            key = f"{m.group(1).upper()}-{int(m.group(2))}"
            rows = _rows(Biro26DB().execute_query(
                "SELECT MAX(COD) COD FROM TMDB_DOCS WHERE SYSFID = 12280 "
                "AND UPPER(TRIM(NRMANUAL)) = :s", {"s": key}))
            return int(rows[0]["cod"]) if rows and rows[0]["cod"] else None
        try:
            n = int(s)
        except (TypeError, ValueError):
            return None
        rows = _rows(Biro26DB().execute_query(
            "SELECT MAX(COD) COD FROM TMDB_DOCS "
            "WHERE SYSFID = 12280 "
            "AND (TRIM(NRMANUAL) = :s "
            "     OR (REGEXP_LIKE(TRIM(NRMANUAL), '^[0-9]+$') "
            "         AND TO_NUMBER(TRIM(NRMANUAL)) = :n))",
            {"s": str(n), "n": n}))
        return int(rows[0]["cod"]) if rows and rows[0]["cod"] else None

    @staticmethod
    def doc_data(cod: int) -> Dict[str, Any]:
        """Collect everything the forms need for one document COD.

        RO: Numarul contului vine DIN pachetul y_ai_BIRO26.get_nrmanual
            (TMDB_DOCS.NRMANUAL) — recompilarea pachetului schimba imediat
            sursa numarului, fara restart app.
        EN: Invoice number comes FROM package y_ai_BIRO26.get_nrmanual
            (TMDB_DOCS.NRMANUAL) — recompile package to change the source
            immediately, no app restart.
        """
        db = Biro26DB()
        # RO: DOAR coloana NRMANUAL (simplu, fara dependenta de pachet in SELECT).
        #     Niciodata NRSET (subset=201).
        # EN: ONLY column NRMANUAL (simple; no package call in SELECT).
        #     Never NRSET (subset=201).
        head = _rows(db.execute_query(
            "SELECT d.COD, "
            "TRIM(d.NRMANUAL) AS NRMANUAL, "
            "TO_CHAR(d.DATAMANUAL,'DD.MM.YYYY') DDATE, "
            "TO_CHAR(d.DATAMANUAL,'DD') DD, TO_CHAR(d.DATAMANUAL,'MM') MM, "
            "TO_CHAR(d.DATAMANUAL,'YYYY') YY, m.DTDEP CLIENT_COD, "
            "u.DENUMIREA CLIENT_NAME "
            "FROM TMDB_DOCS d "
            "JOIN VMDB_ST201M m ON m.NRDOC = d.COD "
            "LEFT JOIN TMS_UNIVERS u ON u.COD = m.DTDEP "
            "WHERE d.COD = :c", {"c": int(cod)}))
        if not head:
            return {"success": False, "error": "document not found"}
        h = head[0]
        # RO: telefon/email daca clientul e din magazinul public
        # EN: phone/email when the client came from the public shop
        extra = _rows(db.execute_query(
            "SELECT phone, email, address, idno FROM YBIRO_CLIENT "
            "WHERE univers_cod = :c", {"c": h["client_cod"]}))
        # RO: rechizitele plătitorului din cartela nativă a contragentului
        #     (TMS_ORG via VMS_UNIV_ORG): cont de decontare (IBAN), banca,
        #     BIC (MFO), cod fiscal, adresa — ca în formularul UNA.md.
        # EN: payer requisites from the native partner card (TMS_ORG).
        org_rows = _rows(db.execute_query(
            "SELECT CODFISCAL, ACCOUNT, BANK, MFO, ADRESS, ORAS, TELEFON "
            "FROM VMS_UNIV_ORG WHERE COD = :c", {"c": h["client_cod"]}))
        org = org_rows[0] if org_rows else {}
        lines = _rows(db.execute_query(
            "SELECT l.CTSC, l.CANT, l.SUMA, l.PRET, u.DENUMIREA, u.UM, u.CODVECHI "
            "FROM VMDB_ST201D l LEFT JOIN TMS_UNIVERS u ON u.COD = l.CTSC "
            "WHERE l.NRDOC = :c ORDER BY l.RROWID", {"c": int(cod)}))
        items, total = [], 0.0
        for ln in lines:
            s = float(ln["suma"] or 0)
            total += s
            qty = float(ln["cant"] or 0)
            items.append({
                "name": ln["denumirea"] or f"#{ln['ctsc']}",
                "cod": ln["codvechi"] or ln["ctsc"],
                "qty": int(qty) if qty == int(qty) else qty,
                "um": ln["um"] or "buc.",
                "price": float(ln["pret"] or 0) or (s / qty if qty else 0),
                "sum": s,
            })
        rate = Config.BIRO26_TVA_RATE
        # RO: modul TVA ales la generarea contului (YBIRO_DOC_META):
        #     'inclus' (implicit, TVA inclusa in pret) / '0' / 'fara'.
        # EN: the VAT mode chosen at generation time.
        from models.biro26_oracle_store import Biro26Store
        tva_mode = Biro26Store.get_doc_tva_mode(cod)
        if tva_mode == "0":
            tva_val, tva_label, tva_text = 0.0, "TVA 0% (НДС):", _fmt(0)
        elif tva_mode == "fara":
            tva_val, tva_label, tva_text = 0.0, "TVA:", "fără TVA / без НДС"
        else:
            tva_val = round(total * rate / (100 + rate), 2)
            rate_txt = int(rate) if float(rate) == int(rate) else rate
            tva_label = f"TVA {rate_txt}% inclus (НДС):"
            tva_text = _fmt(tva_val)
        # RO: numai NRMANUAL. NU NRSET.
        # EN: NRMANUAL only. NEVER NRSET.
        vis_nr = str(h.get("nrmanual") or "").strip()
        if vis_nr.lower() in ("none", "null"):
            vis_nr = ""
        data = {
            # RO/EN: raw number for logic; pdfme header fields filled in _pdfme_inputs
            "cont_number": vis_nr,
            "nrmanual": vis_nr,
            "number": vis_nr,
            "date_short": h["ddate"],
            "date_ro": f"{h['dd']} {_RO_MONTHS[int(h['mm'])]} {h['yy']}",
            "firm": {
                "name": Config.BIRO26_FIRM_NAME,
                "address": Config.BIRO26_FIRM_ADDRESS,
                "fiscal_code": Config.BIRO26_FIRM_FISCAL,
                "iban": Config.BIRO26_FIRM_IBAN,
                "bank": Config.BIRO26_FIRM_BANK,
                "branch": Config.BIRO26_FIRM_BRANCH,
                "phone": Config.BIRO26_FIRM_PHONE,
                "director": Config.BIRO26_FIRM_DIRECTOR,
            },
            "client": {
                "name": h["client_name"] or f"#{h['client_cod']}",
                "cod": h["client_cod"],
                "phone": ((extra[0]["phone"] if extra else None)
                          or org.get("telefon")),
                "email": (extra[0]["email"] if extra else None),
                # RO: rechizitele plătitorului (cartela nativă / magazin)
                "fiscal_code": (org.get("codfiscal")
                                or (extra[0]["idno"] if extra else None)),
                "iban": org.get("account"),
                "bank": org.get("bank"),
                "bic": org.get("mfo"),
                "address": (org.get("adress") or org.get("oras")
                            or (extra[0]["address"] if extra else None)),
            },
            "items": items,
            "total": round(total, 2),
            # RO: TVA dupa modul ales la generare (vezi mai sus)
            # EN: VAT according to the mode chosen at generation time
            "tva": tva_val,
            "tva_mode": tva_mode,
            "tva_label": tva_label,
            "tva_text": tva_text,
            "logo": _logo_data_uri(),
        }
        return {"success": True, "data": data, "client_cod": h["client_cod"]}

    @staticmethod
    def render_doc_xlsx(cod: int,
                        allowed_client_cod: Optional[int] = None) -> Dict[str, Any]:
        """RO: echivalentul EXCEL al contului de plata (bifa «si Excel»):
        - tabelul pozitiilor este TABEL Excel adevarat (ListObject);
        - Suma per rind = FORMULA =Cant*Pret;
        - TOTAL = FORMULA =SUM(...);
        - logo-ul (reports/templates/logo.jpg) la locul lui, ca in PDF.
        EN: XLSX twin of the invoice: real Excel table, =qty*price row
        formulas, =SUM total, embedded logo."""
        d = Biro26Report.doc_data(cod)
        if not d.get("success"):
            return d
        if (allowed_client_cod is not None
                and int(d["client_cod"]) != int(allowed_client_cod)):
            return {"success": False, "error": "document belongs to another client"}
        try:
            return {"success": True,
                    "xlsx": Biro26Report._build_invoice_xlsx(d["data"])}
        except Exception as e:
            return {"success": False, "error": f"xlsx: {e}"}

    @staticmethod
    def _build_invoice_xlsx(data: Dict[str, Any]) -> bytes:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = Workbook()
        ws = wb.active
        ws.title = "Cont de plata"
        widths = {"A": 5, "B": 15, "C": 48, "D": 9, "E": 9, "F": 12, "G": 14}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # ── logo la locul lui (stinga sus, ca in PDF) ──
        logo_path = os.path.join(_TPL_DIR, _LOGO_FILE)
        if os.path.exists(logo_path):
            try:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(logo_path)
                scale = 52.0 / img.height if img.height else 1
                img.height = int(img.height * scale)
                img.width = int(img.width * scale)
                ws.add_image(img, "A1")
                ws.row_dimensions[1].height = 44
            except Exception:
                pass

        bold = Font(bold=True)
        big = Font(bold=True, size=14)
        firm, client = data["firm"], data["client"]

        ws["E1"] = f"Cont de plată Nr. {data['number']}"
        ws["E1"].font = big
        ws["E2"] = f"din {data['date_ro']} · СЧЕТ НА ОПЛАТУ"
        ws["E2"].font = Font(italic=True)

        r = 3
        def put(label, value):
            nonlocal r
            ws.cell(row=r, column=1, value=label).font = bold
            ws.cell(row=r, column=3, value=value)
            r += 1
        put("Furnizor · Поставщик:", firm.get("name"))
        put("Adresa · Адрес:", firm.get("address"))
        put("Cod fiscal:", firm.get("fiscal_code"))
        put("IBAN:", f"{firm.get('iban') or ''}  {firm.get('bank') or ''} "
                     f"{firm.get('branch') or ''}".strip())
        put("Telefon:", firm.get("phone"))
        cl = client.get("name") or ""
        if client.get("phone"):
            cl += f" · tel. {client['phone']}"
        if client.get("email"):
            cl += f" · {client['email']}"
        put("Plătitor · Плательщик:", cl)
        # RO: rechizitele plătitorului — ca in modelul «PRIMARIA JAPCA»
        if client.get("address"):
            put("Adresa plătitor:", client["address"])
        if client.get("iban"):
            put("Cont de decontare nr.:", client["iban"])
        if client.get("bank") or client.get("bic"):
            put("Banca / BIC:", f"{client.get('bank') or ''}  "
                                f"{('BIC: ' + client['bic']) if client.get('bic') else ''}".strip())
        if client.get("fiscal_code"):
            put("Cod fiscal plătitor:", client["fiscal_code"])

        # ── tabelul pozitiilor (TABEL Excel + formule) ──
        head_row = r + 1
        headers = ["Nr", "Cod", "Denumirea · Наименование", "Cant.",
                   "U.M.", "Preț", "Suma"]
        for i, htxt in enumerate(headers, 1):
            c = ws.cell(row=head_row, column=i, value=htxt)
            c.font = bold
            c.alignment = Alignment(horizontal="center")
        first = head_row + 1
        thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
        for i, it in enumerate(data["items"]):
            rr = first + i
            ws.cell(row=rr, column=1, value=i + 1)
            ws.cell(row=rr, column=2, value=str(it.get("cod") or ""))
            ws.cell(row=rr, column=3, value=it.get("name") or "")
            ws.cell(row=rr, column=4, value=it.get("qty"))
            ws.cell(row=rr, column=5, value=it.get("um") or "buc.")
            ws.cell(row=rr, column=6, value=round(float(it.get("price") or 0), 2))
            # RO: suma = FORMULA cant × pret (cerinta: nu valoare "moarta")
            ws.cell(row=rr, column=7, value=f"=D{rr}*F{rr}")
            for cidx in range(1, 8):
                cell = ws.cell(row=rr, column=cidx)
                cell.border = thin
                if cidx in (6, 7):
                    cell.number_format = "#,##0.00"
        last = first + len(data["items"]) - 1
        tbl = Table(displayName="Pozitii",
                    ref=f"A{head_row}:G{last}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True,
            showFirstColumn=False, showLastColumn=False,
            showColumnStripes=False)
        ws.add_table(tbl)

        # ── TOTAL = FORMULA SUM peste coloana Suma ──
        tr = last + 1
        ws.cell(row=tr, column=6, value="TOTAL (ИТОГО):").font = big
        tc = ws.cell(row=tr, column=7, value=f"=SUM(G{first}:G{last})")
        tc.font = big
        tc.number_format = "#,##0.00"
        tvr = tr + 1
        ws.cell(row=tvr, column=6, value=data.get("tva_label") or "TVA:").font = bold
        if data.get("tva_mode") == "fara":
            ws.cell(row=tvr, column=7, value=data.get("tva_text"))
        else:
            tv = ws.cell(row=tvr, column=7, value=round(float(data.get("tva") or 0), 2))
            tv.number_format = "#,##0.00"
        ws.cell(row=tvr + 2, column=1,
                value="Director: " + (firm.get("director") or "_____________"))
        ws.cell(row=tvr + 3, column=1,
                value="Contul este valabil 3 zile · Счёт действителен 3 дня. "
                      "Vă mulțumim pentru achitarea la timp!")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def render(kind: str, data: Dict[str, Any]) -> Dict[str, Any]:
        """POST the template + data to jsReport; returns {'pdf': bytes}."""
        if kind not in REPORT_KINDS:
            return {"success": False, "error": f"unknown report kind: {kind}"}
        try:
            resp = requests.post(
                Config.JSREPORT_URL.rstrip("/") + "/api/report",
                json={
                    "template": {
                        "content": _read(REPORT_KINDS[kind]),
                        "engine": "handlebars",
                        "recipe": "chrome-pdf",
                        "helpers": _read("helpers.js"),
                        "chrome": {"format": "A4",
                                   "marginTop": "8mm", "marginBottom": "10mm",
                                   "marginLeft": "6mm", "marginRight": "6mm"},
                    },
                    "data": data,
                },
                timeout=90)
            if resp.status_code != 200:
                return {"success": False,
                        "error": f"jsreport HTTP {resp.status_code}: {resp.text[:300]}"}
            return {"success": True, "pdf": resp.content}
        except requests.ConnectionError:
            return {"success": False,
                    "error": "report service unavailable (jsreport not running)"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ── engine selection (jsreport | pdfme) per report kind ──

    @staticmethod
    def get_engines() -> Dict[str, Any]:
        import json
        try:
            with open(os.path.join(_TPL_DIR, ENGINES_FILE), encoding="utf-8") as f:
                eng = json.load(f)
        except Exception:
            eng = {}
        return {"success": True,
                "data": {k: (eng.get(k) if eng.get(k) in ("jsreport", "pdfme")
                             else "jsreport") for k in REPORT_KINDS}}

    @staticmethod
    def set_engines(mapping: Dict[str, str]) -> Dict[str, Any]:
        import json
        cur = Biro26Report.get_engines()["data"]
        for k, v in (mapping or {}).items():
            if k in REPORT_KINDS and v in ("jsreport", "pdfme"):
                cur[k] = v
        with open(os.path.join(_TPL_DIR, ENGINES_FILE), "w", encoding="utf-8") as f:
            json.dump(cur, f, indent=2)
        return {"success": True, "data": cur}

    # ── pdfme engine path ──

    @staticmethod
    def _pdfme_inputs(kind: str, data: Dict[str, Any]) -> Dict[str, Any]:
        """Flatten the report data into pdfme text inputs; the items table
        gets the totals appended as extra rows (so long item lists can never
        overlap a fixed-position totals block).

        RO: Sablon activ pdfme_cont_plata.json — campul "invoice_nr"
            = text "CONT DE PLATA № {NRMANUAL}" (NU din NRSET).
        EN: Active template pdfme_cont_plata.json — field "invoice_nr"
            = text "CONT DE PLATA № {NRMANUAL}" (NOT from NRSET).
        """
        import json
        import re
        f, c = data["firm"], data["client"]
        client_line = c["name"] + \
            (f", tel.: {c['phone']}" if c.get("phone") else "")
        # RO: rechizitele plătitorului (ca in formularul nativ UNA.md si in
        #     modelul «PRIMARIA JAPCA»): cont de decontare, banca, BIC,
        #     cod fiscal — doar cele completate in cartela.
        # EN: payer requisites block (only the filled-in ones).
        payer_req = ""
        if c.get("address"):
            payer_req += f", {c['address']}"
        if c.get("iban"):
            payer_req += f"\nCont de decontare nr.: {c['iban']} (Расчетный счет)"
        if c.get("bank"):
            payer_req += f"\n{c['bank']}"
        if c.get("bic"):
            payer_req += f"\nBIC: {c['bic']}"
        if c.get("fiscal_code"):
            payer_req += f"\nCod fiscal: {c['fiscal_code']} (Фискальный код)"
        # RO/EN: bare number from NRMANUAL only
        _nr = str(data.get("nrmanual") or data.get("number")
                  or data.get("cont_number") or data.get("invoice_nr") or "").strip()
        # RO: numerele cu SERIE ('A-23') se pastreaza intregi; doar
        #     etichetele accidentale ('CONT DE PLATA № 23') se reduc la numar
        if not re.fullmatch(r"[A-Za-z]-\d+", _nr):
            m = re.search(r"([A-Za-z]-\d+|\d+)\s*$", _nr)
            if m and not _nr.isdigit():
                _nr = m.group(1)
        if _nr.lower() in ("none", "null"):
            _nr = ""
        if kind == "invoice":
            # RO: coloana Cod (articol) + FARA rinduri de total in tabel —
            #     totalurile stau SUB tabel (campul totals_block, pozitionat
            #     dinamic dupa inaltimea reala a tabelului).
            rows = [[str(i + 1), str(it.get("cod") or ""), it["name"],
                     str(it["qty"]), it["um"],
                     _fmt(it["price"]), _fmt(it["sum"])]
                    for i, it in enumerate(data["items"])]
            _label = f"CONT DE PLATĂ № {_nr}" if _nr else "CONT DE PLATĂ"
            # RO: subtitrarile ruse pe rind separat, SUB textul romanesc
            platitor = ("Platitor, adresa: " + client_line
                        + (f", {c['address']}" if c.get("address") else "")
                        + "\n(Плательщик и его адрес)")
            if c.get("iban"):
                platitor += (f"\nCont de decontare nr.: {c['iban']}"
                             "\n(Расчетный счет)")
            if c.get("bank") or c.get("bic"):
                platitor += "\n" + (c.get("bank") or "")
                if c.get("bic"):
                    platitor += f"  BIC: {c['bic']}"
            if c.get("fiscal_code"):
                platitor += (f"\nCod fiscal: {c['fiscal_code']}"
                             "\n(Фискальный код)")
            totals = (f"Total (Итого):  {_fmt(data['total'])}\n"
                      f"{data.get('tva_label', 'Suma TVA (НДС):')}  "
                      f"{data.get('tva_text', _fmt(data['tva']))}\n"
                      f"SPRE PLATA:  {_fmt(data['total'])}")
            return {
                "furnizor_block":
                    f"Furnizor: {f['name']}\nAdresa: {f['address']}\n"
                    f"Cont de decontare nr.: {f['iban']}\n{f['bank']}\n"
                    f"BRANCH: {f['branch']}\nCod fiscal: {f['fiscal_code']}"
                    + (f"\nTelefon: {f['phone']}" if f.get("phone") else ""),
                # primary field for pdfme_cont_plata.json
                "invoice_nr": _label,
                # aliases for older templates (pdfme_invoice.json)
                "cont_number": _label,
                "title": _label,
                "nrmanual": _nr,
                "number": _nr,
                "date_ro": data["date_ro"],
                "platitor_block": platitor,
                "items": json.dumps(rows, ensure_ascii=False),
                "totals_block": totals,
                "spre_plata": "Spre plata / Всего к оплате: " + _ro_amount(data["total"]),
                "logo": data.get("logo") or "",
            }
        # order
        rows = [[str(i + 1), it["name"], str(it["cod"]), str(it["qty"]),
                 it["um"], _fmt(it["price"]), _fmt(it["sum"])]
                for i, it in enumerate(data["items"])]
        rows += [["", "", "", "", "", "Total:", _fmt(data["total"])],
                 ["", "", "", "", "",
                  data.get("tva_label", "Incl. TVA:"),
                  data.get("tva_text", _fmt(data["tva"]))]]
        _label = (f"Comanda cumpărătorului № {_nr} din {data['date_ro']}"
                  if _nr else f"Comanda cumpărătorului din {data['date_ro']}")
        return {
            "invoice_nr": _label,
            "cont_number": _label,
            "title": _label,
            "nrmanual": _nr,
            "number": _nr,
            "hr": "",
            "executor_block": (f"Executor: {f['name']}, Cod fiscal "
                               f"{f['fiscal_code']}, {f['address']}"
                               f"\nIBAN: {f['iban']}  {f['bank']}"
                               + (f" {f['branch']}" if f.get("branch") else "")),
            "client_block": "Client: " + client_line + payer_req,
            "items": json.dumps(rows, ensure_ascii=False),
            "total_line": f"Total denumiri {len(data['items'])}, în sumă de "
                          f"{_fmt(data['total'])} lei\n{_ro_amount(data['total'])}",
            "logo": data.get("logo") or "",
        }

    @staticmethod
    def _pdfme_invoice_layout(template: Dict[str, Any],
                              data: Dict[str, Any]) -> Dict[str, Any]:
        """RO: ajusteaza sablonul contului la RULARE (fisierul din designer
        ramine neschimbat): tabelul primeste coloana Cod (7 coloane), iar
        totalurile / «Spre plata» / semnaturile se pozitioneaza DINAMIC
        imediat SUB tabel (inaltimea reala dupa numarul de rinduri) — cerinta
        «asta sub tabel» + «sa incape tot». La liste foarte lungi (tabelul
        trece de pagina) se pastreaza asezarea fixa din sablon.
        EN: runtime layout: 7-col table + totals/signatures right below it."""
        import copy
        import math
        tpl = copy.deepcopy(template)
        try:
            schemas = tpl["schemas"][0]
            items = next(s for s in schemas if s.get("name") == "items")
            items["head"] = ["№", "Cod / Артикул", "Denumirea / Предмет счета",
                            "Cant. / Кол-во", "Un. / Ед.",
                            "Pretul / Цена", "Suma / Сумма"]
            items["headWidthPercentages"] = [4, 12, 41, 8, 8, 13, 14]
            items.setdefault("columnStyles", {})["alignment"] = {
                "0": "center", "1": "center", "3": "center",
                "4": "center", "5": "right", "6": "right"}
            # RO: estimarea inaltimii: rind ~6.4mm + 4.3mm per linie extra
            #     (Denumirea are ~44 caractere pe linie la 41% din 190mm)
            y0 = float(items["position"]["y"])
            h = 7.5
            for it in data.get("items", []):
                lines = max(1, math.ceil(len(str(it.get("name") or "")) / 44))
                h += 6.4 + (lines - 1) * 4.3
            bottom = y0 + h + 2
            if bottom <= 236:                    # incape pe pagina -> dinamic
                proto = next((s for s in schemas
                              if s.get("name") == "spre_plata"), None)
                totals = copy.deepcopy(proto) if proto else {
                    "type": "text", "position": {}, "width": 80, "height": 18}
                totals.update({
                    "name": "totals_block",
                    "position": {"x": 120, "y": bottom},
                    "width": 80, "height": 18,
                    "alignment": "right", "fontSize": 9.5,
                    "fontName": "DejaVuSans", "lineHeight": 1.45,
                    "content": ""})
                schemas.append(totals)
                for nm, yy in (("spre_plata", bottom + 21),
                               ("sig1", bottom + 32), ("sig2", bottom + 32)):
                    fld = next((s for s in schemas if s.get("name") == nm), None)
                    if fld:
                        fld["position"]["y"] = yy
        except (KeyError, StopIteration, TypeError, ValueError):
            return template                       # sablon atipic -> neatins
        return tpl

    @staticmethod
    def render_pdfme(kind: str, data: Dict[str, Any],
                     template_json: Optional[str] = None) -> Dict[str, Any]:
        """Render via the pdfme engine of the sidecar (POST /pdfme/generate)."""
        import json
        if kind not in PDFME_KINDS:
            return {"success": False, "error": f"unknown report kind: {kind}"}
        try:
            template = json.loads(template_json or _read(PDFME_KINDS[kind]))
            if kind == "invoice" and not template_json:
                template = Biro26Report._pdfme_invoice_layout(template, data)
            resp = requests.post(
                Config.JSREPORT_URL.rstrip("/") + "/pdfme/generate",
                json={"template": template,
                      "inputs": [Biro26Report._pdfme_inputs(kind, data)]},
                timeout=60)
            if resp.status_code != 200:
                return {"success": False,
                        "error": f"pdfme HTTP {resp.status_code}: {resp.text[:400]}"}
            return {"success": True, "pdf": resp.content}
        except requests.ConnectionError:
            return {"success": False,
                    "error": "report service unavailable (jsreport not running)"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ── template admin (simple editor page in the backoffice) ──

    @staticmethod
    def _safe_tpl(name: str) -> Optional[str]:
        """Whitelist guard: plain file name inside reports/templates only."""
        import re
        if not re.match(r"^[\w.-]+\.(hbs|js|json)$", name or ""):
            return None
        p = os.path.join(_TPL_DIR, name)
        return p if os.path.normpath(p).startswith(_TPL_DIR) else None

    @staticmethod
    def list_templates() -> Dict[str, Any]:
        try:
            out = []
            for f in sorted(os.listdir(_TPL_DIR)):
                if f.endswith((".hbs", ".js", ".json")):
                    st = os.stat(os.path.join(_TPL_DIR, f))
                    out.append({"name": f, "size": st.st_size,
                                "mtime": int(st.st_mtime)})
            return {"success": True, "data": out}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @staticmethod
    def read_template(name: str) -> Dict[str, Any]:
        p = Biro26Report._safe_tpl(name)
        if not p or not os.path.exists(p):
            return {"success": False, "error": "template not found"}
        with open(p, encoding="utf-8") as f:
            return {"success": True, "data": {"name": name, "content": f.read()}}

    @staticmethod
    def save_template(name: str, content: str) -> Dict[str, Any]:
        """Overwrite a template; the previous version goes to <name>.bak.
        NB: edits on the server live until the next code deploy — sync the
        change back into the repo (reports/templates/) to keep it."""
        p = Biro26Report._safe_tpl(name)
        if not p or not os.path.exists(p):
            return {"success": False, "error": "template not found"}
        if not (content or "").strip():
            return {"success": False, "error": "empty content"}
        if name.endswith(".json"):
            import json
            try:
                json.loads(content)
            except Exception as e:
                return {"success": False, "error": f"invalid JSON: {e}"}
        try:
            import shutil
            shutil.copy2(p, p + ".bak")
            with open(p, "w", encoding="utf-8") as f:
                f.write(content)
            return {"success": True, "data": {"name": name, "backup": name + ".bak"}}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @staticmethod
    def preview(content: str, cod: Optional[int] = None,
                name: Optional[str] = None) -> Dict[str, Any]:
        """Render arbitrary (possibly unsaved) template content with the data
        of a real document (cod) or a built-in sample. pdfme_*.json content
        goes through the pdfme engine, everything else through jsReport."""
        if cod:
            d = Biro26Report.doc_data(cod)
            if not d.get("success"):
                return d
            data = d["data"]
        else:
            data = {
                # RO/EN: sample keys for pdfme_cont_plata (invoice_nr) + legacy
                "invoice_nr": "999", "cont_number": "999",
                "nrmanual": "999", "number": "999",
                "date_short": "21.04.2026", "date_ro": "21 aprilie 2026",
                "firm": {"name": Config.BIRO26_FIRM_NAME, "address": Config.BIRO26_FIRM_ADDRESS,
                         "fiscal_code": Config.BIRO26_FIRM_FISCAL, "iban": Config.BIRO26_FIRM_IBAN,
                         "bank": Config.BIRO26_FIRM_BANK, "branch": Config.BIRO26_FIRM_BRANCH,
                         "phone": Config.BIRO26_FIRM_PHONE, "director": Config.BIRO26_FIRM_DIRECTOR},
                "client": {"name": "Client de test S.R.L.", "cod": 0,
                           "phone": "+373 690 00 000", "email": "client@test.md"},
                "items": [
                    {"name": "HELLO! Dosar din plastic cu clapă cu arc, verde",
                     "cod": "GO-00001392", "qty": 10, "um": "șt", "price": 55.0, "sum": 550.0},
                    {"name": "Folii A4 / 80 microni Class Super Clear (pachet de 50)",
                     "cod": "GO-00001647", "qty": 1, "um": "șt", "price": 115.0, "sum": 115.0},
                ],
                "total": 665.0, "tva": 110.84,
                "tva_label": "TVA 20% inclus (НДС):", "tva_text": "110,84",
                "logo": _logo_data_uri(),
            }
        # RO: sabloanele pdfme_*.json merg prin motorul pdfme
        # EN: pdfme_*.json templates go through the pdfme engine
        if name and name.startswith("pdfme") and name.endswith(".json"):
            kind = "order" if "order" in name else "invoice"
            return Biro26Report.render_pdfme(kind, data, template_json=content)
        try:
            resp = requests.post(
                Config.JSREPORT_URL.rstrip("/") + "/api/report",
                json={"template": {"content": content, "engine": "handlebars",
                                   "recipe": "chrome-pdf",
                                   "helpers": _read("helpers.js"),
                                   "chrome": {"format": "A4",
                                              "marginTop": "8mm", "marginBottom": "10mm",
                                              "marginLeft": "6mm", "marginRight": "6mm"}},
                      "data": data},
                timeout=90)
            if resp.status_code != 200:
                return {"success": False,
                        "error": f"jsreport HTTP {resp.status_code}: {resp.text[:400]}"}
            return {"success": True, "pdf": resp.content}
        except requests.ConnectionError:
            return {"success": False,
                    "error": "report service unavailable (jsreport not running)"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ── native attachments (VMDB_DOCS_OLE / TMDB_DOCS_OLE) ──

    @staticmethod
    def attach_pdf(cod: int, kind: str, pdf: bytes) -> Dict[str, Any]:
        """RO: ataseaza PDF-ul generat la document in VMDB_DOCS_OLE (ecranul
        nativ de atasamente OfficePlus). Un atasament per (document, tip de
        formular) — regenerarea inlocuieste fisierul, nu il dubleaza.
        EN: attach the generated PDF to the document via VMDB_DOCS_OLE (the
        native OfficePlus attachments screen). One attachment per (document,
        form kind) — re-rendering replaces the file instead of duplicating."""
        import base64
        names = {"invoice": "Cont_de_plata", "order": "Comanda"}
        fname = f"{names.get(kind, kind)}_{cod}.pdf"
        comment = ("Cont de plata (web/PDF)" if kind == "invoice"
                   else "Comanda cumparatorului (web/PDF)")
        blob = {"__b64__": base64.b64encode(pdf).decode()}
        try:
            r = Biro26DB().execute_script([
                {"sql": "DELETE FROM VMDB_DOCS_OLE WHERE NRDOC = :c AND PFILE = :f",
                 "params": {"c": int(cod), "f": fname}, "kind": "dml"},
                {"sql": "INSERT INTO VMDB_DOCS_OLE "
                        "(NRDOC, NRDOC1, TXTCOMMENT, PFILE, OLEOBJ) "
                        "SELECT :c, NVL(MAX(NRDOC1), 0) + 1, :cm, :f, :b "
                        "FROM VMDB_DOCS_OLE WHERE NRDOC = :c2",
                 "params": {"c": int(cod), "cm": comment, "f": fname,
                            "b": blob, "c2": int(cod)}, "kind": "dml"},
            ])
            if not r.get("success"):
                return {"success": False, "error": r.get("message")}
            return {"success": True, "data": {"pfile": fname}}
        except Exception as e:
            return {"success": False, "error": str(e)}

    @staticmethod
    def render_doc(kind: str, cod: int,
                   allowed_client_cod: Optional[int] = None) -> Dict[str, Any]:
        """Data + render in one call; when allowed_client_cod is given the
        document must belong to that client (public shop session guard)."""
        d = Biro26Report.doc_data(cod)
        if not d.get("success"):
            return d
        if allowed_client_cod is not None and int(d["client_cod"]) != int(allowed_client_cod):
            return {"success": False, "error": "document belongs to another client"}
        # RO: motorul activ per formular (engines.json, editabil in admin)
        # EN: active engine per form kind (engines.json, editable in the admin)
        engine = Biro26Report.get_engines()["data"].get(kind, "jsreport")
        res = (Biro26Report.render_pdfme(kind, d["data"]) if engine == "pdfme"
               else Biro26Report.render(kind, d["data"]))
        if res.get("success"):
            # RO: PDF-ul se ataseaza la document (VMDB_DOCS_OLE) in FUNDAL
            #     (thread daemon, best effort). Atasarea sincrona bloca
            #     raspunsul pina la 300s cind un utilizator al aplicatiei
            #     native tinea rindul OLE blocat (enq: TX row lock) —
            #     descarcarea PDF nu trebuie sa astepte arhivarea.
            # EN: the PDF is attached to the document IN THE BACKGROUND
            #     (daemon thread, best effort). The synchronous attach used
            #     to block the response for up to 300s whenever a native-app
            #     user held the OLE row locked — the download must never
            #     wait for archival.
            import threading
            pdf = res["pdf"]

            def _attach_bg():
                try:
                    Biro26Report.attach_pdf(cod, kind, pdf)
                except Exception:
                    pass

            threading.Thread(target=_attach_bg, daemon=True).start()
            res["attached"] = "background"
        return res

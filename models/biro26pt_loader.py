#!/usr/bin/env python3
"""BIRO26PT loader — xlsx/csv into the raw staging BIRO26PT_RAW/HEADER/FILE.

RO: Incarcator generic: nu interpreteaza structura — doar celule + antet;
    detectia coloanelor se face in DB (pachetul BIRO26PT_importData).
    Ruleaza ca SUBPROCES (thick Oracle init doar aici, ca in
    biro26_worker.py); credentialele si Instant Client vin din config/.env.
EN: Generic loader: no structure interpretation — cells + header only; the
    column detection lives in the DB (package BIRO26PT_importData). Runs as
    a SUBPROCESS (thick Oracle init only here, same rule as
    biro26_worker.py); credentials and the Instant Client path come from
    config/.env. Charset note: the DB is CL8MSWIN1251 — python-oracledb
    converts Unicode correctly; never push Cyrillic via raw SQLcl.

Usage: biro26pt_loader.py <folder_or_file>
stdout: one line per loaded sheet:
  load_id=<n> file='<name>' sheet='<s>' rows=<r> cols=<c>
"""
from __future__ import annotations

import glob
import os
import sys

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import unicodedata  # noqa: E402

import oracledb  # noqa: E402
from config import Config  # noqa: E402

MAXCOL = 16  # c0..c15

# RO: Baza e CL8MSWIN1251 (chirilic) si NU are diacritice romanesti (ă â î ș ț) sau
#     semne tipografice (× — ‑ ² ½ …): Oracle le stocheaza ca '?'. Transliteram tot ce
#     nu incape in cp1251 INAINTE de scriere. Chirilica ramane neatinsa.
# EN: The DB is CL8MSWIN1251 (Cyrillic) and has NO Romanian diacritics (ă â î ș ț) or
#     typographic signs (× — ‑ ² ½ …): Oracle stores them as '?'. We transliterate
#     everything that does not fit cp1251 BEFORE writing. Cyrillic is untouched.
_TRANSLIT = {
    "ă": "a", "Ă": "A", "â": "a", "Â": "A", "î": "i", "Î": "I",
    "ș": "s", "Ș": "S", "ş": "s", "Ş": "S",
    "ț": "t", "Ț": "T", "ţ": "t", "Ţ": "T",
    "×": "x", "÷": ":", "−": "-", "‐": "-", "‑": "-", "‒": "-", "―": "-",
    "≤": "<=", "≥": ">=", "≈": "~", "≠": "!=", "′": "'", "″": '"', "ʼ": "'",
    "½": "1/2", "¼": "1/4", "¾": "3/4", "⁄": "/", "·": ".", "∙": ".",
    "ﬁ": "fi", "ﬂ": "fl", "œ": "oe", "Œ": "OE", "æ": "ae", "Æ": "AE",
    "ß": "ss", "ø": "o", "Ø": "O", "đ": "d", "Đ": "D", "ł": "l", "Ł": "L",
    "​": "", "‌": "", "‍": "", "﻿": "", "­": "",
    " ": " ", " ": " ", " ": " ", " ": " ",
}


def cp1251_safe(s: str) -> str:
    """RO: text care incape garantat in CL8MSWIN1251 / EN: guaranteed cp1251-safe text."""
    out = []
    for ch in s:
        repl = _TRANSLIT.get(ch)
        if repl is not None:
            out.append(repl)
            continue
        try:
            ch.encode("cp1251")
            out.append(ch)
            continue
        except UnicodeEncodeError:
            pass
        # RO: descompunere Unicode: litera de baza fara semne / EN: strip combining marks
        dec = unicodedata.normalize("NFKD", ch)
        dec = "".join(c for c in dec if not unicodedata.combining(c))
        dec = "".join(_TRANSLIT.get(c, c) for c in dec)
        try:
            dec.encode("cp1251")
            out.append(dec)
        except UnicodeEncodeError:
            out.append("")  # RO: nerecuperabil (emoji etc.) / EN: unrecoverable
    return "".join(out)


def cells(row):
    """RO: intoarce (valori transliterate, originale). Originalul se pastreaza DOAR
    unde transliterarea a schimbat ceva (diacritice / semne non-cp1251).
    EN: returns (transliterated values, originals); the original is kept ONLY where
    transliteration actually changed something."""
    out, orig = [], {}
    for i, v in enumerate(row):
        if v is None:
            out.append(None)
            continue
        raw = str(v).strip()
        s = cp1251_safe(raw)
        out.append(s[:1000] if s != "" else None)
        if raw and s != raw:
            orig[i] = raw[:4000]
    return out, orig


def read_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        rows = [cells(r) for r in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(x is not None for x in r[0])]
        if rows:
            # RO: numele foii devine GRUPA -> trebuie si el cp1251-safe
            # EN: the sheet name becomes GRUPA -> must be cp1251-safe too
            yield cp1251_safe(sh), rows


def read_csv(path):
    import csv
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096); f.seek(0)
        delim = ";" if sample.count(";") >= sample.count(",") else ","
        rows = [cells(r) for r in csv.reader(f, delimiter=delim)]
    rows = [r for r in rows if any(x is not None for x in r[0])]
    if rows:
        yield cp1251_safe(os.path.basename(path)), rows


def main():
    if len(sys.argv) < 2:
        print("usage: biro26pt_loader.py <folder_or_file>")
        sys.exit(1)
    target = sys.argv[1]
    files = []
    if os.path.isdir(target):
        for ext in ("*.xlsx", "*.xls", "*.csv"):
            files += glob.glob(os.path.join(target, ext))
    else:
        files = [target]
    files = sorted(f for f in files if not os.path.basename(f).startswith("~"))
    if not files:
        print("no files found")
        sys.exit(1)

    oracledb.init_oracle_client(lib_dir=Config.BIRO26_INSTANT_CLIENT)
    con = oracledb.connect(user=Config.BIRO26_DB_USER,
                           password=Config.BIRO26_DB_PASSWORD,
                           dsn=Config.BIRO26_DB_DSN)
    cur = con.cursor()
    cur.execute("SELECT NVL(MAX(load_id),0) FROM biro26pt_file")
    load_id = cur.fetchone()[0]

    for path in files:
        base = cp1251_safe(os.path.basename(path))[:200]
        reader = read_csv if path.lower().endswith(".csv") else read_xlsx
        for sheet, rows in reader(path):
            load_id += 1
            header = rows[0][0]
            n_cols = min(max(len(r[0]) for r in rows), MAXCOL)
            cur.executemany(
                "INSERT INTO biro26pt_header(load_id,src_file,col_idx,header_text) "
                "VALUES(:1,:2,:3,:4)",
                [(load_id, base, i, (header[i] if i < len(header) else None))
                 for i in range(n_cols)])
            buf, blobs = [], []
            for rno, (rvals, rorig) in enumerate(rows[1:], start=1):
                vals = [load_id, base, sheet[:120], rno] + \
                       [(rvals[i] if i < len(rvals) else None) for i in range(MAXCOL)]
                buf.append(tuple(vals))
                # RO: originalul UTF-8 doar unde transliterarea a schimbat ceva
                # EN: UTF-8 original only where transliteration changed something
                for ci, ov in rorig.items():
                    if ci < MAXCOL:
                        blobs.append((load_id, rno, ci, ov.encode("utf-8")))
            ph = ",".join(":%d" % i for i in range(1, 4 + MAXCOL + 1))
            cur.executemany(
                "INSERT INTO biro26pt_raw(load_id,src_file,sheet,row_no," +
                ",".join("c%d" % i for i in range(MAXCOL)) + ") VALUES(" + ph + ")", buf)
            if blobs:
                cur.executemany(
                    "INSERT INTO biro26pt_raw_blob(load_id,row_no,col_idx,val_blob) "
                    "VALUES(:1,:2,:3,:4)", blobs)
            cur.execute(
                "INSERT INTO biro26pt_file(load_id,src_file,sheet,n_rows,n_cols) "
                "VALUES(:1,:2,:3,:4,:5)",
                (load_id, base, sheet[:120], len(rows) - 1, n_cols))
            con.commit()
            print(f"load_id={load_id} file='{base}' sheet='{sheet}' "
                  f"rows={len(rows)-1} cols={n_cols}", flush=True)
    con.close()


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# RO: Reparare completa a textelor stricate ('?') din BD, folosind TOATE fisierele sursa.
#     Doua metode: (1) dictionar mangled->corect (cp1251 roundtrip);
#                  (2) potrivire prin masca: '?' = orice caracter (sursa e deja ASCII).
# EN: Full repair of '?'-mangled DB text using ALL source files.
#     (1) mangled->fixed dictionary; (2) mask matching where '?' matches any char.
import glob, os, sys, csv, json, unicodedata, re, oracledb
from collections import defaultdict

GO = '--go' in sys.argv
SCRATCH = '/private/tmp/claude-501/-Users-pt-Projects-AI-BIRO26/fc666e55-371e-4d3a-986d-ddffe457696f/scratchpad'

PRE = {'×':'x','÷':':','−':'-','‐':'-','‑':'-','‒':'-','―':'-','≤':'<=','≥':'>=','≈':'~','≠':'!=',
       '∅':'0','∙':'.','·':'.','½':'1/2','¼':'1/4','¾':'3/4','⁄':'/','′':"'",'″':'"','ʼ':"'",
       'ˮ':'"','‟':'"','​':'','‌':'','‍':'','﻿':'','­':'',' ':' ',' ':' ',' ':' ',' ':' ',
       'ﬁ':'fi','ﬂ':'fl','œ':'oe','Œ':'OE','æ':'ae','Æ':'AE','ß':'ss','ø':'o','Ø':'O',
       'đ':'d','Đ':'D','ł':'l','Ł':'L','ș':'s','Ș':'S','ş':'s','Ş':'S','ț':'t','Ț':'T',
       'ţ':'t','Ţ':'T','ă':'a','Ă':'A','â':'a','Â':'A','î':'i','Î':'I'}

def mangle(s): return s.encode('cp1251', errors='replace').decode('cp1251')

def fix(s):
    out = []
    for ch in s:
        if ch in PRE: out.append(PRE[ch]); continue
        try: ch.encode('cp1251'); out.append(ch); continue
        except UnicodeEncodeError: pass
        d = unicodedata.normalize('NFKD', ch)
        d = ''.join(c for c in d if not unicodedata.combining(c))
        d = ''.join(PRE.get(c, c) for c in d)
        try: d.encode('cp1251'); out.append(d)
        except UnicodeEncodeError: out.append('')
    return mangle(''.join(out))

# ---------- 1) colecteaza toate textele sursa / collect all source strings ----------
DICT = {}                     # mangled -> corect (pentru texte cu diacritice in sursa)
CLEAN = set()                 # texte sursa deja cp1251-safe (candidati pt. masca)

def add(s):
    if not isinstance(s, str): return
    s = s.strip()
    if len(s) < 5 or not any(c.isalpha() for c in s): return
    m = mangle(s)
    if '?' in m:
        c = fix(s)
        if '?' not in c and c != m: DICT[m] = c
    else:
        CLEAN.add(s)

def scan_xlsx(p):
    import openpyxl
    try: wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    except Exception: return
    for sh in wb.sheetnames:
        add(sh)
        try:
            for row in wb[sh].iter_rows(values_only=True):
                for v in row: add(v)
        except Exception: pass

def scan_xls(p):
    try:
        import xlrd; b = xlrd.open_workbook(p, on_demand=True)
    except Exception: return
    for n in b.sheet_names():
        add(n)
        try:
            sh = b.sheet_by_name(n)
            for r in range(sh.nrows):
                for v in sh.row_values(r): add(v)
        except Exception: pass

def scan_csv(p):
    for enc in ('utf-8-sig','utf-8','cp1251','latin-1'):
        try:
            with open(p, newline='', encoding=enc) as f:
                smp = f.read(8192); f.seek(0)
                d = ';' if smp.count(';') >= smp.count(',') else ','
                for row in csv.reader(f, delimiter=d):
                    for v in row: add(v)
            return
        except (UnicodeDecodeError, csv.Error): continue
        except Exception: return

roots = [a for a in sys.argv[1:] if not a.startswith('--')] or ['/Users/pt/Projects.AI/BIRO26']
files = []
for r in roots:
    if os.path.isfile(r): files.append(r)
    else:
        for e in ('xlsx','xls','csv'):
            files += glob.glob(os.path.join(r, '**', f'*.{e}'), recursive=True)
files = [f for f in sorted(set(files)) if not os.path.basename(f).startswith('~$')]
print(f"scanning {len(files)} source files...", file=sys.stderr)
for f in files:
    low = f.lower()
    (scan_xlsx if low.endswith('.xlsx') else scan_xls if low.endswith('.xls') else scan_csv)(f)
print(f"dict={len(DICT)}  clean-source-strings={len(CLEAN)}", file=sys.stderr)

# ---------- 2) index pentru potrivire prin masca / index for mask matching ----------
by_pref = defaultdict(list)          # primele 4 caractere -> liste de texte
for s in CLEAN:
    by_pref[s[:4]].append(s)

def mask_match(v):
    """RO: '?' = orice caracter; sursa poate fi mai lunga (BD trunchiaza)."""
    n = len(v)
    pref = v[:4]
    if '?' in pref:                                  # prefix incert -> cauta dupa prefixul curat
        cands = []
        head = v.split('?')[0]
        if len(head) >= 2:
            for k, lst in by_pref.items():
                if k.startswith(head) or head.startswith(k): cands.extend(lst)
        else:
            return None
    else:
        cands = by_pref.get(pref, [])
    if not cands: return None
    rx = re.compile('^' + ''.join('.' if c == '?' else re.escape(c) for c in v))
    hits = {s[:n] for s in cands if len(s) >= n and rx.match(s)}
    if len(hits) == 1:
        h = hits.pop()
        return h if '?' not in h else None
    return None                                       # ambiguu -> nu atingem

def repair(v):
    if not v or '?' not in v: return None
    s = v.strip()
    if s in DICT: return DICT[s]
    # prefix in dictionar (BD trunchiata)
    for k, c in ():                                   # (placeholder, dictionarul e exact)
        pass
    r = mask_match(s)
    return r

# ---------- 3) aplica pe BD / apply to DB ----------
oracledb.init_oracle_client(lib_dir='/Users/pt/Downloads/instantclient_23_26')
con = oracledb.connect(user='officeplus', password='officeplus26', dsn='orange.una.md:4024/cloudbd.world')
cur = con.cursor()
TARGETS = [('TMS_UNIVERS','DENUMIREA'), ('BIRO26_GOODS','DENUMIRE'), ('BIRO26_GOODS','GRUPA'),
           ('BIRO26_GOODS','CATEGORIE'), ('TMS_SYSGRPH','COMENT'), ('TMS_MPT_BARCODE','COMENT'),
           ('BIRO26PT_STG','DENUMIRE'), ('BIRO26PT_STG','GRUPA'), ('BIRO26PT_STG','CATEG')]
LOG=[]
print(f"\n{'TABLE.COLUMN':30} {'distinct?':>10} {'reparabile':>11} {'rinduri':>9}")
print('-'*64)
tot = 0
for tbl, col in TARGETS:
    cur.execute(f"SELECT {col}, COUNT(*) FROM {tbl} WHERE {col} LIKE '%?%' GROUP BY {col}")
    rows = cur.fetchall()
    mp, nr = [], 0
    for val, cnt in rows:
        r = repair(val)
        if r and r != val:
            mp.append((r, val)); nr += cnt
    print(f"{tbl+'.'+col:30} {len(rows):>10} {len(mp):>11} {nr:>9}")
    tot += nr
    LOG.extend([{'t':tbl,'c':col,'old':o,'new':n} for n,o in mp])
    if GO and mp:
        cur.executemany(f"UPDATE {tbl} SET {col} = :1 WHERE {col} = :2", mp)
        con.commit()
print('-'*64)
print(f"{'TOTAL':30} {'':>10} {'':>11} {tot:>9}")
if GO:
    print("\n=== dupa aplicare / after apply ===")
    for tbl, col in TARGETS:
        cur.execute(f"SELECT COUNT(*) FROM {tbl} WHERE {col} LIKE '%?%'")
        print(f"  {tbl}.{col}: {cur.fetchone()[0]} ramase")
else:
    print("\nDRY-RUN — ruleaza cu --go pentru aplicare.")
json.dump(LOG, open(SCRATCH+'/replacements.json','w'), ensure_ascii=False)
print(f'log: {len(LOG)} replacements saved')
con.close()

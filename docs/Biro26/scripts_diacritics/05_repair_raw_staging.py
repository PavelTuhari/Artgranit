#!/usr/bin/env python3
# RO: Curata stagin-ul brut BIRO26PT_RAW (c0..c15) - altfel un re-import al unui load
#     vechi ar readuce '?'. Foloseste corpusul de cuvinte curate (BD + fisiere),
#     doar '?' INCONJURAT de litere/cifre (semnele de intrebare reale se pastreaza).
# EN: Clean the raw staging BIRO26PT_RAW so a re-import of an old load cannot bring '?' back.
import sys, re, json, glob, os, csv, unicodedata, oracledb
from collections import defaultdict, Counter

GO = '--go' in sys.argv
MIN_RATIO = 4
WORD_RX = re.compile(r'[0-9A-Za-zА-Яа-яЁё?]+')
PRE = {'ș':'s','Ș':'S','ş':'s','Ş':'S','ț':'t','Ț':'T','ţ':'t','Ţ':'T','ă':'a','Ă':'A',
       'â':'a','Â':'A','î':'i','Î':'I','×':'x','−':'-','‑':'-','–':'-','—':'-'}

def translit(s):
    out=[]
    for ch in s:
        if ch in PRE: out.append(PRE[ch]); continue
        try: ch.encode('cp1251'); out.append(ch)
        except UnicodeEncodeError:
            d=unicodedata.normalize('NFKD',ch)
            d=''.join(c for c in d if not unicodedata.combining(c))
            try: d.encode('cp1251'); out.append(d)
            except UnicodeEncodeError: out.append('')
    return ''.join(out)

WORDS = Counter()
def feed(t):
    if not isinstance(t,str): return
    for w in WORD_RX.findall(translit(t)):
        if '?' not in w and len(w)>=3 and any(c.isalpha() for c in w): WORDS[w]+=1

def scan(p):
    low=p.lower()
    try:
        if low.endswith('.xlsx'):
            import openpyxl
            wb=openpyxl.load_workbook(p, read_only=True, data_only=True)
            for sh in wb.sheetnames:
                feed(sh)
                for row in wb[sh].iter_rows(values_only=True):
                    for v in row: feed(v)
        elif low.endswith('.xls'):
            import xlrd
            b=xlrd.open_workbook(p, on_demand=True)
            for n in b.sheet_names():
                feed(n); sh=b.sheet_by_name(n)
                for r in range(sh.nrows):
                    for v in sh.row_values(r): feed(v)
        else:
            for enc in ('utf-8-sig','utf-8','cp1251','latin-1'):
                try:
                    with open(p,newline='',encoding=enc) as f:
                        smp=f.read(8192); f.seek(0)
                        d=';' if smp.count(';')>=smp.count(',') else ','
                        for row in csv.reader(f,delimiter=d):
                            for v in row: feed(v)
                    break
                except (UnicodeDecodeError, csv.Error): continue
    except Exception: pass

for e in ('xlsx','xls','csv'):
    for f in glob.glob(f'/Users/pt/Projects.AI/BIRO26/**/*.{e}', recursive=True):
        if not os.path.basename(f).startswith('~$'): scan(f)

oracledb.init_oracle_client(lib_dir='/Users/pt/Downloads/instantclient_23_26')
con=oracledb.connect(user='officeplus',password='officeplus26',dsn='orange.una.md:4024/cloudbd.world')
cur=con.cursor(); cur.arraysize=5000
for sql in ["SELECT denumirea FROM tms_univers WHERE denumirea NOT LIKE '%?%'",
            "SELECT denumire FROM biro26_goods WHERE denumire NOT LIKE '%?%'",
            "SELECT categorie FROM biro26_goods WHERE categorie NOT LIKE '%?%'"]:
    cur.execute(sql)
    for (v,) in cur: feed(v)
print(f"corpus: {len(WORDS)} words", file=sys.stderr)

by_len=defaultdict(list)
for w in WORDS: by_len[len(w)].append(w)
CACHE={}

REAL_Q_MIN = 3   # RO: daca restul cuvintului e frecvent -> semn de intrebare real

def fix_word(w):
    if '?' not in w: return None
    if w in CACHE: return CACHE[w]
    # protectie: '?' final si restul e cuvint real frecvent -> nu atingem
    if w.endswith('?') and WORDS.get(w[:-1], 0) >= REAL_Q_MIN:
        CACHE[w]=None; return None
    rx=re.compile('^'+''.join('.' if c=='?' else re.escape(c) for c in w)+'$')
    c=[(WORDS[x],x) for x in by_len.get(len(w),()) if rx.match(x)]
    r=None
    if c:
        c.sort(reverse=True)
        if len(c)==1 or c[0][0] >= MIN_RATIO*c[1][0]: r=c[0][1]
    CACHE[w]=r; return r

def fix_text(s):
    if not s or '?' not in s: return None
    # RO: nu atingem URL-urile ('?' = separator query) / EN: never touch URLs
    if s.lstrip()[:8].lower().startswith(('http://','https:/','www.')): return None
    out=[]; pos=0; ch=False
    for m in WORD_RX.finditer(s):
        w=m.group()
        if '?' not in w: continue
        r=fix_word(w)
        if r: out.append(s[pos:m.start()]); out.append(r); pos=m.end(); ch=True
    if not ch: return None
    out.append(s[pos:]); r=''.join(out)
    return r if r!=s else None

total=0
print(f"\n{'COLUMN':8} {'distinct?':>10} {'reparate':>9} {'rinduri':>9}")
print('-'*40)
for i in range(16):
    col=f'c{i}'
    cur.execute(f"SELECT {col}, COUNT(*) FROM biro26pt_raw WHERE {col} LIKE '%?%' GROUP BY {col}")
    rows=cur.fetchall()
    if not rows: continue
    mp=[]; nr=0
    for val,cnt in rows:
        r=fix_text(val)
        if r and r!=val: mp.append((r,val)); nr+=cnt
    print(f"{col:8} {len(rows):>10} {len(mp):>9} {nr:>9}")
    total+=nr
    if GO and mp:
        cur.executemany(f"UPDATE biro26pt_raw SET {col}=:1 WHERE {col}=:2", mp)
        con.commit()
print('-'*40); print(f"TOTAL rinduri: {total}")
if GO:
    cur.execute("SELECT COUNT(*) FROM biro26pt_raw WHERE " + " OR ".join(f"c{i} LIKE '%?%'" for i in range(16)))
    print(f"RAW rows still containing '?': {cur.fetchone()[0]}")
else:
    print("DRY-RUN — --go pentru aplicare.")
con.close()

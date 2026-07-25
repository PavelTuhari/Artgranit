#!/usr/bin/env python3
# RO: Valul 4 - '?' la INCEPUT/SFIRSIT de cuvint (Șnur->'?nur', Coș->'Co?').
#     Protectie contra semnelor de intrebare REALE: daca restul cuvintului exista
#     in corpus ca cuvint de sine statator si e frecvent -> e semn de intrebare, nu-l atingem.
# EN: Wave 4 - '?' at word START/END. Guard against REAL question marks: if the word
#     without '?' exists in the corpus as a frequent standalone word -> leave it alone.
import sys, re, json, glob, os, csv, unicodedata, oracledb
from collections import defaultdict, Counter

GO = '--go' in sys.argv
MIN_RATIO = 4          # candidatul top trebuie sa domine
REAL_Q_MIN = 3         # daca restul cuvintului apare de >= N ori -> semn de intrebare real
WORD_RX = re.compile(r'[0-9A-Za-zА-Яа-яЁё?]+')
PRE = {'ș':'s','Ș':'S','ş':'s','Ş':'S','ț':'t','Ț':'T','ţ':'t','Ţ':'T','ă':'a','Ă':'A',
       'â':'a','Â':'A','î':'i','Î':'I','×':'x','−':'-','‑':'-','–':'-','—':'-'}

def translit(s):
    out=[]
    for ch in s:
        if ch in PRE: out.append(PRE[ch]); continue
        try: ch.encode('cp1251'); out.append(ch)
        except UnicodeEncodeError:
            d=unicodedata.normalize('NFKD',ch)
            d=''.join(c for c in d if not unicodedata.combining(c))
            try: d.encode('cp1251'); out.append(d)
            except UnicodeEncodeError: out.append('')
    return ''.join(out)

WORDS=Counter()
def feed(t):
    if not isinstance(t,str): return
    for w in WORD_RX.findall(translit(t)):
        if '?' not in w and len(w)>=2 and any(c.isalpha() for c in w): WORDS[w]+=1

def scan(p):
    low=p.lower()
    try:
        if low.endswith('.xlsx'):
            import openpyxl
            wb=openpyxl.load_workbook(p, read_only=True, data_only=True)
            for sh in wb.sheetnames:
                feed(sh)
                for row in wb[sh].iter_rows(values_only=True):
                    for v in row: feed(v)
        elif low.endswith('.xls'):
            import xlrd
            b=xlrd.open_workbook(p, on_demand=True)
            for n in b.sheet_names():
                feed(n); sh=b.sheet_by_name(n)
                for r in range(sh.nrows):
                    for v in sh.row_values(r): feed(v)
        else:
            for enc in ('utf-8-sig','utf-8','cp1251','latin-1'):
                try:
                    with open(p,newline='',encoding=enc) as f:
                        smp=f.read(8192); f.seek(0)
                        d=';' if smp.count(';')>=smp.count(',') else ','
                        for row in csv.reader(f,delimiter=d):
                            for v in row: feed(v)
                    break
                except (UnicodeDecodeError,csv.Error): continue
    except Exception: pass

for e in ('xlsx','xls','csv'):
    for f in glob.glob(f'/Users/pt/Projects.AI/BIRO26/**/*.{e}', recursive=True):
        if not os.path.basename(f).startswith('~$'): scan(f)
oracledb.init_oracle_client(lib_dir='/Users/pt/Downloads/instantclient_23_26')
con=oracledb.connect(user='officeplus',password='officeplus26',dsn='orange.una.md:4024/cloudbd.world')
cur=con.cursor(); cur.arraysize=5000
for sql in ["SELECT denumirea FROM tms_univers WHERE denumirea NOT LIKE '%?%'",
            "SELECT denumire FROM biro26_goods WHERE denumire NOT LIKE '%?%'",
            "SELECT categorie FROM biro26_goods WHERE categorie NOT LIKE '%?%'"]:
    cur.execute(sql)
    for (v,) in cur: feed(v)
print(f"corpus: {len(WORDS)} words", file=sys.stderr)
by_len=defaultdict(list)
for w in WORDS: by_len[len(w)].append(w)
CACHE={}

def fix_word(w):
    """RO: repara '?' oriunde in cuvint, cu protectie pentru semnul de intrebare real."""
    if '?' not in w: return None
    if w in CACHE: return CACHE[w]
    res=None
    # protectie: '?' final si restul e un cuvint real frecvent -> semn de intrebare
    if w.endswith('?') and WORDS.get(w[:-1], 0) >= REAL_Q_MIN:
        CACHE[w]=None; return None
    rx=re.compile('^'+''.join('.' if c=='?' else re.escape(c) for c in w)+'$')
    c=[(WORDS[x],x) for x in by_len.get(len(w),()) if rx.match(x)]
    if c:
        c.sort(reverse=True)
        if len(c)==1 or c[0][0] >= MIN_RATIO*c[1][0]:
            res=c[0][1]
    CACHE[w]=res; return res

def fix_text(s):
    if not s or '?' not in s: return None
    out=[]; pos=0; ch=False
    for m in WORD_RX.finditer(s):
        w=m.group()
        if '?' not in w: continue
        r=fix_word(w)
        if r: out.append(s[pos:m.start()]); out.append(r); pos=m.end(); ch=True
    if not ch: return None
    out.append(s[pos:]); r=''.join(out)
    return r if r!=s else None

TARGETS=[('TMS_UNIVERS','DENUMIREA'),('BIRO26_GOODS','DENUMIRE'),('BIRO26_GOODS','CATEGORIE'),
         ('BIRO26_GOODS','GRUPA'),('TMS_SYSGRPH','COMENT'),('TMS_MPT_BARCODE','COMENT'),
         ('BIRO26PT_STG','DENUMIRE'),('BIRO26PT_STG','CATEG')]
LOG=[]
print(f"\n{'TABLE.COLUMN':30} {'distinct?':>10} {'reparate':>9} {'rinduri':>9}")
print('-'*62)
for tbl,col in TARGETS:
    cur.execute(f"SELECT {col}, COUNT(*) FROM {tbl} WHERE {col} LIKE '%?%' GROUP BY {col}")
    rows=cur.fetchall(); mp=[]; nr=0
    for val,cnt in rows:
        r=fix_text(val)
        if r and r!=val: mp.append((r,val)); nr+=cnt
    print(f"{tbl+'.'+col:30} {len(rows):>10} {len(mp):>9} {nr:>9}")
    LOG.extend([{'t':tbl,'c':col,'old':o,'new':n} for n,o in mp])
    if GO and mp:
        cur.executemany(f"UPDATE {tbl} SET {col}=:1 WHERE {col}=:2", mp)
        con.commit()
print('-'*62)
if GO:
    print("\n=== ramase ===")
    for tbl,col in TARGETS:
        cur.execute(f"SELECT COUNT(*) FROM {tbl} WHERE {col} LIKE '%?%'")
        print(f"  {tbl}.{col}: {cur.fetchone()[0]}")
    json.dump(LOG, open('/private/tmp/claude-501/-Users-pt-Projects-AI-BIRO26/fc666e55-371e-4d3a-986d-ddffe457696f/scratchpad/replacements_w4.json','w'), ensure_ascii=False)
else:
    print("\nDRY-RUN. Sample:")
    for e in LOG[:14]: print(f"  {e['old'][:58]}\n  -> {e['new'][:58]}")
con.close()

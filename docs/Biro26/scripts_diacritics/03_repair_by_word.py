#!/usr/bin/env python3
# RO: Valul 3 - repara '?' la nivel de CUVINT (nu de propozitie), folosind corpusul de
#     cuvinte curate din BD + fisiere. Un cuvint stricat ("patra?ele") se potriveste prin
#     masca cu un cuvint cunoscut ("patratele"). Se aplica doar cind rezultatul e clar.
# EN: Wave 3 - WORD-level '?' repair using the clean word corpus from DB + files.
import sys, re, json, glob, os, csv, unicodedata, oracledb
from collections import defaultdict, Counter

GO = '--go' in sys.argv
MIN_RATIO = 4        # RO: candidatul top trebuie sa fie de N ori mai frecvent / top must dominate
SCRATCH = '/private/tmp/claude-501/-Users-pt-Projects-AI-BIRO26/fc666e55-371e-4d3a-986d-ddffe457696f/scratchpad'
WORD_RX = re.compile(r'[0-9A-Za-zА-Яа-яЁё?]+')

PRE = {'ș':'s','Ș':'S','ş':'s','Ş':'S','ț':'t','Ț':'T','ţ':'t','Ţ':'T','ă':'a','Ă':'A',
       'â':'a','Â':'A','î':'i','Î':'I','×':'x','−':'-','‑':'-','–':'-','—':'-'}
def translit(s):
    out=[]
    for ch in s:
        if ch in PRE: out.append(PRE[ch]); continue
        try: ch.encode('cp1251'); out.append(ch)
        except UnicodeEncodeError:
            d=unicodedata.normalize('NFKD',ch)
            d=''.join(c for c in d if not unicodedata.combining(c))
            try: d.encode('cp1251'); out.append(d)
            except UnicodeEncodeError: out.append('')
    return ''.join(out)

WORDS = Counter()
def feed(text):
    if not isinstance(text,str): return
    t = translit(text)
    for w in WORD_RX.findall(t):
        if '?' not in w and len(w) >= 3 and any(c.isalpha() for c in w):
            WORDS[w] += 1

# ---- corpus din fisiere / from source files ----
def scan_file(p):
    low=p.lower()
    try:
        if low.endswith('.xlsx'):
            import openpyxl
            wb=openpyxl.load_workbook(p, read_only=True, data_only=True)
            for sh in wb.sheetnames:
                feed(sh)
                for row in wb[sh].iter_rows(values_only=True):
                    for v in row: feed(v)
        elif low.endswith('.xls'):
            import xlrd
            b=xlrd.open_workbook(p, on_demand=True)
            for n in b.sheet_names():
                feed(n); sh=b.sheet_by_name(n)
                for r in range(sh.nrows):
                    for v in sh.row_values(r): feed(v)
        else:
            for enc in ('utf-8-sig','utf-8','cp1251','latin-1'):
                try:
                    with open(p, newline='', encoding=enc) as f:
                        smp=f.read(8192); f.seek(0)
                        d=';' if smp.count(';')>=smp.count(',') else ','
                        for row in csv.reader(f, delimiter=d):
                            for v in row: feed(v)
                    break
                except (UnicodeDecodeError, csv.Error): continue
    except Exception: pass

files=[]
for e in ('xlsx','xls','csv'):
    files += glob.glob(f'/Users/pt/Projects.AI/BIRO26/**/*.{e}', recursive=True)
files=[f for f in files if not os.path.basename(f).startswith('~$')]
for f in files: scan_file(f)
print(f"corpus after files: {len(WORDS)} words", file=sys.stderr)

# ---- corpus din BD / from the DB ----
oracledb.init_oracle_client(lib_dir='/Users/pt/Downloads/instantclient_23_26')
con=oracledb.connect(user='officeplus',password='officeplus26',dsn='orange.una.md:4024/cloudbd.world')
cur=con.cursor(); cur.arraysize=5000
for sql in ["SELECT denumirea FROM tms_univers WHERE denumirea NOT LIKE '%?%'",
            "SELECT denumire FROM biro26_goods WHERE denumire NOT LIKE '%?%'",
            "SELECT categorie FROM biro26_goods WHERE categorie NOT LIKE '%?%'"]:
    cur.execute(sql)
    for (v,) in cur: feed(v)
print(f"corpus total: {len(WORDS)} words", file=sys.stderr)

by_len=defaultdict(list)
for w in WORDS: by_len[len(w)].append(w)

CACHE={}
def has_inner_q(w):
    # RO: '?' considerat stricat DOAR daca are litera/cifra de AMBELE parti
    # EN: '?' is mangled ONLY when flanked by alphanumerics on BOTH sides
    return any(w[i]=='?' and i>0 and i<len(w)-1 and w[i-1].isalnum() and w[i+1].isalnum()
               for i in range(len(w)))

def fix_word(w):
    if '?' not in w or not has_inner_q(w): return None
    if w in CACHE: return CACHE[w]
    rx=re.compile('^'+''.join('.' if c=='?' else re.escape(c) for c in w)+'$')
    cands=[(WORDS[c],c) for c in by_len.get(len(w),()) if rx.match(c)]
    res=None
    if cands:
        cands.sort(reverse=True)
        if len(cands)==1 or cands[0][0] >= MIN_RATIO*cands[1][0]:
            res=cands[0][1]
    CACHE[w]=res
    return res

def fix_text(s):
    if not s or '?' not in s: return None
    out=[]; pos=0; changed=False
    for m in WORD_RX.finditer(s):
        if '?' not in m.group() or not has_inner_q(m.group()): continue
        r=fix_word(m.group())
        if r:
            out.append(s[pos:m.start()]); out.append(r); pos=m.end(); changed=True
    if not changed: return None
    out.append(s[pos:])
    r=''.join(out)
    return r if r!=s else None

TARGETS=[('TMS_UNIVERS','DENUMIREA'),('BIRO26_GOODS','DENUMIRE'),('BIRO26_GOODS','CATEGORIE'),
         ('TMS_MPT_BARCODE','COMENT'),('BIRO26PT_STG','DENUMIRE'),('BIRO26PT_STG','CATEG')]
LOG=[]
print(f"\n{'TABLE.COLUMN':30} {'distinct?':>10} {'reparate':>9} {'partial':>8} {'rinduri':>9}")
print('-'*70)
for tbl,col in TARGETS:
    cur.execute(f"SELECT {col}, COUNT(*) FROM {tbl} WHERE {col} LIKE '%?%' GROUP BY {col}")
    rows=cur.fetchall(); mp=[]; nr=0; partial=0
    for val,cnt in rows:
        r=fix_text(val)
        if r and r!=val:
            mp.append((r,val)); nr+=cnt
            if '?' in r: partial+=1
    print(f"{tbl+'.'+col:30} {len(rows):>10} {len(mp):>9} {partial:>8} {nr:>9}")
    LOG.extend([{'t':tbl,'c':col,'old':o,'new':n} for n,o in mp])
    if GO and mp:
        cur.executemany(f"UPDATE {tbl} SET {col} = :1 WHERE {col} = :2", mp)
        con.commit()
print('-'*70)
if GO:
    print("\n=== ramase cu '?' ===")
    for tbl,col in TARGETS:
        cur.execute(f"SELECT COUNT(*) FROM {tbl} WHERE {col} LIKE '%?%'")
        print(f"  {tbl}.{col}: {cur.fetchone()[0]}")
    json.dump(LOG, open(SCRATCH+'/replacements_w3.json','w'), ensure_ascii=False)
    print(f"log: {len(LOG)}")
else:
    print("\nDRY-RUN — --go pentru aplicare.")
    for e in LOG[:10]:
        print(f"  {e['old'][:60]}\n  -> {e['new'][:60]}\n")
con.close()

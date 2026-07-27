#!/usr/bin/env python3
# RO: Incarcator generic de fisiere (xlsx/csv) in stagin brut BIRO26PT_RAW/HEADER/FILE.
#     Nu interpreteaza structura - doar celule + antet. Detectia coloanelor se face in DB
#     (pachetul BIRO26PT_importData).
# EN: Generic file loader (xlsx/csv) into the raw staging BIRO26PT_RAW/HEADER/FILE.
#     It does NOT interpret structure - only cells + header row. Column detection is done
#     in the DB (package BIRO26PT_importData).
#
# Usage: biro26pt_loader.py <folder_or_file> [--sheet-first-row-header]
import sys, os, glob, unicodedata, oracledb

LIB = '/Users/pt/Downloads/instantclient_23_26'
DSN = 'orange.una.md:4024/cloudbd.world'
USER, PWD = 'officeplus', 'officeplus26'
MAXCOL = 16  # c0..c15

# RO: Baza e CL8MSWIN1251 (chirilic) si NU are diacritice romanesti (ă â î ș ț) sau
#     semne tipografice (× — ‑ ² ½ …): Oracle le stocheaza ca '?'. Transliteram tot ce
#     nu incape in cp1251 INAINTE de scriere. Chirilica ramane neatinsa.
# EN: The DB is CL8MSWIN1251 (Cyrillic) and has NO Romanian diacritics (ă â î ș ț) or
#     typographic signs (× — ‑ ² ½ …): Oracle stores them as '?'. We transliterate
#     everything that does not fit cp1251 BEFORE writing. Cyrillic is untouched.
_TRANSLIT = {
    "ă": "a", "Ă": "A", "â": "a", "Â": "A", "î": "i", "Î": "I",
    "ș": "s", "Ș": "S", "ş": "s", "Ş": "S",
    "ț": "t", "Ț": "T", "ţ": "t", "Ţ": "T",
    "×": "x", "÷": ":", "−": "-", "‐": "-", "‑": "-", "‒": "-", "―": "-",
    "≤": "<=", "≥": ">=", "≈": "~", "≠": "!=", "′": "'", "″": '"', "ʼ": "'",
    "½": "1/2", "¼": "1/4", "¾": "3/4", "⁄": "/", "·": ".", "∙": ".",
    "ﬁ": "fi", "ﬂ": "fl", "œ": "oe", "Œ": "OE", "æ": "ae", "Æ": "AE",
    "ß": "ss", "ø": "o", "Ø": "O", "đ": "d", "Đ": "D", "ł": "l", "Ł": "L",
    "​": "", "‌": "", "‍": "", "﻿": "", "­": "",
    " ": " ", " ": " ", " ": " ", " ": " ",
}


def cp1251_safe(s: str) -> str:
    """RO: text care incape garantat in CL8MSWIN1251 / EN: guaranteed cp1251-safe text."""
    out = []
    for ch in s:
        repl = _TRANSLIT.get(ch)
        if repl is not None:
            out.append(repl)
            continue
        try:
            ch.encode("cp1251")
            out.append(ch)
            continue
        except UnicodeEncodeError:
            pass
        # RO: descompunere Unicode: litera de baza fara semne / EN: strip combining marks
        dec = unicodedata.normalize("NFKD", ch)
        dec = "".join(c for c in dec if not unicodedata.combining(c))
        dec = "".join(_TRANSLIT.get(c, c) for c in dec)
        try:
            dec.encode("cp1251")
            out.append(dec)
        except UnicodeEncodeError:
            out.append("")  # RO: nerecuperabil (emoji etc.) / EN: unrecoverable
    return "".join(out)


def cells(row):
    out = []
    for v in row:
        if v is None:
            out.append(None)
        else:
            s = cp1251_safe(str(v).strip())
            out.append(s[:1000] if s != '' else None)
    return out


def read_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        rows = [cells(r) for r in ws.iter_rows(values_only=True)]
        # RO: sari peste randuri complet goale la coada / EN: skip fully-empty trailing rows
        rows = [r for r in rows if any(x is not None for x in r)]
        if rows:
            yield cp1251_safe(sh), rows

def read_csv(path):
    import csv
    with open(path, newline='', encoding='utf-8-sig') as f:
        # RO: detecteaza separatorul ; sau , / EN: sniff ; or , delimiter
        sample = f.read(4096); f.seek(0)
        delim = ';' if sample.count(';') >= sample.count(',') else ','
        rows = [cells(r) for r in csv.reader(f, delimiter=delim)]
    rows = [r for r in rows if any(x is not None for x in r)]
    if rows:
        yield cp1251_safe(os.path.basename(path)), rows

def main():
    if len(sys.argv) < 2:
        print('usage: biro26pt_loader.py <folder_or_file>'); sys.exit(1)
    target = sys.argv[1]
    files = []
    if os.path.isdir(target):
        for ext in ('*.xlsx', '*.xls', '*.csv'):
            files += glob.glob(os.path.join(target, ext))
    else:
        files = [target]
    files = sorted(f for f in files if not os.path.basename(f).startswith('~'))
    if not files:
        print('no files found'); sys.exit(1)

    oracledb.init_oracle_client(lib_dir=LIB)
    con = oracledb.connect(user=USER, password=PWD, dsn=DSN); cur = con.cursor()
    cur.execute("SELECT NVL(MAX(load_id),0) FROM biro26pt_file")
    load_id = cur.fetchone()[0]

    for path in files:
        base = cp1251_safe(os.path.basename(path))[:200]
        reader = read_csv if path.lower().endswith('.csv') else read_xlsx
        for sheet, rows in reader(path):
            load_id += 1
            header = rows[0]
            n_cols = min(max(len(r) for r in rows), MAXCOL)
            # header row
            cur.executemany(
                "INSERT INTO biro26pt_header(load_id,src_file,col_idx,header_text) VALUES(:1,:2,:3,:4)",
                [(load_id, base, i, (header[i] if i < len(header) else None)) for i in range(n_cols)])
            # data rows (row_no starts at 1; header excluded)
            buf = []
            for rno, r in enumerate(rows[1:], start=1):
                vals = [load_id, base, sheet[:120], rno] + [(r[i] if i < len(r) else None) for i in range(MAXCOL)]
                buf.append(tuple(vals))
            ph = ",".join(":%d" % i for i in range(1, 4 + MAXCOL + 1))
            cur.executemany(
                "INSERT INTO biro26pt_raw(load_id,src_file,sheet,row_no," +
                ",".join("c%d" % i for i in range(MAXCOL)) + ") VALUES(" + ph + ")", buf)
            cur.execute(
                "INSERT INTO biro26pt_file(load_id,src_file,sheet,n_rows,n_cols) VALUES(:1,:2,:3,:4,:5)",
                (load_id, base, sheet[:120], len(rows) - 1, n_cols))
            con.commit()
            print(f"load_id={load_id} file='{base}' sheet='{sheet}' rows={len(rows)-1} cols={n_cols}")
    con.close()

if __name__ == '__main__':
    main()
